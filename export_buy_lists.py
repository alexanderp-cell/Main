#!/usr/bin/env python3
"""
Рекомендации к выкупу со склада АТИ:
- внутренний файл: сервис (только rotables), расходка, ансервис, неизвестное
- клиентские: ликвидный сервис (rotables), ликвидный ансервис, ликвидная расходка
  (формат склада АТИ, одна строка на P/N)
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import analyze_liquidity as al

OUT_DIR = Path("/workspace/output")
ART_DIR = Path("/opt/cursor/artifacts")

INTERNAL_NAME = "ATI_buyout_recommendations_internal.xlsx"
CLIENT_SERVICEABLE = "ATI_buyout_client_serviceable.xlsx"
CLIENT_UNSERVICEABLE = "ATI_buyout_client_unserviceable.xlsx"
CLIENT_EXPENDABLES = "ATI_buyout_client_expendables.xlsx"
# старый файл unknown больше не генерируем
LEGACY_CLIENT_UNKNOWN = "ATI_buyout_client_unknown.xlsx"


def is_buy_candidate(r: dict) -> bool:
    """Позиции для внутреннего списка рекомендаций."""
    g = r["liquidity_grade"]
    if g in {"A", "B"}:
        return True
    if g == "C":
        demand = (r.get("taz_orders") or 0) + (r.get("requests") or 0)
        rev = r.get("potential_revenue_usd") or 0
        return demand >= 2 or rev >= 3000
    return False


def is_liquid(r: dict) -> bool:
    return r["liquidity_grade"] in {"A", "B"}


def sort_key(r: dict):
    """Сначала ликвидность (A→B→C), затем потенц. выручка."""
    return (
        al.GRADE_ORDER[r["liquidity_grade"]],
        -(r.get("potential_revenue_usd") or 0),
        -r["liquidity_score"],
        -r["qty"],
        r["partno"],
    )


def is_expendable_part(row: dict, market: al.MarketAgg) -> bool:
    """
    Расходка (expendables): в основном в таблицах EXP, редко в ТУЗ,
    обычно NEW, редко по 1 шт, обычно недорогие.
    Дорогие позиции даже при EXP-рынке оставляем как rotables.
    """
    n_tuz = len(market.request_keys_tuz)
    n_exp = len(market.request_keys_exp)
    price = row.get("price_ref_usd")
    if price is None:
        price = 0.0
    cond = (row.get("condition") or "").upper()
    qty = row.get("qty") or 0

    # Дорогие — скорее rotables (даже если встречаются в EXP-рынке)
    if price >= 2000:
        return False

    # Чистый EXP-рынок без ТУЗ
    if n_exp > 0 and n_tuz == 0:
        return True

    # EXP доминирует и позиция недорогая
    if n_exp >= 2 and n_exp > n_tuz and price > 0 and price < 1000:
        return True

    # Сильное доминирование EXP
    if n_exp >= 3 and n_exp >= 3 * max(n_tuz, 1) and price < 1500:
        return True

    # Нет рыночных запросов, но по складу похоже на расходку
    if n_exp == 0 and n_tuz == 0:
        if "NEW" in cond and price > 0 and price < 300 and qty >= 2:
            return True

    return False


def load_all_market_and_score():
    taz_files = [
        al.DATA / "TAZ_17.07.2026.xlsx",
        al.DATA / "TA3 2025.xlsx",
        al.DATA / "TA3-Архив 2024-01-26.xlsx",
    ]
    tuz_files = [
        al.DATA / "TUZ_17.07.2026.xlsx",
        al.DATA / "ТУЗ 2025 Jan - June.xlsx",
        al.DATA / "ТУЗ 6_19.xlsx",
    ]

    print("Loading market + ATI for buyout lists...")
    taz: list[al.Event] = []
    for path in taz_files:
        if path.exists():
            part = al.load_taz(path)
            print(f"  TAZ {path.name}: {len(part)}")
            taz.extend(part)

    tuz: list[al.Event] = []
    for path in tuz_files:
        if path.exists():
            part = al.load_tuz(path)
            print(f"  TUZ {path.name}: {len(part)}")
            tuz.extend(part)

    exp: list[al.Event] = []
    exp_xlsx = al.DATA / "EXPENDABLES.xlsx"
    if exp_xlsx.exists():
        part = al.load_exp(exp_xlsx)
        print(f"  EXP {exp_xlsx.name}: {len(part)}")
        exp.extend(part)
    for path in sorted(al.DATA.glob("*.csv")):
        part = al.load_exp_csv(path)
        print(f"  EXP CSV {path.name}: {len(part)}")
        exp.extend(part)

    ati = al.load_ati(al.DATA / "ATI.xlsx")
    print(f"  ATI rows: {len(ati)}")

    def event_global_key(e: al.Event):
        if e.kind == "order":
            inv = al.normalize_invoice(e.request_no) or (
                f"NO_INV|{al.client_key(e.client)}|{al.day_key(e.date)}|{round(e.qty or 0, 4)}"
            )
            return ("O", e.pn, inv)
        return (
            "R",
            e.pn,
            al.client_key(e.client),
            al.day_key(e.date),
            round(e.qty or 0, 4),
            round(e.price or 0, 2),
            (e.description or "")[:40],
        )

    seen = set()
    all_events = []
    for e in taz + tuz + exp:
        k = event_global_key(e)
        if k in seen:
            continue
        seen.add(k)
        all_events.append(e)

    by_pn, soft_to_pns, alt_to_pns = al.build_market_index(all_events)
    stock = al.aggregate_ati_stock(ati)
    scored = []
    for item in stock:
        market = al.resolve_market(item["pn"], by_pn, soft_to_pns, alt_to_pns)
        row = al.build_row_from_stock(item, market)
        row["is_expendable"] = is_expendable_part(row, market)
        scored.append(row)
    return ati, scored


def write_internal(path: Path, scored_buy: list[dict], ati_n: int):
    svc = [r for r in scored_buy if r["section"] == 3 and not r.get("is_expendable")]
    exp = [r for r in scored_buy if r["section"] == 3 and r.get("is_expendable")]
    us = [r for r in scored_buy if r["section"] == 2]
    unk = [r for r in scored_buy if r["section"] == 1]

    sheets = [
        ("1. Сервис ликвидный", sorted([r for r in svc if is_liquid(r)], key=sort_key), "1F7A4D"),
        ("2. Сервис неликвидный", sorted([r for r in svc if not is_liquid(r)], key=sort_key), "2F5D9F"),
        ("3. Расходка ликвидная", sorted([r for r in exp if is_liquid(r)], key=sort_key), "0E7490"),
        ("4. Расходка неликвидная", sorted([r for r in exp if not is_liquid(r)], key=sort_key), "117A65"),
        ("5. Ансервис ликвидный", sorted([r for r in us if is_liquid(r)], key=sort_key), "B33B3B"),
        ("6. Ансервис неликвидный", sorted([r for r in us if not is_liquid(r)], key=sort_key), "C47F00"),
        ("7. Неизвестное", sorted(unk, key=sort_key), "6B4C9A"),
    ]

    wb = openpyxl.Workbook()
    ws0 = wb.active
    ws0.title = "0. Сводка"
    ws0["A1"] = "Рекомендации к выкупу — склад АТИ (внутренний)"
    ws0["A1"].font = Font(bold=True, size=16, name="Calibri")
    ws0["A2"] = (
        f"Дата: {datetime.now():%Y-%m-%d %H:%M} | "
        "Ликвидные = A/B; неликвидные = C (сильный C: спрос ≥2 или выручка ≥ $3 000)"
    )
    ws0.merge_cells("A2:F2")

    total_rev = sum(r.get("potential_revenue_usd") or 0 for r in scored_buy)
    liquid_n = sum(1 for r in scored_buy if is_liquid(r))
    rows_sum = [
        ("Строк в исходном АТИ", ati_n),
        ("Рекомендовано позиций (P/N×состояние)", len(scored_buy)),
        ("Ликвидные (A/B)", liquid_n),
        ("Неликвидные (C)", len(scored_buy) - liquid_n),
        ("Сервис ликвидный (rotables)", len(sheets[0][1])),
        ("Сервис неликвидный (rotables)", len(sheets[1][1])),
        ("Расходка ликвидная", len(sheets[2][1])),
        ("Расходка неликвидная", len(sheets[3][1])),
        ("Ансервис ликвидный", len(sheets[4][1])),
        ("Ансервис неликвидный", len(sheets[5][1])),
        ("Неизвестное состояние", len(sheets[6][1])),
        ("Суммарное кол-во Utair в рекомендациях, шт.", sum(r["qty"] for r in scored_buy)),
        ("Суммарная потенц. выручка (где есть цена), USD", round(total_rev, 2)),
    ]
    ws0["A4"] = "Сводка"
    ws0["A4"].font = Font(bold=True, size=13)
    for i, (k, v) in enumerate(rows_sum, 5):
        ws0.cell(i, 1, k)
        cell = ws0.cell(i, 2, v)
        cell.font = Font(bold=True)
        if "выручка" in k.lower() and isinstance(v, (int, float)):
            cell.number_format = '"$"#,##0.00'

    ws0["A20"] = "Методика"
    ws0["A20"].font = Font(bold=True, size=13)
    method = [
        "Ликвидный = оценка A или B (устойчивый рыночный спрос).",
        "Неликвидный = C и ниже в этом файле (практически сильный C; D не включаем).",
        "Сервис = только rotables (ремонтируемые/заменяемые узлы).",
        "Расходка = expendables: в основном рынок EXP (не ТУЗ), обычно NEW, недорогие, редко по 1 шт.",
        "Дорогие (≥ $2 000) даже при EXP-рынке остаются в Сервисе как rotables.",
        "Сортировка внутри листа: ликвидность (A→B→C) → потенц. выручка → балл → qty.",
        "Клиенту: ликвидный сервис (rotables), ликвидный ансервис, ликвидная расходка — 1 строка на P/N.",
    ]
    for i, t in enumerate(method, 21):
        ws0.cell(i, 1, t)
        ws0.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)
    ws0.column_dimensions["A"].width = 70
    ws0.column_dimensions["B"].width = 22

    for title, rows, color in sheets:
        al.write_rows(wb.create_sheet(title), rows, color)

    wsl = wb.create_sheet("Легенда")
    wsl["A1"] = "Ликвидность"
    wsl["A1"].font = Font(bold=True, size=13)
    for i, (g, name) in enumerate([("A", "Высокая"), ("B", "Средняя"), ("C", "Низкая")], 3):
        c = wsl.cell(i, 1, g)
        c.fill = al.GRADE_FILL[g]
        c.font = al.GRADE_FONT
        wsl.cell(i, 2, name)
    wsl["A7"] = (
        "Внутренний файл. Клиенту — ликвидный сервис (rotables), "
        "ликвидный ансервис и ликвидная расходка (агрегат по P/N)."
    )
    wsl.column_dimensions["A"].width = 12
    wsl.column_dimensions["B"].width = 50

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"Saved internal: {path}")
    return {title: len(rows) for title, rows, _ in sheets}


def aggregate_ati_by_pn(ati_rows: list[dict], keys: set[tuple]) -> list[dict]:
    """Одна строка на P/N внутри раздела состояния; qty = сумма; serialno пустой."""
    groups: dict[str, dict] = {}
    order: list[str] = []
    for r in ati_rows:
        sec = al.condition_section(r.get("condition") or "")
        key = (r["pn"], sec)
        if key not in keys:
            continue
        pn = r["pn"]
        if pn not in groups:
            groups[pn] = {
                "partno": r.get("partno") or pn,
                "serialno": "",
                "description": r.get("description") or "",
                "ata": r.get("ata") or "",
                "ac_typs": set(),
                "conditions": set(),
                "qty": 0.0,
                "pn": pn,
            }
            order.append(pn)
        g = groups[pn]
        q = r.get("qty") if r.get("qty") is not None else 1.0
        g["qty"] += float(q) if q else 1.0
        if r.get("description") and not g["description"]:
            g["description"] = r["description"]
        if r.get("ata") and not g["ata"]:
            g["ata"] = r["ata"]
        if r.get("ac_typ"):
            g["ac_typs"].add(str(r["ac_typ"]).strip())
        cond = (r.get("condition") or "").strip()
        if cond:
            g["conditions"].add(cond.upper())

    out = []
    for pn in order:
        g = groups[pn]
        out.append(
            {
                "partno": g["partno"],
                "serialno": "",
                "description": g["description"],
                "ata": g["ata"],
                "ac_typ": ", ".join(sorted(x for x in g["ac_typs"] if x)),
                "condition": ", ".join(sorted(g["conditions"])) if g["conditions"] else "",
                "qty": g["qty"],
                "pn": pn,
            }
        )
    return out


def write_client_ati_aggregated(
    path: Path,
    ati_rows: list[dict],
    scored_subset: list[dict],
) -> int:
    """Клиентский файл в формате склада АТИ, без дублей P/N."""
    keys = {(r["pn"], r["section"]) for r in scored_subset}
    # порядок как во внутреннем топе (ликвидность → выручка)
    rank = {r["pn"]: i for i, r in enumerate(sorted(scored_subset, key=sort_key))}
    aggregated = aggregate_ati_by_pn(ati_rows, keys)
    aggregated.sort(key=lambda x: rank.get(x["pn"], 10**9))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Лист2"
    headers = ["partno", "serialno", "description", "ata_chapter", "ac_typ", "condition", "qty"]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    thin = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, r in enumerate(aggregated):
        vals = [
            r["partno"],
            r["serialno"],
            r["description"],
            r["ata"],
            r["ac_typ"],
            r["condition"],
            r["qty"],
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(i + 2, col, val)
            cell.border = thin
            cell.alignment = Alignment(vertical="center")

    widths = [18, 16, 36, 10, 12, 12, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    n = len(aggregated)
    if n:
        ws.auto_filter.ref = f"A1:G{n + 1}"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"Saved client ({n} unique P/N): {path}")
    return n


def main():
    ati, scored = load_all_market_and_score()
    buy = [r for r in scored if is_buy_candidate(r)]
    liquid_svc = [
        r for r in buy if r["section"] == 3 and is_liquid(r) and not r.get("is_expendable")
    ]
    liquid_exp = [
        r for r in buy if r["section"] == 3 and is_liquid(r) and r.get("is_expendable")
    ]
    liquid_us = [r for r in buy if r["section"] == 2 and is_liquid(r)]
    print(
        f"Buy candidates: {len(buy)}; "
        f"liquid rotables={len(liquid_svc)}, liquid exp={len(liquid_exp)}, "
        f"liquid us={len(liquid_us)}; "
        f"exp total in buy={sum(1 for r in buy if r.get('is_expendable'))}"
    )

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    ART_DIR.mkdir(parents=True, exist_ok=True)

    internal = OUT_DIR / INTERNAL_NAME
    counts = write_internal(internal, buy, len(ati))
    openpyxl.load_workbook(internal).save(ART_DIR / INTERNAL_NAME)

    n1 = write_client_ati_aggregated(OUT_DIR / CLIENT_SERVICEABLE, ati, liquid_svc)
    openpyxl.load_workbook(OUT_DIR / CLIENT_SERVICEABLE).save(ART_DIR / CLIENT_SERVICEABLE)
    n2 = write_client_ati_aggregated(OUT_DIR / CLIENT_UNSERVICEABLE, ati, liquid_us)
    openpyxl.load_workbook(OUT_DIR / CLIENT_UNSERVICEABLE).save(ART_DIR / CLIENT_UNSERVICEABLE)
    n3 = write_client_ati_aggregated(OUT_DIR / CLIENT_EXPENDABLES, ati, liquid_exp)
    openpyxl.load_workbook(OUT_DIR / CLIENT_EXPENDABLES).save(ART_DIR / CLIENT_EXPENDABLES)

    for folder in (OUT_DIR, ART_DIR):
        legacy = folder / LEGACY_CLIENT_UNKNOWN
        if legacy.exists():
            legacy.unlink()
            print(f"Removed legacy: {legacy}")

    print("Done.", counts, f"client_svc={n1}", f"client_us={n2}", f"client_exp={n3}")


if __name__ == "__main__":
    main()
