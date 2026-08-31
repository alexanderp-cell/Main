#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Формирование файла ГОТОВЫЕ_ОСТАТКИ.xlsx.

Скрипт берёт актуальные складские остатки из ОСТАТКИ.xlsx (неструктурированный
формат с цветовой разметкой) и приводит их к виду файла-образца ОБРАЗЕЦ.xlsx.

Что делает скрипт:
  1. Оставляет только "зелёные" строки (заливка 00FF00) с чёрным текстом.
     Красные, жёлтые, фиолетовые и прочие не-зелёные строки, а также строки
     с не-чёрным шрифтом отбрасываются.
  2. Раскладывает данные по столбцам образца:
        Part Number | Alt. PN | DESCRIPTION | Quantity | Condition | AC Type
  3. Чистит партийные номера (основной и альтернативные): убирает служебные
     слова (STK, MOW, STOCK, MOSCOW, PN, P/N, ALT, QTY, PCS, EA, названия
     складов и т.п.), комментарии и любую кириллицу. Несколько альтернативных
     номеров записываются через запятую.
  4. Удаляет кириллицу из всех значений (в т.ч. переводит "В737" -> "B737").
  5. Сортирует позиции по типам самолётов в том же порядке, что и в образце,
     а внутри каждого типа — по основному Part Number. Позиции с неопределимым
     типом помещаются в конец.
  6. Полностью сохраняет оформление образца (заголовки, шрифты, заливки,
     границы, ширину столбцов, высоту строк, выравнивание, форматы чисел),
     используя сам образец как шаблон вывода.

Запуск:
    python3 generate_ready_stock.py

Исходные файлы ОБРАЗЕЦ.xlsx и ОСТАТКИ.xlsx НЕ изменяются.
"""

import datetime
import re
import sys
import warnings
from collections import Counter, OrderedDict
from copy import copy

import openpyxl

# Безобидное предупреждение openpyxl о неподдерживаемом расширении Data Validation.
warnings.filterwarnings("ignore", message="Data Validation extension is not supported")

SAMPLE_FILE = "ОБРАЗЕЦ.xlsx"      # эталон структуры и оформления
STOCK_FILE = "ОСТАТКИ.xlsx"       # актуальные остатки (источник данных)
OUTPUT_FILE = "ГОТОВЫЕ_ОСТАТКИ.xlsx"

# Порядок столбцов в ИТОГОВОМ файле (как в образце):
# 1 Part Number | 2 Alt PN | 3 DESCRIPTION | 4 Quantity | 5 Condition | 6 AC Type
COL_PN, COL_ALT, COL_DESC, COL_QTY, COL_COND, COL_AC = 1, 2, 3, 4, 5, 6
NCOLS = 6

# Как распознать столбцы во ВХОДНОМ файле остатков по тексту заголовка.
# Структура выгрузки может меняться (например, столбец состояния может
# отсутствовать, а тип ВС — сдвинуться), поэтому столбцы ищем по названию,
# а не по фиксированному номеру. Ключи совпадают с полями записи.
STOCK_HEADER_PATTERNS = {
    "pn":   [r"P/?N", r"PART\s*NUMBER", r"ПАРТ", r"ПАРТ\w*"],
    "alt":  [r"ALT"],
    "desc": [r"DESCR", r"НАИМЕН", r"ОПИСАН"],
    "qty":  [r"QTY", r"QUANT", r"КОЛ[-\s]?ВО", r"КОЛИЧ"],
    "cond": [r"СОСТОЯН", r"CONDITION", r"COND\b"],
    "ac":   [r"ТИП\s*ВС", r"AC\s*TYPE", r"ТИП"],
}

GREEN = "FF00FF00"  # цвет "хороших" (зелёных) строк

# Транслитерация кириллических букв-двойников в латиницу
# (нужна, например, для типа ВС "В737" -> "B737").
CYR_TO_LAT = {
    "А": "A", "В": "B", "С": "C", "Е": "E", "Н": "H", "К": "K", "М": "M",
    "О": "O", "Р": "P", "Т": "T", "Х": "X", "У": "Y", "І": "I",
    "а": "a", "в": "b", "с": "c", "е": "e", "н": "h", "к": "k", "м": "m",
    "о": "o", "р": "p", "т": "t", "х": "x", "у": "y",
}
CYRILLIC_RE = re.compile(r"[А-Яа-яЁё]")

# Служебные слова/обозначения, которые не являются частью партийного номера.
# Список сознательно консервативен: сюда входят слова из ТЗ и реально
# встречающиеся в данных обозначения складов. Короткие двойники, которые
# часто являются частью настоящих партномеров (например, "PC" в "PC-1067"),
# намеренно НЕ включены.
NOISE_WORDS = [
    "PART NUMBER", "P/N", "PN", "ALT PN", "ALT", "QTY", "PCS", "EA",
    "STOCK", "STK", "MOSCOW", "MOW", "SIRIUS", "SIR", "VKO", "FAI", "FEI",
]
# Регэксп для служебных слов как отдельных токенов (без учёта регистра).
NOISE_RE = re.compile(
    r"(?<![A-Za-z0-9])(?:" + "|".join(re.escape(w) for w in NOISE_WORDS) + r")(?![A-Za-z0-9])",
    re.IGNORECASE,
)

# Значения типа ВС, которые считаем "неопределимыми" (уходят в конец таблицы).
UNKNOWN_AC = {"", "#N/A", "N/A", "0", "00-01-1900"}


def transliterate(text):
    """Переводит кириллические буквы-двойники в латиницу."""
    return "".join(CYR_TO_LAT.get(ch, ch) for ch in text)


def strip_cyrillic(text):
    """Удаляет из строки все кириллические символы."""
    return CYRILLIC_RE.sub("", text)


def cell_str(value):
    """Приводит значение ячейки к строке без крайних пробелов.

    Даты форматируются как ГГГГ-ММ-ДД: Excel часто превращает партийные номера
    вида "8296-01-01" в даты, и их нужно вернуть в исходный текстовый вид
    (а не "8296-01-01 00:00:00").
    """
    if value is None:
        return ""
    if isinstance(value, datetime.date):
        text = "%04d-%02d-%02d" % (value.year, value.month, value.day)
        if isinstance(value, datetime.datetime) and (value.hour or value.minute or value.second):
            text += " %02d:%02d:%02d" % (value.hour, value.minute, value.second)
        return text
    return str(value).strip()


def is_green(cell):
    """True, если заливка ячейки — зелёная (00FF00)."""
    fill = cell.fill
    if fill is None or fill.patternType is None:
        return False
    fg = fill.fgColor
    return fg is not None and fg.type == "rgb" and str(fg.rgb) == GREEN


def is_black_font(cell):
    """True, если шрифт ячейки чёрный (или цвет не задан = по умолчанию чёрный)."""
    color = cell.font.color
    if color is None:
        return True
    if color.type == "theme" and color.theme == 1:  # theme 1 = Text 1 (чёрный)
        return True
    if color.type == "rgb" and str(color.rgb) in ("FF000000", "00000000"):
        return True
    return False


def clean_token(token):
    """Чистит один партийный номер: убирает служебные слова и мусор."""
    token = NOISE_RE.sub(" ", token)
    # Убираем повторяющиеся пробелы и служебные разделители по краям.
    token = re.sub(r"\s+", " ", token).strip()
    token = token.strip(" .-/;,")
    return token.strip()


def is_valid_pn(token):
    """Настоящий партийный номер содержит хотя бы одну цифру и одну букву/цифру."""
    if not token:
        return False
    return bool(re.search(r"\d", token))


def clean_main_pn(raw):
    """Чистит основной Part Number (одно значение)."""
    text = cell_str(raw)
    if not text:
        return ""
    text = transliterate(text)
    text = strip_cyrillic(text)
    text = text.replace("\n", " ")
    text = clean_token(text)
    return text


def clean_alt_pn(raw, main_pn=""):
    """Чистит альтернативные Part Numbers (может быть несколько -> через запятую).

    Токены, повторяющие основной Part Number, не являются альтернативами
    и отбрасываются.
    """
    text = cell_str(raw)
    if not text:
        return ""
    # Ячейки с кириллицей в столбце Alt PN — это комментарии, а не номера.
    if CYRILLIC_RE.search(text):
        return ""
    main_key = main_pn.strip().upper()
    parts = re.split(r"[\n,;]+", text)
    result = []
    for part in parts:
        token = clean_token(part)
        if not is_valid_pn(token):
            continue
        if token.upper() == main_key:  # повтор основного номера — не альтернатива
            continue
        if token not in result:
            result.append(token)
    return ", ".join(result)


def clean_text_value(raw):
    """Чистит обычное текстовое значение (описание, состояние): без кириллицы."""
    text = cell_str(raw)
    if not text:
        return ""
    text = strip_cyrillic(text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def clean_quantity(raw):
    """Возвращает количество как int, если возможно, иначе очищенную строку."""
    text = cell_str(raw)
    if not text:
        return ""
    # \xa0 — неразрывный пробел, часто используется как разделитель тысяч.
    text = strip_cyrillic(text).replace("\xa0", " ").strip()
    # Целое число, возможно с пробелами-разделителями тысяч ("1 500" -> 1500).
    compact = text.replace(" ", "")
    if re.fullmatch(r"-?\d+", compact):
        return int(compact)
    if re.fullmatch(r"-?\d+[.,]\d+", compact):
        return float(compact.replace(",", "."))
    return text


def normalize_ac(raw):
    """Нормализует тип ВС: транслитерация, удаление кириллицы, схлопывание пробелов.

    Заведомо мусорные значения ("0", "00-01-1900", "#N/A" и т.п.) не являются
    типами ВС и приводятся к пустой строке (позиция уйдёт в конец таблицы).
    """
    text = cell_str(raw)
    if not text:
        return ""
    text = transliterate(text)
    text = strip_cyrillic(text)
    text = re.sub(r"\s+", " ", text).strip()
    if text.upper() in UNKNOWN_AC:
        return ""
    return text


def canon_condition(cond):
    """Каноническое состояние для группировки/вывода.

    NE, NEW и New обозначают одно и то же ("новое") -> приводим к "NEW".
    Остальные состояния сохраняются (в верхнем регистре).
    """
    u = (cond or "").strip().upper()
    if u in ("NE", "NEW"):
        return "NEW"
    return u


def _uniq_join(values):
    """Объединяет непустые значения в порядке появления, без повторов, через запятую."""
    out = []
    for v in values:
        v = (v or "").strip()
        if v and v not in out:
            out.append(v)
    return ", ".join(out)


def consolidate_records(records):
    """Объединяет строки с одинаковыми Part Number И состоянием в одну.

    Ключ группировки — пара (основной Part Number, каноническое состояние).
    Позиции с одинаковым партномером, но разными состояниями (например NEW и OH)
    остаются отдельными строками. NE и NEW считаются одним состоянием.

    В рамках группы:
    - Quantity суммируется;
    - Alt PN / AC Type объединяются как уникальные значения через запятую;
    - DESCRIPTION берётся самое частое непустое (при равенстве — первое встреченное);
    - Condition приводится к каноническому виду (NE/New -> NEW).
    Строки без Part Number не объединяются и сохраняются как есть (состояние
    тоже приводится к каноническому виду).
    """
    groups = OrderedDict()
    no_key = []
    for rec in records:
        pn_key = rec["pn"].strip().upper()
        if not pn_key:
            rec = dict(rec, cond=canon_condition(rec["cond"]))
            no_key.append(rec)
            continue
        groups.setdefault((pn_key, canon_condition(rec["cond"])), []).append(rec)

    result = []
    for (pn_key, cond_key), recs in groups.items():
        # Количество: сумма числовых значений.
        numeric = [r["qty"] for r in recs if isinstance(r["qty"], (int, float))]
        total = sum(numeric)
        if numeric and all(isinstance(q, int) for q in numeric):
            total = int(total)
        # Нечисловые значения (например "3.6 FT" — метраж) сохраняем текстом.
        non_numeric = [str(r["qty"]).strip() for r in recs
                       if not isinstance(r["qty"], (int, float)) and str(r["qty"]).strip()]
        if numeric and non_numeric:
            qty = _uniq_join([str(total)] + non_numeric)
        elif numeric:
            qty = total
        else:
            qty = _uniq_join(non_numeric)

        # Alt PN: все уникальные альтернативные номера группы (без основного).
        alt_tokens = []
        for r in recs:
            for tok in re.split(r"[\n,;]+", str(r["alt"])):
                tok = tok.strip()
                if tok and tok.upper() != pn_key and tok not in alt_tokens:
                    alt_tokens.append(tok)
        alt = ", ".join(alt_tokens)

        # Описание: самое частое непустое значение.
        descs = [r["desc"].strip() for r in recs if r["desc"].strip()]
        desc = Counter(descs).most_common(1)[0][0] if descs else ""

        result.append({
            "pn": recs[0]["pn"],
            "alt": alt,
            "desc": desc,
            "qty": qty,
            "cond": cond_key,
            "ac": _uniq_join(r["ac"] for r in recs),
        })

    result.extend(no_key)
    return result


def detect_stock_columns(stock_ws):
    """Определяет номера столбцов входного файла остатков по их заголовкам.

    Возвращает словарь {pn, alt, desc, qty, cond, ac -> номер столбца или None}.
    Порядок важен: столбец "Alt PN" не должен перехватить столбец "PN",
    поэтому длинные/специфичные шаблоны проверяются первыми, а уже занятые
    столбцы повторно не используются.
    """
    headers = {}
    for c in range(1, stock_ws.max_column + 1):
        val = stock_ws.cell(1, c).value
        if val not in (None, ""):
            headers[c] = str(val).strip().upper()

    mapping = {key: None for key in STOCK_HEADER_PATTERNS}
    used = set()
    # Сначала более специфичные поля (alt, cond, ac, qty, desc), затем pn —
    # чтобы "ALT PN" не был ошибочно принят за "P/N".
    for key in ("alt", "cond", "ac", "qty", "desc", "pn"):
        for pattern in STOCK_HEADER_PATTERNS[key]:
            rx = re.compile(pattern)
            for col, text in headers.items():
                if col in used:
                    continue
                if rx.search(text):
                    mapping[key] = col
                    used.add(col)
                    break
            if mapping[key] is not None:
                break
    return mapping


def build_sample_order(sample_ws):
    """Строит порядок типов ВС по первому появлению в образце (нормализованный ключ)."""
    order = {}
    for row in range(2, sample_ws.max_row + 1):
        key = normalize_ac(sample_ws.cell(row, COL_AC).value).upper()
        if key and key not in order:
            order[key] = len(order)
    return order


def main():
    # --- Загружаем образец (он же станет шаблоном оформления вывода) ---
    out_wb = openpyxl.load_workbook(SAMPLE_FILE)
    out_ws = out_wb.active

    sample_order = build_sample_order(out_ws)

    # Сохраняем стиль строки-данных образца (строка 2) для каждого столбца,
    # чтобы применить его ко всем новым строкам.
    template_style = {}
    for c in range(1, NCOLS + 1):
        src = out_ws.cell(2, c)
        template_style[c] = {
            "font": copy(src.font),
            "fill": copy(src.fill),
            "border": copy(src.border),
            "alignment": copy(src.alignment),
            "number_format": src.number_format,
        }
    data_row_height = out_ws.row_dimensions[2].height or out_ws.sheet_format.defaultRowHeight

    # --- Читаем остатки ---
    stock_wb = openpyxl.load_workbook(STOCK_FILE)
    stock_ws = stock_wb.active

    # Определяем столбцы входного файла по заголовкам (структура может меняться).
    cols = detect_stock_columns(stock_ws)
    if cols["pn"] is None:
        raise SystemExit("Не найден столбец с Part Number в файле %s" % STOCK_FILE)
    missing = [k for k in ("alt", "desc", "qty", "cond", "ac") if cols[k] is None]
    if missing:
        names = {"alt": "Alt PN", "desc": "DESCRIPTION", "qty": "Quantity",
                 "cond": "Condition", "ac": "AC Type"}
        print("  ВНИМАНИЕ: во входном файле нет столбцов: %s "
              "(в результате они будут пустыми)" % ", ".join(names[k] for k in missing))

    def stock_val(row, key):
        col = cols[key]
        return stock_ws.cell(row, col).value if col else None

    max_stock_col = max(c for c in cols.values() if c)

    records = []
    stats = {"total": 0, "kept": 0, "dropped_color": 0, "dropped_font": 0}
    for row in range(2, stock_ws.max_row + 1):
        pn_cell = stock_ws.cell(row, cols["pn"])
        # Пропускаем полностью пустые строки.
        if all(stock_ws.cell(row, c).value in (None, "") for c in range(1, max_stock_col + 1)):
            continue
        stats["total"] += 1
        if not is_green(pn_cell):
            stats["dropped_color"] += 1
            continue
        if not is_black_font(pn_cell):
            stats["dropped_font"] += 1
            continue

        pn = clean_main_pn(pn_cell.value)
        alt = clean_alt_pn(stock_val(row, "alt"), pn)
        desc = clean_text_value(stock_val(row, "desc"))
        qty = clean_quantity(stock_val(row, "qty"))
        cond = clean_text_value(stock_val(row, "cond"))
        ac = normalize_ac(stock_val(row, "ac"))

        records.append({"pn": pn, "alt": alt, "desc": desc, "qty": qty,
                        "cond": cond, "ac": ac})
        stats["kept"] += 1

    # --- Объединение строк с одинаковым Part Number ---
    rows_before_merge = len(records)
    records = consolidate_records(records)
    rows_after_merge = len(records)

    # --- Сортировка ---
    def sort_key(rec):
        # У объединённых позиций тип ВС может содержать несколько значений —
        # для сортировки используем первый (основной) тип.
        key = rec["ac"].split(",")[0].strip().upper()
        if key in UNKNOWN_AC or not key:
            group = (2, "")               # неопределимые — в самый конец
        elif key in sample_order:
            group = (0, sample_order[key])  # порядок как в образце
        else:
            group = (1, key)              # определимые, но не из образца — после образца
        return (group[0], group[1], rec["pn"].upper(), rec["cond"].upper())

    records.sort(key=sort_key)

    # --- Записываем данные в шаблон ---
    # Удаляем старые строки-данные образца, оставляя заголовок (строка 1).
    if out_ws.max_row >= 2:
        out_ws.delete_rows(2, out_ws.max_row - 1)

    def write_cell(row, col, value):
        cell = out_ws.cell(row, col, value)
        st = template_style[col]
        cell.font = copy(st["font"])
        cell.fill = copy(st["fill"])
        cell.border = copy(st["border"])
        cell.alignment = copy(st["alignment"])
        cell.number_format = "General" if col == COL_QTY else "@"
        return cell

    for i, rec in enumerate(records):
        r = i + 2
        write_cell(r, COL_PN, rec["pn"])
        write_cell(r, COL_ALT, rec["alt"])
        write_cell(r, COL_DESC, rec["desc"])
        write_cell(r, COL_QTY, rec["qty"] if rec["qty"] != "" else None)
        write_cell(r, COL_COND, rec["cond"])
        write_cell(r, COL_AC, rec["ac"])
        out_ws.row_dimensions[r].height = data_row_height

    out_wb.save(OUTPUT_FILE)

    print("Готово: %s" % OUTPUT_FILE)
    print("  строк в источнике (непустых):      %d" % stats["total"])
    print("  отброшено по цвету заливки:        %d" % stats["dropped_color"])
    print("  отброшено по цвету шрифта:         %d" % stats["dropped_font"])
    print("  строк после фильтрации:            %d" % rows_before_merge)
    print("  объединено дубликатов PN:          %d" % (rows_before_merge - rows_after_merge))
    print("  записано в результат (строк):       %d" % rows_after_merge)


if __name__ == "__main__":
    sys.exit(main())
