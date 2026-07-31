#!/usr/bin/env python3
"""
Сверка склада поставщика (GEM LLP + Available Units) с рынком ТУЗ / ТАЗ.

Источники:
  - data/supplier_LLP.xlsx          — GEM LLP Inventory Report (серийный)
  - data/supplier_Available_Units.xlsx — агрегированные Available Units (PN/QTY)
  - data/TAZ_27.07.2026.xlsx         — заказы (ORDERS)
  - data/TUZ_31.07.2026.xlsx         — запросы клиентов

Результат: Excel с матчами по P/N (+ ALT P/N), счётчиками заказов/запросов и ценами.
"""
from __future__ import annotations

import re
import warnings
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from statistics import median
from typing import Any, Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

DATA = Path("/workspace/data")
OUT = Path("/workspace/output/Supplier_vs_TAZ_TUZ_sverka.xlsx")
OUT_ART = Path("/opt/cursor/artifacts/Supplier_vs_TAZ_TUZ_sverka.xlsx")

SAMPLE_PN = {
    "SAMPLE", "ОБРАЗЕЦ", "DEFAULT", "P/N", "PN", "PART NUMBER", "PARTNUMBER",
    "-", "—", "N/A", "NA", "NONE", "NULL", "TEST", "NSN",
}
SAMPLE_CLIENT = {"DEFAULT", "SAMPLE", "ОБРАЗЕЦ", "TBA", "TEST", "CLIENT"}
INTERNAL_CLIENT_MARKERS = (
    "закупка на склад", "на склад", "warehouse", "stock purchase", "internal",
)

AS_OF = date(2026, 7, 31)


# ---------------------------------------------------------------------------
# normalization helpers
# ---------------------------------------------------------------------------

def norm_pn(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value != value:
            return ""
        if value == int(value) and abs(value) < 1e15:
            s = str(int(value))
        else:
            s = str(value).strip().upper()
    elif isinstance(value, int):
        s = str(value)
    else:
        s = str(value).strip().upper()
    if not s:
        return ""
    s = (
        s.replace("\u2212", "-").replace("\u2013", "-").replace("\u2014", "-")
        .replace("\u00a0", "").replace("\u200b", "")
        .replace("\t", "").replace("\n", "").replace("\r", "")
    )
    s = re.sub(r"\s+", "", s)
    if re.fullmatch(r"\d+\.0+", s):
        s = s.split(".", 1)[0]
    return s.strip(".,;:|/\\")


def soft_pn_key(pn: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", pn)


def is_sample_pn(pn: str) -> bool:
    if not pn or len(soft_pn_key(pn)) < 3:
        return True
    return pn.upper() in SAMPLE_PN


def norm_client(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def is_market_client(client: str) -> bool:
    if not client:
        return False
    c = client.strip().lower()
    if client.upper() in SAMPLE_CLIENT or c in {x.lower() for x in SAMPLE_CLIENT}:
        return False
    return not any(m in c for m in INTERNAL_CLIENT_MARKERS)


def is_plausible_client(client: str, pn: str = "") -> bool:
    if not client or not is_market_client(client):
        return False
    c = client.strip()
    if re.fullmatch(r"[A-Za-zА-Яа-я]{1,4}\d{0,2}", c):
        return True
    if re.fullmatch(r"\d+(\.0+)?", c.replace(" ", "").replace("\u00a0", "")):
        return False
    if pn and soft_pn_key(c.upper()) == soft_pn_key(pn):
        return False
    return len(c) >= 2


def parse_money(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if value != value:
            return None
        return float(value) if value != 0 else None
    s = str(value).replace("\xa0", "").replace(" ", "").strip()
    if not s or s.upper() in {"TBA", "N/A", "NA", "-", "—"}:
        return None
    # min/max вроде "1000/1500" или "1000-1500" → берём min
    for sep in ("/", "-", "–"):
        if sep in s and re.fullmatch(r"[\d.,]+" + re.escape(sep) + r"[\d.,]+", s):
            parts = s.split(sep)
            vals = []
            for p in parts:
                try:
                    vals.append(float(p.replace(",", ".")))
                except ValueError:
                    pass
            return min(vals) if vals else None
    s = s.replace(",", ".")
    s = re.sub(r"[^\d.\-]", "", s)
    try:
        v = float(s)
        return v if v != 0 else None
    except ValueError:
        return None


def parse_qty(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value) if value == value else 0.0
    s = str(value).replace("\xa0", "").replace(" ", "").replace(",", ".").strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


def day_key(value: Any) -> str:
    d = parse_date(value)
    return d.isoformat() if d else ""


def parse_date(value: Any) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s[:19], fmt).date()
        except ValueError:
            continue
    return None


def normalize_invoice(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    s = str(value).strip()
    if re.fullmatch(r"\d+\.0+", s):
        s = s.split(".", 1)[0]
    return s


def clean_header_cell(h: Any) -> str:
    if h is None:
        return ""
    return re.sub(r"\s+", " ", str(h).replace("\n", " ").replace("\xa0", " ")).strip().lower()


def client_key(client: str) -> str:
    return re.sub(r"\s+", " ", client.strip().lower())


# ---------------------------------------------------------------------------
# market events
# ---------------------------------------------------------------------------

@dataclass
class Event:
    source: str
    kind: str
    pn: str
    alt: str = ""
    client: str = ""
    qty: float = 0.0
    price: Optional[float] = None
    condition: str = ""
    date: Any = None
    request_no: str = ""
    sheet: str = ""
    description: str = ""
    ac_type: str = ""


@dataclass
class MarketAgg:
    order_clients: set = field(default_factory=set)
    request_clients: set = field(default_factory=set)
    order_invoices: set = field(default_factory=set)
    request_keys: set = field(default_factory=set)
    order_prices: list = field(default_factory=list)
    request_prices: list = field(default_factory=list)
    order_qtys: list = field(default_factory=list)
    request_qtys: list = field(default_factory=list)
    last_order: Optional[date] = None
    last_request: Optional[date] = None
    sample_desc: str = ""
    matched_via: set = field(default_factory=set)

    @property
    def n_orders(self) -> int:
        return len(self.order_invoices)

    @property
    def n_requests(self) -> int:
        return len(self.request_keys)

    def merge(self, e: Event, via: str):
        self.matched_via.add(via)
        if e.description and not self.sample_desc:
            self.sample_desc = e.description
        d = parse_date(e.date)
        if e.kind == "order":
            if e.request_no:
                self.order_invoices.add(e.request_no)
            else:
                self.order_invoices.add(f"NOINV|{client_key(e.client)}|{day_key(e.date)}")
            if e.client and is_market_client(e.client):
                self.order_clients.add(client_key(e.client))
            if e.price:
                self.order_prices.append(e.price)
            if e.qty:
                self.order_qtys.append(e.qty)
            if d and (self.last_order is None or d > self.last_order):
                self.last_order = d
        else:
            rk = (client_key(e.client), day_key(e.date), e.request_no or "")
            self.request_keys.add(rk)
            if e.client and is_market_client(e.client):
                self.request_clients.add(client_key(e.client))
            if e.price:
                self.request_prices.append(e.price)
            if e.qty:
                self.request_qtys.append(e.qty)
            if d and (self.last_request is None or d > self.last_request):
                self.last_request = d


def _header_map_from_row(row: tuple, aliases: dict[str, list[str]]) -> dict[str, int]:
    mapping = {}
    cleaned = [clean_header_cell(c) for c in row]
    for key, names in aliases.items():
        for i, h in enumerate(cleaned):
            if not h:
                continue
            if any(h == n or h.startswith(n) for n in names):
                mapping[key] = i
                break
    return mapping


def find_header_map(rows: list[tuple], aliases: dict[str, list[str]], scan: int = 5):
    best_idx, best_map, best_score = 0, {}, -1
    for idx, row in enumerate(rows[:scan]):
        mapping = _header_map_from_row(row, aliases)
        score = len(mapping)
        if "pn" in mapping:
            score += 5
        if "client" in mapping or "customer" in mapping:
            score += 2
        if score > best_score:
            best_idx, best_map, best_score = idx, mapping, score
    return best_idx, best_map


TAZ_ALIASES = {
    "invoice": ["номер счета", "номер счёта", "счет", "счёт"],
    "customer": ["customer"],
    "pn": ["p/n", "part number", "partnumber"],
    "alt": ["alt p/n", "alt. p/n"],
    "desc": ["description"],
    "qty": ["qty in po", "qtyinpo"],
    "date": ["заказ взят в работу", "заказ клиента взят"],
    "cond": ["condition"],
    "sell": ["продажная, ед"],
    "status": ["status"],
    "ac": ["тип вс", "тип вс (продажи)", "a/c", "aircraft"],
}


def load_taz(path: Path) -> list[Event]:
    events: list[Event] = []
    seen = set()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "ORDERS" not in wb.sheetnames:
        wb.close()
        return events
    ws = wb["ORDERS"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return events

    mapping = _header_map_from_row(rows[0], TAZ_ALIASES)
    if "pn" not in mapping:
        for idx in range(min(3, len(rows))):
            mapping = _header_map_from_row(rows[idx], TAZ_ALIASES)
            if "pn" in mapping:
                break
    if "pn" not in mapping:
        print(f"  WARNING: no P/N in {path.name} ORDERS")
        return events

    start = 1
    for i, row in enumerate(rows[:5]):
        if row and any(clean_header_cell(c) in {"p/n", "part number", "pn"} for c in row[:15] if c):
            start = i + 1

    def cell(r, key, default=None):
        idx = mapping.get(key)
        if idx is None or idx >= len(r):
            return default
        return r[idx]

    for r in rows[start:]:
        if not r:
            continue
        pn = norm_pn(cell(r, "pn"))
        if is_sample_pn(pn):
            continue
        client = norm_client(cell(r, "customer"))
        if client.upper() in SAMPLE_CLIENT:
            client = ""
        # Внутренние «закупка на склад» — не рыночный спрос
        if client and not is_market_client(client):
            continue
        invoice = normalize_invoice(cell(r, "invoice"))
        qty = parse_qty(cell(r, "qty"))
        price = parse_money(cell(r, "sell"))
        cond = str(cell(r, "cond") or "").strip()
        alt = norm_pn(cell(r, "alt"))
        desc = str(cell(r, "desc") or "").strip()
        dt = cell(r, "date")
        ac_raw = cell(r, "ac")
        ac_type = str(ac_raw).strip() if ac_raw is not None else ""
        if ac_type.lower() in {"nan", "none", "тип вс", "a/c", "n/a", "-"}:
            ac_type = ""
        key = ("TAZ", pn, invoice or f"NO_INV|{client_key(client)}|{day_key(dt)}|{round(qty, 4)}")
        if key in seen:
            continue
        seen.add(key)
        if not client and not invoice and qty == 0:
            continue
        events.append(Event(
            source="TAZ", kind="order", pn=pn,
            alt=alt if alt and alt != pn else "",
            client=client, qty=qty, price=price, condition=cond,
            date=dt, request_no=invoice, sheet="ORDERS", description=desc,
            ac_type=ac_type,
        ))
    return events


TUZ_SKIP_SHEETS = {
    "suppliers list", "dropdown", "customer list",
    "transport rates and terms", "assets", "troubles",
}

TUZ_ALIASES = {
    "pn": ["p/n", "part number", "partnumber"],
    "alt": ["alt. p/n", "alt p/n", "alt pn", "alt.pn"],
    "client": ["client", "customer"],
    "ac": ["a/c", "ac", "aircraft", "тип вс"],
    "desc": ["description"],
    "qty": ["qty"],
    "cond": ["cond", "condition"],
    "date": ["request date"],
    "req": ["request №", "request no", "request #"],
    "price": ["offered\nper unit", "offered per unit", "offered"],
    "invoice": ["invoice to customer"],
}


def _tuz_layouts(primary: dict[str, int]) -> list[dict[str, int]]:
    classic_24 = {
        "date": 1, "req": 2, "client": 4, "ac": 5, "pn": 6, "alt": 7, "desc": 8, "qty": 9,
        "cond": 18, "price": 24, "invoice": 27,
    }
    classic_27 = {
        "date": 1, "req": 2, "client": 4, "ac": 5, "pn": 6, "alt": 7, "desc": 8, "qty": 9,
        "cond": 21, "price": 27, "invoice": 30,
    }
    shifted = {
        "date": 3, "req": 4, "client": 6, "ac": 7, "pn": 8, "alt": 9, "desc": 10, "qty": 11,
        "cond": 21, "price": 27, "invoice": 30,
    }
    qv2 = {
        "date": 3, "req": 4, "client": 6, "ac": 7, "pn": 8, "alt": 9, "desc": 10, "qty": 11,
        "cond": 20, "price": 26, "invoice": 29,
    }
    layouts = []
    if primary:
        layouts.append(dict(primary))
    for lay in (classic_27, classic_24, shifted, qv2):
        if lay not in layouts:
            layouts.append(lay)
    return layouts


def _extract_tuz_row(r: tuple, layouts: list[dict[str, int]]) -> Optional[dict]:
    for mapping in layouts:
        def get(key, default=None, _m=mapping):
            idx = _m.get(key)
            if idx is None or idx >= len(r):
                return default
            return r[idx]

        pn_raw = get("pn")
        pn = norm_pn(pn_raw)
        if is_sample_pn(pn):
            continue
        if isinstance(pn_raw, str) and " " in pn_raw.strip() and len(pn_raw) > 40:
            continue

        client = norm_client(get("client"))
        if client.upper() in SAMPLE_CLIENT:
            continue
        if not is_plausible_client(client, pn):
            continue

        alt = norm_pn(get("alt"))
        desc = str(get("desc") or "").strip()
        qty = parse_qty(get("qty"))
        price = parse_money(get("price"))
        cond = str(get("cond") or "").strip()
        if cond.lower() in {"cond", "condition"}:
            cond = ""
        ac_raw = get("ac")
        ac_type = str(ac_raw).strip() if ac_raw is not None else ""
        if ac_type.lower() in {"nan", "none", "a/c", "ac", "n/a", "-", "тип вс"}:
            ac_type = ""
        # не путать сдвинутый P/N / клиент с типом ВС
        if ac_type and (soft_pn_key(ac_type.upper()) == soft_pn_key(pn) or len(ac_type) > 40):
            ac_type = ""
        dt = get("date")
        if parse_date(dt) is None:
            for idx in (1, 3, mapping.get("date")):
                if idx is not None and idx < len(r) and parse_date(r[idx]) is not None:
                    dt = r[idx]
                    break
        req = str(get("req") or "").strip()
        invoice = str(get("invoice") or "").strip()
        return {
            "pn": pn,
            "alt": alt if alt and alt != pn else "",
            "client": client,
            "desc": desc,
            "qty": qty,
            "price": price,
            "cond": cond,
            "date": dt,
            "req": req or invoice,
            "ac_type": ac_type,
        }
    return None


def load_tuz(path: Path) -> list[Event]:
    events: list[Event] = []
    seen = set()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sheet in wb.sheetnames:
        if sheet.strip().lower() in TUZ_SKIP_SHEETS:
            continue
        ws = wb[sheet]
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            continue
        preview = all_rows[:5]
        hdr_idx, mapping = find_header_map(preview, TUZ_ALIASES, scan=5)
        if "pn" not in mapping:
            continue
        mapping.setdefault("client", 4)
        mapping.setdefault("ac", mapping["pn"] - 1 if mapping["pn"] > 0 else 5)
        mapping.setdefault("alt", mapping["pn"] + 1)
        mapping.setdefault("desc", mapping["pn"] + 2)
        mapping.setdefault("qty", mapping["pn"] + 3)
        mapping.setdefault("price", 27)

        layouts = _tuz_layouts(mapping)
        start = 0
        for i, row in enumerate(all_rows[:6]):
            if row and any(str(c).strip().lower() in {"p/n", "part number"} for c in row[:15] if c):
                start = i + 1
        start = max(start, hdr_idx + 1)

        for r in all_rows[start:]:
            if not r:
                continue
            parsed = _extract_tuz_row(r, layouts)
            if not parsed:
                continue
            dedupe_key = (
                parsed["pn"], client_key(parsed["client"]), day_key(parsed["date"]),
                parsed["req"] or "", round(parsed["qty"], 4),
                round(parsed["price"] or 0, 2), parsed["desc"][:40], parsed["cond"],
            )
            if dedupe_key in seen:
                continue
            seen.add(dedupe_key)
            events.append(Event(
                source="TUZ", kind="request", pn=parsed["pn"], alt=parsed["alt"],
                client=parsed["client"], qty=parsed["qty"], price=parsed["price"],
                condition=parsed["cond"], date=parsed["date"],
                request_no=parsed["req"], sheet=sheet, description=parsed["desc"],
                ac_type=parsed.get("ac_type") or "",
            ))
    wb.close()
    return events


# ---------------------------------------------------------------------------
# supplier stock
# ---------------------------------------------------------------------------

@dataclass
class StockLine:
    source: str          # LLP / Available
    pn: str
    description: str = ""
    condition: str = ""
    qty: float = 1.0
    serial: str = ""
    cycles_rem: str = ""
    engine_app: str = ""
    location: str = ""
    status: str = ""
    is_llp: bool = False


def parse_cycles_rem(raw: Any) -> str:
    if raw is None:
        return ""
    s = str(raw).strip()
    if not s or s.lower().startswith("cfm56"):
        return ""
    return s


def load_llp(path: Path) -> list[StockLine]:
    """Parse GEM LLP Inventory Report (sparse layout)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet1"]
    items: list[StockLine] = []
    # First pass: collect item rows
    item_rows = []
    for r in range(1, ws.max_row + 1):
        pn = ws.cell(r, 1).value
        serial = ws.cell(r, 13).value
        status = ws.cell(r, 21).value
        if pn and serial and status and str(status).strip() == "Available":
            item_rows.append(r)

    for r in item_rows:
        pn = norm_pn(ws.cell(r, 1).value)
        if is_sample_pn(pn):
            continue
        serial = str(ws.cell(r, 13).value or "").strip()
        if serial.upper() in SAMPLE_PN:
            serial = ""
        desc = str(ws.cell(r, 5).value or "").strip()
        cond = str(ws.cell(r, 15).value or "").strip()
        loc = str(ws.cell(r, 18).value or "").strip()
        cycles = parse_cycles_rem(ws.cell(r, 20).value)
        # engine app often 2 rows below in col 20
        engine = ""
        for rr in range(r + 1, min(r + 4, ws.max_row + 1)):
            v = ws.cell(rr, 20).value
            if v and str(v).strip().upper().startswith("CFM56"):
                engine = str(v).strip()
                break
        items.append(StockLine(
            source="LLP", pn=pn, description=desc, condition=cond,
            qty=1.0, serial=serial, cycles_rem=cycles, engine_app=engine,
            location=loc, status="Available", is_llp=True,
        ))
    wb.close()
    return items


def load_available_units(path: Path) -> list[StockLine]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    items: list[StockLine] = []
    # header row 1: PN, Description, Condition, QTY, PPW
    for r in range(2, ws.max_row + 1):
        pn_raw = ws.cell(r, 1).value
        pn = norm_pn(pn_raw)
        if is_sample_pn(pn):
            continue
        desc = str(ws.cell(r, 2).value or "").strip()
        cond = str(ws.cell(r, 3).value or "").strip()
        qty = parse_qty(ws.cell(r, 4).value)
        items.append(StockLine(
            source="Available", pn=pn, description=desc, condition=cond,
            qty=qty if qty else 1.0, is_llp=False, status="Available",
        ))
    wb.close()
    return items


# ---------------------------------------------------------------------------
# matching
# ---------------------------------------------------------------------------

def build_market_index(events: list[Event]) -> tuple[dict[str, MarketAgg], dict[str, set[str]]]:
    by_pn: dict[str, MarketAgg] = defaultdict(MarketAgg)
    soft_to_pns: dict[str, set[str]] = defaultdict(set)

    for e in events:
        by_pn[e.pn].merge(e, "exact")
        soft_to_pns[soft_pn_key(e.pn)].add(e.pn)
        if e.alt and e.alt != e.pn:
            by_pn[e.alt].merge(e, "alt")
            soft_to_pns[soft_pn_key(e.alt)].add(e.alt)

    return by_pn, soft_to_pns


def lookup_market(pn: str, by_pn: dict[str, MarketAgg], soft_to_pns: dict[str, set[str]]) -> tuple[MarketAgg, str]:
    if pn in by_pn:
        return by_pn[pn], "exact"
    soft = soft_pn_key(pn)
    cands = soft_to_pns.get(soft, set())
    if cands:
        # merge all soft matches
        agg = MarketAgg()
        for c in cands:
            src = by_pn[c]
            for inv in src.order_invoices:
                agg.order_invoices.add(inv)
            for k in src.request_keys:
                agg.request_keys.add(k)
            agg.order_clients |= src.order_clients
            agg.request_clients |= src.request_clients
            agg.order_prices.extend(src.order_prices)
            agg.request_prices.extend(src.request_prices)
            agg.order_qtys.extend(src.order_qtys)
            agg.request_qtys.extend(src.request_qtys)
            if src.last_order and (agg.last_order is None or src.last_order > agg.last_order):
                agg.last_order = src.last_order
            if src.last_request and (agg.last_request is None or src.last_request > agg.last_request):
                agg.last_request = src.last_request
            if src.sample_desc and not agg.sample_desc:
                agg.sample_desc = src.sample_desc
            agg.matched_via.add("soft")
        return agg, "soft"
    return MarketAgg(), "none"


def price_med(prices: list[float]) -> Optional[float]:
    if not prices:
        return None
    return round(median(prices), 2)


def indicative_price(m: MarketAgg) -> tuple[Optional[float], str]:
    if m.order_prices:
        return price_med(m.order_prices), "ТАЗ sell (median)"
    if m.request_prices:
        return price_med(m.request_prices), "ТУЗ offered (median)"
    return None, "нет цены"


def liquidity_hint(m: MarketAgg) -> str:
    n_ord = m.n_orders
    n_req = m.n_requests
    n_oc = len(m.order_clients)
    n_rc = len(m.request_clients)
    score = n_ord * 12 + n_oc * 10 + n_req * 3 + n_rc * 5
    if n_ord == 0 and n_req == 0:
        return "D — нет рынка"
    if score >= 55 or (n_ord >= 2 and n_oc >= 2):
        return "A — высокий спрос"
    if score >= 25 or n_ord >= 1 or n_rc >= 3:
        return "B — средний спрос"
    if n_req >= 1:
        return "C — слабый спрос (только запросы)"
    return "D — нет рынка"


# ---------------------------------------------------------------------------
# report
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
HIT_FILL = PatternFill("solid", fgColor="C6EFCE")
MISS_FILL = PatternFill("solid", fgColor="F2F2F2")
LLP_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)


def write_report(
    stock: list[StockLine],
    by_pn: dict[str, MarketAgg],
    soft_to_pns: dict[str, set[str]],
    taz_n: int,
    tuz_n: int,
    out: Path,
) -> dict:
    # Aggregate stock by PN for summary sheet.
    # Если PN есть и в LLP, и в Available — qty берём из LLP (серии),
    # Available qty показываем отдельно (возможен overlap одной и той же партии).
    by_stock_pn: dict[str, dict] = {}
    for s in stock:
        g = by_stock_pn.setdefault(s.pn, {
            "pn": s.pn,
            "desc": s.description,
            "sources": set(),
            "conds": set(),
            "qty_llp": 0.0,
            "qty_avail": 0.0,
            "serials": 0,
            "is_llp": False,
            "cycles": [],
            "engines": set(),
            "lines": [],
            "overlap": False,
        })
        g["sources"].add(s.source)
        if s.condition:
            g["conds"].add(s.condition)
        if s.is_llp:
            g["is_llp"] = True
            g["qty_llp"] += s.qty
            if s.serial:
                g["serials"] += 1
        else:
            g["qty_avail"] += s.qty
        if s.cycles_rem:
            g["cycles"].append(s.cycles_rem)
        if s.engine_app:
            g["engines"].add(s.engine_app)
        if not g["desc"] and s.description:
            g["desc"] = s.description
        g["lines"].append(s)

    for g in by_stock_pn.values():
        if g["qty_llp"] and g["qty_avail"]:
            g["overlap"] = True
            # не суммируем — берём LLP как основное кол-во
            g["qty"] = g["qty_llp"]
            g["qty_note"] = f"LLP={g['qty_llp']:.0f}; Available={g['qty_avail']:.0f} (не суммированы)"
        elif g["qty_llp"]:
            g["qty"] = g["qty_llp"]
            g["qty_note"] = ""
        else:
            g["qty"] = g["qty_avail"]
            g["qty_note"] = ""

    rows_out = []
    hits_taz = hits_tuz = hits_both = misses = 0
    for pn, g in sorted(by_stock_pn.items(), key=lambda x: x[0]):
        m, via = lookup_market(pn, by_pn, soft_to_pns)
        has_taz = m.n_orders > 0
        has_tuz = m.n_requests > 0
        if has_taz and has_tuz:
            hits_both += 1
            flag = "ТАЗ+ТУЗ"
        elif has_taz:
            hits_taz += 1
            flag = "только ТАЗ"
        elif has_tuz:
            hits_tuz += 1
            flag = "только ТУЗ"
        else:
            misses += 1
            flag = "нет в рынке"
        price, price_src = indicative_price(m)
        pot = round(price * g["qty"], 2) if price else None
        rows_out.append({
            "flag": flag,
            "match_via": via if via != "none" else "",
            "source": "+".join(sorted(g["sources"])),
            "is_llp": "LLP" if g["is_llp"] else "Rotable/прочее",
            "pn": pn,
            "desc": g["desc"],
            "conds": ", ".join(sorted(g["conds"])),
            "qty": g["qty"],
            "qty_note": g.get("qty_note") or "",
            "overlap": "да" if g.get("overlap") else "",
            "serials": g["serials"] or "",
            "cycles": "; ".join(g["cycles"][:5]) + ("…" if len(g["cycles"]) > 5 else ""),
            "engine": ", ".join(sorted(g["engines"])),
            "n_orders": m.n_orders,
            "n_order_clients": len(m.order_clients),
            "n_requests": m.n_requests,
            "n_req_clients": len(m.request_clients),
            "last_order": m.last_order.isoformat() if m.last_order else "",
            "last_request": m.last_request.isoformat() if m.last_request else "",
            "price": price,
            "price_src": price_src if price else "",
            "potential": pot,
            "hint": liquidity_hint(m),
            "market_desc": m.sample_desc,
        })

    # sort: hits first by potential desc, then misses
    def sort_key(r):
        hit = 0 if r["flag"] != "нет в рынке" else 1
        pot = r["potential"] or 0
        return (hit, -pot, r["pn"])

    rows_out.sort(key=sort_key)

    wb = openpyxl.Workbook()

    # Summary
    ws = wb.active
    ws.title = "Сводка"
    summary_lines = [
        ("Сверка склада поставщика (GEM) ↔ ТАЗ / ТУЗ", ""),
        ("Дата сверки", AS_OF.isoformat()),
        ("ТАЗ", "ТАЗ полный файл 27.07.2026.xlsx"),
        ("ТУЗ", "ТУЗ 2026 полный файл 31.07.2026.xlsx"),
        ("Склад LLP", "CFM56-3 5 7 LLP TXG.xlsx"),
        ("Склад Available", "Availabe Units.xlsx"),
        ("", ""),
        ("Событий ТАЗ (уник. заказы по P/N+счёт)", taz_n),
        ("Событий ТУЗ (уник. запросы)", tuz_n),
        ("Уникальных P/N на складе", len(by_stock_pn)),
        ("  — из них LLP (серийные)", sum(1 for g in by_stock_pn.values() if g["is_llp"])),
        ("  — строк Available Units", sum(1 for s in stock if s.source == "Available")),
        ("", ""),
        ("Найдены в ТАЗ+ТУЗ", hits_both),
        ("Только в ТАЗ (заказы)", hits_taz),
        ("Только в ТУЗ (запросы)", hits_tuz),
        ("Нет в рынке", misses),
        ("Покрытие (есть хоть 1 сигнал)", f"{(hits_both+hits_taz+hits_tuz)/max(len(by_stock_pn),1)*100:.1f}%"),
        ("", ""),
        ("Потенц. выручка (qty × медианная цена), USD",
         round(sum(r["potential"] or 0 for r in rows_out), 2)),
        ("  — по позициям с ценой", sum(1 for r in rows_out if r["potential"])),
    ]
    ws["A1"] = "Параметр"
    ws["B1"] = "Значение"
    for c in ("A1", "B1"):
        ws[c].fill = HEADER_FILL
        ws[c].font = HEADER_FONT
    for i, (a, b) in enumerate(summary_lines, start=2):
        ws.cell(i, 1, a)
        ws.cell(i, 2, b)
    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 45

    # Main match sheet
    headers = [
        "Сигнал рынка", "Матч", "Источник склада", "Тип", "P/N", "Описание (склад)",
        "Condition", "Кол-во (для оценки)", "Overlap LLP/Avail", "Серий (LLP)",
        "Cycles Rem.", "Engine App",
        "Заказов ТАЗ", "Клиентов (заказы)", "Запросов ТУЗ", "Клиентов (запросы)",
        "Посл. заказ", "Посл. запрос",
        "Цена индикат. USD", "Источник цены", "Потенц. выручка USD",
        "Оценка спроса", "Описание (рынок)", "Примечание qty",
    ]
    ws2 = wb.create_sheet("Сверка по P_N")
    for i, h in enumerate(headers, 1):
        cell = ws2.cell(1, i, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for ri, r in enumerate(rows_out, start=2):
        vals = [
            r["flag"], r["match_via"], r["source"], r["is_llp"], r["pn"], r["desc"],
            r["conds"], r["qty"], r["overlap"], r["serials"], r["cycles"], r["engine"],
            r["n_orders"], r["n_order_clients"], r["n_requests"], r["n_req_clients"],
            r["last_order"], r["last_request"],
            r["price"], r["price_src"], r["potential"],
            r["hint"], r["market_desc"], r["qty_note"],
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws2.cell(ri, ci, v)
            cell.border = THIN
            cell.alignment = Alignment(vertical="center", wrap_text=ci in {5, 6, 11, 23, 24})
        fill = HIT_FILL if r["flag"] != "нет в рынке" else MISS_FILL
        if r["is_llp"].startswith("LLP") and r["flag"] != "нет в рынке":
            fill = LLP_FILL
        for ci in range(1, len(headers) + 1):
            ws2.cell(ri, ci).fill = fill

    widths = [14, 8, 14, 14, 18, 36, 12, 12, 12, 10, 16, 12, 10, 12, 10, 12, 12, 12, 12, 16, 14, 22, 36, 36]
    for i, w in enumerate(widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows_out)+1}"
    ws2.freeze_panes = "E2"

    # Detail serial sheet for LLP hits
    ws3 = wb.create_sheet("LLP серии (детально)")
    det_headers = [
        "Сигнал рынка", "P/N", "Serial", "Описание", "Condition", "Cycles Rem.",
        "Engine", "Location", "Заказов ТАЗ", "Запросов ТУЗ", "Цена USD", "Оценка спроса",
    ]
    for i, h in enumerate(det_headers, 1):
        cell = ws3.cell(1, i, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    dri = 2
    for s in sorted([x for x in stock if x.is_llp], key=lambda x: (x.pn, x.serial)):
        m, via = lookup_market(s.pn, by_pn, soft_to_pns)
        has = m.n_orders > 0 or m.n_requests > 0
        if not has:
            continue
        price, _ = indicative_price(m)
        flag = (
            "ТАЗ+ТУЗ" if m.n_orders and m.n_requests else
            "только ТАЗ" if m.n_orders else "только ТУЗ"
        )
        vals = [
            flag, s.pn, s.serial, s.description, s.condition, s.cycles_rem,
            s.engine_app, s.location, m.n_orders, m.n_requests, price, liquidity_hint(m),
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws3.cell(dri, ci, v)
            cell.border = THIN
            cell.fill = LLP_FILL
        dri += 1
    for i, w in enumerate([14, 18, 16, 36, 10, 14, 12, 10, 10, 10, 12, 22], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.freeze_panes = "C2"

    # No market sheet
    ws4 = wb.create_sheet("Нет в рынке")
    nm_headers = ["Источник", "Тип", "P/N", "Описание", "Condition", "Кол-во", "Серий", "Cycles Rem.", "Примечание qty"]
    for i, h in enumerate(nm_headers, 1):
        cell = ws4.cell(1, i, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    ni = 2
    for r in rows_out:
        if r["flag"] != "нет в рынке":
            continue
        vals = [r["source"], r["is_llp"], r["pn"], r["desc"], r["conds"], r["qty"], r["serials"], r["cycles"], r["qty_note"]]
        for ci, v in enumerate(vals, 1):
            cell = ws4.cell(ni, ci, v)
            cell.border = THIN
            cell.fill = MISS_FILL
        ni += 1
    for i, w in enumerate([14, 14, 18, 36, 12, 10, 10, 16, 36], 1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    # Method note
    ws5 = wb.create_sheet("Методика")
    notes = [
        "Методика сверки",
        "",
        "1. Склад: LLP (серийный GEM LLP Inventory) + Available Units (агрегат по PN).",
        "2. Рынок: ТАЗ ORDERS (заказы) + ТУЗ (все рабочие листы запросов).",
        "3. Нормализация P/N: upper, без пробелов/nbsp, унификация тире; soft-match без разделителей.",
        "4. ALT P/N из ТАЗ/ТУЗ индексируется как отдельный ключ (matched_via=alt).",
        "5. Заказ = уникальная пара P/N + № счёта; запрос = P/N + клиент + день (+ request №).",
        "6. Внутренние «закупка на склад» и SAMPLE/TBA не считаются рыночным спросом.",
        "7. Цена: медиана sell из ТАЗ; если нет — медиана Offered из ТУЗ.",
        "8. Потенц. выручка = кол-во на складе × индикативная цена (грубая оценка).",
        "9. Оценка спроса A–D: заказы×12 + клиенты_заказов×10 + запросы×3 + клиенты_запросов×5.",
        "10. LLP выделены отдельно; Cycles Rem. — как в выгрузке GEM (сырых TSN/CSN нет).",
        "11. Если P/N есть и в LLP, и в Available Units — qty для оценки = число LLP-серий (Available не суммируется, overlap помечен).",
        "12. Строки ТАЗ с Customer «Закупка на склад» исключены из рыночного спроса.",
    ]
    for i, line in enumerate(notes, 1):
        ws5.cell(i, 1, line)
        if i == 1:
            ws5.cell(i, 1).font = Font(bold=True, size=14)
    ws5.column_dimensions["A"].width = 110

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    OUT_ART.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_ART)

    stats = {
        "stock_pn": len(by_stock_pn),
        "hits_both": hits_both,
        "hits_taz": hits_taz,
        "hits_tuz": hits_tuz,
        "misses": misses,
        "potential": round(sum(r["potential"] or 0 for r in rows_out), 2),
        "top": rows_out[:15],
    }
    return stats


def main():
    llp_path = DATA / "supplier_LLP.xlsx"
    avail_path = DATA / "supplier_Available_Units.xlsx"
    taz_path = DATA / "TAZ_27.07.2026.xlsx"
    tuz_path = DATA / "TUZ_31.07.2026.xlsx"

    print("Loading supplier LLP...")
    llp = load_llp(llp_path)
    print(f"  LLP serials: {len(llp)}, unique PN: {len({s.pn for s in llp})}")

    print("Loading Available Units...")
    avail = load_available_units(avail_path)
    print(f"  Available lines: {len(avail)}, qty sum: {sum(s.qty for s in avail)}")

    stock = llp + avail

    print("Loading TAZ ORDERS...")
    taz = load_taz(taz_path)
    print(f"  TAZ events: {len(taz)}")

    print("Loading TUZ requests...")
    tuz = load_tuz(tuz_path)
    print(f"  TUZ events: {len(tuz)}")

    by_pn, soft_to_pns = build_market_index(taz + tuz)
    print(f"  Market index PNs: {len(by_pn)}")

    print("Writing report...")
    stats = write_report(stock, by_pn, soft_to_pns, len(taz), len(tuz), OUT)
    print(f"Saved: {OUT}")
    print(f"Artifact: {OUT_ART}")
    print(
        f"Coverage: both={stats['hits_both']} taz_only={stats['hits_taz']} "
        f"tuz_only={stats['hits_tuz']} miss={stats['misses']} / {stats['stock_pn']}"
    )
    print(f"Potential USD: {stats['potential']:,.2f}")
    print("\nTop hits:")
    for r in stats["top"]:
        if r["flag"] == "нет в рынке":
            break
        print(
            f"  {r['flag']:12} {r['pn']:20} qty={r['qty']:<6} "
            f"ord={r['n_orders']:<3} req={r['n_requests']:<3} "
            f"price={r['price'] or '-'} pot={r['potential'] or '-'}  {r['hint']}"
        )


if __name__ == "__main__":
    main()
