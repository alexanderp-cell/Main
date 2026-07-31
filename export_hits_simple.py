#!/usr/bin/env python3
"""
Упрощённый отчёт: только позиции склада, которые встречаются в ТАЗ и/или ТУЗ.
По каждой позиции — детальные появления: когда, где (лист/счёт), клиент, цена.
"""
from __future__ import annotations

from collections import defaultdict
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from reconcile_supplier_vs_market import (
    DATA,
    Event,
    build_market_index,
    indicative_price,
    load_available_units,
    load_llp,
    load_taz,
    load_tuz,
    lookup_market,
    norm_pn,
    parse_date,
    soft_pn_key,
)

OUT = Path("/workspace/output/Supplier_hits_TAZ_TUZ.xlsx")
OUT_ART = Path("/opt/cursor/artifacts/Supplier_hits_TAZ_TUZ.xlsx")

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
TAZ_FILL = PatternFill("solid", fgColor="DDEBF7")
TUZ_FILL = PatternFill("solid", fgColor="E2EFDA")
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)


def stock_by_pn(stock):
    """pn -> {desc, qty, sources, conds, is_llp, serials} без двойного подсчёта overlap."""
    out = {}
    for s in stock:
        g = out.setdefault(s.pn, {
            "pn": s.pn,
            "desc": s.description,
            "qty_llp": 0.0,
            "qty_avail": 0.0,
            "sources": set(),
            "conds": set(),
            "is_llp": False,
            "serials": [],
        })
        g["sources"].add(s.source)
        if s.condition:
            g["conds"].add(s.condition)
        if s.is_llp:
            g["is_llp"] = True
            g["qty_llp"] += s.qty
            if s.serial:
                g["serials"].append(s.serial)
        else:
            g["qty_avail"] += s.qty
        if not g["desc"] and s.description:
            g["desc"] = s.description
    for g in out.values():
        if g["qty_llp"] and g["qty_avail"]:
            g["qty"] = g["qty_llp"]
            g["qty_note"] = f"LLP={g['qty_llp']:.0f}; Avail={g['qty_avail']:.0f}"
        elif g["qty_llp"]:
            g["qty"] = g["qty_llp"]
            g["qty_note"] = ""
        else:
            g["qty"] = g["qty_avail"]
            g["qty_note"] = ""
    return out


def events_for_pn(pn: str, events: list[Event], by_pn, soft_to_pns) -> list[Event]:
    """Все рыночные события, матчащиеся на складской P/N (exact / soft / alt уже в индексе)."""
    keys = {pn}
    soft = soft_pn_key(pn)
    keys |= soft_to_pns.get(soft, set())
    # также события, где наш pn был alt у другого ключа — они уже смержены в by_pn[pn] via alt,
    # но для детализации берём сырые события по exact pn и soft-совпадениям
    matched = []
    for e in events:
        if e.pn in keys or (e.alt and e.alt in keys) or soft_pn_key(e.pn) == soft:
            matched.append(e)
        elif e.alt and soft_pn_key(e.alt) == soft:
            matched.append(e)
    # dedupe
    seen = set()
    out = []
    for e in matched:
        d = parse_date(e.date)
        key = (
            e.source, e.kind, e.pn, e.alt, e.client, e.request_no,
            d.isoformat() if d else "", round(e.qty or 0, 4),
            round(e.price or 0, 2), e.sheet, e.condition,
        )
        if key in seen:
            continue
        seen.add(key)
        out.append(e)
    out.sort(key=lambda e: (
        parse_date(e.date) is None,
        parse_date(e.date) or parse_date("1900-01-01"),
        e.source,
        e.client,
    ))
    return out


def style_header(ws, n_cols):
    for i in range(1, n_cols + 1):
        c = ws.cell(1, i)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(wrap_text=True, vertical="center")


def write_report(stock_map, events, by_pn, soft_to_pns, path: Path):
    # only PNs with market signal
    hits = []
    for pn, g in stock_map.items():
        m, via = lookup_market(pn, by_pn, soft_to_pns)
        if m.n_orders == 0 and m.n_requests == 0:
            continue
        evs = events_for_pn(pn, events, by_pn, soft_to_pns)
        hits.append((pn, g, m, via, evs))
    hits.sort(key=lambda x: (-(x[2].n_orders + x[2].n_requests), x[0]))

    wb = openpyxl.Workbook()

    # --- Sheet 1: Positions ---
    ws = wb.active
    ws.title = "Позиции"
    headers1 = [
        "P/N", "Описание", "Тип", "Condition на складе", "Кол-во",
        "Источник склада", "Заказов ТАЗ", "Запросов ТУЗ",
        "Клиенты (заказы)", "Клиенты (запросы)",
        "Цена индикат. USD", "Источник цены",
        "Посл. заказ", "Посл. запрос", "Примечание qty",
    ]
    for i, h in enumerate(headers1, 1):
        ws.cell(1, i, h)
    style_header(ws, len(headers1))

    for ri, (pn, g, m, via, evs) in enumerate(hits, start=2):
        price, psrc = indicative_price(m)
        # клиенты в исходном регистре из появлений
        ord_clients = sorted({e.client for e in evs if e.source == "TAZ" and e.client}, key=str.lower)
        req_clients = sorted({e.client for e in evs if e.source == "TUZ" and e.client}, key=str.lower)
        vals = [
            pn,
            g["desc"],
            "LLP" if g["is_llp"] else "Rotable",
            ", ".join(sorted(g["conds"])),
            g["qty"],
            "+".join(sorted(g["sources"])),
            m.n_orders,
            m.n_requests,
            ", ".join(ord_clients),
            ", ".join(req_clients),
            price,
            psrc if price else "",
            m.last_order.isoformat() if m.last_order else "",
            m.last_request.isoformat() if m.last_request else "",
            g.get("qty_note") or "",
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(ri, ci, v)
            cell.border = THIN
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    for i, w in enumerate([18, 36, 10, 14, 8, 14, 10, 10, 28, 28, 12, 18, 12, 12, 28], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f"A1:O{len(hits)+1}"
    ws.freeze_panes = "B2"

    # --- Sheet 2: Appearances detail ---
    ws2 = wb.create_sheet("Появления")
    headers2 = [
        "P/N склада", "Описание склада", "Тип", "Кол-во на складе",
        "Где", "Лист / источник", "Дата", "Клиент",
        "P/N в документе", "ALT P/N", "Описание (рынок)",
        "Condition", "Qty (заказ/запрос)", "Цена USD",
        "№ счёта / Request №",
    ]
    for i, h in enumerate(headers2, 1):
        ws2.cell(1, i, h)
    style_header(ws2, len(headers2))

    ri = 2
    for pn, g, m, via, evs in hits:
        for e in evs:
            d = parse_date(e.date)
            where = "ТАЗ (заказ)" if e.source == "TAZ" else "ТУЗ (запрос)"
            vals = [
                pn,
                g["desc"],
                "LLP" if g["is_llp"] else "Rotable",
                g["qty"],
                where,
                e.sheet or e.source,
                d.isoformat() if d else "",
                e.client,
                e.pn,
                e.alt,
                e.description,
                e.condition,
                e.qty or "",
                e.price if e.price is not None else "",
                e.request_no,
            ]
            for ci, v in enumerate(vals, 1):
                cell = ws2.cell(ri, ci, v)
                cell.border = THIN
                cell.fill = TAZ_FILL if e.source == "TAZ" else TUZ_FILL
                cell.alignment = Alignment(vertical="center", wrap_text=ci in {2, 11})
            ri += 1

    for i, w in enumerate([18, 32, 10, 10, 14, 16, 12, 22, 18, 14, 32, 10, 10, 12, 22], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.auto_filter.ref = f"A1:O{ri-1}"
    ws2.freeze_panes = "B2"

    # --- Sheet 3: short note ---
    ws3 = wb.create_sheet("Пояснение")
    lines = [
        "Упрощённая сверка: только позиции склада поставщика, которые есть в ТАЗ и/или ТУЗ.",
        "",
        "Лист «Позиции» — сводка по P/N (сколько заказов/запросов, клиенты, индикативная цена).",
        "Лист «Появления» — каждое появление в ТАЗ/ТУЗ: дата, клиент, цена, № счёта или Request.",
        "",
        "ТАЗ: полный файл 27.07.2026 (ORDERS), без «Закупка на склад».",
        "ТУЗ: полный файл 31.07.2026.",
        "Склад: GEM LLP + Available Units; при overlap qty = число LLP-серий.",
        "Синий = ТАЗ, зелёный = ТУЗ.",
    ]
    for i, line in enumerate(lines, 1):
        ws3.cell(i, 1, line)
        if i == 1:
            ws3.cell(i, 1).font = Font(bold=True, size=13)
    ws3.column_dimensions["A"].width = 100

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    OUT_ART.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_ART)
    return len(hits), ri - 2


def main():
    print("Loading stock...")
    stock = load_llp(DATA / "supplier_LLP.xlsx") + load_available_units(DATA / "supplier_Available_Units.xlsx")
    smap = stock_by_pn(stock)
    print(f"  stock PNs: {len(smap)}")

    print("Loading TAZ/TUZ...")
    taz = load_taz(DATA / "TAZ_27.07.2026.xlsx")
    tuz = load_tuz(DATA / "TUZ_31.07.2026.xlsx")
    events = taz + tuz
    print(f"  TAZ={len(taz)} TUZ={len(tuz)}")

    by_pn, soft_to_pns = build_market_index(events)
    n_pos, n_rows = write_report(smap, events, by_pn, soft_to_pns, OUT)
    print(f"Saved {OUT} — positions={n_pos}, appearance rows={n_rows}")
    print(f"Artifact: {OUT_ART}")


if __name__ == "__main__":
    main()
