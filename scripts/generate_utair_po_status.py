#!/usr/bin/env python3
"""Generate METEOR-UTAIR PO STATUS report from TAZ source data."""

from __future__ import annotations

import re
import shutil
from copy import copy
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

TAZ_PATH = Path(
    "/home/ubuntu/.cursor/projects/workspace/uploads/__________17.07.2026_4a9d.xlsx"
)
TEMPLATE_PATH = Path(
    "/home/ubuntu/.cursor/projects/workspace/uploads/METEOR-UTAIR_PO_STATUS__1__082a.xlsx"
)
OUTPUT_PATH = Path("/workspace/METEOR-UTAIR PO STATUS 17.07.2026.xlsx")

TODAY = date(2026, 7, 17)
PRE_TRANSIT_ROW = 420

STATUS_ORDER = [
    "Ожидает аванса",
    "Требует внимания",
    "Принят в работу",
    "Отгрузка от поставщика",
    "Лид тайм",
    "В транзите",
    "Отгружен",
    "Завершен",
    "Отменен",
]

STATUS_RANK = {
    "4 finished": 100,
    "3 shipped": 90,
    "5 cancelled": 80,
    "6 trouble": 70,
    "2 paid": 60,
    "11 accepted": 55,
    "1 not paid": 40,
    "7 refund": 30,
    "8 warranty": 20,
}

TRANSIT_COMMENTS = {
    "consolidation": "На складе консолидации, готовится к отправке в транзитную точку",
    "transit_point": "В транзитной точке, готовится к отправке в РФ",
    "transit_delay": "Задержка в транзитной точке, отправка в РФ в ближайшее время",
    "shop": "Работы в шопе завершены, ожидаем отгрузку от поставщика",
    "default": "Груз в пути, уточняем детали транзита",
}

LEAD_TIME_FILL = PatternFill("solid", fgColor="FFFF00")


def norm_inv(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text or None


def norm_pn(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].replace(".", "", 1).isdigit():
        text = text[:-2]
    return text


def is_utair_customer(customer) -> bool:
    if not customer:
        return False
    text = str(customer).lower()
    return "utair" in text or text == "utg" or "utg" in text


def to_date(value) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def fmt_date(value) -> str | None:
    parsed = to_date(value)
    if not parsed:
        return None
    return parsed.strftime("%d.%m.%Y")


def fmt_money(value) -> str | float | None:
    if value is None:
        return None
    if isinstance(value, str):
        return value
    try:
        number = float(value)
    except (TypeError, ValueError):
        return value
    if number == int(number):
        text = f"{int(number)}"
    else:
        text = f"{number:.2f}".replace(".", ",")
    return text


def taz_status_key(status) -> str:
    return str(status or "").strip().lower()


def status_rank(status) -> int:
    key = taz_status_key(status)
    if key in STATUS_RANK:
        return STATUS_RANK[key]
    if key.startswith("4"):
        return 100
    if key.startswith("3"):
        return 90
    if key.startswith("5"):
        return 80
    if key.startswith("6"):
        return 70
    if key.startswith("2"):
        return 60
    if key.startswith("1"):
        return 40
    return 10


def is_lead_time_number(value) -> bool:
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return True
    text = str(value).strip().upper()
    if text in {"", "STK", "N/A", "NA"}:
        return False
    try:
        float(text.replace(",", "."))
        return True
    except ValueError:
        return False


def has_client_payment(row, ws) -> bool:
    status = taz_status_key(ws.cell(row, 5).value)
    if status.startswith("2") or "paid" in status:
        return True
    for col in (45, 47):  # AS, AU
        value = ws.cell(row, col).value
        if value not in (None, "", 0, 0.0):
            try:
                if float(value) != 0:
                    return True
            except (TypeError, ValueError):
                return True
    for col in (44, 46):  # AR, AT
        if ws.cell(row, col).value:
            return True
    return False


def days_in_work(work_date) -> int | None:
    parsed = to_date(work_date)
    if not parsed:
        return None
    return (TODAY - parsed).days


def clean_order_number(comment) -> str | None:
    if not comment:
        return None
    lines = [line.strip() for line in str(comment).splitlines() if line.strip()]
    kept = []
    for line in lines:
        upper = line.upper()
        if "ПЕРЕД ОТПРАВКОЙ" in upper or "ПРОВЕРИТЬ МПТ" in upper:
            continue
        if upper.startswith("SAMPLE"):
            continue
        kept.append(line)
        if re.search(r"P\d+", line):
            break
    return "\n".join(kept) if kept else str(comment)


def compute_status(row: int, ws, below_pre: bool) -> str:
    status = taz_status_key(ws.cell(row, 5).value)

    if "finished" in status or status.startswith("4"):
        return "Завершен"
    if "shipped" in status or status.startswith("3"):
        return "Отгружен"
    if "cancelled" in status or status.startswith("5"):
        return "Отменен"
    if "trouble" in status or status.startswith("6"):
        return "Требует внимания"
    if below_pre:
        return "В транзите"
    if is_lead_time_number(ws.cell(row, 19).value):
        return "Лид тайм"

    days = days_in_work(ws.cell(row, 17).value)
    paid = has_client_payment(row, ws)

    if days is not None and days < 10:
        return "Ожидает аванса"
    if days is not None and days > 15:
        return "Отгрузка от поставщика"
    if (days is not None and days > 10) or paid:
        return "Принят в работу"
    return "Ожидает аванса"


def planned_delivery(status: str, deadline, fact_date) -> str | None:
    if status == "Ожидает аванса":
        return None
    if status == "Требует внимания":
        return "N/A"
    if status == "Принят в работу":
        return fmt_date(TODAY + timedelta(days=20))
    if status == "Отгрузка от поставщика":
        return fmt_date(TODAY + timedelta(days=15))
    if status == "В транзите":
        return fmt_date(TODAY + timedelta(days=10))
    if status in {"Отгружен", "Завершен"}:
        return fmt_date(fact_date) or fmt_date(deadline)
    if status == "Лид тайм":
        return None
    return fmt_date(deadline)


def transit_comment(row: int, ws, existing: str | None = None) -> str:
    if existing:
        return existing

    comment = str(ws.cell(row, 6).value or "").lower()
    destination = str(ws.cell(row, 20).value or "").lower()
    supplier = str(ws.cell(row, 26).value or "").lower()
    status = taz_status_key(ws.cell(row, 5).value)

    if any(word in comment for word in ("шоп", "shop", "стекл")):
        return TRANSIT_COMMENTS["shop"]
    if "транзитн" in comment or "transit" in destination:
        if "задерж" in comment:
            return TRANSIT_COMMENTS["transit_delay"]
        return TRANSIT_COMMENTS["transit_point"]
    if status.startswith("2") or "consolid" in supplier or "склад" in comment:
        return TRANSIT_COMMENTS["consolidation"]
    if status.startswith("3"):
        return "Отгружено, ожидаем поступление на склад в Москве"
    return TRANSIT_COMMENTS["default"]


def pn_matches(template_pn: str, taz_pn: str) -> bool:
    left = template_pn.replace(" ", "").upper()
    right = str(taz_pn).replace(" ", "").upper()
    if left == right:
        return True
    if left.endswith(".0"):
        left = left[:-2]
    if right.endswith(".0"):
        right = right[:-2]
    return left == right


def pick_best_row(rows: list[int], ws, template_pn: str | None = None) -> int:
    def sort_key(row: int):
        below = row > PRE_TRANSIT_ROW
        pn_match = 0
        if template_pn:
            taz_pn = norm_pn(ws.cell(row, 11).value) or ""
            pn_match = 1 if pn_matches(template_pn, taz_pn) else 0
        return (
            1 if below else 0,
            pn_match,
            status_rank(ws.cell(row, 5).value),
            row,
        )

    return sorted(rows, key=sort_key, reverse=True)[0]


def load_taz_index(ws):
    index: dict[tuple[str, str], list[int]] = {}
    for row in range(2, ws.max_row + 1):
        invoice = norm_inv(ws.cell(row, 1).value)
        pn = norm_pn(ws.cell(row, 11).value)
        if not invoice or not pn or not is_utair_customer(ws.cell(row, 7).value):
            continue
        index.setdefault((invoice, pn), []).append(row)
    return index


def build_record(key: tuple[str, str], row: int, ws, existing_comment: str | None = None) -> dict:
    below_pre = row > PRE_TRANSIT_ROW
    status = compute_status(row, ws, below_pre)
    deadline = ws.cell(row, 22).value
    fact_date = ws.cell(row, 23).value
    invoice = norm_inv(ws.cell(row, 1).value)

    # invoice numeric when possible
    invoice_value: str | float = invoice
    if invoice and re.fullmatch(r"\d+", invoice):
        invoice_value = float(invoice)

    record = {
        "key": key,
        "status": status,
        "invoice": invoice_value,
        "order_no": clean_order_number(ws.cell(row, 6).value),
        "pn": key[1],
        "alt_pn": ws.cell(row, 12).value,
        "description": ws.cell(row, 13).value,
        "qty": ws.cell(row, 14).value,
        "uom": ws.cell(row, 15).value,
        "work_date": fmt_date(ws.cell(row, 17).value),
        "lead_days": ws.cell(row, 18).value,
        "deadline": fmt_date(deadline),
        "fact_date": fmt_date(fact_date),
        "condition": ws.cell(row, 25).value,
        "serial": ws.cell(row, 30).value,
        "unit_price": fmt_money(ws.cell(row, 33).value),
        "total_price": fmt_money(ws.cell(row, 34).value),
        "payment_type": ws.cell(row, 43).value,
        "balance": ws.cell(row, 48).value,
        "advance_date": fmt_date(ws.cell(row, 44).value),
        "advance_amount": ws.cell(row, 45).value,
        "final_payment_date": fmt_date(ws.cell(row, 46).value),
        "final_payment_amount": ws.cell(row, 47).value,
        "planned_date": planned_delivery(status, deadline, fact_date),
        "comment": None,
        "taz_row": row,
    }

    if status == "В транзите":
        record["comment"] = transit_comment(row, ws, existing_comment)

    return record


def find_taz_rows(key: tuple[str, str], taz_index: dict, taz_ws) -> list[int]:
    rows = taz_index.get(key, [])
    if rows:
        return rows

    invoice, pn = key
    # Fallback: match by invoice and comparable part number.
    fallback = []
    for (inv, part), row_list in taz_index.items():
        if inv != invoice:
            continue
        if part == pn or pn_matches(pn, part):
            fallback.extend(row_list)
    if fallback:
        return fallback

    # Last resort for legacy template rows: any row with the same invoice.
    for (inv, _part), row_list in taz_index.items():
        if inv == invoice:
            fallback.extend(row_list)
    return fallback


def collect_orders(taz_ws, template_ws):
    taz_index = load_taz_index(taz_ws)
    orders: dict[tuple[str, str], dict] = {}
    existing_comments: dict[tuple[str, str], str] = {}
    template_keys: set[tuple[str, str]] = set()

    for row in range(2, template_ws.max_row + 1):
        invoice = norm_inv(template_ws.cell(row, 4).value)
        pn = norm_pn(template_ws.cell(row, 6).value)
        if not invoice or not pn:
            continue
        key = (invoice, pn)
        template_keys.add(key)
        comment = template_ws.cell(row, 1).value
        if comment:
            existing_comments[key] = str(comment)

    selected_keys = set(template_keys)

    # Add new active Utair orders from the upper part of TAZ.
    for row in range(2, PRE_TRANSIT_ROW + 1):
        invoice = norm_inv(taz_ws.cell(row, 1).value)
        pn = norm_pn(taz_ws.cell(row, 11).value)
        if not invoice or not pn or not is_utair_customer(taz_ws.cell(row, 7).value):
            continue
        selected_keys.add((invoice, pn))

    for key in sorted(selected_keys):
        rows = find_taz_rows(key, taz_index, taz_ws)
        if not rows:
            continue
        best_row = pick_best_row(rows, taz_ws, template_pn=key[1])
        orders[key] = build_record(
            key,
            best_row,
            taz_ws,
            existing_comments.get(key),
        )

    return orders


def clone_cell_style(source, target):
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.number_format = source.number_format
    target.protection = copy(source.protection)


def build_workbook(orders: dict[tuple[str, str], dict]):
    shutil.copy2(TEMPLATE_PATH, OUTPUT_PATH)
    wb = openpyxl.load_workbook(OUTPUT_PATH)
    ws = wb["Main"]
    template_ws = openpyxl.load_workbook(TEMPLATE_PATH)["Main"]

    header_style = {col: template_ws.cell(1, col) for col in range(1, 25)}
    section_style = template_ws.cell(3, 6)
    data_style = {col: template_ws.cell(16, col) for col in range(1, 25)}

    max_row = ws.max_row
    if max_row > 1:
        ws.delete_rows(2, max_row - 1)

    grouped: dict[str, list[dict]] = {status: [] for status in STATUS_ORDER}
    for record in orders.values():
        grouped.setdefault(record["status"], []).append(record)

    for status in grouped:
        grouped[status].sort(
            key=lambda item: (
                str(item["invoice"]),
                str(item["pn"]),
            )
        )

    current_row = 2
    data_rows_for_validation: dict[str, list[int]] = {status: [] for status in STATUS_ORDER}

    for status in STATUS_ORDER:
        records = grouped.get(status, [])
        if not records and status not in {
            "Ожидает аванса",
            "Требует внимания",
            "Принят в работу",
            "Отгрузка от поставщика",
            "Лид тайм",
            "В транзите",
            "Отгружен",
            "Завершен",
            "Отменен",
        }:
            continue

        section_cell = ws.cell(current_row, 6, status)
        clone_cell_style(section_style, section_cell)
        section_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[current_row].height = 28.5
        current_row += 1

        for record in records:
            values = {
                1: record["comment"],
                2: record["planned_date"],
                3: record["status"],
                4: record["invoice"],
                5: record["order_no"],
                6: record["pn"],
                7: record["alt_pn"],
                8: record["description"],
                9: record["qty"],
                10: record["uom"],
                11: record["work_date"],
                12: record["lead_days"],
                13: record["deadline"],
                14: record["fact_date"],
                15: record["condition"],
                16: record["serial"],
                17: record["unit_price"],
                18: record["total_price"],
                19: record["payment_type"],
                20: record["balance"],
                21: record["advance_date"],
                22: record["advance_amount"],
                23: record["final_payment_date"],
                24: record["final_payment_amount"],
            }

            for col, value in values.items():
                cell = ws.cell(current_row, col, value)
                clone_cell_style(data_style[col], cell)
                if status == "Лид тайм" and col == 2:
                    cell.fill = LEAD_TIME_FILL

            ws.row_dimensions[current_row].height = 28.5
            data_rows_for_validation[status].append(current_row)
            current_row += 1

    last_row = max(current_row - 1, 1)
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"$A$1:$X${last_row}"

    ws.data_validations.dataValidation.clear()
    if last_row > 1:
        from openpyxl.worksheet.datavalidation import DataValidation

        status_validation = DataValidation(
            type="list",
            formula1="Dropdown!$A$2:$A$34",
            allow_blank=True,
        )
        status_validation.add(f"C2:C{last_row}")
        ws.add_data_validation(status_validation)

        payment_validation = DataValidation(
            type="list",
            formula1="Dropdown!$B$2:$B$4",
            allow_blank=True,
        )
        payment_validation.add(f"S2:S{last_row}")
        ws.add_data_validation(payment_validation)

    wb.save(OUTPUT_PATH)
    return grouped


def main():
    taz_wb = openpyxl.load_workbook(TAZ_PATH, data_only=True)
    template_wb = openpyxl.load_workbook(TEMPLATE_PATH, data_only=True)
    taz_ws = taz_wb["Лист1"]
    template_ws = template_wb["Main"]

    orders = collect_orders(taz_ws, template_ws)
    grouped = build_workbook(orders)

    print(f"Generated: {OUTPUT_PATH}")
    for status in STATUS_ORDER:
        print(f"  {status}: {len(grouped.get(status, []))}")


if __name__ == "__main__":
    main()
