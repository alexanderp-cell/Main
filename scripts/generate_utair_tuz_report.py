#!/usr/bin/env python3
"""Generate Utair TUZ performance HTML report (group sheets only)."""

from __future__ import annotations

import html
import json
import re
import statistics
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl

TUZ_PATH = Path("/tmp/utair_analysis/TUZ_ТУЗ полный файл 01.09.2026.xlsx")
TAZ_PATH = Path("/tmp/utair_analysis/TAZ_ТАЗ полный файл 28.08.2026.xlsx")
OUTPUT_PATH = Path("/workspace/UTAIR_TUZ_performance_report.html")
ZIP_PATH = Path("/workspace/UTAIR_TUZ_performance_report.zip")

EARLY_STATUSES = {
    "0. Начальный этап",
    "1. Проценка у поставщиков",
    "8. Backup",
    "11. Внесено в ТАЗ",
}

CONDITION_ONLY = {"SV", "N", "OH", "R", "HT", "NEW"}

HOLD_REASON_RULES = [
    ("В активах", r"актив"),
    ("Документы / сертификаты", r"(жд[её]м|нужн|запраш|ждут|provide).*(док|doc|btb|arc|nis|серт|cert|ppwk|ах|approval)"),
    ("Габариты / транспорт", r"(габарит|транспорт|доставк|exw|размер|вес)"),
    ("Отмена / не отправляем", r"(отмен|не предлаг|не отправ|sold|зина)"),
    ("Обсуждение / риски", r"(обсуд|риск|целесообраз)"),
    ("Ждём ответ поставщика", r"(жд[её]м|жду|апдейт|квота от)"),
]

TROUBLE_RE = re.compile(r"trouble|трабл", re.IGNORECASE)
ORDER_REF_RE = re.compile(r"(P\d{5,}|Q\d{5,}|\d{4,}\.0)", re.IGNORECASE)
TAZ_EXCLUDED_STATUSES = {"5 CANCELLED", "7 REFUND"}
TAZ_WARRANTY_STATUSES = {"8 WARRANTY"}
WARRANTY_TEXT_RE = re.compile(r"гарант|warranty", re.IGNORECASE)

# Forgotten invoice in TUZ — treat as present for TAZ matching.
TUZ_INVOICE_OVERRIDES: dict[tuple[str, str], str] = {
    ("4232", "761574B"): "16042615005",
}

# Quoted in TUZ 2025; TAZ 2026 order is expected without a 2026 TUZ win.
TAZ_PRIOR_TUZ_PN = frozenset({"2410M50P02", "811390-3"})
TROUBLE_ONLY_REQUEST_RE = re.compile(r"^troubles?$", re.IGNORECASE)
ORDER_NUM_DT_RE = re.compile(r"^№\s*(\d+)$", re.IGNORECASE)

GROUP_SHEETS = [
    "Группа A",
    "Группа B",
    "Группа C",
    "Группа 2 (old)",
    "Группа 3 (old)",
    "Группа 5 (old)",
]
OLD_SHEETS = {"Группа 2 (old)", "Группа 3 (old)", "Группа 5 (old)"}

PRICE_BUCKETS = [
    ("0–10 тыс. $", 0, 10_000),
    ("11–25 тыс. $", 10_001, 25_000),
    ("26–50 тыс. $", 25_001, 50_000),
    ("51–75 тыс. $", 50_001, 75_000),
    ("76–100 тыс. $", 75_001, 100_000),
    ("101–150 тыс. $", 100_001, 150_000),
    ("151–200 тыс. $", 150_001, 200_000),
    ("200+ тыс. $", 200_001, 9_999_999),
]

NOT_FOUND_STATUSES = {
    "1. Не нашли",
    "2. Не будет проквотировано",
    "1. Вне компетенции",
    "1. Пропуск",
}


def is_utair(customer) -> bool:
    if not customer:
        return False
    text = str(customer).lower()
    return "utair" in text or text == "utg" or "компонентс (utg)" in text


def parse_num(value) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if text.startswith("="):
        return None
    text = text.upper()
    if text in {"", "TBA", "N/A", "NA", "#VALUE!", "-", "БЕЗ ЦЕНЫ"}:
        return None
    match = re.search(r"[-+]?\d[\d\s.,]*", text)
    if not match:
        return None
    token = match.group().replace(" ", "").replace(",", "")
    try:
        return float(token)
    except ValueError:
        return None


def to_dt(value) -> datetime | None:
    return value if isinstance(value, datetime) else None


def hours_between(start: datetime | None, end: datetime | None) -> float | None:
    if not start or not end:
        return None
    delta = (end - start).total_seconds() / 3600
    if delta < 0:
        return None
    return delta


def fmt_hours(value: float | None) -> str:
    if value is None:
        return "—"
    if value < 24:
        return f"{value:.0f} ч"
    return f"{value / 24:.1f} д"


def fmt_money(value: float | None) -> str:
    if value is None:
        return "—"
    return f"${value:,.0f}"


def fmt_pct(value: float | None) -> str:
    if value is None:
        return "—"
    return f"{value:.0f}%"


def clean_text(value) -> str | None:
    if value in (None, ""):
        return None
    text = str(value).strip()
    return text if text and text != "-" else None


def meaningful_remarks(text: str | None) -> str | None:
    if not text:
        return None
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    if not lines:
        return None
    first = lines[0].upper()
    if first in CONDITION_ONLY:
        if len(lines) == 1:
            return None
        rest = "\n".join(lines[1:]).strip()
        return rest or None
    return text.strip()


def is_workflow_root(text: str) -> bool:
    low = text.lower()
    if re.search(r"[а-яё]", text, re.IGNORECASE):
        return True
    if len(text) > 35:
        return True
    return any(token in low for token in ("жд", "квота", "апдейт", "док", "актив", "таргет"))


def classify_hold_reason(text: str | None) -> str:
    if not text:
        return "Без комментария"
    low = text.lower()
    for label, pattern in HOLD_REASON_RULES:
        if re.search(pattern, low, re.IGNORECASE):
            return label
    if re.search(r"[а-яё]", text, re.IGNORECASE):
        return "Прочий комментарий"
    return "Без категории"


def request_dt_order_ref(value) -> str | None:
    if value is None or isinstance(value, datetime):
        return None
    text = str(value).strip()
    match = ORDER_NUM_DT_RE.match(text)
    if match:
        return f"№{match.group(1)}"
    return None


def extract_order_refs(*texts: str | None) -> list[str]:
    refs: list[str] = []
    seen: set[str] = set()
    for text in texts:
        if not text:
            continue
        for match in ORDER_REF_RE.finditer(str(text)):
            token = match.group(1)
            if token not in seen:
                seen.add(token)
                refs.append(token)
    return refs


def normalize_invoice(value) -> str | None:
    if value in (None, ""):
        return None
    raw = str(value).strip()
    if " " in raw and len(re.sub(r"\D", "", raw)) <= 8:
        return None
    digits = re.sub(r"\D", "", raw.replace(".0", ""))
    if not digits:
        return None
    if len(digits) < 10 or len(digits) > 14:
        return None
    return digits


def line_invoices(line: RequestLine) -> set[str]:
    invs: set[str] = set()
    for offer in line.offers:
        inv = normalize_invoice(offer.invoice)
        if inv:
            invs.add(inv)
    override = TUZ_INVOICE_OVERRIDES.get((line.request_no, line.pn))
    if override:
        invs.add(override)
    return invs


def normalize_ref_token(value: str | None) -> str | None:
    if not value:
        return None
    text = str(value).strip().splitlines()[0].strip()
    return text or None


def is_taz_warranty_row(row: tuple) -> bool:
    status = str(row[4]).strip() if row[4] else ""
    if status in TAZ_WARRANTY_STATUSES:
        return True
    comment = str(row[5]).strip() if len(row) > 5 and row[5] else ""
    delivery = str(row[23]).strip() if len(row) > 23 and row[23] else ""
    return bool(WARRANTY_TEXT_RE.search(f"{comment} {delivery}"))


def is_prior_year_tuz_pn(pn: str, lines: list[RequestLine]) -> bool:
    if pn in TAZ_PRIOR_TUZ_PN:
        return True
    pn_lines = [line for line in lines if line.pn == pn]
    if not pn_lines or any(line.won() for line in pn_lines):
        return False
    return all(line.sheet in OLD_SHEETS for line in pn_lines)


def pick_canonical_won_line(lines: list[RequestLine]) -> RequestLine:
    def latest_dt(line: RequestLine) -> datetime:
        return max(
            (o.accepted_dt or o.sent_dt or o.quote_dt or datetime.min for o in line.offers),
            default=datetime.min,
        )

    def score(line: RequestLine) -> tuple:
        return (latest_dt(line), len(line_invoices(line)), not line.is_trouble_search())

    return max(lines, key=score)


def build_won_deals(won_lines: list[RequestLine]) -> list[dict[str, Any]]:
    if not won_lines:
        return []

    parent = {id(line): id(line) for line in won_lines}
    req_map: dict[str, list[int]] = defaultdict(list)
    for line in won_lines:
        token = normalize_ref_token(line.request_no) or line.request_no
        req_map[token].append(id(line))

    def find(node: int) -> int:
        while parent[node] != node:
            parent[node] = parent[parent[node]]
            node = parent[node]
        return node

    def union(a: int, b: int) -> None:
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[rb] = ra

    for line in won_lines:
        lid = id(line)
        invs = line_invoices(line)
        for other in won_lines:
            oid = id(other)
            if invs & line_invoices(other):
                union(lid, oid)
            if line.pn == other.pn and (line.is_trouble_search() or other.is_trouble_search()):
                union(lid, oid)
        ref = normalize_ref_token(line.trouble_order_ref())
        if ref:
            for oid in req_map.get(ref, []):
                union(lid, oid)

    groups: dict[int, list[RequestLine]] = defaultdict(list)
    for line in won_lines:
        groups[find(id(line))].append(line)

    deals: list[dict[str, Any]] = []
    for members in groups.values():
        canonical = pick_canonical_won_line(members)
        invoices: set[str] = set()
        for member in members:
            invoices |= line_invoices(member)
        deals.append(
            {
                "lines": members,
                "canonical": canonical,
                "invoices": sorted(invoices),
                "pn": canonical.pn,
                "renegotiation_count": len(members) - 1,
                "tuz_offered": canonical.order_value() or 0,
                "has_trouble": any(member.is_trouble_search() for member in members),
            }
        )
    deals.sort(key=lambda deal: deal["tuz_offered"], reverse=True)
    return deals


def collect_won_invoices(won_lines: list[RequestLine]) -> set[str]:
    invs: set[str] = set()
    for line in won_lines:
        invs |= line_invoices(line)
    invs.discard(None)
    return invs


def taz_row_brief(row: tuple) -> dict[str, Any]:
    work_dt = row[16]
    return {
        "invoice": normalize_invoice(row[0]) or str(row[0]).strip() if row[0] else "—",
        "pn": str(row[10]).strip() if row[10] else "—",
        "description": str(row[12]).strip() if len(row) > 12 and row[12] else "—",
        "amount": taz_line_amount(row),
        "status": str(row[4]).strip() if row[4] else "—",
        "delivery": str(row[23]).strip() if len(row) > 23 and row[23] else "—",
        "comment": str(row[5]).strip().splitlines()[0] if row[5] else "—",
        "work_dt": work_dt.strftime("%d.%m.%Y") if isinstance(work_dt, datetime) else "—",
    }


def taz_line_amount(row: tuple) -> float:
    sale_total = parse_num(row[33])
    if sale_total is not None:
        return sale_total
    qty = parse_num(row[13]) or 1
    sale_ea = parse_num(row[32])
    if sale_ea is not None:
        return sale_ea * qty
    return 0.0


def taz_row_costs(row: tuple) -> dict[str, float]:
    return {
        "revenue": taz_line_amount(row),
        "purchase": parse_num(row[31]) or 0.0,
        "transport_plan": parse_num(row[34]) or 0.0 if len(row) > 34 else 0.0,
        "transport_fact": parse_num(row[35]) or 0.0 if len(row) > 35 else 0.0,
        "fee": parse_num(row[36]) or 0.0 if len(row) > 36 else 0.0,
        "customs": (parse_num(row[37]) or 0.0 if len(row) > 37 else 0.0)
        + (parse_num(row[38]) or 0.0 if len(row) > 38 else 0.0),
    }


def load_taz_money_2026(path: Path) -> dict[str, Any]:
    """Utair TAZ 2026 P&L from ORDERS sheet (excl. Cancel/Refund; warranty separate)."""
    totals = Counter()
    warranty = Counter()
    lines = 0
    warranty_lines = 0
    invoices: set[str] = set()

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["ORDERS"] if "ORDERS" in wb.sheetnames else wb[wb.sheetnames[0]]

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not is_utair(row[6]):
            continue
        work_dt = row[16]
        if not isinstance(work_dt, datetime) or work_dt.year != 2026:
            continue
        status = str(row[4]).strip() if row[4] else "—"
        costs = taz_row_costs(row)
        if status in TAZ_EXCLUDED_STATUSES:
            totals["excluded_revenue"] += costs["revenue"]
            continue
        if is_taz_warranty_row(row):
            warranty_lines += 1
            for key, value in costs.items():
                warranty[key] += value
            continue
        for key, value in costs.items():
            totals[key] += value
        lines += 1
        invoice = normalize_invoice(row[0])
        if invoice:
            invoices.add(invoice)

    wb.close()

    revenue = totals["revenue"]
    margin = (
        revenue
        - totals["purchase"]
        - totals["transport_fact"]
        - totals["fee"]
        - totals["customs"]
    )
    return {
        "revenue": revenue,
        "purchase": totals["purchase"],
        "transport_plan": totals["transport_plan"],
        "transport_fact": totals["transport_fact"],
        "fee": totals["fee"],
        "customs": totals["customs"],
        "margin": margin,
        "margin_pct": (margin / revenue * 100) if revenue else 0.0,
        "lines": lines,
        "invoices": len(invoices),
        "excluded_revenue": totals["excluded_revenue"],
        "warranty_revenue": warranty["revenue"],
        "warranty_lines": warranty_lines,
    }


def bucket_for_sale(value: float | None) -> str | None:
    if value is None:
        return None
    for label, low, high in PRICE_BUCKETS:
        if low <= value <= high:
            return label
    return None


def sheet_columns(sheet: str) -> dict[str, int]:
    if sheet in OLD_SHEETS:
        return {
            "status": 1,
            "request_dt": 2,
            "request_no": 3,
            "urgency": 4,
            "client": 5,
            "pn": 7,
            "alt_pn": 8,
            "description": 9,
            "qty": 10,
            "quote_dt": 12,
            "root_price": 14,
            "root": 15,
            "supplier_price": 16,
            "supplier": 17,
            "transit": 18,
            "cond": 19,
            "ppwk": 23,
            "remarks": 24,
            "offered": 25,
            "sent_dt": 26,
            "accepted_dt": 27,
            "invoice": 28,
        }
    return {
        "status": 1,
        "request_dt": 2,
        "request_no": 3,
        "urgency": 4,
        "client": 5,
        "pn": 7,
        "alt_pn": 8,
        "description": 9,
        "qty": 10,
        "quote_dt": 15,
        "root_price": 17,
        "root": 18,
        "supplier_price": 19,
        "supplier": 20,
        "transit": 21,
        "cond": 22,
        "ppwk": 26,
        "remarks": 27,
        "offered": 28,
        "sent_dt": 29,
        "accepted_dt": 30,
        "invoice": 31,
    }


def sheet_columns_shifted(sheet: str) -> dict[str, int]:
    """Layout with Sales ID / Purch ID inserted before Request date (+2 cols)."""
    if sheet in OLD_SHEETS:
        return sheet_columns(sheet)
    return {
        "status": 1,
        "request_dt": 4,
        "request_no": 5,
        "urgency": 3,
        "client": 7,
        "pn": 9,
        "alt_pn": 10,
        "description": 11,
        "qty": 12,
        "quote_dt": 15,
        "root_price": 17,
        "root": 18,
        "supplier_price": 19,
        "supplier": 20,
        "transit": 21,
        "cond": 22,
        "ppwk": 26,
        "remarks": 27,
        "offered": 28,
        "sent_dt": 29,
        "accepted_dt": 30,
        "invoice": 31,
    }


def format_request_no(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def is_shifted_row(values: list) -> bool:
    """Detect Mary/KDV-style layout: Client in col G, Request № in col E."""
    if len(values) < 8:
        return False
    client_normal = values[4]
    client_shifted = values[6]
    return is_utair(client_shifted) and not is_utair(client_normal)


def resolve_row_columns(sheet: str, values: list) -> dict[str, int]:
    if sheet in {"Группа A", "Группа B", "Группа C"} and is_shifted_row(values):
        return sheet_columns_shifted(sheet)
    return sheet_columns(sheet)


@dataclass
class OfferRow:
    sheet: str
    status: str | None
    request_dt: datetime | None
    quote_dt: datetime | None
    sent_dt: datetime | None
    accepted_dt: datetime | None
    root_price: float | None
    supplier_price: float | None
    transit: float | None
    offered: float | None
    cond: str | None
    invoice: str | None
    request_dt_text: str | None = None
    urgency: str | None = None
    remarks: str | None = None
    ppwk: str | None = None
    root_text: str | None = None
    supplier: str | None = None
    pn_bold: bool = False
    alt_bold: bool = False


@dataclass
class RequestLine:
    sheet: str
    request_no: str
    pn: str
    alt_pn: str | None
    description: str | None
    qty: float | None
    offers: list[OfferRow] = field(default_factory=list)

    @property
    def key(self) -> tuple[str, str, str]:
        return (self.sheet, self.request_no, self.pn)

    def primary_status(self) -> str:
        priority = [
            "7. Клиент согласовал",
            "5. Есть интерес",
            "4. Квотация направлена клиенту",
            "3. Цена на деталь получена",
            "1. Проценка у поставщиков",
            "6. Клиент отказал",
            "8. Backup",
        ]
        statuses = {str(o.status) for o in self.offers if o.status}
        for label in priority:
            if label in statuses:
                return label
        return next(iter(statuses), "—")

    def market_rows(self) -> list[OfferRow]:
        """Supplier quotes on meaningful rows (exclude Backup / not-found statuses)."""
        rows = []
        for offer in self.offers:
            if offer.status in NOT_FOUND_STATUSES or offer.status == "8. Backup":
                continue
            if offer.supplier_price is not None:
                rows.append(offer)
        return rows

    def quote_rows(self) -> list[OfferRow]:
        """All market quotes for the request, including Backup (procurement effort)."""
        rows = []
        for offer in self.offers:
            if offer.status in NOT_FOUND_STATUSES:
                continue
            if offer.supplier_price is not None:
                rows.append(offer)
        return rows

    def has_supplier_price(self) -> bool:
        return len(self.market_rows()) > 0

    def supplier_without_offer(self) -> bool:
        """Supplier quote exists, status past procurement, but no client Offered × Qty."""
        if not self.market_rows() or self.sale_value() is not None:
            return False
        return self.primary_status() not in EARLY_STATUSES

    def workflow_notes(self) -> str | None:
        """Operational comments from PPWK, Remarks, Root (when used as a note)."""
        parts: list[str] = []
        seen: set[str] = set()

        def add(label: str, text: str | None) -> None:
            if not text:
                return
            key = text.strip()
            if key in seen:
                return
            seen.add(key)
            parts.append(f"{label}: {key}")

        for offer in self.offers:
            add("PPWK", clean_text(offer.ppwk))
            add("Remarks", meaningful_remarks(clean_text(offer.remarks)))
            root = clean_text(offer.root_text)
            if root and is_workflow_root(root):
                add("Root", root)

        return " · ".join(parts) if parts else None

    def hold_reason(self) -> str:
        return classify_hold_reason(self.workflow_notes())

    def is_trouble_search(self) -> bool:
        """Repeat procurement when the first source was lost (Trouble / order ref in left cols)."""
        if TROUBLE_RE.search(self.request_no):
            return True
        for offer in self.offers:
            if offer.urgency and TROUBLE_RE.search(offer.urgency):
                return True
            if offer.request_dt_text and request_dt_order_ref(offer.request_dt_text):
                return True
        return False

    def trouble_order_ref(self) -> str | None:
        refs = extract_order_refs(self.request_no)
        if refs:
            return refs[0]
        if self.request_no and not TROUBLE_ONLY_REQUEST_RE.match(self.request_no.strip()):
            if any(offer.urgency and TROUBLE_RE.search(offer.urgency) for offer in self.offers):
                token = self.request_no.strip().splitlines()[0].strip()
                if ORDER_REF_RE.fullmatch(token) or ORDER_NUM_DT_RE.match(token):
                    return token
        for offer in self.offers:
            dt_ref = request_dt_order_ref(offer.request_dt_text)
            if dt_ref:
                return dt_ref
        return None

    def trouble_via(self) -> str | None:
        if TROUBLE_RE.search(self.request_no):
            return "Request №"
        for offer in self.offers:
            if offer.urgency and TROUBLE_RE.search(offer.urgency):
                return "Urgency"
            if offer.request_dt_text and request_dt_order_ref(offer.request_dt_text):
                return "Request date"
        return None

    def trouble_tag_html(self) -> str:
        if not self.is_trouble_search():
            return ""
        ref = self.trouble_order_ref()
        label = "Trouble"
        if ref:
            label += f" → {ref}"
        return f" <span class='tag trouble' title='Повторный поиск: сохраняем уже согласованный заказ'>{html.escape(label)}</span>"

    def has_root_price(self) -> bool:
        return any(o.root_price is not None for o in self.offers)

    def is_not_found(self) -> bool:
        if self.found_on_market():
            return False
        statuses = {o.status for o in self.offers if o.status}
        return bool(statuses & NOT_FOUND_STATUSES) or self.primary_status() in NOT_FOUND_STATUSES

    def is_in_procurement(self) -> bool:
        return self.primary_status() == "1. Проценка у поставщиков" and not self.found_on_market()

    def selected_offer(self) -> OfferRow | None:
        sent = [o for o in self.offers if o.sent_dt and o.offered is not None]
        if sent:
            return sorted(sent, key=lambda o: o.sent_dt)[0]
        priced = [o for o in self.offers if o.offered is not None and o.status not in {"8. Backup"}]
        if priced:
            return sorted(priced, key=lambda o: o.offered)[0]
        priced = [o for o in self.offers if o.offered is not None]
        return priced[0] if priced else None

    def sale_value(self) -> float | None:
        offer = self.selected_offer()
        if not offer or offer.offered is None or not self.qty:
            return None
        return offer.offered * self.qty

    def markup_pct(self) -> float | None:
        offer = self.selected_offer()
        if not offer or offer.offered is None:
            return None
        cost = (offer.supplier_price or 0) + (offer.transit or 0)
        if cost <= 0:
            cost = offer.root_price
        if not cost or cost <= 0:
            return None
        return (offer.offered - cost) / cost * 100

    def won(self) -> bool:
        if any(o.status == "7. Клиент согласовал" for o in self.offers):
            return True
        return any(o.invoice for o in self.offers)

    def found_on_market(self) -> bool:
        """Found = есть Supplier Price на строке, которая не Backup и не «не нашли»."""
        return len(self.market_rows()) > 0

    def order_value(self) -> float | None:
        if not self.won():
            return None
        sale = self.sale_value()
        if sale is not None:
            return sale
        qty = self.qty or 1
        for offer in self.offers:
            if offer.status == "7. Клиент согласовал" and offer.offered is not None:
                return offer.offered * qty
        for offer in self.offers:
            if offer.offered is not None:
                return offer.offered * qty
        for offer in self.market_rows():
            return offer.supplier_price * qty
        return None

    def timing(self) -> dict[str, float | None]:
        request_dt = next((o.request_dt for o in self.offers if o.request_dt), None)
        quote_dts = [o.quote_dt for o in self.offers if o.quote_dt]
        sent_dts = [o.sent_dt for o in self.offers if o.sent_dt]
        quote_dt = min(quote_dts) if quote_dts else None
        sent_dt = min(sent_dts) if sent_dts else None
        return {
            "procurement": hours_between(request_dt, quote_dt),
            "sales": hours_between(quote_dt, sent_dt),
            "total": hours_between(request_dt, sent_dt),
        }


def load_request_lines(path: Path) -> list[RequestLine]:
    wb_vals = openpyxl.load_workbook(path, read_only=False, data_only=True)
    wb_fmt = openpyxl.load_workbook(path, read_only=False, data_only=False)
    grouped: dict[tuple[str, str, str], RequestLine] = {}

    for sheet in GROUP_SHEETS:
        if sheet not in wb_vals.sheetnames:
            continue
        ws_vals = wb_vals[sheet]
        ws_fmt = wb_fmt[sheet]

        for row_vals, row_fmt in zip(ws_vals.iter_rows(min_row=2), ws_fmt.iter_rows(min_row=2)):
            values = [cell.value for cell in row_vals]
            if not values:
                continue

            cols = resolve_row_columns(sheet, values)

            def val(key: str):
                index = cols[key] - 1
                return values[index] if index < len(values) else None

            client = val("client")
            if not is_utair(client):
                continue

            request_no = val("request_no")
            pn = val("pn")
            if request_no in (None, "") or pn in (None, ""):
                continue

            request_no_s = format_request_no(request_no)
            pn_s = str(pn).strip()
            key = (sheet, request_no_s, pn_s)
            if key not in grouped:
                grouped[key] = RequestLine(
                    sheet=sheet,
                    request_no=request_no_s,
                    pn=pn_s,
                    alt_pn=str(val("alt_pn")).strip() if val("alt_pn") not in (None, "") else None,
                    description=str(val("description")).strip() if val("description") else None,
                    qty=parse_num(val("qty")),
                    offers=[],
                )

            pn_cell = row_fmt[cols["pn"] - 1]
            alt_cell = row_fmt[cols["alt_pn"] - 1]
            invoice_raw = val("invoice")
            invoice = normalize_invoice(invoice_raw)

            raw_request_dt = val("request_dt")
            request_dt = to_dt(raw_request_dt)
            request_dt_text = clean_text(raw_request_dt) if request_dt is None else None

            grouped[key].offers.append(
                OfferRow(
                    sheet=sheet,
                    status=str(val("status")).strip() if val("status") else None,
                    request_dt=request_dt,
                    request_dt_text=request_dt_text,
                    quote_dt=to_dt(val("quote_dt")),
                    sent_dt=to_dt(val("sent_dt")),
                    accepted_dt=to_dt(val("accepted_dt")),
                    root_price=parse_num(val("root_price")),
                    supplier_price=parse_num(val("supplier_price")),
                    transit=parse_num(val("transit")),
                    offered=parse_num(val("offered")),
                    cond=str(val("cond")).strip() if val("cond") else None,
                    remarks=clean_text(val("remarks")),
                    ppwk=clean_text(val("ppwk")),
                    root_text=clean_text(val("root")),
                    supplier=clean_text(val("supplier")),
                    invoice=invoice,
                    urgency=clean_text(val("urgency")),
                    pn_bold=bool(pn_cell.font and pn_cell.font.bold),
                    alt_bold=bool(alt_cell.font and alt_cell.font.bold),
                )
            )

    wb_vals.close()
    wb_fmt.close()
    return list(grouped.values())


def load_taz_orders_2026(path: Path) -> dict[str, Any]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["ORDERS"] if "ORDERS" in wb.sheetnames else wb[wb.sheetnames[0]]
    totals = Counter()
    category_totals = Counter()
    status_totals = Counter()
    lines = 0
    invoices: set[str] = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        customer = row[6]
        if not is_utair(customer):
            continue
        work_dt = row[16]
        if not isinstance(work_dt, datetime) or work_dt.year != 2026:
            continue
        amount = taz_line_amount(row)
        status = str(row[4]).strip() if row[4] else "—"
        category = str(row[15]).strip() if row[15] else "—"
        totals["gross"] += amount
        status_totals[status] += amount
        category_totals[category] += amount
        if status in TAZ_EXCLUDED_STATUSES:
            totals[status] += amount
            continue
        totals["net"] += amount
        lines += 1
        if category == "ROTABLE":
            totals["rotable_net"] += amount
        elif category == "EXPENDABLE":
            totals["expendable_net"] += amount
        if row[0] not in (None, ""):
            invoices.add(str(row[0]).strip())

    wb.close()
    return {
        "total": totals["net"],
        "gross_total": totals["gross"],
        "cancelled": totals.get("5 CANCELLED", 0),
        "refund": totals.get("7 REFUND", 0),
        "rotable_net": totals["rotable_net"],
        "expendable_net": totals["expendable_net"],
        "lines": lines,
        "invoices": len(invoices),
        "by_category": dict(category_totals),
        "by_status": dict(status_totals.most_common()),
    }


def reconcile_taz_tuz(lines: list[RequestLine], path: Path) -> dict[str, Any]:
    won_lines = [line for line in lines if line.won()]
    won_deals = build_won_deals(won_lines)
    won_invoices = collect_won_invoices(won_lines)
    won_pn = {line.pn for line in won_lines}
    all_pn = {line.pn for line in lines}

    buckets = Counter()
    counts = Counter()
    bucket_rows: dict[str, list[dict[str, Any]]] = defaultdict(list)
    tuz_taz_2026 = 0.0
    tuz_taz_any_year = 0.0
    invoice_amounts_2026: dict[str, float] = defaultdict(float)
    invoice_amounts_any: dict[str, float] = defaultdict(float)

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["ORDERS"] if "ORDERS" in wb.sheetnames else wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not is_utair(row[6]):
            continue
        amount = taz_line_amount(row)
        invoice = normalize_invoice(row[0])
        work_dt = row[16]
        if invoice:
            invoice_amounts_any[invoice] += amount
            if isinstance(work_dt, datetime) and work_dt.year == 2026:
                invoice_amounts_2026[invoice] += amount
        if not isinstance(work_dt, datetime) or work_dt.year != 2026:
            continue
        status = str(row[4]).strip() if row[4] else ""
        category = str(row[15]).strip() if row[15] else ""
        if category != "ROTABLE" or status in TAZ_EXCLUDED_STATUSES:
            continue
        pn = str(row[10]).strip() if row[10] else ""
        brief = taz_row_brief(row)
        if is_taz_warranty_row(row):
            key = "warranty"
        elif invoice in won_invoices:
            key = "tuz_won_invoice"
        elif is_prior_year_tuz_pn(pn, lines):
            key = "tuz_prior_year"
        elif pn in won_pn:
            key = "tuz_won_pn_no_invoice"
        elif pn in all_pn:
            key = "tuz_open_or_lost"
        else:
            key = "not_in_tuz"
        buckets[key] += amount
        counts[key] += 1
        bucket_rows[key].append(brief)

    for invoice in won_invoices:
        tuz_taz_any_year += invoice_amounts_any.get(invoice, 0)
        tuz_taz_2026 += invoice_amounts_2026.get(invoice, 0)

    wb.close()
    rotable_net = sum(buckets.values())
    return {
        "tuz_won_offered": sum(line.order_value() or 0 for line in won_lines),
        "tuz_won_offered_deals": sum(deal["tuz_offered"] for deal in won_deals),
        "tuz_won_count": len(won_lines),
        "tuz_deal_count": len(won_deals),
        "tuz_taz_2026": tuz_taz_2026,
        "tuz_taz_any_year": tuz_taz_any_year,
        "taz_rotable_buckets": dict(buckets),
        "taz_rotable_counts": dict(counts),
        "taz_rotable_rows": {key: rows for key, rows in bucket_rows.items()},
        "taz_rotable_net": rotable_net,
        "won_deals": won_deals,
    }


def compute_funnel(lines: list[RequestLine]) -> dict[str, Any]:
    active = [line for line in lines if not line.is_trouble_search()]
    total = len(active)
    won_deals = build_won_deals([line for line in lines if line.won()])
    primary_deals = [
        deal
        for deal in won_deals
        if any(not member.is_trouble_search() for member in deal["lines"])
    ]

    def is_sent(line: RequestLine) -> bool:
        return any(offer.sent_dt for offer in line.offers)

    def has_interest(line: RequestLine) -> bool:
        if line.won():
            return True
        return any(offer.status == "5. Есть интерес" for offer in line.offers)

    found = sum(1 for line in active if line.found_on_market())
    sent = sum(1 for line in active if is_sent(line))
    interest = sum(1 for line in active if has_interest(line))
    won = len(primary_deals)

    proc = [line.timing()["procurement"] for line in active if line.timing()["procurement"] is not None]
    sales = [line.timing()["sales"] for line in active if line.timing()["sales"] is not None]
    total_t = [line.timing()["total"] for line in active if line.timing()["total"] is not None]

    def step_pct(value: int) -> float:
        return (value / total * 100) if total else 0

    def conv(value: int, base: int) -> float:
        return (value / base * 100) if base else 0

    return {
        "total": total,
        "found": found,
        "sent": sent,
        "interest": interest,
        "won": won,
        "won_raw": len(won_deals),
        "won_renegotiations": sum(deal["renegotiation_count"] for deal in won_deals),
        "found_pct": step_pct(found),
        "sent_pct": step_pct(sent),
        "interest_pct": step_pct(interest),
        "won_pct": step_pct(won),
        "sent_conv": conv(sent, found),
        "interest_conv": conv(interest, sent),
        "won_conv": conv(won, interest),
        "median_proc": median(proc),
        "median_sales": median(sales),
        "median_total": median(total_t),
    }


def load_won_money_reconciliation(lines: list[RequestLine], path: Path) -> dict[str, Any]:
    won_lines = [line for line in lines if line.won()]
    won_deals = build_won_deals(won_lines)
    primary_deals = won_deals
    won_invoices = collect_won_invoices(won_lines)

    tuz_offered_raw = sum(line.order_value() or 0 for line in won_lines)
    tuz_offered = sum(deal["tuz_offered"] for deal in won_deals)
    tuz_offered_matched = 0.0
    tuz_offered_no_taz = 0.0
    for deal in won_deals:
        value = deal["tuz_offered"]
        if set(deal["invoices"]) & won_invoices:
            tuz_offered_matched += value
        else:
            tuz_offered_no_taz += value

    taz_gross = 0.0
    taz_net = 0.0
    taz_cancelled = 0.0
    cancelled_invoices: list[tuple[str, float]] = []

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["ORDERS"] if "ORDERS" in wb.sheetnames else wb[wb.sheetnames[0]]
    per_inv: dict[str, dict[str, float]] = defaultdict(lambda: {"gross": 0.0, "net": 0.0, "cancelled": 0.0})

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not is_utair(row[6]):
            continue
        work_dt = row[16]
        if not isinstance(work_dt, datetime) or work_dt.year != 2026:
            continue
        invoice = normalize_invoice(row[0])
        if not invoice or invoice not in won_invoices:
            continue
        amount = taz_line_amount(row)
        status = str(row[4]).strip() if row[4] else ""
        per_inv[invoice]["gross"] += amount
        taz_gross += amount
        if status in TAZ_EXCLUDED_STATUSES:
            per_inv[invoice]["cancelled"] += amount
            taz_cancelled += amount
        else:
            per_inv[invoice]["net"] += amount
            taz_net += amount

    for invoice, amounts in per_inv.items():
        if amounts["cancelled"] > 0:
            cancelled_invoices.append((invoice, amounts["cancelled"]))

    wb.close()
    cancelled_invoices.sort(key=lambda item: item[1], reverse=True)

    return {
        "tuz_won_count": len(won_lines),
        "tuz_deal_count": len(won_deals),
        "tuz_won_offered_raw": tuz_offered_raw,
        "tuz_won_offered": tuz_offered,
        "tuz_won_offered_matched": tuz_offered_matched,
        "tuz_won_offered_no_taz": tuz_offered_no_taz,
        "taz_won_gross": taz_gross,
        "taz_won_net": taz_net,
        "taz_won_cancelled": taz_cancelled,
        "cancelled_invoices": cancelled_invoices,
        "won_invoice_count": len(won_invoices),
        "won_deals": won_deals,
        "renegotiation_deals": [deal for deal in won_deals if deal["renegotiation_count"] > 0],
    }


def median(values: list[float]) -> float | None:
    return statistics.median(values) if values else None


def load_taz_invoice_money(path: Path) -> dict[str, dict[str, float]]:
    """Per-invoice TAZ 2026 money (excl. Cancel/Refund and warranty)."""
    by_invoice: dict[str, dict[str, float]] = defaultdict(
        lambda: {
            "revenue": 0.0,
            "purchase": 0.0,
            "transport_fact": 0.0,
            "fee": 0.0,
            "customs": 0.0,
            "margin": 0.0,
        }
    )
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["ORDERS"] if "ORDERS" in wb.sheetnames else wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not is_utair(row[6]):
            continue
        work_dt = row[16]
        if not isinstance(work_dt, datetime) or work_dt.year != 2026:
            continue
        status = str(row[4]).strip() if row[4] else ""
        if status in TAZ_EXCLUDED_STATUSES or is_taz_warranty_row(row):
            continue
        invoice = normalize_invoice(row[0])
        if not invoice:
            continue
        costs = taz_row_costs(row)
        bucket = by_invoice[invoice]
        for key in ("revenue", "purchase", "transport_fact", "fee", "customs"):
            bucket[key] += costs[key]
        bucket["margin"] = (
            bucket["revenue"]
            - bucket["purchase"]
            - bucket["transport_fact"]
            - bucket["fee"]
            - bucket["customs"]
        )
    wb.close()
    return dict(by_invoice)


def normalize_supplier_name(value: str | None) -> str:
    if not value:
        return "—"
    text = str(value).strip()
    if not text or text.upper() in {"TBA", "N/A", "NA", "-", "БЕЗ ЦЕНЫ"}:
        return "—"
    return text


def supplier_mix(group: list[RequestLine], top_n: int = 6) -> list[dict[str, Any]]:
    counts: Counter[str] = Counter()
    for line in group:
        for offer in line.quote_rows():
            counts[normalize_supplier_name(offer.supplier)] += 1
    total = sum(counts.values())
    if not total:
        return []
    ranked = counts.most_common()
    top = ranked[:top_n]
    other = sum(count for _, count in ranked[top_n:])
    rows = [
        {"name": name, "count": count, "pct": count / total * 100}
        for name, count in top
    ]
    if other:
        rows.append({"name": "прочие", "count": other, "pct": other / total * 100})
    return rows


def bucket_money_from_taz(
    group: list[RequestLine], invoice_money: dict[str, dict[str, float]]
) -> dict[str, float]:
    invoices: set[str] = set()
    for line in group:
        if not line.won():
            continue
        invoices |= line_invoices(line)
    revenue = purchase = transport = fee = customs = 0.0
    matched = 0
    for invoice in invoices:
        row = invoice_money.get(invoice)
        if not row:
            continue
        matched += 1
        revenue += row["revenue"]
        purchase += row["purchase"]
        transport += row["transport_fact"]
        fee += row["fee"]
        customs += row["customs"]
    margin = revenue - purchase - transport - fee - customs
    return {
        "revenue": revenue,
        "purchase": purchase,
        "transport_fact": transport,
        "fee": fee,
        "customs": customs,
        "margin": margin,
        "margin_pct": (margin / revenue * 100) if revenue else 0.0,
        "matched_invoices": matched,
    }


def aggregate(lines: list[RequestLine]) -> dict[str, Any]:
    by_bucket: dict[str, list[RequestLine]] = {label: [] for label, _, _ in PRICE_BUCKETS}
    unpriced: list[RequestLine] = []
    invoice_money = load_taz_invoice_money(TAZ_PATH) if TAZ_PATH.exists() else {}

    for line in lines:
        sale = line.sale_value()
        bucket = bucket_for_sale(sale)
        if bucket:
            by_bucket[bucket].append(line)
        else:
            unpriced.append(line)

    def summarize(group: list[RequestLine]) -> dict[str, Any]:
        proc = [t["procurement"] for line in group for t in [line.timing()] if t["procurement"] is not None]
        sales = [t["sales"] for line in group for t in [line.timing()] if t["sales"] is not None]
        total = [t["total"] for line in group for t in [line.timing()] if t["total"] is not None]
        found = sum(1 for line in group if line.found_on_market())
        not_found = sum(1 for line in group if line.is_not_found())
        in_procurement = sum(1 for line in group if line.is_in_procurement())
        supplier_no_offer = sum(1 for line in group if line.supplier_without_offer())
        won = sum(1 for line in group if line.won())
        tuz_orders_total = sum(line.order_value() or 0 for line in group if line.won())
        sent = sum(1 for line in group if any(o.sent_dt for o in line.offers))
        pending_proc = in_procurement
        quote_counts = [len(line.quote_rows()) for line in group]
        quotes_total = sum(quote_counts)
        markups = [line.markup_pct() for line in group if line.markup_pct() is not None]
        money = bucket_money_from_taz(group, invoice_money)
        mix = supplier_mix(group)

        return {
            "count": len(group),
            "found": found,
            "found_pct": (found / len(group) * 100) if group else 0,
            "not_found": not_found,
            "in_procurement": in_procurement,
            "supplier_no_offer": supplier_no_offer,
            "won": won,
            "won_pct": (won / len(group) * 100) if group else 0,
            "tuz_orders_total": tuz_orders_total,
            "revenue": money["revenue"],
            "margin": money["margin"],
            "margin_pct": money["margin_pct"],
            "matched_invoices": money["matched_invoices"],
            "sent": sent,
            "pending_proc": pending_proc,
            "quotes_total": quotes_total,
            "avg_quotes": (quotes_total / len(group)) if group else 0,
            "avg_offers": statistics.mean([c for c in quote_counts if c]) if any(quote_counts) else 0,
            "supplier_mix": mix,
            "median_markup": median(markups),
            "median_proc": median(proc),
            "median_sales": median(sales),
            "median_total": median(total),
            "lines": group,
        }

    taz_money = load_taz_money_2026(TAZ_PATH) if TAZ_PATH.exists() else None
    taz_orders = load_taz_orders_2026(TAZ_PATH) if TAZ_PATH.exists() else None
    reconciliation = reconcile_taz_tuz(lines, TAZ_PATH) if TAZ_PATH.exists() else None
    won_money = load_won_money_reconciliation(lines, TAZ_PATH) if TAZ_PATH.exists() else None
    funnel = compute_funnel(lines)

    overall = summarize(lines)
    if taz_orders:
        overall["orders_total"] = taz_orders["total"]
        overall["orders_gross"] = taz_orders["gross_total"]
        overall["orders_cancelled"] = taz_orders["cancelled"]
        overall["orders_refund"] = taz_orders["refund"]
        overall["orders_rotable"] = taz_orders["rotable_net"]
        overall["orders_expendable"] = taz_orders["expendable_net"]
        overall["taz_lines_2026"] = taz_orders["lines"]
        overall["taz_invoices_2026"] = taz_orders["invoices"]
    else:
        overall["orders_total"] = overall["tuz_orders_total"]
        overall["orders_gross"] = None
        overall["orders_cancelled"] = None
        overall["orders_refund"] = None
        overall["orders_rotable"] = None
        overall["orders_expendable"] = None
        overall["taz_lines_2026"] = None
        overall["taz_invoices_2026"] = None
    if won_money:
        overall["tuz_orders_total"] = won_money["tuz_won_offered"]
        overall["taz_won_net"] = won_money["taz_won_net"]
        overall["taz_won_gross"] = won_money["taz_won_gross"]
        overall["taz_won_cancelled"] = won_money["taz_won_cancelled"]
    buckets = {label: summarize(by_bucket[label]) for label, _, _ in PRICE_BUCKETS}
    buckets["Без продажной оценки"] = summarize(unpriced)

    priced_revenue = sum(buckets[label]["revenue"] for label, _, _ in PRICE_BUCKETS)
    for label, _, _ in PRICE_BUCKETS:
        buckets[label]["revenue_share"] = (
            buckets[label]["revenue"] / priced_revenue * 100 if priced_revenue else 0
        )
    buckets["Без продажной оценки"]["revenue_share"] = 0

    return {
        "generated_at": datetime.now().strftime("%d.%m.%Y %H:%M"),
        "source": TUZ_PATH.name,
        "taz_source": TAZ_PATH.name if TAZ_PATH.exists() else None,
        "overall": overall,
        "buckets": buckets,
        "funnel": funnel,
        "taz_money": taz_money,
        "won_money": won_money,
        "reconciliation": reconciliation,
        "taz_orders": taz_orders,
        "won_deals": reconciliation.get("won_deals") if reconciliation else [],
    }


def render_table_rows(lines: list[RequestLine]) -> str:
    rows = sorted(lines, key=lambda line: line.sale_value() or 0, reverse=True)
    html_rows = []
    for line in rows[:200]:
        offer = line.selected_offer()
        best_supplier = min((o.supplier_price for o in line.market_rows()), default=None)
        display_offer = offer.offered if offer else None
        timing = line.timing()
        alt_flag = ""
        if any(o.alt_bold or o.pn_bold for o in line.offers):
            alt_flag = " <span class='tag alt'>alt P/N</span>"
        html_rows.append(
            "<tr>"
            f"<td>{html.escape(line.request_no)}</td>"
            f"<td><b>{html.escape(line.pn)}</b>{alt_flag}"
            f"{('<div class=\"muted\">ALT: ' + html.escape(line.alt_pn) + '</div>') if line.alt_pn else ''}</td>"
            f"<td>{html.escape(line.description or '—')}</td>"
            f"<td>{line.qty or '—'}</td>"
            f"<td>{html.escape(line.primary_status())}</td>"
            f"<td>{len(line.quote_rows())}</td>"
            f"<td>{fmt_money(display_offer)}"
            f"{('<div class=\"muted\">Supp: ' + fmt_money(best_supplier) + '</div>') if (display_offer is None and best_supplier is not None) else ''}</td>"
            f"<td>{fmt_money(line.sale_value())}</td>"
            f"<td>{fmt_pct(line.markup_pct())}</td>"
            f"<td>{fmt_hours(timing['procurement'])}</td>"
            f"<td>{fmt_hours(timing['sales'])}</td>"
            f"<td>{fmt_hours(timing['total'])}</td>"
            f"<td>{'✓' if line.won() else '—'}</td>"
            "</tr>"
        )
    if len(rows) > 200:
        html_rows.append(
            f"<tr><td colspan='13' class='muted'>… ещё {len(rows)-200} строк</td></tr>"
        )
    return "\n".join(html_rows)


def render_funnel_section(data: dict[str, Any]) -> str:
    f = data["funnel"]
    steps = [
        ("total", "Всего запросов", f["total"], None),
        ("found", "Найдено на рынке", f["found"], f["found_pct"]),
        ("sent", "Отправлено клиенту", f["sent"], f["sent_pct"]),
        ("interest", "Есть интерес", f["interest"], f["interest_pct"]),
        ("won", "Заказов получено", f["won"], f["won_pct"]),
    ]
    parts = []
    for idx, (key, label, value, pct) in enumerate(steps):
        conv = ""
        if idx > 0:
            prev = steps[idx - 1][2]
            c = (value / prev * 100) if prev else 0
            conv = f'<div class="funnel-conv">{c:.0f}% от пред. шага</div>'
        pct_html = f'<div class="funnel-pct">{pct:.0f}% от всех</div>' if pct is not None else ""
        parts.append(
            f"""
            <div class="funnel-step">
              <div class="funnel-num">{value}</div>
              <div class="funnel-label">{html.escape(label)}</div>
              {pct_html}
              {conv}
            </div>
            """
        )
        if idx < len(steps) - 1:
            parts.append('<div class="funnel-arrow" aria-hidden="true">→</div>')

    return f"""
  <div class="panel funnel-panel">
    <h2>Воронка конверсии</h2>
    <p class="lead-sm">Без повторных поисков (Trouble). «Заказов получено» — уникальные сделки (пересогласования Trouble не дублируют). «Найдено» — Supplier Price; «Отправлено» — дата в AC; «Интерес» — «5. Есть интерес» или заказ.</p>
    <div class="funnel">{''.join(parts)}</div>
  </div>
"""


def render_speed_section(data: dict[str, Any]) -> str:
    f = data["funnel"]
    return f"""
  <div class="panel speed-panel">
    <h2>Скорость</h2>
    <p class="lead-sm">Медиана по запросам без Trouble, где есть обе даты на этапе.</p>
    <div class="speed-grid">
      <div class="speed-card">
        <div class="speed-k">B → O</div>
        <div class="speed-v">{fmt_hours(f['median_proc'])}</div>
        <div class="speed-h">внесли запрос → нашли на рынке</div>
      </div>
      <div class="speed-card">
        <div class="speed-k">O → AC</div>
        <div class="speed-v">{fmt_hours(f['median_sales'])}</div>
        <div class="speed-h">нашли на рынке → отправили клиенту</div>
      </div>
      <div class="speed-card accent">
        <div class="speed-k">B → AC</div>
        <div class="speed-v">{fmt_hours(f['median_total'])}</div>
        <div class="speed-h">полный цикл до отправки оффера</div>
      </div>
    </div>
  </div>
"""


def render_money_section(data: dict[str, Any]) -> str:
    money = data.get("taz_money")
    if not money:
        return ""

    excluded = money.get("excluded_revenue") or 0
    warranty_note = ""
    if money.get("warranty_lines"):
        warranty_note = (
            f" Гарантийные поставки ({money['warranty_lines']} строк, "
            f"{fmt_money(money.get('warranty_revenue'))}) — отдельный блок ниже."
        )

    return f"""
  <div class="panel money-panel">
    <h2>Деньги</h2>
    <p class="lead-sm">Источник — <b>ТАЗ</b>, Utair, work date 2026, без Cancel/Refund и гарантийных поставок ({money.get('warranty_lines', 0)} строк, {fmt_money(money.get('warranty_revenue'))}). Исключено отмен: {fmt_money(excluded)}.</p>
    <div class="money-grid money-grid-6">
      <div class="money-card accent">
        <div class="money-k">1. Выручка</div>
        <div class="money-v">{fmt_money(money['revenue'])}</div>
        <div class="money-h">{money['lines']} строк · {money['invoices']} счетов · продажная AH</div>
      </div>
      <div class="money-card">
        <div class="money-k">2. Закупка</div>
        <div class="money-v">{fmt_money(money['purchase'])}</div>
        <div class="money-h">закупка, итого</div>
      </div>
      <div class="money-card">
        <div class="money-k">3. Транспорт</div>
        <div class="money-v">{fmt_money(money['transport_fact'])}</div>
        <div class="money-h">факт · план {fmt_money(money['transport_plan'])}</div>
      </div>
      <div class="money-card">
        <div class="money-k">4. Fee</div>
        <div class="money-v">{fmt_money(money['fee'])}</div>
        <div class="money-h">transaction fee</div>
      </div>
      <div class="money-card">
        <div class="money-k">5. Таможня</div>
        <div class="money-v">{fmt_money(money['customs'])}</div>
        <div class="money-h">пошлина + СВХ</div>
      </div>
      <div class="money-card accent">
        <div class="money-k">6. Маржа</div>
        <div class="money-v">{fmt_money(money['margin'])}</div>
        <div class="money-h">{money['margin_pct']:.1f}% · выручка − закупка − транспорт факт − fee − таможня</div>
      </div>
    </div>
  </div>
"""


def render_taz_rows_table(rows: list[dict[str, Any]]) -> str:
    if not rows:
        return ""
    body = []
    for row in rows:
        body.append(
            "<tr>"
            f"<td>{html.escape(str(row['invoice']))}</td>"
            f"<td><b>{html.escape(str(row['pn']))}</b></td>"
            f"<td>{html.escape(str(row['description']))}</td>"
            f"<td>{fmt_money(row['amount'])}</td>"
            f"<td>{html.escape(str(row['status']))}</td>"
            f"<td>{html.escape(str(row['delivery']))}</td>"
            f"<td class='note-cell'>{html.escape(str(row['comment']))}</td>"
            "</tr>"
        )
    return (
        "<div class='table-wrap'><table>"
        "<thead><tr><th>Invoice</th><th>P/N</th><th>Description</th><th>ТАЗ $</th>"
        "<th>Status</th><th>Тип поставки</th><th>Комментарий</th></tr></thead>"
        f"<tbody>{''.join(body)}</tbody></table></div>"
    )


def render_warranty_section(data: dict[str, Any]) -> str:
    recon = data.get("reconciliation")
    if not recon:
        return ""
    rows = recon.get("taz_rotable_rows", {}).get("warranty", [])
    if not rows:
        return ""
    total = sum(row["amount"] for row in rows)
    return f"""
  <div class="panel warranty-panel">
    <h2>Гарантийные поставки (TAZ ROTABLE 2026)</h2>
    <p class="lead-sm">Замена юнита клиенту вместо ранее поставленного, который не приняли. В ТУЗ это не новый RFQ — отдельный блок, не «открытые заказы».</p>
    <div class="mini-grid" style="margin-bottom:14px">
      <div class="mini warn"><div class="k">Строк</div><div class="v">{len(rows)}</div></div>
      <div class="mini"><div class="k">Сумма</div><div class="v">{fmt_money(total)}</div></div>
    </div>
    {render_taz_rows_table(rows)}
  </div>
"""


def render_renegotiations_section(data: dict[str, Any]) -> str:
    wm = data.get("won_money") or {}
    deals = wm.get("renegotiation_deals") or []
    if not deals:
        return ""

    cards = []
    for deal in deals:
        members = []
        for line in deal["lines"]:
            tag = " <span class='tag trouble'>Trouble</span>" if line.is_trouble_search() else ""
            invs = ", ".join(deal["invoices"]) if deal["invoices"] else "—"
            members.append(
                "<tr>"
                f"<td>{html.escape(line.request_no)}{tag}</td>"
                f"<td>{html.escape(line.primary_status())}</td>"
                f"<td>{fmt_money(line.order_value())}</td>"
                f"<td>{html.escape(', '.join(sorted(line_invoices(line))) or '—')}</td>"
                "</tr>"
            )
        cards.append(
            f"""
            <details class="bucket">
              <summary>
                <span class="bucket-title">{html.escape(deal['pn'])}</span>
                <span class="bucket-meta">{len(deal['lines'])} строк ТУЗ · invoice {html.escape(', '.join(deal['invoices']) or '—')} · итог {fmt_money(deal['tuz_offered'])}</span>
              </summary>
              <div class="bucket-body">
                <div class="table-wrap">
                  <table>
                    <thead><tr><th>Request №</th><th>Статус</th><th>Offered×Qty</th><th>Invoice</th></tr></thead>
                    <tbody>{''.join(members)}</tbody>
                  </table>
                </div>
              </div>
            </details>
            """
        )

    return f"""
  <div class="panel">
    <h2>Пересогласования (Trouble → новый «выигрыш»)</h2>
    <p class="lead-sm">В ТУЗ виден процесс: несколько зелёных строк на одну сделку. В ТАЗ — один итоговый заказ. Здесь сгруппировано {len(deals)} сделок с повторным согласованием; в метриках считается одна сделка.</p>
    {''.join(cards)}
  </div>
"""


def render_supplier_mix(mix: list[dict[str, Any]]) -> str:
    if not mix:
        return "<div class='muted'>Нет квот с заполненным Supplier (T)</div>"
    rows = []
    for item in mix:
        rows.append(
            "<tr>"
            f"<td>{html.escape(item['name'])}</td>"
            f"<td>{item['count']}</td>"
            f"<td>{item['pct']:.0f}%</td>"
            "</tr>"
        )
    return (
        "<div class='table-wrap supplier-mix'>"
        "<table><thead><tr><th>Supplier (T)</th><th>Квот</th><th>Доля</th></tr></thead>"
        f"<tbody>{''.join(rows)}</tbody></table></div>"
    )


def render_bucket_section(label: str, bucket: dict[str, Any], open_first: bool) -> str:
    mix_html = render_supplier_mix(bucket.get("supplier_mix") or [])
    revenue = bucket.get("revenue") or 0
    margin = bucket.get("margin") or 0
    share = bucket.get("revenue_share") or 0
    return f"""
            <details class="bucket" {'open' if open_first else ''}>
              <summary>
                <div class="bucket-head">
                  <div class="bucket-title">{html.escape(label)}</div>
                  <div class="bucket-kpis">
                    <div class="bk"><span class="bk-k">Запросов</span><span class="bk-v">{bucket['count']}</span></div>
                    <div class="bk"><span class="bk-k">Заказов</span><span class="bk-v">{bucket['won']}</span></div>
                    <div class="bk accent"><span class="bk-k">Выручка ТАЗ</span><span class="bk-v">{fmt_money(revenue)}</span></div>
                    <div class="bk accent"><span class="bk-k">Маржа</span><span class="bk-v">{fmt_money(margin)} <small>{bucket.get('margin_pct', 0):.0f}%</small></span></div>
                    <div class="bk"><span class="bk-k">Доля выручки</span><span class="bk-v">{share:.0f}%</span></div>
                    <div class="bk"><span class="bk-k">Квот / запрос</span><span class="bk-v">{bucket.get('avg_quotes', 0):.1f}</span></div>
                    <div class="bk"><span class="bk-k">B→AC</span><span class="bk-v">{fmt_hours(bucket['median_total'])}</span></div>
                  </div>
                </div>
              </summary>
              <div class="bucket-body">
                <div class="mini-grid bucket-summary">
                  <div class="mini"><div class="k">Отправлено клиенту</div><div class="v">{bucket['sent']}</div></div>
                  <div class="mini"><div class="k">Всего квот (T)</div><div class="v">{bucket.get('quotes_total', 0)}</div></div>
                  <div class="mini"><div class="k">Ср. квот на запрос</div><div class="v">{bucket.get('avg_quotes', 0):.1f}</div></div>
                  <div class="mini accent"><div class="k">Выручка ТАЗ</div><div class="v">{fmt_money(revenue)}</div><div class="sub">{bucket.get('matched_invoices', 0)} invoice</div></div>
                  <div class="mini accent"><div class="k">Маржа ТАЗ</div><div class="v">{fmt_money(margin)} <span class="sub">({bucket.get('margin_pct', 0):.0f}%)</span></div></div>
                  <div class="mini"><div class="k">Offered×Qty (ТУЗ)</div><div class="v">{fmt_money(bucket['tuz_orders_total'])}</div></div>
                  <div class="mini"><div class="k">B→O / O→AC / B→AC</div><div class="v">{fmt_hours(bucket['median_proc'])} / {fmt_hours(bucket['median_sales'])} / {fmt_hours(bucket['median_total'])}</div></div>
                  <div class="mini"><div class="k">В проценке</div><div class="v">{bucket['in_procurement']}</div></div>
                </div>
                <div class="supplier-block">
                  <h3>Квоты по Supplier (T)</h3>
                  <p class="lead-sm">Сколько рыночных квот (включая Backup) пришло через какого западного поставщика/канал (колонка T). Root (R) — конечный источник.</p>
                  {mix_html}
                </div>
                <div class="table-wrap">
                  <table>
                    <thead>
                      <tr>
                        <th>Request №</th><th>P/N</th><th>Description</th><th>Qty</th><th>Статус</th>
                        <th>Квот</th><th>Offered/ea</th><th>Sale $</th><th>Наценка</th>
                        <th>B→O</th><th>O→AC</th><th>B→AC</th><th>Заказ</th>
                      </tr>
                    </thead>
                    <tbody>
                      {render_table_rows(bucket['lines'])}
                    </tbody>
                  </table>
                </div>
              </div>
            </details>
            """


def render_html(data: dict[str, Any]) -> str:
    overall = data["overall"]
    bucket_order = [label for label, _, _ in PRICE_BUCKETS] + ["Без продажной оценки"]

    bucket_sections = []
    first = True
    for label in bucket_order:
        bucket = data["buckets"][label]
        if bucket["count"] == 0:
            continue
        bucket_sections.append(render_bucket_section(label, bucket, open_first=first))
        first = False

    return f"""<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="utf-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1"/>
<title>UTair — оценка ТУЗ</title>
<link rel="preconnect" href="https://fonts.googleapis.com"/>
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin/>
<link href="https://fonts.googleapis.com/css2?family=Manrope:wght@400;500;600;700;800&family=Fraunces:opsz,wght@9..144,600;9..144,700&display=swap" rel="stylesheet"/>
<style>
:root {{
  --ink:#1a2332; --ink-soft:#3d4a5c; --muted:#6b7a8f; --line:rgba(26,35,50,.10);
  --paper:#f3f6f8; --panel:rgba(255,255,255,.84); --teal:#0b6e6b; --teal-deep:#084e4c;
  --amber:#c9782a; --danger:#b42318; --ok:#0f7a4c; --shadow:0 16px 40px rgba(26,35,50,.08); --radius:18px;
}}
* {{ box-sizing:border-box; }}
body {{
  margin:0; color:var(--ink); font-family:"Manrope",system-ui,sans-serif;
  background:radial-gradient(1000px 500px at 0% 0%, rgba(11,110,107,.12), transparent 55%), var(--paper);
}}
.wrap {{ max-width:1180px; margin:0 auto; padding:36px 18px 72px; }}
h1 {{ font-family:"Fraunces",Georgia,serif; font-size:clamp(2rem,4vw,3rem); color:var(--teal-deep); margin:0 0 8px; }}
.lead {{ color:var(--ink-soft); max-width:46rem; line-height:1.55; margin:0 0 18px; }}
.meta {{ color:var(--muted); font-size:.88rem; margin-bottom:24px; }}
.funnel-panel {{ margin-top:8px; }}
.funnel {{
  display:flex; flex-wrap:wrap; align-items:stretch; gap:8px; margin-top:8px;
}}
.funnel-step {{
  flex:1 1 140px; min-width:120px; background:#fff; border:1px solid var(--line);
  border-radius:14px; padding:14px 12px; text-align:center;
}}
.funnel-num {{
  font-family:"Fraunces",Georgia,serif; font-size:1.75rem; font-weight:700; color:var(--teal-deep);
}}
.funnel-label {{ font-size:.82rem; font-weight:700; color:var(--ink-soft); margin-top:4px; line-height:1.3; }}
.funnel-pct {{ font-size:.75rem; color:var(--muted); margin-top:6px; }}
.funnel-conv {{ font-size:.72rem; color:var(--teal); font-weight:700; margin-top:4px; }}
.funnel-arrow {{
  display:flex; align-items:center; color:var(--muted); font-size:1.2rem; font-weight:700;
  padding:0 2px;
}}
@media(max-width:720px){{ .funnel-arrow {{ display:none; }} }}
.speed-grid {{ display:grid; grid-template-columns:repeat(3,1fr); gap:12px; }}
@media(max-width:720px){{ .speed-grid {{ grid-template-columns:1fr; }} }}
.speed-card {{
  background:#fff; border:1px solid var(--line); border-radius:14px; padding:16px;
}}
.speed-card.accent {{ background:linear-gradient(135deg,#d7efed,#fff); }}
.speed-k {{ font-size:.72rem; text-transform:uppercase; letter-spacing:.05em; color:var(--muted); font-weight:700; }}
.speed-v {{ font-family:"Fraunces",Georgia,serif; font-size:1.6rem; color:var(--teal-deep); margin-top:4px; }}
.speed-h {{ font-size:.84rem; color:var(--ink-soft); margin-top:6px; line-height:1.4; }}
.money-grid {{ display:grid; grid-template-columns:repeat(3,1fr); gap:12px; margin:14px 0; }}
.money-grid-6 {{ grid-template-columns:repeat(3,1fr); }}
@media(max-width:980px){{ .money-grid, .money-grid-6 {{ grid-template-columns:repeat(2,1fr); }} }}
@media(max-width:560px){{ .money-grid, .money-grid-6 {{ grid-template-columns:1fr; }} }}
.money-card {{
  background:#fff; border:1px solid var(--line); border-radius:14px; padding:14px;
}}
.money-card.accent {{ background:linear-gradient(135deg,#d7efed,#fff); }}
.money-card.warn {{ background:linear-gradient(135deg,#f8ead8,#fff); }}
.money-k {{ font-size:.68rem; text-transform:uppercase; letter-spacing:.04em; color:var(--muted); font-weight:700; }}
.money-v {{ font-family:"Fraunces",Georgia,serif; font-size:1.35rem; color:var(--teal-deep); margin-top:4px; }}
.money-h {{ font-size:.8rem; color:var(--ink-soft); margin-top:4px; }}
.panel {{ background:var(--panel); border:1px solid var(--line); border-radius:var(--radius); padding:18px; box-shadow:var(--shadow); margin:18px 0; }}
.panel h2 {{ font-family:"Fraunces",Georgia,serif; margin:0 0 12px; font-size:1.35rem; }}
.lead-sm {{ color:var(--ink-soft); font-size:.92rem; line-height:1.5; margin:0 0 14px; max-width:52rem; }}
.note {{ font-size:.9rem; color:var(--ink-soft); border-left:3px solid var(--teal); padding:10px 12px; background:rgba(255,255,255,.65); }}
.note-cell {{ max-width:320px; font-size:.8rem; line-height:1.35; white-space:pre-wrap; }}
details.bucket {{ background:var(--panel); border:1px solid var(--line); border-radius:18px; box-shadow:var(--shadow); margin:18px 0; overflow:hidden; }}
details.bucket > summary {{
  list-style:none; cursor:pointer; padding:20px 22px; background:linear-gradient(180deg,#fff,rgba(255,255,255,.72));
}}
details.bucket > summary::-webkit-details-marker {{ display:none; }}
details.bucket[open] > summary {{ border-bottom:1px solid var(--line); }}
.bucket-head {{ display:flex; flex-direction:column; gap:14px; width:100%; }}
.bucket-title {{ font-family:"Fraunces",Georgia,serif; font-size:1.55rem; font-weight:700; color:var(--teal-deep); }}
.bucket-kpis {{ display:grid; grid-template-columns:repeat(7,minmax(96px,1fr)); gap:10px; }}
@media(max-width:980px){{ .bucket-kpis {{ grid-template-columns:repeat(3,1fr); }} }}
@media(max-width:560px){{ .bucket-kpis {{ grid-template-columns:repeat(2,1fr); }} }}
.bk {{ background:#fff; border:1px solid var(--line); border-radius:14px; padding:12px 12px 10px; min-height:74px; }}
.bk.accent {{ background:linear-gradient(135deg,#d7efed,#fff); }}
.bk-k {{ display:block; font-size:.68rem; text-transform:uppercase; letter-spacing:.04em; color:var(--muted); font-weight:700; }}
.bk-v {{ display:block; margin-top:6px; font-size:1.2rem; font-weight:800; color:var(--ink); line-height:1.15; }}
.bk-v small {{ font-size:.78rem; color:var(--muted); font-weight:700; }}
.bucket-meta {{ color:var(--muted); font-size:.86rem; }}
.bucket-body {{ padding:16px 18px 22px; }}
.supplier-block {{ margin:8px 0 18px; }}
.supplier-block h3 {{ margin:0 0 6px; font-size:1rem; color:var(--teal-deep); }}
.supplier-mix {{ max-width:520px; }}
.mini-grid {{ display:grid; grid-template-columns:repeat(4,1fr); gap:10px; margin-bottom:14px; }}
@media(max-width:860px){{ .mini-grid {{ grid-template-columns:repeat(2,1fr); }} }}
.mini {{ background:#fff; border:1px solid var(--line); border-radius:12px; padding:12px 14px; }}
.mini.warn {{ background:#fff7ed; }}
.mini.accent {{ background:linear-gradient(135deg,#d7efed,#fff); }}
.mini .k {{ font-size:.68rem; text-transform:uppercase; letter-spacing:.04em; color:var(--muted); font-weight:700; }}
.mini .v {{ font-size:1.12rem; font-weight:800; margin-top:4px; color:var(--ink); }}
.mini .sub {{ font-size:.82rem; color:var(--muted); font-weight:600; }}
.table-wrap {{ overflow:auto; border:1px solid var(--line); border-radius:12px; }}
table {{ width:100%; border-collapse:collapse; font-size:.82rem; background:#fff; }}
th,td {{ padding:8px 7px; border-bottom:1px solid var(--line); text-align:left; vertical-align:top; }}
th {{ font-size:.68rem; text-transform:uppercase; letter-spacing:.04em; color:var(--muted); background:#f8fafb; position:sticky; top:0; }}
tr:hover td {{ background:#fafcfd; }}
.muted {{ color:var(--muted); font-size:.82rem; }}
.tag {{ display:inline-block; font-size:.68rem; font-weight:700; padding:2px 6px; border-radius:999px; }}
.tag.alt {{ background:#fde68a; color:#7c5a00; }}
.tag.hold {{ background:#e8f0fe; color:#1e4a8a; }}
.tag.trouble {{ background:#fde8e8; color:#9b1c1c; }}
</style>
</head>
<body>
<div class="wrap">
  <h1>UTair — оценка работы по ТУЗ</h1>
  <p class="lead">Анализ проработки запросов Ютэйр по групповым листам ТУЗ: скорость закупок и продаж, качество проценки, наценка и конверсия в заказы. Категории — по продажной стоимости (Offered × Qty).</p>
  <div class="meta">Источник: {html.escape(data['source'])} · сформировано {html.escape(data['generated_at'])} · Questions v2 не включён</div>

  {render_funnel_section(data)}

  {render_speed_section(data)}

  {render_money_section(data)}

  <div class="panel">
    <h2>Как читать отчёт</h2>
    <div class="note">
      <b>Найдено на рынке</b> — есть <b>Supplier Price per unit</b> на строке, которая не Backup и не «не нашли». Root Price <u>не считается</u> рыночным оффером.<br/>
      <b>B→O</b> — от внесения запроса (B) до получения цены с рынка (O). · <b>O→AC</b> — от цены до отправки оффера (AC). · <b>B→AC</b> — полный цикл до отправки.<br/>
      <b>Деньги</b> — только ТАЗ (продажная AH, закупка, транспорт, fee, таможня, маржа). Offered×Qty из ТУЗ в шапке не используется.<br/>
      <b>Категории</b> — по Offered×Qty запроса. В резюме: выручка/маржа ТАЗ по выигранным invoice, ср. квот на запрос (включая Backup), доля Supplier (T).<br/>
      <b>alt P/N</b> — P/N выделен жирным в ТУЗ (часто предложен альтернативный номер).
    </div>
  </div>

  <div class="panel">
    <h2>Категории по продажной стоимости</h2>
    <p class="lead-sm">Сравнивайте не только число заказов, но выручку и маржу: 30 мелких заказов могут давать меньше денег, чем 3 крупных.</p>
    {''.join(bucket_sections)}
  </div>
</div>
</body>
</html>"""


def main():
    lines = load_request_lines(TUZ_PATH)
    data = aggregate(lines)
    OUTPUT_PATH.write_text(render_html(data), encoding="utf-8")
    with zipfile.ZipFile(ZIP_PATH, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.write(OUTPUT_PATH, OUTPUT_PATH.name)
    print(f"Generated {OUTPUT_PATH}")
    print(f"Generated {ZIP_PATH}")
    print(json.dumps({k: data['overall'][k] for k in ['count','found','won','pending_proc']}, ensure_ascii=False))


if __name__ == "__main__":
    main()
