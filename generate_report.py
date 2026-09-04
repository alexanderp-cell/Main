#!/usr/bin/env python3
"""Generate a client status Excel report from a TAZ export."""

from __future__ import annotations

import argparse
import hashlib
import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# Internal TAZ column names
COL_INVOICE = "Номер счета"
COL_INVOICER = "Invoicer"
COL_STATUS = "Status"
COL_CUSTOMER = "Customer"
EXCLUDED_INVOICERS = {"ФЭ"}
COL_COMMENT = "Комментарии"
COL_PURCHASE = "Закупка, итого"
COL_SALE = "Продажная, итого"
COL_TRANSPORT = (
    "Стоимость доставки ПЛАН, за весь счет! Если в счете несколько строк, "
    'то "размазываем" равномерно планируюмую стоиомость транспорта на все позиции из счета.'
)
COL_FEE = "Transaction fee"
COL_PAY1 = "Оплачено клиентом, USD (cntr+shift+V)\nПервая группа платежей"
COL_PAY1_DATE = "Дата оплаты Клиентом\n\nПервая группа платежей"
COL_PAY2 = "Оплачено клиентом, USD (cntr+shift+V)\nЗавершающий платеж"
COL_PAY2_DATE = "Дата оплаты Клиентом\n\nЗавершающий платеж"
COL_BALANCE = "Остаток к оплате клиентом, USD"
COL_PN = "p/n"
COL_DESC = "DESCRIPTION"
COL_DELIVERY_ACTUAL = "ФАКТИЧЕСКАЯ ДАТА ПОСТАВКИ (СОГЛАСНО УСЛОВИЯМ ПОСТАВКИ)"
COL_EXPECTED_ARRIVAL = (
    "ОЖИДАЕМАЯ ДАТА ПРИХОДА (ставим планируемую, и обновляем ее. "
    "Когда появляется # AWB корректируем на дату из AWB)"
)
COL_ORDER_DATE = "ЗАКАЗ ВЗЯТ В РАБОТУ (ДАТА) ОТ КЛИЕНТА"
COL_CATEGORY = "Category"
COL_QTY = "QTY IN PO"
COL_UNIT_PRICE = "Продажная, ед."
COL_DAYS_TO_DELIVER = "Дней на поставку (ЧИСЛО)"
COL_LEAD_TIME = "Lead time"
COL_DEADLINE = "КРАЙНЯЯ ДАТА ПОСТАВКИ"

STATUS_SHIPPED = "3 SHIPPED"
STATUS_FINISHED = "4 FINISHED"
STATUSES_IN_WORK = {"1 NOT PAID", "2 PAID", "6 TROUBLE"}
STATUSES_SHIPPED = {STATUS_SHIPPED, STATUS_FINISHED}
# SHIPPED = still in work; only FINISHED counts as shipped (Aeroflot, S7 Engineering).
CLIENTS_SHIPPED_COUNTS_AS_IN_WORK = {"Аэрофлот", "S7 Engineering"}
CUTE_THEMES = {"lavender_raf", "como_prosecco", "montecarlo_ferrari", "tropical_yacht", "bmw_night", "deep_space"}

COLUMN_RENAME = {
    COL_INVOICE: "№ счета",
    COL_COMMENT: "Комментарий",
    "Sgmt": "Номер группы",
    "p/n": "P/N",
    "QTY IN PO": "QTY",
    "Units (UOM) for customer": "UOM",
    COL_ORDER_DATE: "ЗАКАЗ ВЗЯТ В РАБОТУ",
    "Дней на поставку (ЧИСЛО)": "Дней на поставку",
    COL_SALE: "Продажная, итого (без НДС)",
    COL_TRANSPORT: "Стоимость доставки ПЛАН, за весь счет! ",
    "Пошлина,сбор (ФАКТ) USD.": "Пошлина, сбор(ФАКТ) USD",
    "Дата оплаты Клиентом\n\nПервая группа платежей": "Дата оплаты\n\nПервая группа платежей",
    COL_PAY1: "Оплачено\n\nПервая группа платежей",
    "Дата оплаты Клиентом\n\nЗавершающий платеж": "Дата оплаты\n\nЗавершающий платеж",
    COL_PAY2: "Оплачено\n\nЗавершающий платеж",
    COL_BALANCE: "Остаток к оплате",
}

TOTAL_HEADERS = [
    "Закупка",
    "Транспорт",
    "FEE",
    "Продажа",
    "Маржа",
    "Оплачено клиентом",
    "К оплате клиентом",
]

# --- Styling ---
COLOR_NAVY = "1F3864"
COLOR_BLUE = "2F5597"
COLOR_LIGHT_BLUE = "D9E2F3"
COLOR_GREEN = "548235"
COLOR_LIGHT_GREEN = "E2EFDA"
COLOR_ORANGE = "C55A11"
COLOR_LIGHT_ORANGE = "FCE4D6"
COLOR_YELLOW = "FFF2CC"
COLOR_RED = "C00000"
COLOR_LIGHT_RED = "F8CBAD"
COLOR_WHITE = "FFFFFF"
COLOR_ZEBRA = "F7F9FC"
COLOR_SUBTOTAL = "FFF2CC"
COLOR_BORDER = "B4C6E7"


@dataclass(frozen=True)
class ReportTheme:
    name: str
    navy: str
    blue: str
    light_blue: str
    green: str
    light_green: str
    orange: str
    light_orange: str
    zebra: str
    subtotal: str
    border: str
    tab_total: str
    tab_shipped: str
    tab_in_work: str
    tab_summary: str
    title_font: str = "Calibri"


THEMES = {
    "default": ReportTheme(
        name="default",
        navy=COLOR_NAVY,
        blue=COLOR_BLUE,
        light_blue=COLOR_LIGHT_BLUE,
        green=COLOR_GREEN,
        light_green=COLOR_LIGHT_GREEN,
        orange=COLOR_ORANGE,
        light_orange=COLOR_LIGHT_ORANGE,
        zebra=COLOR_ZEBRA,
        subtotal=COLOR_SUBTOTAL,
        border=COLOR_BORDER,
        tab_total=COLOR_NAVY,
        tab_shipped=COLOR_GREEN,
        tab_in_work=COLOR_ORANGE,
        tab_summary="7030A0",
    ),
    # Soft lavender Raf / milk-lavender palette for Aeroflot.
    "lavender_raf": ReportTheme(
        name="lavender_raf",
        navy="6B5B7A",
        blue="B59AC7",
        light_blue="F3EAF8",
        green="C4A8D4",
        light_green="F7F1FA",
        orange="D4B8D8",
        light_orange="FFF8F2",
        zebra="FCFAFD",
        subtotal="F6ECF8",
        border="E2D4EC",
        tab_total="B59AC7",
        tab_shipped="C4A8D4",
        tab_in_work="D8C4E8",
        tab_summary="E0CDEF",
        title_font="Georgia",
    ),
    # Pastel prosecco on Lake Como promenade — Aeroflot.
    "como_prosecco": ReportTheme(
        name="como_prosecco",
        navy="4A5D6A",
        blue="8FB8C4",
        light_blue="E8F2F4",
        green="A8C5B0",
        light_green="F3F7F2",
        orange="C9A46A",
        light_orange="FBF4EA",
        zebra="FBF8F3",
        subtotal="F3E8D8",
        border="D9D0C4",
        tab_total="8FB8C4",
        tab_shipped="A8C5B0",
        tab_in_work="D4B48A",
        tab_summary="E5D5B8",
        title_font="Georgia",
    ),
    # Ferrari 250 GTO at Casino de Monte-Carlo — Aeroflot weekly theme.
    "montecarlo_ferrari": ReportTheme(
        name="montecarlo_ferrari",
        navy="3B0A14",
        blue="8B1E2D",
        light_blue="F4E7D8",
        green="B08D57",
        light_green="F7F0E4",
        orange="C9A227",
        light_orange="F8EED9",
        zebra="FBF6EE",
        subtotal="F0E0C8",
        border="D4C2A5",
        tab_total="8B1E2D",
        tab_shipped="B08D57",
        tab_in_work="C9A227",
        tab_summary="3B0A14",
        title_font="Georgia",
    ),
    # Luxury sailing yacht in a tropical bay — Aeroflot weekly theme.
    "tropical_yacht": ReportTheme(
        name="tropical_yacht",
        navy="0B3D4C",
        blue="1A7A8C",
        light_blue="E0F4F7",
        green="2E8B6E",
        light_green="E8F6F0",
        orange="D4A017",
        light_orange="FBF3E0",
        zebra="F5FAFB",
        subtotal="E8F2E8",
        border="B8D4D9",
        tab_total="1A7A8C",
        tab_shipped="2E8B6E",
        tab_in_work="D4A017",
        tab_summary="0B3D4C",
        title_font="Georgia",
    ),
    # BMW 8 Series at night by the river — S7 / AK Sibir weekly theme.
    "bmw_night": ReportTheme(
        name="bmw_night",
        navy="0D1B2A",
        blue="1B3A5C",
        light_blue="E8EEF4",
        green="3D5A80",
        light_green="EDF1F7",
        orange="5C7AEA",
        light_orange="E8EDFA",
        zebra="F4F6FA",
        subtotal="DDE4F0",
        border="A8B8CC",
        tab_total="1B3A5C",
        tab_shipped="3D5A80",
        tab_in_work="5C7AEA",
        tab_summary="0D1B2A",
        title_font="Georgia",
    ),
    # Deep space: black holes, galaxies, stars — Aeroflot weekly theme.
    "deep_space": ReportTheme(
        name="deep_space",
        navy="070B1A",
        blue="1A2A6B",
        light_blue="E8ECF8",
        green="3D6B8C",
        light_green="EDF2F8",
        orange="C9A227",
        light_orange="F8F1DC",
        zebra="F3F5FA",
        subtotal="DDE3F2",
        border="A8B0CC",
        tab_total="1A2A6B",
        tab_shipped="3D6B8C",
        tab_in_work="C9A227",
        tab_summary="070B1A",
        title_font="Georgia",
    ),
}


def activate_theme(theme: ReportTheme) -> None:
    global COLOR_NAVY, COLOR_BLUE, COLOR_LIGHT_BLUE, COLOR_GREEN, COLOR_LIGHT_GREEN
    global COLOR_ORANGE, COLOR_LIGHT_ORANGE, COLOR_ZEBRA, COLOR_SUBTOTAL, COLOR_BORDER
    global FONT_TITLE, FONT_SUBTITLE, FONT_SECTION, FONT_HEADER, FONT_BODY, FONT_BODY_BOLD, FONT_KPI
    global FILL_SECTION_WORK, FILL_SECTION_SHIPPED, FILL_HEADER, FILL_SUBTOTAL, FILL_TOTAL_ROW, FILL_KPI, FILL_ALERT
    global THIN, BORDER_THIN, BORDER_BOTTOM

    COLOR_NAVY = theme.navy
    COLOR_BLUE = theme.blue
    COLOR_LIGHT_BLUE = theme.light_blue
    COLOR_GREEN = theme.green
    COLOR_LIGHT_GREEN = theme.light_green
    COLOR_ORANGE = theme.orange
    COLOR_LIGHT_ORANGE = theme.light_orange
    COLOR_ZEBRA = theme.zebra
    COLOR_SUBTOTAL = theme.subtotal
    COLOR_BORDER = theme.border

    FONT_TITLE = Font(name=theme.title_font, size=16, bold=True, color=COLOR_NAVY)
    subtitle_color = {
        "lavender_raf": "6B5B73",
        "como_prosecco": "6A7B78",
        "montecarlo_ferrari": "5C3B2E",
        "tropical_yacht": "2E6B7A",
        "bmw_night": "4A6080",
        "deep_space": "6B7AA8",
    }.get(theme.name, "595959")
    FONT_SUBTITLE = Font(name="Calibri", size=11, color=subtitle_color)
    FONT_SECTION = Font(name="Calibri", size=12, bold=True, color=COLOR_WHITE)
    FONT_HEADER = Font(name="Calibri", size=10, bold=True, color=COLOR_WHITE)
    FONT_BODY = Font(name="Calibri", size=10)
    FONT_BODY_BOLD = Font(name="Calibri", size=10, bold=True)
    FONT_KPI = Font(name=theme.title_font, size=14, bold=True, color=COLOR_NAVY)

    FILL_SECTION_WORK = PatternFill("solid", fgColor=COLOR_BLUE)
    FILL_SECTION_SHIPPED = PatternFill("solid", fgColor=COLOR_GREEN)
    FILL_HEADER = PatternFill("solid", fgColor=COLOR_NAVY)
    FILL_SUBTOTAL = PatternFill("solid", fgColor=COLOR_SUBTOTAL)
    FILL_TOTAL_ROW = PatternFill("solid", fgColor=COLOR_LIGHT_BLUE)
    FILL_KPI = PatternFill("solid", fgColor=COLOR_LIGHT_BLUE)
    FILL_ALERT = PatternFill("solid", fgColor=COLOR_LIGHT_ORANGE)

    THIN = Side(style="thin", color=COLOR_BORDER)
    BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    BORDER_BOTTOM = Border(bottom=Side(style="medium", color=COLOR_NAVY))


activate_theme(THEMES["default"])

FONT_TITLE = Font(name="Calibri", size=16, bold=True, color=COLOR_NAVY)
FONT_SUBTITLE = Font(name="Calibri", size=11, color="595959")
FONT_SECTION = Font(name="Calibri", size=12, bold=True, color=COLOR_WHITE)
FONT_HEADER = Font(name="Calibri", size=10, bold=True, color=COLOR_WHITE)
FONT_BODY = Font(name="Calibri", size=10)
FONT_BODY_BOLD = Font(name="Calibri", size=10, bold=True)
FONT_KPI = Font(name="Calibri", size=14, bold=True, color=COLOR_NAVY)

FILL_SECTION_WORK = PatternFill("solid", fgColor=COLOR_BLUE)
FILL_SECTION_SHIPPED = PatternFill("solid", fgColor=COLOR_GREEN)
FILL_HEADER = PatternFill("solid", fgColor=COLOR_NAVY)
FILL_SUBTOTAL = PatternFill("solid", fgColor=COLOR_SUBTOTAL)
FILL_TOTAL_ROW = PatternFill("solid", fgColor=COLOR_LIGHT_BLUE)
FILL_KPI = PatternFill("solid", fgColor=COLOR_LIGHT_BLUE)
FILL_ALERT = PatternFill("solid", fgColor=COLOR_LIGHT_ORANGE)

THIN = Side(style="thin", color=COLOR_BORDER)
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_BOTTOM = Border(bottom=Side(style="medium", color=COLOR_NAVY))

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

NUM_FMT = '#,##0.00'
DATE_FMT = "DD.MM.YYYY"

CURRENCY_COLUMNS = {
    "Закупка, ед.",
    "Закупка, итого",
    "Продажная, ед.",
    "Продажная, итого (без НДС)",
    "Стоимость доставки ПЛАН, за весь счет! ",
    "Стоимость доставки факт",
    "Transaction fee",
    "Пошлина, сбор(ФАКТ) USD",
    "СВХ стоимость (ФАКТ)",
    "Оплачено\n\nПервая группа платежей",
    "Оплачено\n\nЗавершающий платеж",
    "Остаток к оплате",
}

DATE_COLUMNS = {
    "ЗАКАЗ ВЗЯТ В РАБОТУ",
    "КРАЙНЯЯ ДАТА ПОСТАВКИ",
    "ФАКТИЧЕСКАЯ ДАТА ПОСТАВКИ (СОГЛАСНО УСЛОВИЯМ ПОСТАВКИ)",
    "Дата оплаты\n\nПервая группа платежей",
    "Дата оплаты\n\nЗавершающий платеж",
}

STATUS_FILLS = {
    "1 NOT PAID": PatternFill("solid", fgColor="FFF2CC"),
    "2 PAID": PatternFill("solid", fgColor="E2F0D9"),
    "3 SHIPPED": PatternFill("solid", fgColor="C6E0B4"),
    "4 FINISHED": PatternFill("solid", fgColor="BDD7EE"),
    "6 TROUBLE": PatternFill("solid", fgColor="F8CBAD"),
}

COLUMN_WIDTHS = {
    "№ счета": 14,
    "Invoicer": 8,
    "Менеджер продажи": 14,
    "Менеджер закупки": 14,
    "Status": 12,
    "Комментарий": 24,
    "Customer": 12,
    "Chnl": 8,
    "Номер группы": 10,
    "ТИП ВС (Продажи)": 12,
    "P/N": 16,
    "ALT P/N": 14,
    "DESCRIPTION": 28,
    "QTY": 8,
    "UOM": 6,
    "Category": 12,
    "ЗАКАЗ ВЗЯТ В РАБОТУ": 12,
    "Дней на поставку": 10,
    "Lead time": 10,
    "Destination": 12,
    "КРАЙНЯЯ ДАТА ПОСТАВКИ": 12,
    "ФАКТИЧЕСКАЯ ДАТА ПОСТАВКИ (СОГЛАСНО УСЛОВИЯМ ПОСТАВКИ)": 12,
    "Поставщик": 16,
    "Root supplier": 18,
    "PO #": 14,
    "Закупка, итого": 14,
    "Продажная, итого (без НДС)": 16,
    "Остаток к оплате": 14,
    "Тип оплаты": 18,
}

WEEKLY_NEW_HEADERS = [
    "№ счета",
    "P/N",
    "DESCRIPTION",
    "Category",
    "QTY",
    "Цена за шт, USD",
    "Сумма, USD",
    "Дата взятия в работу",
    "Примечание",
]
WEEKLY_SHIPPED_HEADERS = [
    "№ счета",
    "P/N",
    "DESCRIPTION",
    "Category",
    "QTY",
    "Цена за шт, USD",
    "Сумма, USD",
    "Дней на поставку",
    "Lead time",
    "Крайняя дата поставки",
    "Фактическая дата поставки",
    "Поставка",
    "Примечание",
]
WEEKLY_PAID_HEADERS = [
    "№ счета",
    "P/N",
    "DESCRIPTION",
    "Category",
    "Дата оплаты этап 1",
    "Оплата этап 1, USD",
    "Дата оплаты этап 2",
    "Оплата этап 2, USD",
    "Оплачено за неделю, USD",
    "Остаток к оплате, USD",
    "Примечание",
]
WEEKLY_MONEY_HEADERS = {
    "Цена за шт, USD",
    "Сумма, USD",
    "Оплата этап 1, USD",
    "Оплата этап 2, USD",
    "Оплачено за неделю, USD",
    "Остаток к оплате, USD",
}
WEEKLY_DATE_HEADERS = {
    "Дата взятия в работу",
    "Крайняя дата поставки",
    "Фактическая дата поставки",
    "Дата оплаты этап 1",
    "Дата оплаты этап 2",
}


@dataclass
class WeeklySection:
    title: str
    count: int
    total: float
    rows: list[dict[str, Any]]
    headers: list[str]


@dataclass
class WeeklySummary:
    week_start: date
    week_end: date
    new_orders: WeeklySection
    shipped_orders: WeeklySection
    paid_orders: WeeklySection


def merge_weekly_summaries(items: list[tuple[str, WeeklySummary]]) -> WeeklySummary:
    """Combine weekly sections from several clients; adds Customer column."""
    if not items:
        raise ValueError("merge_weekly_summaries requires at least one summary")
    week_start = items[0][1].week_start
    week_end = items[0][1].week_end

    def merge_section(section_attr: str) -> WeeklySection:
        prototype: WeeklySection = getattr(items[0][1], section_attr)
        headers = ["Customer", *prototype.headers]
        rows: list[dict[str, Any]] = []
        total = 0.0
        count = 0
        for client_name, summary in items:
            section: WeeklySection = getattr(summary, section_attr)
            for row in section.rows:
                rows.append({"Customer": client_name, **row})
            total += section.total
            count += section.count
        return WeeklySection(
            title=prototype.title,
            count=count,
            total=total,
            rows=rows,
            headers=headers,
        )

    return WeeklySummary(
        week_start=week_start,
        week_end=week_end,
        new_orders=merge_section("new_orders"),
        shipped_orders=merge_section("shipped_orders"),
        paid_orders=merge_section("paid_orders"),
    )


def parse_numeric(value: Any) -> float:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text or text.startswith("#"):
        return 0.0
    text = text.replace("\xa0", "").replace(" ", "")
    text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def parse_date(value: Any) -> date | None:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text or text.upper() == "N/A":
        return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d.%m.%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def load_taz(path: Path, excluded_invoicers: set[str] | None = None) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name=0)
    df.columns = [col.strip() if isinstance(col, str) else col for col in df.columns]
    invoice_col = get_invoice_col(df)
    if invoice_col != COL_INVOICE:
        df = df.rename(columns={invoice_col: COL_INVOICE})
    df = df[df[COL_STATUS].notna()]
    df = df[~((df[COL_STATUS] == "1 NOT PAID") & (df[COL_COMMENT] == "SAMPLE"))]
    excluded = EXCLUDED_INVOICERS if excluded_invoicers is None else excluded_invoicers
    if COL_INVOICER in df.columns and excluded:
        invoicer = df[COL_INVOICER].astype(str).str.strip()
        df = df[~invoicer.isin(excluded)]
    return df


def get_invoice_col(df: pd.DataFrame) -> str:
    for col in (COL_INVOICE, "счет", "№ счета"):
        if col in df.columns:
            return col
    raise KeyError(f"Invoice column not found. Columns: {list(df.columns)[:10]}")


def normalize_invoice(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def add_row_key(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out["_key"] = (
        out[COL_INVOICE].map(normalize_invoice)
        + "|"
        + out[COL_PN].astype(str).str.strip()
        + "|"
        + out[COL_DESC].astype(str).str.strip().str.slice(0, 60)
    )
    return out


def report_snapshot(df: pd.DataFrame, client: str = "Utair") -> pd.DataFrame:
    parts = [filter_in_work(df, client), filter_shipped(df, client)]
    if not parts:
        return pd.DataFrame()
    snapshot = pd.concat(parts, ignore_index=True)
    snapshot = add_row_key(snapshot)
    return snapshot.drop_duplicates(subset="_key")


def report_snapshot_keys(df: pd.DataFrame, client: str = "Utair") -> set[str]:
    snap = report_snapshot(df, client)
    if snap.empty:
        return set()
    return set(snap["_key"])


def date_in_range(value: Any, start: date, end: date) -> bool:
    parsed = parse_date(value)
    return parsed is not None and start <= parsed <= end


def row_sale(row: pd.Series) -> float:
    return parse_numeric(row.get(COL_SALE))


def row_qty(row: pd.Series) -> float | int | None:
    value = row.get(COL_QTY)
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    qty = parse_numeric(value)
    return int(qty) if qty.is_integer() else qty


def row_unit_price(row: pd.Series) -> float:
    unit = parse_numeric(row.get(COL_UNIT_PRICE))
    if unit:
        return unit
    qty = parse_numeric(row.get(COL_QTY))
    sale = row_sale(row)
    if qty:
        return sale / qty
    return sale


def resolve_actual_delivery(row: pd.Series) -> tuple[date | None, str]:
    actual = parse_date(row.get(COL_DELIVERY_ACTUAL))
    if actual is not None:
        return actual, "факт"
    expected = parse_date(row.get(COL_EXPECTED_ARRIVAL))
    if expected is not None:
        return expected, "ожидаемая (BC)"
    return None, ""


def delivery_timing_note(deadline: date | None, actual: date | None) -> str:
    if deadline is None or actual is None:
        return "нет данных для сравнения"
    delta = (actual - deadline).days
    if delta > 0:
        return f"просрочка {delta} дн."
    if delta < 0:
        return f"досрочно на {abs(delta)} дн."
    return "в срок"


def build_weekly_summary(
    current_df: pd.DataFrame,
    previous_df: pd.DataFrame,
    week_start: date,
    week_end: date,
    client: str = "Utair",
) -> WeeklySummary:
    current = add_row_key(current_df)
    previous = add_row_key(previous_df)
    prev_by_key = previous.set_index("_key", drop=False)

    new_rows: list[dict[str, Any]] = []
    for _, row in current.iterrows():
        order_date = parse_date(row.get(COL_ORDER_DATE))
        if order_date is None or not (week_start <= order_date <= week_end):
            continue
        if not row_counts_as_in_work(row, client):
            continue
        new_rows.append(
            {
                "№ счета": row[COL_INVOICE],
                "P/N": row[COL_PN],
                "DESCRIPTION": row[COL_DESC],
                "Category": row.get(COL_CATEGORY, ""),
                "QTY": row_qty(row),
                "Цена за шт, USD": row_unit_price(row),
                "Сумма, USD": row_sale(row),
                "Дата взятия в работу": order_date,
                "Примечание": row.get(COL_COMMENT, ""),
            }
        )

    shipped_rows: list[dict[str, Any]] = []
    for _, row in current.iterrows():
        if not row_counts_as_shipped_status(row, client):
            continue
        actual_date, actual_source = resolve_actual_delivery(row)
        delivery_in_week = actual_date is not None and week_start <= actual_date <= week_end

        prev_status = None
        prev_was_shipped = False
        if row["_key"] in prev_by_key.index:
            prev_row = prev_by_key.loc[row["_key"]]
            if isinstance(prev_row, pd.DataFrame):
                prev_row = prev_row.iloc[0]
            prev_status = prev_row[COL_STATUS]
            prev_was_shipped = row_counts_as_shipped_status(prev_row, client)

        status_became_shipped = (not prev_was_shipped) and row_counts_as_shipped_status(row, client)
        if prev_status is None and not delivery_in_week:
            # No previous snapshot match: keep delivery-in-week as the trigger.
            status_became_shipped = False
        if not (status_became_shipped or delivery_in_week):
            continue

        deadline = parse_date(row.get(COL_DEADLINE))
        timing = delivery_timing_note(deadline, actual_date)
        note_parts = []
        if status_became_shipped:
            note_parts.append("статус → отгружено")
        if delivery_in_week:
            note_parts.append("дата поставки в периоде")
        if actual_source == "ожидаемая (BC)":
            note_parts.append("дата из BC")
        shipped_rows.append(
            {
                "№ счета": row[COL_INVOICE],
                "P/N": row[COL_PN],
                "DESCRIPTION": row[COL_DESC],
                "Category": row.get(COL_CATEGORY, ""),
                "QTY": row_qty(row),
                "Цена за шт, USD": row_unit_price(row),
                "Сумма, USD": row_sale(row),
                "Дней на поставку": row.get(COL_DAYS_TO_DELIVER, ""),
                "Lead time": row.get(COL_LEAD_TIME, ""),
                "Крайняя дата поставки": deadline,
                "Фактическая дата поставки": actual_date,
                "Поставка": timing,
                "Примечание": "; ".join(note_parts),
            }
        )

    paid_rows: list[dict[str, Any]] = []
    for _, row in current.iterrows():
        pay1_date = parse_date(row.get(COL_PAY1_DATE))
        pay2_date = parse_date(row.get(COL_PAY2_DATE))
        pay1_amount = parse_numeric(row.get(COL_PAY1))
        pay2_amount = parse_numeric(row.get(COL_PAY2))

        week_pay1 = pay1_amount if pay1_date is not None and week_start <= pay1_date <= week_end else 0.0
        week_pay2 = pay2_amount if pay2_date is not None and week_start <= pay2_date <= week_end else 0.0
        week_paid = week_pay1 + week_pay2
        if week_paid <= 0:
            continue

        notes = []
        if week_pay1:
            notes.append("этап 1 в периоде")
        if week_pay2:
            notes.append("этап 2 в периоде")
        paid_rows.append(
            {
                "№ счета": row[COL_INVOICE],
                "P/N": row[COL_PN],
                "DESCRIPTION": row[COL_DESC],
                "Category": row.get(COL_CATEGORY, ""),
                "Дата оплаты этап 1": pay1_date,
                "Оплата этап 1, USD": pay1_amount,
                "Дата оплаты этап 2": pay2_date,
                "Оплата этап 2, USD": pay2_amount,
                "Оплачено за неделю, USD": week_paid,
                "Остаток к оплате, USD": parse_numeric(row.get(COL_BALANCE)),
                "Примечание": "; ".join(notes),
            }
        )

    return WeeklySummary(
        week_start=week_start,
        week_end=week_end,
        new_orders=WeeklySection(
            title="Новые заказы",
            count=len(new_rows),
            total=sum(r["Сумма, USD"] for r in new_rows),
            rows=new_rows,
            headers=WEEKLY_NEW_HEADERS,
        ),
        shipped_orders=WeeklySection(
            title="Отгруженные заказы",
            count=len(shipped_rows),
            total=sum(r["Сумма, USD"] for r in shipped_rows),
            rows=shipped_rows,
            headers=WEEKLY_SHIPPED_HEADERS,
        ),
        paid_orders=WeeklySection(
            title="Оплаченные клиентом",
            count=len(paid_rows),
            total=sum(r["Оплачено за неделю, USD"] for r in paid_rows),
            rows=paid_rows,
            headers=WEEKLY_PAID_HEADERS,
        ),
    )


def write_weekly_summary_sheet(
    ws,
    client: str,
    summary: WeeklySummary,
    sheet_title: str = "Сводка за неделю",
    cute_comments: bool = False,
    theme_name: str = "default",
) -> None:
    ws.sheet_view.showGridLines = False
    widths = {
        "A": 14,
        "B": 16,
        "C": 28,
        "D": 12,
        "E": 14,
        "F": 14,
        "G": 14,
        "H": 14,
        "I": 14,
        "J": 14,
        "K": 16,
        "L": 18,
        "M": 22,
    }
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width

    ws.merge_cells("A1:M1")
    title = ws["A1"]
    title.value = f"{sheet_title} — {client}"
    if cute_comments and theme_name == "montecarlo_ferrari":
        title.value = f"♦ {sheet_title} — {client} · Monte-Carlo · Ferrari 250 GTO"
    elif cute_comments and theme_name == "como_prosecco":
        title.value = f"🥂 {sheet_title} — {client} · просекко на Комо"
    elif cute_comments and theme_name == "tropical_yacht":
        title.value = f"⛵ {sheet_title} — {client} · яхта · тропический рай · отдых на богатом"
    elif cute_comments and theme_name == "bmw_night":
        title.value = f"🚗 {sheet_title} — {client} · BMW 8 · набережная Казани · кальянная"
    elif cute_comments and theme_name == "deep_space":
        title.value = f"✦ {sheet_title} — {client} · космос · чёрные дыры · галактики · покорение дальних миров"
    elif cute_comments:
        title.value = f"♡ {sheet_title} — {client} · нежный отчёт ♡"
    _style_cell(title, font=FONT_TITLE, alignment=ALIGN_LEFT)

    ws.merge_cells("A2:M2")
    subtitle = ws["A2"]
    subtitle.value = (
        f"Период: {summary.week_start.strftime('%d.%m.%Y')} — "
        f"{summary.week_end.strftime('%d.%m.%Y')}"
    )
    if cute_comments and theme_name == "montecarlo_ferrari":
        epigraphs = [
            "красная 250 GTO у казино — и неделя, которая знает цену себе",
            "золотой свет фасада, влажный асфальт и цифры без блефа",
            "Monte-Carlo не прощает суеты: считаем спокойно, как ставка наверняка",
            "шампанское в купе, ключ в замке зажигания, отчёт на высокой передаче",
            "богатый тон недели: кармин, золото и лёгкий запах бензина класса гран-при",
        ]
        epi = epigraphs[
            int(hashlib.md5(summary.week_end.isoformat().encode()).hexdigest()[:8], 16)
            % len(epigraphs)
        ]
        subtitle.value = f"{subtitle.value}  ·  {epi}"
    elif cute_comments and theme_name == "como_prosecco":
        epigraphs = [
            "бокал на перилах, озеро дышит, цифры сверкают",
            "cin cin: считаем неделю, не расплёскивая пузырьки",
            "набережная Комо, пастель и лёгкий золотой отсвет",
            "сегодня отчёт пахнет лимоном, камнем и просекко",
            "виллы стоят веками — мы просто поднимаем бокал за неделю",
        ]
        epi = epigraphs[
            int(hashlib.md5(summary.week_end.isoformat().encode()).hexdigest()[:8], 16)
            % len(epigraphs)
        ]
        subtitle.value = f"{subtitle.value}  ·  {epi}"
    elif cute_comments and theme_name == "tropical_yacht":
        epigraphs = [
            "парусная яхта в лазурной бухте — и неделя, которая отдыхает на богатом",
            "фуршет на палубе, купюры рядом, цифры без шторма",
            "тропический остров не торопит: считаем спокойно, как капитан в гамаке",
            "шампанское на борт, бриз в парусах, отчёт на круиз-контроле",
            "богатый тон недели: бирюза, золото и лёгкий запах морского бриза",
        ]
        epi = epigraphs[
            int(hashlib.md5(summary.week_end.isoformat().encode()).hexdigest()[:8], 16)
            % len(epigraphs)
        ]
        subtitle.value = f"{subtitle.value}  ·  {epi}"
    elif cute_comments and theme_name == "bmw_night":
        epigraphs = [
            "восьмёрка на кремлёвской набережной — и кальян, который никто не тушит",
            "Куул-Шариф в отражениях, угли тлеют, цифры без суеты",
            "Казань ночью: фары, дым и спокойный круиз-контроль",
            "сапфировый кузов, яблоко-мята, отчёт не спешит",
            "богатый тон недели: тёмно-синий лак, янтарь кальяна и Волга",
        ]
        epi = epigraphs[
            int(hashlib.md5(summary.week_end.isoformat().encode()).hexdigest()[:8], 16)
            % len(epigraphs)
        ]
        subtitle.value = f"{subtitle.value}  ·  {epi}"
    elif cute_comments and theme_name == "deep_space":
        epigraphs = [
            "чёрная дыра на горизонте — и неделя, которая тянет гравитацией цифр",
            "спираль галактики, золотые звёзды и спокойный курс к дальним мирам",
            "покорение дальнего космоса: считаем без паники, как экипаж у шлюза",
            "туманность светится, отчёт держит орбиту, Аэрофлот в гиперпространстве",
            "богатый тон недели: индиго, золото звёзд и лёгкий запах межзвёздной пыли",
        ]
        epi = epigraphs[
            int(hashlib.md5(summary.week_end.isoformat().encode()).hexdigest()[:8], 16)
            % len(epigraphs)
        ]
        subtitle.value = f"{subtitle.value}  ·  {epi}"
    elif cute_comments:
        epigraphs = [
            "сегодняшний раф — с ноткой перрона и терпения",
            "помешиваем цифры медленно, чтобы не сбить пенку",
            "нежный отчёт на связи: дышим, считаем, хвалим",
            "курс лавандовый, турбулентность эмоциональная — низкая",
            "в кадре только забота, заказы и чуть-чуть сливок",
        ]
        epi = epigraphs[
            int(hashlib.md5(summary.week_end.isoformat().encode()).hexdigest()[:8], 16)
            % len(epigraphs)
        ]
        subtitle.value = f"{subtitle.value}  ·  {epi}"
    _style_cell(subtitle, font=FONT_SUBTITLE, alignment=ALIGN_LEFT)

    mood_cards = {}
    mood_images: dict[str, Path] = {}
    if cute_comments:
        if theme_name == "montecarlo_ferrari":
            mood_cards = build_montecarlo_mood_cards(summary)
            cycle = [MONTE_FERRARI, MONTE_CASINO, MONTE_CHAMPAGNE, MONTE_ROULETTE, MONTE_GOLD]
            seed = int(hashlib.md5(summary.week_end.isoformat().encode()).hexdigest()[:8], 16)
            mood_images = {
                "Новые заказы": cycle[seed % len(cycle)],
                "Отгруженные заказы": cycle[(seed + 1) % len(cycle)],
                "Оплаченные клиентом": cycle[(seed + 2) % len(cycle)],
            }
        elif theme_name == "como_prosecco":
            mood_cards = build_como_mood_cards(summary)
            cycle = [COMO_GLASS, COMO_LAKE, COMO_LEMON, COMO_VILLA, COMO_TERRACE]
            seed = int(hashlib.md5(summary.week_end.isoformat().encode()).hexdigest()[:8], 16)
            mood_images = {
                "Новые заказы": cycle[seed % len(cycle)],
                "Отгруженные заказы": cycle[(seed + 1) % len(cycle)],
                "Оплаченные клиентом": cycle[(seed + 2) % len(cycle)],
            }
        elif theme_name == "tropical_yacht":
            mood_cards = build_yacht_mood_cards(summary, client)
            cycle = [YACHT_SAIL, YACHT_BUFFET, YACHT_MONEY, YACHT_BAY, YACHT_CHAMPAGNE]
            seed = int(hashlib.md5(summary.week_end.isoformat().encode()).hexdigest()[:8], 16)
            mood_images = {
                "Новые заказы": cycle[seed % len(cycle)],
                "Отгруженные заказы": cycle[(seed + 1) % len(cycle)],
                "Оплаченные клиентом": cycle[(seed + 2) % len(cycle)],
            }
        elif theme_name == "bmw_night":
            mood_cards = build_bmw_mood_cards(summary, client)
            cycle = [BMW8, BMW_HOOKAH, BMW_RIVER, BMW_LIGHTS]
            seed = int(hashlib.md5(summary.week_end.isoformat().encode()).hexdigest()[:8], 16)
            mood_images = {
                "Новые заказы": cycle[seed % len(cycle)],
                "Отгруженные заказы": cycle[(seed + 1) % len(cycle)],
                "Оплаченные клиентом": cycle[(seed + 2) % len(cycle)],
            }
        elif theme_name == "deep_space":
            mood_cards = build_space_mood_cards(summary, client)
            cycle = [SPACE_BLACKHOLE, SPACE_GALAXY, SPACE_STARS, SPACE_VOYAGE]
            seed = int(hashlib.md5(summary.week_end.isoformat().encode()).hexdigest()[:8], 16)
            mood_images = {
                "Новые заказы": cycle[seed % len(cycle)],
                "Отгруженные заказы": cycle[(seed + 1) % len(cycle)],
                "Оплаченные клиентом": cycle[(seed + 2) % len(cycle)],
            }
        else:
            mood_cards = build_section_mood_cards(summary)
            week_seed = int(hashlib.md5(summary.week_end.isoformat().encode()).hexdigest()[:8], 16)
            image_cycle = [LAVENDER_CUP, LAVENDER_HEART, LAVENDER_CUP]
            if week_seed % 2:
                image_cycle = [LAVENDER_HEART, LAVENDER_CUP, LAVENDER_HEART]
            mood_images = {
                "Новые заказы": image_cycle[0],
                "Отгруженные заказы": image_cycle[1],
                "Оплаченные клиентом": image_cycle[2],
            }

    row = 4
    for section in (summary.new_orders, summary.shipped_orders, summary.paid_orders):
        section_start = row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)
        section_cell = ws.cell(row, 1)
        section_cell.value = section.title
        fill = FILL_SECTION_WORK if section.title == "Новые заказы" else (
            FILL_SECTION_SHIPPED if section.title == "Отгруженные заказы" else PatternFill("solid", fgColor=COLOR_LIGHT_ORANGE)
        )
        font = FONT_SECTION if section.title != "Оплаченные клиентом" else Font(
            name="Calibri", size=12, bold=True, color=COLOR_NAVY
        )
        _style_cell(section_cell, font=font, fill=fill, alignment=ALIGN_LEFT)
        row += 1

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        totals = ws.cell(row, 1)
        label = "Оплачено за период" if section.title == "Оплаченные клиентом" else "Сумма"
        totals.value = f"Количество: {section.count}   |   {label}: {section.total:,.2f} USD"
        _style_cell(totals, font=FONT_BODY_BOLD, fill=FILL_KPI, alignment=ALIGN_LEFT, border=BORDER_THIN)
        row += 1

        headers = section.headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row, col_idx)
            cell.value = header
            _style_cell(cell, font=FONT_HEADER, fill=FILL_HEADER, alignment=ALIGN_CENTER, border=BORDER_THIN)
        row += 1

        if not section.rows:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
            empty = ws.cell(row, 1)
            empty.value = "Нет данных за период"
            _style_cell(empty, font=FONT_BODY, alignment=ALIGN_LEFT, border=BORDER_THIN)
            row += 2
        else:
            for item in section.rows:
                for col_idx, header in enumerate(headers, start=1):
                    cell = ws.cell(row, col_idx)
                    value = item.get(header)
                    cell.value = value
                    number_format = None
                    if header in WEEKLY_MONEY_HEADERS:
                        number_format = NUM_FMT
                        align = ALIGN_RIGHT
                    elif header in WEEKLY_DATE_HEADERS:
                        number_format = DATE_FMT if value else None
                        align = ALIGN_CENTER
                    elif header == "QTY":
                        align = ALIGN_RIGHT
                    else:
                        align = ALIGN_LEFT
                    fill = None
                    if header == "Поставка" and isinstance(value, str) and value.startswith("просрочка"):
                        fill = FILL_ALERT
                    elif header == "Поставка" and isinstance(value, str) and value.startswith("досрочно"):
                        fill = PatternFill("solid", fgColor=COLOR_LIGHT_GREEN)
                    _style_cell(cell, font=FONT_BODY, fill=fill, alignment=align, border=BORDER_THIN, number_format=number_format)
                row += 1
            row += 2

        if cute_comments and section.title in mood_cards:
            default_img = {
                "montecarlo_ferrari": MONTE_FERRARI,
                "como_prosecco": COMO_GLASS,
                "tropical_yacht": YACHT_SAIL,
                "bmw_night": BMW8,
                "deep_space": SPACE_GALAXY,
            }.get(theme_name, LAVENDER_HEART)
            _write_mood_card(
                ws,
                section_start,
                mood_cards[section.title],
                mood_images.get(section.title, default_img),
                theme_name=theme_name,
            )
            row = max(row, section_start + 6)
            if theme_name in {"bmw_night", "tropical_yacht", "deep_space"}:
                row = max(row, section_start + 9)

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)
    note = ws.cell(row, 1)
    note.value = (
        "Оплаты за неделю: AR/AS — дата и сумма этапа 1, AT/AU — дата и сумма этапа 2. "
        "В итог попадает только сумма платежей с датой в периоде. "
        "Для отгрузок при пустой фактической дате используется BC (ожидаемая дата прихода)."
    )
    if cute_comments and theme_name == "montecarlo_ferrari":
        note.value += "  ·  справа — тост за неделю у казино Монте-Карло ♦"
    elif cute_comments and theme_name == "como_prosecco":
        note.value += "  ·  справа — тост за неделю на набережной Комо 🥂"
    elif cute_comments and theme_name == "tropical_yacht":
        note.value += "  ·  справа — тост за неделю на палубе яхты в тропическом раю ⛵"
    elif cute_comments and theme_name == "bmw_night":
        note.value += "  ·  справа — тост за неделю: BMW на набережной Казани и кальянная 🚗"
    elif cute_comments and theme_name == "deep_space":
        note.value += "  ·  справа — тост за неделю в дальнем космосе: чёрные дыры, галактики, звёзды ✦"
    elif cute_comments:
        note.value += "  ·  комментарии справа — нежное резюме недели ♡"
    _style_cell(note, font=FONT_SUBTITLE, alignment=ALIGN_LEFT)


def comment_has_ddp_mow(value: Any) -> bool:
    return bool(re.search(r"\bDDP\s*MOW\b", str(value or ""), flags=re.IGNORECASE))


def client_shipped_counts_as_in_work(client: str) -> bool:
    """AFL / S7: SHIPPED stays in «В работе» until FINISHED."""
    return client in CLIENTS_SHIPPED_COUNTS_AS_IN_WORK


def row_counts_as_in_work(row: pd.Series, client: str) -> bool:
    """Utair: DDP MOW in SHIPPED stays in work until FINISHED.
    Aeroflot / S7: SHIPPED counts as in work; only FINISHED is shipped.
    """
    status = row.get(COL_STATUS)
    if client_shipped_counts_as_in_work(client):
        return status in STATUSES_IN_WORK or status == STATUS_SHIPPED
    if client == "Utair" and status == STATUS_SHIPPED and comment_has_ddp_mow(row.get(COL_COMMENT)):
        return True
    return status in STATUSES_IN_WORK


def row_counts_as_shipped_status(row: pd.Series, client: str) -> bool:
    """Utair: SHIPPED+FINISHED are shipped, except DDP MOW which needs FINISHED.
    Aeroflot / S7: only FINISHED counts as shipped.
    """
    status = row.get(COL_STATUS)
    if client_shipped_counts_as_in_work(client):
        return status == STATUS_FINISHED
    if client == "Utair" and status == STATUS_SHIPPED and comment_has_ddp_mow(row.get(COL_COMMENT)):
        return False
    return status in STATUSES_SHIPPED


def filter_client(df: pd.DataFrame, client: str) -> pd.DataFrame:
    return df[df[COL_CUSTOMER] == client].copy()


def filter_clients(df: pd.DataFrame, clients: list[str]) -> pd.DataFrame:
    return df[df[COL_CUSTOMER].isin(clients)].copy()


def row_customer(row: pd.Series) -> str:
    return str(row.get(COL_CUSTOMER) or "").strip()


def filter_in_work_rows(df: pd.DataFrame) -> pd.DataFrame:
    """In-work filter using each row's Customer (multi-client reports)."""
    if df.empty:
        return df.copy()
    mask = df.apply(lambda row: row_counts_as_in_work(row, row_customer(row)), axis=1)
    return df[mask].copy()


def filter_shipped_rows(df: pd.DataFrame) -> pd.DataFrame:
    """Shipped filter using each row's Customer (multi-client reports)."""
    if df.empty:
        return df.copy()
    mask = df.apply(lambda row: row_counts_as_shipped_status(row, row_customer(row)), axis=1)
    shipped = df[mask].copy()
    shipped["_balance"] = shipped[COL_BALANCE].map(parse_numeric)
    return shipped[shipped["_balance"] != 0].drop(columns="_balance")


def filter_in_work(df: pd.DataFrame, client: str = "Utair") -> pd.DataFrame:
    if df.empty:
        return df.copy()
    mask = df.apply(lambda row: row_counts_as_in_work(row, client), axis=1)
    return df[mask].copy()


def filter_shipped(df: pd.DataFrame, client: str = "Utair") -> pd.DataFrame:
    if df.empty:
        return df.copy()
    mask = df.apply(lambda row: row_counts_as_shipped_status(row, client), axis=1)
    shipped = df[mask].copy()
    shipped["_balance"] = shipped[COL_BALANCE].map(parse_numeric)
    return shipped[shipped["_balance"] != 0].drop(columns="_balance")


def aggregate(df: pd.DataFrame) -> dict[str, float]:
    purchase = df[COL_PURCHASE].map(parse_numeric).sum()
    transport = df[COL_TRANSPORT].map(parse_numeric).sum()
    fee = df[COL_FEE].map(parse_numeric).sum()
    sale = df[COL_SALE].map(parse_numeric).sum()
    paid = df[COL_PAY1].map(parse_numeric).sum() + df[COL_PAY2].map(parse_numeric).sum()
    due = df[COL_BALANCE].map(parse_numeric).sum()
    margin = sale - purchase - transport - fee
    return {
        "Закупка": purchase,
        "Транспорт": transport,
        "FEE": fee,
        "Продажа": sale,
        "Маржа": margin,
        "Оплачено клиентом": paid,
        "К оплате клиентом": due,
    }


def shipped_balance_over_30_days(df: pd.DataFrame, report_date: date) -> float:
    total = 0.0
    for _, row in df.iterrows():
        delivery = parse_date(row.get(COL_DELIVERY_ACTUAL))
        if delivery is None:
            continue
        days = (report_date - delivery).days
        if days > 30:
            total += parse_numeric(row.get(COL_BALANCE))
    return total


def prepare_output_frame(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=list(COLUMN_RENAME.values()))
    out = df.rename(columns=COLUMN_RENAME)
    sort_by: list[str] = []
    if "ЗАКАЗ ВЗЯТ В РАБОТУ" in out.columns:
        out["_sort_order_date"] = pd.to_datetime(out["ЗАКАЗ ВЗЯТ В РАБОТУ"], errors="coerce")
        sort_by.append("_sort_order_date")
    if "№ счета" in out.columns:
        out["_sort_invoice"] = out["№ счета"].astype(str)
        sort_by.append("_sort_invoice")
    if sort_by:
        out = out.sort_values(sort_by, na_position="last")
        out = out.drop(columns=sort_by)
    return out.reset_index(drop=True)


def build_subtotal_row(totals: dict[str, float], columns: list[str]) -> dict[str, Any]:
    row: dict[str, Any] = {col: None for col in columns}
    row["Status"] = "ИТОГО"
    row["Комментарий"] = "Сводка по разделу"
    row["Chnl"] = ""
    row["Номер группы"] = ""
    if "UNIQUE\nUNIT\nCODE" in row:
        row["UNIQUE\nUNIT\nCODE"] = "-"
    value_map = {
        "Закупка, итого": totals["Закупка"],
        "Продажная, итого (без НДС)": totals["Продажа"],
        "Стоимость доставки ПЛАН, за весь счет! ": totals["Транспорт"],
        "Transaction fee": totals["FEE"],
        "Оплачено\n\nПервая группа платежей": totals["Оплачено клиентом"],
        "Остаток к оплате": totals["К оплате клиентом"],
    }
    for col, value in value_map.items():
        if col in row:
            row[col] = value
    return row


def _style_cell(cell, font=None, fill=None, alignment=None, border=None, number_format=None) -> None:
    if font is not None:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if alignment is not None:
        cell.alignment = alignment
    if border is not None:
        cell.border = border
    if number_format is not None:
        cell.number_format = number_format


def _apply_range_border(ws, min_row, max_row, min_col, max_col) -> None:
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = BORDER_THIN


def _set_column_widths(ws, columns: list[str]) -> None:
    for idx, name in enumerate(columns, start=1):
        letter = get_column_letter(idx)
        ws.column_dimensions[letter].width = COLUMN_WIDTHS.get(name, 12)


def _format_detail_sheet(ws, columns: list[str], data_rows: int) -> None:
    last_col = len(columns)
    col_index = {name: idx for idx, name in enumerate(columns, start=1)}

    # Header row
    for col in range(1, last_col + 1):
        cell = ws.cell(1, col)
        _style_cell(cell, font=FONT_HEADER, fill=FILL_HEADER, alignment=ALIGN_CENTER, border=BORDER_THIN)

    # Subtotal row
    for col in range(1, last_col + 1):
        cell = ws.cell(2, col)
        _style_cell(cell, font=FONT_BODY_BOLD, fill=FILL_SUBTOTAL, alignment=ALIGN_LEFT, border=BORDER_THIN)
        header = columns[col - 1]
        if header in CURRENCY_COLUMNS and isinstance(cell.value, (int, float)):
            cell.number_format = NUM_FMT

    # Data rows
    for row_idx in range(3, data_rows + 2):
        zebra = PatternFill("solid", fgColor=COLOR_ZEBRA) if row_idx % 2 == 1 else None
        status = ws.cell(row_idx, col_index.get("Status", 0)).value if "Status" in col_index else None
        status_fill = STATUS_FILLS.get(str(status))

        for col in range(1, last_col + 1):
            cell = ws.cell(row_idx, col)
            header = columns[col - 1]
            fill = status_fill if header == "Status" and status_fill else zebra
            align = ALIGN_RIGHT if header in CURRENCY_COLUMNS or header == "QTY" else ALIGN_LEFT
            number_format = None
            if header in CURRENCY_COLUMNS:
                number_format = NUM_FMT
            elif header in DATE_COLUMNS and isinstance(cell.value, (datetime, date)):
                number_format = DATE_FMT
            _style_cell(
                cell,
                font=FONT_BODY,
                fill=fill,
                alignment=align,
                border=BORDER_THIN,
                number_format=number_format,
            )

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A1:{get_column_letter(last_col)}{data_rows + 1}"
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 22
    _set_column_widths(ws, columns)


def write_detail_sheet(ws, df: pd.DataFrame, totals: dict[str, float]) -> None:
    columns = list(df.columns)
    subtotal = build_subtotal_row(totals, columns)
    ws.append(columns)
    ws.append([subtotal.get(col) for col in columns])
    for row in dataframe_to_rows(df, index=False, header=False):
        ws.append(row)
    _format_detail_sheet(ws, columns, len(df))


def write_total_sheet(
    ws,
    client: str,
    report_date: date,
    in_work_totals: dict[str, float],
    shipped_totals: dict[str, float],
    shipped_over_30: float,
    in_work_count: int,
    shipped_count: int,
    client_stats: list[tuple[str, int, int]] | None = None,
) -> None:
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    for col in range(2, 9):
        ws.column_dimensions[get_column_letter(col)].width = 16

    # Title block
    ws.merge_cells("A1:H1")
    title = ws["A1"]
    if FONT_TITLE.name == "Georgia":
        title.value = f"♡ Отчёт по заказам — {client} ♡"
    else:
        title.value = f"Отчёт по заказам — {client}"
    _style_cell(title, font=FONT_TITLE, alignment=Alignment(horizontal="left", vertical="center"))

    ws.merge_cells("A2:H2")
    subtitle = ws["A2"]
    if FONT_TITLE.name == "Georgia":
        subtitle.value = (
            f"Дата отчёта: {report_date.strftime('%d.%m.%Y')}   |   "
            f"В работе: {in_work_count} поз.   |   Отгружено: {shipped_count} поз.   |   лавандовый раф"
        )
    else:
        subtitle.value = (
            f"Дата отчёта: {report_date.strftime('%d.%m.%Y')}   |   "
            f"В работе: {in_work_count} поз.   |   Отгружено: {shipped_count} поз."
        )
    if client_stats:
        breakdown = " · ".join(f"{name}: {iw} / {sh}" for name, iw, sh in client_stats)
        subtitle.value += f"   |   {breakdown}"
    _style_cell(subtitle, font=FONT_SUBTITLE, alignment=ALIGN_LEFT)

    # KPI cards
    kpi_row = 4
    kpis = [
        ("Продажа (в работе)", in_work_totals["Продажа"], COLOR_LIGHT_BLUE),
        ("К оплате (в работе)", in_work_totals["К оплате клиентом"], COLOR_LIGHT_ORANGE),
        ("Продажа (отгружено)", shipped_totals["Продажа"], COLOR_LIGHT_GREEN),
        ("К оплате (отгружено)", shipped_totals["К оплате клиентом"], COLOR_LIGHT_ORANGE),
    ]
    for i, (label, value, color) in enumerate(kpis):
        col = 1 + i * 2
        label_cell = ws.cell(kpi_row, col)
        value_cell = ws.cell(kpi_row + 1, col)
        ws.merge_cells(start_row=kpi_row, start_column=col, end_row=kpi_row, end_column=col + 1)
        ws.merge_cells(start_row=kpi_row + 1, start_column=col, end_row=kpi_row + 1, end_column=col + 1)
        label_cell.value = label
        value_cell.value = value
        fill = PatternFill("solid", fgColor=color)
        _style_cell(label_cell, font=FONT_BODY_BOLD, fill=fill, alignment=ALIGN_CENTER, border=BORDER_THIN)
        _style_cell(value_cell, font=FONT_KPI, fill=fill, alignment=ALIGN_CENTER, border=BORDER_THIN, number_format=NUM_FMT)

    def write_block(start_row: int, title: str, totals: dict[str, float], section_fill: PatternFill) -> int:
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=8)
        section_cell = ws.cell(start_row, 1)
        section_cell.value = title
        _style_cell(section_cell, font=FONT_SECTION, fill=section_fill, alignment=ALIGN_LEFT)

        header_row = start_row + 1
        ws.cell(header_row, 1).value = "Раздел"
        for idx, header in enumerate(TOTAL_HEADERS, start=2):
            cell = ws.cell(header_row, idx)
            cell.value = header
            _style_cell(cell, font=FONT_HEADER, fill=FILL_HEADER, alignment=ALIGN_CENTER, border=BORDER_THIN)
        _style_cell(ws.cell(header_row, 1), font=FONT_HEADER, fill=FILL_HEADER, alignment=ALIGN_CENTER, border=BORDER_THIN)

        data_row = header_row + 1
        ws.cell(data_row, 1).value = title
        for idx, header in enumerate(TOTAL_HEADERS, start=2):
            cell = ws.cell(data_row, idx)
            cell.value = totals[header]
            fill = FILL_ALERT if header == "К оплате клиентом" else None
            _style_cell(
                cell,
                font=FONT_BODY_BOLD,
                fill=fill,
                alignment=ALIGN_RIGHT,
                border=BORDER_THIN,
                number_format=NUM_FMT,
            )
        _style_cell(ws.cell(data_row, 1), font=FONT_BODY_BOLD, alignment=ALIGN_LEFT, border=BORDER_THIN)
        _apply_range_border(ws, header_row, data_row, 1, 8)
        return data_row + 2

    row = write_block(8, "В работе", in_work_totals, FILL_SECTION_WORK)
    row = write_block(row, "Отгружено", shipped_totals, FILL_SECTION_SHIPPED)

    note_row = row
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=7)
    note = ws.cell(note_row, 1)
    note.value = "Из них отгружено более 30 дней назад (остаток к оплате)"
    _style_cell(note, font=FONT_BODY_BOLD, fill=FILL_ALERT, alignment=ALIGN_RIGHT, border=BORDER_THIN)

    alert_cell = ws.cell(note_row, 8)
    alert_cell.value = shipped_over_30
    _style_cell(alert_cell, font=FONT_KPI, fill=FILL_ALERT, alignment=ALIGN_RIGHT, border=BORDER_THIN, number_format=NUM_FMT)


def load_report_snapshot_as_previous(path: Path) -> pd.DataFrame:
    frames: list[pd.DataFrame] = []
    for sheet in ("В работе", "Отгружено"):
        df = pd.read_excel(path, sheet_name=sheet)
        status_col = "Status" if "Status" in df.columns else COL_STATUS
        df = df[df[status_col] != "ИТОГО"]
        rename_map: dict[str, str] = {status_col: COL_STATUS}
        for src, dst in (
            ("№ счета", COL_INVOICE),
            ("счет", COL_INVOICE),
            ("P/N", COL_PN),
            ("Комментарий", COL_COMMENT),
            ("Продажная, итого (без НДС)", COL_SALE),
            ("Остаток к оплате", COL_BALANCE),
            ("ЗАКАЗ ВЗЯТ В РАБОТУ", COL_ORDER_DATE),
            ("ФАКТИЧЕСКАЯ ДАТА ПОСТАВКИ (СОГЛАСНО УСЛОВИЯМ ПОСТАВКИ)", COL_DELIVERY_ACTUAL),
        ):
            if src in df.columns:
                rename_map[src] = dst
        frames.append(df.rename(columns=rename_map))
    if not frames:
        return pd.DataFrame()
    merged = pd.concat(frames, ignore_index=True)
    if COL_INVOICE not in merged.columns:
        raise KeyError(f"Previous report {path} does not contain invoice column")
    return merged


ASSETS_DIR = Path(__file__).resolve().parent / "assets"
LAVENDER_BANNER = ASSETS_DIR / "banner_wide.png"
LAVENDER_CUP = ASSETS_DIR / "cup_small.png"
LAVENDER_PATTERN = ASSETS_DIR / "pattern_bg.png"
LAVENDER_HEART = ASSETS_DIR / "heart_small.png"
COMO_BANNER = ASSETS_DIR / "como-banner-wide.jpg"
COMO_PROMENADE = ASSETS_DIR / "como-promenade.jpg"
COMO_PATTERN = ASSETS_DIR / "como-pattern.jpg"
COMO_GLASS = ASSETS_DIR / "como-glass-small.jpg"
COMO_LAKE = ASSETS_DIR / "como-lake-small.jpg"
COMO_LEMON = ASSETS_DIR / "como-lemon-small.jpg"
COMO_VILLA = ASSETS_DIR / "como-villa-small.jpg"
COMO_TERRACE = ASSETS_DIR / "como-terrace-small.jpg"
MONTE_BANNER = ASSETS_DIR / "montecarlo-banner-wide.jpg"
MONTE_PROMENADE = ASSETS_DIR / "montecarlo-promenade.jpg"
MONTE_FERRARI = ASSETS_DIR / "montecarlo-ferrari-small.jpg"
MONTE_CASINO = ASSETS_DIR / "montecarlo-casino-small.jpg"
MONTE_CHAMPAGNE = ASSETS_DIR / "montecarlo-champagne-small.jpg"
MONTE_ROULETTE = ASSETS_DIR / "montecarlo-roulette-small.jpg"
MONTE_GOLD = ASSETS_DIR / "montecarlo-gold-small.jpg"
YACHT_BANNER = ASSETS_DIR / "yacht-banner-wide.jpg"
YACHT_SAIL = ASSETS_DIR / "yacht-sail-small.jpg"
YACHT_BUFFET = ASSETS_DIR / "yacht-buffet-small.jpg"
YACHT_MONEY = ASSETS_DIR / "yacht-money-small.jpg"
YACHT_BAY = ASSETS_DIR / "yacht-bay-small.jpg"
YACHT_CHAMPAGNE = ASSETS_DIR / "yacht-champagne-small.jpg"
BMW_BANNER = ASSETS_DIR / "bmw-banner-wide.jpg"
BMW8 = ASSETS_DIR / "bmw8-small.jpg"
BMW_RIVER = ASSETS_DIR / "bmw-kazan-small.jpg"
BMW_LIGHTS = ASSETS_DIR / "bmw-lights-small.jpg"
BMW_HOOKAH = ASSETS_DIR / "bmw-hookah-small.jpg"
BMW_PROMENADE = ASSETS_DIR / "bmw-promenade.jpg"
SPACE_BANNER = ASSETS_DIR / "space-banner-wide.jpg"
SPACE_BLACKHOLE = ASSETS_DIR / "space-blackhole-small.jpg"
SPACE_GALAXY = ASSETS_DIR / "space-galaxy-small.jpg"
SPACE_STARS = ASSETS_DIR / "space-stars-small.jpg"
SPACE_VOYAGE = ASSETS_DIR / "space-voyage-small.jpg"
SPACE_NEBULA = ASSETS_DIR / "space-nebula-wide.jpg"


def _pick_variant(seed: str, options: list[tuple[str, str]]) -> tuple[str, str]:
    """Stable per-week pick so comments change from week to week."""
    digest = hashlib.md5(seed.encode("utf-8")).hexdigest()
    idx = int(digest[:8], 16) % len(options)
    return options[idx]


def _fmt_money(value: float) -> str:
    return f"{value:,.0f}".replace(",", " ")


def _avg_late_days(shipped_rows: list[dict[str, Any]]) -> tuple[int, int, int]:
    late_days: list[int] = []
    for row in shipped_rows:
        note = str(row.get("Поставка", ""))
        if note.startswith("просрочка"):
            m = re.search(r"(\d+)", note)
            if m:
                late_days.append(int(m.group(1)))
    if not late_days:
        return 0, 0, 0
    return len(late_days), int(sum(late_days) / len(late_days)), max(late_days)


def _top_sale_line(rows: list[dict[str, Any]]) -> tuple[str, float]:
    best_desc, best_amt = "", 0.0
    for row in rows:
        amt = float(row.get("Сумма, USD") or 0)
        if amt >= best_amt:
            best_amt = amt
            desc = str(row.get("DESCRIPTION") or row.get("P/N") or "позиция").strip()
            best_desc = re.sub(r"\s+", " ", desc)[:42]
    return best_desc, best_amt


def _category_mix(rows: list[dict[str, Any]]) -> str:
    cats: dict[str, int] = {}
    for row in rows:
        cat = str(row.get("Category") or "").strip().upper() or "OTHER"
        cats[cat] = cats.get(cat, 0) + 1
    if not cats:
        return ""
    top = sorted(cats.items(), key=lambda x: (-x[1], x[0]))
    bits = [f"{name.lower()} ×{n}" for name, n in top[:2]]
    return ", ".join(bits)


def build_section_mood_cards(summary: WeeklySummary) -> dict[str, dict[str, str]]:
    """Cute narrative cards for weekly sections — fresh copy every week."""
    cards: dict[str, dict[str, str]] = {}
    week_tag = summary.week_end.strftime("%Y-%m-%d")
    period = f"{summary.week_start.strftime('%d.%m')}–{summary.week_end.strftime('%d.%m')}"

    new = summary.new_orders
    amounts = [float(r.get("Сумма, USD") or 0) for r in new.rows]
    max_amount = max(amounts) if amounts else 0.0
    avg_check = (new.total / new.count) if new.count else 0.0
    top_desc, top_amt = _top_sale_line(new.rows)
    mix = _category_mix(new.rows)
    money = _fmt_money(new.total)
    max_m = _fmt_money(max_amount)
    avg_m = _fmt_money(avg_check)

    if new.count == 0:
        options = [
            (
                "☕ Тихая чашечка",
                f"За {period} новых заказов нет — бывает и такая погода. "
                "Идеальный момент прогреть лавандовый раф и набраться сил 💜",
            ),
            (
                "🌸 Пауза между рейсами",
                "Новых позиций на этой неделе нет. Нежный отчёт шепчет: "
                "даже Аэрофлоту иногда нужна мягкая посадка ✈️",
            ),
            (
                "🫧 Пена без заказа",
                "Тишина в продажах. Завариваем раф, разминаем крылья — "
                "следующая неделя уже рулит к нам 💜",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|sales|empty", options)
    elif max_amount >= 30000 or new.total >= 150000:
        options = [
            (
                "✨ Мы молодцы!",
                f"Продажи за {period}: {money} USD · {new.count} поз. "
                f"Звезда недели — «{top_desc}» на {max_m} USD. "
                "Нежный отчёт аплодирует стоя 🥰✈️",
            ),
            (
                "🛫 Взлётная полоса продаж",
                f"{new.count} новых заказов на {money} USD (средний чек ~{avg_m}). "
                f"Крупняк до {max_m} USD — крылья Аэрофлота ловят попутный ветер 💜",
            ),
            (
                "💜 Раф с двойной порцией",
                f"Неделя со вкусом победы: {money} USD. "
                f"В меню — {mix or 'микс позиций'}, а топ — «{top_desc}». "
                "Так держать, команда! ✨",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|sales|big", options)
    elif max_amount < 5000 and new.total < 25000:
        options = [
            (
                "🪙 Копейка рубль бережёт",
                f"Аккуратная неделя: {new.count} поз. на {money} USD. "
                "Мелкие заказы — как сахарная пудра на рафе: мало, но мило 🥛",
            ),
            (
                "🌱 Маленький, но свой урожай",
                f"{money} USD по {new.count} позициям. Без фанфар — зато честно. "
                "Нежный отчёт шепчет: тихий рост тоже рост 🌷",
            ),
            (
                "🧁 Мини-десерт недели",
                f"Чек скромный ({avg_m} USD в среднем), но портфель не пустой. "
                f"За {period} набралось {money} USD — и это уже повод улыбнуться 💜",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|sales|small", options)
    else:
        options = [
            (
                "🌷 Ровный эшелон",
                f"{new.count} новых позиций на {money} USD за {period}. "
                f"Средний чек ~{avg_m}; топ — «{top_desc}» ({_fmt_money(top_amt)} USD). "
                "Без турбулентности — только лавандовый курс 💜",
            ),
            (
                "✈️ Крейсерский режим",
                f"Продажи: {money} USD · {new.count} поз."
                + (f" · микс: {mix}." if mix else ".")
                + " Не взрыв, не штиль — комфортная высота для нежного отчёта ✨",
            ),
            (
                "🥛 Пенка средней плотности",
                f"За неделю набралось {money} USD. "
                f"Самый заметный след — «{top_desc}» на {_fmt_money(top_amt)} USD. "
                "Аэрофлот в деле, раф остывать не успевает 🥰",
            ),
            (
                "🌤️ Мягкая ясная неделя",
                f"{new.count} заказов, {money} USD. "
                "Ни бури, ни скуки — именно такой прогноз любит нежный отчёт 💜🌸",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|sales|mid", options)
    cards["Новые заказы"] = {"title": title, "body": body}

    shipped = summary.shipped_orders
    late = sum(1 for r in shipped.rows if str(r.get("Поставка", "")).startswith("просрочка"))
    early = sum(1 for r in shipped.rows if str(r.get("Поставка", "")).startswith("досрочно"))
    on_time = max(shipped.count - late - early, 0)
    late_n, late_avg, late_max = _avg_late_days(shipped.rows)
    ship_money = _fmt_money(shipped.total)

    if shipped.count == 0:
        options = [
            (
                "📦 Склад в режиме spa",
                "Отгрузок за период нет. Всё на полочках, лаванда в воздухе — можно выдохнуть 🌸",
            ),
            (
                "🛋️ Выходной у рампы",
                "Ни одной отгрузки. Нежный отчёт ставит чашку на стол и говорит: отдых тоже логистика ☕",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|ship|empty", options)
    elif shipped.count and late / shipped.count >= 0.5:
        options = [
            (
                "🌙 Главное — долетели",
                f"Из {shipped.count} отгрузок ({ship_money} USD) с опозданием {late}"
                + (f", в среднем на ~{late_avg} дн." if late_avg else "")
                + (f", рекорд {late_max} дн." if late_max else "")
                + ". Дышим глубже: детали всё равно нашли путь к клиенту. "
                "Лаванда не торопит — она сопровождает 🫶💜",
            ),
            (
                "🛟 Спокойно, мы на связи",
                f"{late} из {shipped.count} пришли позже плана — бывает на больших маршрутах. "
                f"Отгружено на {ship_money} USD. Нежный отчёт обнимает команду и шепчет: "
                "лучше поздно с заботой, чем рано с тревогой 🌸",
            ),
            (
                "🌧️ Дождик на перроне",
                f"Просрочек много ({late}), но самолёты тоже ждут погоду. "
                f"За неделю ушло {shipped.count} поз. на {ship_money} USD. "
                "Мы рядом, клиент в курсе, раф тёплый — значит, всё поправимо ☔️💜",
            ),
            (
                "🕯️ Мягкий свет в конце перрона",
                f"Да, график гуляет (до {late_max} дн. у самой «задумчивой» позиции). "
                f"Но {shipped.count} отгрузок уже в пути/у клиента. "
                "Не ругаем календарь — хвалим упорство ✨",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|ship|late", options)
    elif early > late:
        options = [
            (
                "🏁 Быстрее пенки на рафе!",
                f"Досрочно: {early} из {shipped.count}. "
                f"Ещё {on_time} в срок, опозданий всего {late}. "
                "Неделя на позитиве — как утренний лавандовый раф перед вылетом ☀️💜",
            ),
            (
                "🚀 Ранний слот",
                f"{early} позиций приехали раньше срока. "
                f"Отгружено на {ship_money} USD — нежный отчёт ставит лайк экипажу ✈️✨",
            ),
            (
                "💎 Пунктуальность со сливками",
                f"График почти мурлычет: досрочно {early}, в срок {on_time}. "
                "Так и хочется добавить второй шотик эспрессо в раф ☕🥰",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|ship|early", options)
    else:
        options = [
            (
                "🚚 Ровный ритм перрона",
                f"Отгружено {shipped.count} поз. на {ship_money} USD "
                f"(досрочно {early}, в срок {on_time}, позже {late}). "
                "Мягкий, уверенный ход — фирменный стиль нежного отчёта ✨",
            ),
            (
                "🧭 Курс стабильный",
                f"{shipped.count} отгрузок за {period}. Без драм, с результатом на {ship_money} USD. "
                "Иногда лучший комментарий — «всё штатно» 💜",
            ),
            (
                "🌸 Логистика с ароматом",
                f"Неделя отгрузок: {ship_money} USD. "
                f"Баланс сроков — рано {early} / вовремя {on_time} / позже {late}. "
                "Дышим ровно, летим дальше ✈️",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|ship|ok", options)
    cards["Отгруженные заказы"] = {"title": title, "body": body}

    paid = summary.paid_orders
    paid_money = _fmt_money(paid.total)
    pay_amounts = [float(r.get("Оплачено за неделю, USD") or 0) for r in paid.rows]
    max_pay = max(pay_amounts) if pay_amounts else 0.0
    pay_share = (paid.total / new.total * 100) if new.total > 0 and paid.total > 0 else 0.0

    if paid.count == 0 or paid.total <= 0:
        options = [
            (
                "🫧 Оплаты на паузе",
                "Пока без платежей за период. Не грустим — "
                "тишина нужна, чтобы следующий раф был ещё вкуснее ☕",
            ),
            (
                "🎧 Ждём радиообмена",
                "Касса молчит, но это не финал. Нежный отчёт ставит напоминалку "
                "и наливает ещё глоточек лаванды 💜",
            ),
            (
                "🌙 Ночная стоянка денег",
                "Платежей нет — бывает. Зато есть повод пересчитать звёзды и планы на следующую неделю ✨",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|pay|empty", options)
    elif paid.total >= 100000:
        options = [
            (
                "💸 Касса звенит в лаванде!",
                f"За {period} пришло {paid_money} USD по {paid.count} позициям "
                f"(крупнейший платёж ~{_fmt_money(max_pay)}). "
                "Самый сливочный момент нежного отчёта — спасибо, Аэрофлот 🥰💜",
            ),
            (
                "🏦 Дождь из бонусных миль… почти",
                f"{paid_money} USD на счёте настроения. "
                f"{paid.count} платежей, и каждый — как тёплый след на перроне. "
                "Мы это отмечаем двойной пенкой ✨☕",
            ),
            (
                "🥳 День зарплаты рафа",
                f"Оплаты на {paid_money} USD. Нежный отчёт кружит вальс вокруг кассы "
                "и шепчет команде: вы это заслужили ✈️💜",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|pay|big", options)
    else:
        options = [
            (
                "🥛 Деньги капают — и греют",
                f"Оплачено {paid_money} USD ({paid.count} поз., максимум {_fmt_money(max_pay)}). "
                "Даже не рекордный платёж согревает, как лавандовый раф в прохладный день 🌸",
            ),
            (
                "💳 Мягкий приход",
                f"{paid.count} платежей на {paid_money} USD за {period}."
                + (f" Это ~{pay_share:.0f}% от новых продаж недели." if pay_share else "")
                + " Нежный отчёт кивает одобрительно: движение есть 💜",
            ),
            (
                "🍯 Ложечка сливок в кассу",
                f"Пришло {paid_money} USD. Не фейерверк — зато честный сладкий момент. "
                f"Самый заметный платёж: {_fmt_money(max_pay)} USD. Спасибо клиенту ✨",
            ),
            (
                "📻 Позывной: оплата принята",
                f"На частоте кассы — {paid_money} USD ({paid.count} поз.). "
                "Связь устойчивая, настроение взлётное, раф не остыл ✈️🥰",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|pay|mid", options)
    cards["Оплаченные клиентом"] = {"title": title, "body": body}

    return cards


def build_como_mood_cards(summary: WeeklySummary) -> dict[str, dict[str, str]]:
    """Pastel prosecco / Lake Como narrative cards — fresh copy each week."""
    cards: dict[str, dict[str, str]] = {}
    week_tag = summary.week_end.strftime("%Y-%m-%d")
    period = f"{summary.week_start.strftime('%d.%m')}–{summary.week_end.strftime('%d.%m')}"

    new = summary.new_orders
    amounts = [float(r.get("Сумма, USD") or 0) for r in new.rows]
    max_amount = max(amounts) if amounts else 0.0
    avg_check = (new.total / new.count) if new.count else 0.0
    top_desc, top_amt = _top_sale_line(new.rows)
    mix = _category_mix(new.rows)
    money = _fmt_money(new.total)
    max_m = _fmt_money(max_amount)
    avg_m = _fmt_money(avg_check)

    if new.count == 0:
        options = [
            (
                "🥂 Тихий бокал на перилах",
                f"За {period} новых заказов нет. Озеро всё равно сверкает — "
                "идеальный момент смотреть на виллы и не торопить пузырьки ✨",
            ),
            (
                "🍋 Пауза с лимоном",
                "Продажи молчат, как полдень в Белладжо. Cin cin за терпение: "
                "следующая неделя уже плещется у набережной 🌿",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|como|sales|empty", options)
    elif max_amount >= 30000 or new.total >= 150000:
        options = [
            (
                "✨ Cin cin — мы молодцы!",
                f"Продажи за {period}: {money} USD · {new.count} поз. "
                f"Звезда террасы — «{top_desc}» на {max_m} USD. "
                "Просекко само поднимается в бокале 🥂💛",
            ),
            (
                "🌅 Закат над Комо и крупный чек",
                f"{new.count} заказов на {money} USD. Средний чек ~{avg_m}, "
                f"а топ сияет как вилла на воде. Аэрофлот, это было красиво ✈️",
            ),
            (
                "🍾 Бутылка открыта не зря",
                f"Неделя со вкусом праздника: {money} USD"
                + (f", микс {mix}" if mix else "")
                + f". «{top_desc}» — тот самый золотой блик на стекле ✨",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|como|sales|big", options)
    elif max_amount < 5000 and new.total < 25000:
        options = [
            (
                "🪙 Копейка рубль бережёт",
                f"Аккуратные {new.count} поз. на {money} USD. "
                "Как мелкие пузырьки: по одному почти не слышно, вместе — уже игристое 🥂",
            ),
            (
                "🍋 Долька лимона в бокале",
                f"Чек скромный (в среднем ~{avg_m} USD), зато живой. "
                f"За {period} набралось {money} USD — лёгкий аперитив недели 🌿",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|como|sales|small", options)
    else:
        options = [
            (
                "🌿 Крейсер по озеру",
                f"{new.count} новых позиций на {money} USD за {period}. "
                f"Топ — «{top_desc}» ({_fmt_money(top_amt)} USD). "
                "Ни шторма, ни штиля — ровная гладь Комо ✨",
            ),
            (
                "🥂 Аперитив средней крепости",
                f"Продажи: {money} USD · {new.count} поз."
                + (f" · {mix}." if mix else ".")
                + " Садимся на набережную, не спешим, пьём маленькими глотками 💛",
            ),
            (
                "🌤️ Мягкий свет вилл",
                f"{new.count} заказов, {money} USD, средний чек ~{avg_m}. "
                "Именно такой пастельный прогноз любит отчёт с просекко 🍋",
            ),
            (
                "⛵ Парус на горизонте",
                f"За неделю {money} USD. Самый заметный блик — «{top_desc}». "
                "Аэрофлот держит курс вдоль берега, бокал не проливается 🥂",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|como|sales|mid", options)
    cards["Новые заказы"] = {"title": title, "body": body}

    shipped = summary.shipped_orders
    late = sum(1 for r in shipped.rows if str(r.get("Поставка", "")).startswith("просрочка"))
    early = sum(1 for r in shipped.rows if str(r.get("Поставка", "")).startswith("досрочно"))
    on_time = max(shipped.count - late - early, 0)
    late_n, late_avg, late_max = _avg_late_days(shipped.rows)
    ship_money = _fmt_money(shipped.total)

    if shipped.count == 0:
        options = [
            (
                "🏡 Виллы пока закрыты",
                "Отгрузок-FINISHED за период нет. Озеро спокойно, бокал на столе — "
                "SHIPPED гуляют по набережной «в работе», и это по правилам 🌿",
            ),
            (
                "🫧 Пузырьки отдыхают",
                "На этой неделе finished-отгрузок не случилось. "
                "Не грустим: Комо умеет ждать красиво ✨",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|como|ship|empty", options)
    elif shipped.count and late / shipped.count >= 0.5:
        options = [
            (
                "🌙 Озеро никуда не спешит",
                f"Из {shipped.count} отгрузок (только FINISHED, {ship_money} USD) "
                f"с опозданием {late}"
                + (f", в среднем ~{late_avg} дн." if late_avg else "")
                + (f", рекорд {late_max} дн." if late_max else "")
                + ". Виллы стоят веками — и поставки тоже доходят. "
                "Дышим, смотрим на воду, поднимаем бокал за терпение 🫶🥂",
            ),
            (
                "🛟 Спокойно, мы на набережной",
                f"{late} из {shipped.count} пришли позже плана. "
                f"На {ship_money} USD всё равно уже у клиента. "
                "Просекко не кипит — оно играет. Мы рядом 💛",
            ),
            (
                "🌧️ Лёгкий дождик над Комо",
                f"График гуляет, но берег на месте: {shipped.count} поз. на {ship_money} USD. "
                "Лучше опоздать с заботой, чем спешить без неё. Cin cin 🍋",
            ),
            (
                "🕯️ Золотой час всё равно будет",
                f"Да, сроки плавают (до {late_max} дн. у самой задумчивой). "
                f"Но {shipped.count} finished-отгрузок уже сверкают, как стёкла на закате ✨",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|como|ship|late", options)
    elif early > late:
        options = [
            (
                "🏁 Быстрее пузырьков!",
                f"Досрочно {early} из {shipped.count}, в срок {on_time}. "
                "Неделя как утренний просекко на террасе: холодно, золото, восторг ☀️🥂",
            ),
            (
                "⛵ Ранний катер к вилле",
                f"{early} позиций приехали раньше срока, отгружено на {ship_money} USD. "
                "Наливаем второй бокал — за экипаж ✨",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|como|ship|early", options)
    else:
        options = [
            (
                "🚶 Ровный шаг по набережной",
                f"FINISHED: {shipped.count} поз. на {ship_money} USD "
                f"(рано {early} / вовремя {on_time} / позже {late}). "
                "Пастельный ритм — фирменный стиль Комо 🌿",
            ),
            (
                "🥂 Тост за ровную логистику",
                f"{shipped.count} отгрузок за {period} на {ship_money} USD. "
                "Без драмы, с видом на воду. Иногда лучший комментарий — cin cin ✨",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|como|ship|ok", options)
    cards["Отгруженные заказы"] = {"title": title, "body": body}

    paid = summary.paid_orders
    paid_money = _fmt_money(paid.total)
    pay_amounts = [float(r.get("Оплачено за неделю, USD") or 0) for r in paid.rows]
    max_pay = max(pay_amounts) if pay_amounts else 0.0
    pay_share = (paid.total / new.total * 100) if new.total > 0 and paid.total > 0 else 0.0

    if paid.count == 0 or paid.total <= 0:
        options = [
            (
                "🫧 Оплаты ещё в бутылке",
                "Платежей за период нет. Не грустим: просекко тоже сначала лежит на льду, "
                "а потом открывается с праздником 🥂",
            ),
            (
                "🌙 Ночная гладь кассы",
                "Пока тихо. Смотрим на огни вилл и верим, что следующий перевод "
                "придёт с лимонной цедрой 🍋",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|como|pay|empty", options)
    elif paid.total >= 100000:
        options = [
            (
                "🍾 Касса играет, как просекко!",
                f"За {period} пришло {paid_money} USD по {paid.count} позициям "
                f"(крупнейший ~{_fmt_money(max_pay)}). "
                "Это тот самый бокал, который звякает о перила 💛🥂",
            ),
            (
                "💛 Золото на воде",
                f"{paid_money} USD — закат полностью наш. "
                f"{paid.count} платежей, и каждый пузырёк на месте. Спасибо, Аэрофлот ✨",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|como|pay|big", options)
    else:
        options = [
            (
                "🥂 Ещё по глотку",
                f"Оплачено {paid_money} USD ({paid.count} поз., максимум {_fmt_money(max_pay)}). "
                "Не фейерверк — зато честный аперитив. Касса мурлычет, озеро довольно 🍋",
            ),
            (
                "💳 Мягкий приход на террасу",
                f"{paid.count} платежей на {paid_money} USD за {period}."
                + (f" Это ~{pay_share:.0f}% от новых продаж недели." if pay_share else "")
                + " Поднимаем бокал: движение есть 🌿",
            ),
            (
                "🍋 Цедра в кассе",
                f"Пришло {paid_money} USD. Самый заметный платёж: {_fmt_money(max_pay)} USD. "
                "Спасибо клиенту — просекко сегодня с характером ✨",
            ),
            (
                "⛵ Позывной: оплата принята",
                f"На частоте кассы — {paid_money} USD ({paid.count} поз.). "
                "Связь устойчивая, вид на Комо отличный, бокал не пустой 🥂",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|como|pay|mid", options)
    cards["Оплаченные клиентом"] = {"title": title, "body": body}
    return cards


def build_montecarlo_mood_cards(summary: WeeklySummary) -> dict[str, dict[str, str]]:
    """Ferrari 250 GTO / Monte-Carlo casino narrative cards."""
    cards: dict[str, dict[str, str]] = {}
    week_tag = summary.week_end.strftime("%Y-%m-%d")
    period = f"{summary.week_start.strftime('%d.%m')}–{summary.week_end.strftime('%d.%m')}"

    new = summary.new_orders
    amounts = [float(r.get("Сумма, USD") or 0) for r in new.rows]
    max_amount = max(amounts) if amounts else 0.0
    avg_check = (new.total / new.count) if new.count else 0.0
    top_desc, top_amt = _top_sale_line(new.rows)
    mix = _category_mix(new.rows)
    money = _fmt_money(new.total)
    max_m = _fmt_money(max_amount)
    avg_m = _fmt_money(avg_check)

    if new.count == 0:
        options = [
            (
                "♦ Тихий заезд к казино",
                f"За {period} новых заказов нет. Ferrari остывает у фасада — "
                "иногда роскошь в том, чтобы просто стоять и блестеть ✨",
            ),
            (
                "🕯️ Антракт перед ставкой",
                "Продажи молчат, как зал перед рулеткой. Не блефуем — "
                "ждём следующую партию с достоинством ♠️",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|mc|sales|empty", options)
    elif max_amount >= 30000 or new.total >= 150000:
        options = [
            (
                "🏎️ Grand Prix продаж!",
                f"Продажи за {period}: {money} USD · {new.count} поз. "
                f"Звезда вечера — «{top_desc}» на {max_m} USD. "
                "Красная 250 GTO аплодирует фарами ✨♦",
            ),
            (
                "💎 Ставка сыграла",
                f"{new.count} заказов на {money} USD. Средний чек ~{avg_m}. "
                "Это уже не аперитив — это джекпот у казино Монте-Карло 🥂",
            ),
            (
                "🔑 Ключ от вечера",
                f"Неделя класса гран-при: {money} USD"
                + (f", микс {mix}" if mix else "")
                + f". «{top_desc}» сверкает как лак на капоте 🔴",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|mc|sales|big", options)
    elif max_amount < 5000 and new.total < 25000:
        options = [
            (
                "🪙 Мелкая, но честная ставка",
                f"Аккуратные {new.count} поз. на {money} USD. "
                "В Монте-Карло уважают и скромный чип — если он на своём месте ♠️",
            ),
            (
                "🥂 Купе наполовину",
                f"Чек скромный (~{avg_m} USD), зато живой. "
                f"За {period} — {money} USD. Не блеф, а разгон перед кругом ✨",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|mc|sales|small", options)
    else:
        options = [
            (
                "🌴 Круиз по набережной",
                f"{new.count} новых позиций на {money} USD за {period}. "
                f"Топ — «{top_desc}» ({_fmt_money(top_amt)} USD). "
                "Ровный ход, золотой свет, без лишнего шума ♦",
            ),
            (
                "♠️ Игра средней руки",
                f"Продажи: {money} USD · {new.count} поз."
                + (f" · {mix}." if mix else ".")
                + " Не джекпот и не пас — уверенная ставка Аэрофлота ✨",
            ),
            (
                "🔴 Кармин недели",
                f"{new.count} заказов, {money} USD, средний чек ~{avg_m}. "
                "Именно такой тон любит Ferrari у казино: богатство без крика 🏎️",
            ),
            (
                "✨ Фары на казино",
                f"За неделю {money} USD. Самый яркий блик — «{top_desc}». "
                "Аэрофлот держит курс вдоль Place du Casino 💛",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|mc|sales|mid", options)
    cards["Новые заказы"] = {"title": title, "body": body}

    shipped = summary.shipped_orders
    late = sum(1 for r in shipped.rows if str(r.get("Поставка", "")).startswith("просрочка"))
    early = sum(1 for r in shipped.rows if str(r.get("Поставка", "")).startswith("досрочно"))
    on_time = max(shipped.count - late - early, 0)
    late_n, late_avg, late_max = _avg_late_days(shipped.rows)
    ship_money = _fmt_money(shipped.total)

    if shipped.count == 0:
        options = [
            (
                "🏛️ Казино закрыто на час",
                "FINISHED-отгрузок за период нет. SHIPPED гуляет по залу как «в работе» — "
                "по правилам стола. Ferrari ждёт у подъезда ✨",
            ),
            (
                "🕯️ Пауза у рулетки",
                "Пока без finished-отгрузок. В Монте-Карло умеют ждать красиво ♦",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|mc|ship|empty", options)
    elif shipped.count and late / shipped.count >= 0.5:
        options = [
            (
                "🌙 Ночь длинная — дорога короче",
                f"Из {shipped.count} finished-отгрузок ({ship_money} USD) "
                f"с опозданием {late}"
                + (f", в среднем ~{late_avg} дн." if late_avg else "")
                + (f", рекорд {late_max} дн." if late_max else "")
                + ". Даже 250 GTO иногда стоит в пробке. "
                "Главное — доехать с достоинством ♠️🫶",
            ),
            (
                "🛟 Спокойно, дилер рядом",
                f"{late} из {shipped.count} позже плана, на {ship_money} USD всё равно у клиента. "
                "Не блефуем нервами — поднимаем купе за терпение 🥂",
            ),
            (
                "🌧️ Дождь на Place du Casino",
                f"График гуляет, асфальт блестит: {shipped.count} поз. на {ship_money} USD. "
                "Лучше опоздать роскошно, чем спешить дёшево ✨",
            ),
            (
                "🔑 Золотой час всё равно наш",
                f"Да, сроки плавают (до {late_max} дн.). "
                f"Но {shipped.count} finished-отгрузок уже сияют, как лак на капоте 🔴",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|mc|ship|late", options)
    elif early > late:
        options = [
            (
                "🏁 Быстрее рулетки!",
                f"Досрочно {early} из {shipped.count}, в срок {on_time}. "
                "Неделя как красная GTO на пустой набережной — чистый кайф 🏎️✨",
            ),
            (
                "💎 Ранний валет",
                f"{early} позиций раньше срока, отгружено на {ship_money} USD. "
                "Дилер кивает: красивая комбинация ♠️",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|mc|ship|early", options)
    else:
        options = [
            (
                "🚶 Ровный круг у казино",
                f"FINISHED: {shipped.count} поз. на {ship_money} USD "
                f"(рано {early} / вовремя {on_time} / позже {late}). "
                "Богатый, спокойный ритм — фирменный стиль вечера ♦",
            ),
            (
                "🥂 Тост за логистику гран-при",
                f"{shipped.count} отгрузок за {period} на {ship_money} USD. "
                "Без драмы, с золотым светом. Иногда лучший комментарий — «банк наш» ✨",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|mc|ship|ok", options)
    cards["Отгруженные заказы"] = {"title": title, "body": body}

    paid = summary.paid_orders
    paid_money = _fmt_money(paid.total)
    pay_amounts = [float(r.get("Оплачено за неделю, USD") or 0) for r in paid.rows]
    max_pay = max(pay_amounts) if pay_amounts else 0.0
    pay_share = (paid.total / new.total * 100) if new.total > 0 and paid.total > 0 else 0.0

    if paid.count == 0 or paid.total <= 0:
        options = [
            (
                "🫧 Касса на паузе",
                "Платежей за период нет. Не грустим: даже в Монте-Карло "
                "сначала показывают карты, потом считают банк ♠️",
            ),
            (
                "🌙 Ночная тишина стола",
                "Пока без оплат. Смотрим на фасад казино и верим, "
                "что следующий перевод приедет на красной Ferrari ✨",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|mc|pay|empty", options)
    elif paid.total >= 100000:
        options = [
            (
                "💎 Джекпот кассы!",
                f"За {period} пришло {paid_money} USD по {paid.count} позициям "
                f"(крупнейший ~{_fmt_money(max_pay)}). "
                "Это тот самый звук, когда фишки сыплются на сукно 🥂♦",
            ),
            (
                "🏎️ Банк на высокой передаче",
                f"{paid_money} USD — вечер полностью наш. "
                f"{paid.count} платежей, и каждый блестит как лак 250 GTO. Спасибо, Аэрофлот ✨",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|mc|pay|big", options)
    else:
        options = [
            (
                "🥂 Ещё по купе",
                f"Оплачено {paid_money} USD ({paid.count} поз., максимум {_fmt_money(max_pay)}). "
                "Не джекпот — зато честный, золотой приход. Касса довольна 💛",
            ),
            (
                "💳 Мягкий приход к столу",
                f"{paid.count} платежей на {paid_money} USD за {period}."
                + (f" Это ~{pay_share:.0f}% от новых продаж недели." if pay_share else "")
                + " Дилер кивает: движение есть ♠️",
            ),
            (
                "🔑 Золотой ключ кассы",
                f"Пришло {paid_money} USD. Самый заметный платёж: {_fmt_money(max_pay)} USD. "
                "Спасибо клиенту — ставка сыграла красиво ✨",
            ),
            (
                "🔴 Позывной: оплата принята",
                f"На частоте кассы — {paid_money} USD ({paid.count} поз.). "
                "Связь устойчивая, Ferrari у подъезда, шампанское холодное 🏎️🥂",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|mc|pay|mid", options)
    cards["Оплаченные клиентом"] = {"title": title, "body": body}
    return cards


def _remap_mood_cards(cards: dict[str, dict[str, str]], pairs: list[tuple[str, str]]) -> dict[str, dict[str, str]]:
    out: dict[str, dict[str, str]] = {}
    for key, card in cards.items():
        title, body = card["title"], card["body"]
        for old, new in pairs:
            title = title.replace(old, new)
            body = body.replace(old, new)
        out[key] = {"title": title, "body": body}
    return out


def build_yacht_mood_cards(summary: WeeklySummary, client: str = "Аэрофлот") -> dict[str, dict[str, str]]:
    """Tropical yacht / paradise vacation on a budget of 'rich' — original weekly copy."""
    cards: dict[str, dict[str, str]] = {}
    week_tag = summary.week_end.strftime("%Y-%m-%d")
    period = f"{summary.week_start.strftime('%d.%m')}–{summary.week_end.strftime('%d.%m')}"

    new = summary.new_orders
    amounts = [float(r.get("Сумма, USD") or 0) for r in new.rows]
    max_amount = max(amounts) if amounts else 0.0
    avg_check = (new.total / new.count) if new.count else 0.0
    top_desc, top_amt = _top_sale_line(new.rows)
    mix = _category_mix(new.rows)
    money = _fmt_money(new.total)
    max_m = _fmt_money(max_amount)
    avg_m = _fmt_money(avg_check)

    if new.count == 0:
        options = [
            (
                "🌴 Тихая бухта",
                f"За {period} новых заказов нет. Яхта покачивается на якоре, "
                "фуршет накрыт на всякий случай — иногда роскошь в том, чтобы просто отдыхать ⛵",
            ),
            (
                "🥂 Антракт между коктейлями",
                "Продажи молчат, как лагуна в штиль. Не тревожим капитана — "
                "ждём следующую волну с бирюзовой водой ✨",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|yacht|sales|empty", options)
    elif max_amount >= 30000 or new.total >= 150000:
        options = [
            (
                "⛵ Regatta продаж!",
                f"Продажи за {period}: {money} USD · {new.count} поз. "
                f"Звезда вечера — «{top_desc}» на {max_m} USD. "
                "Палуба аплодирует, шампанское уже открыто ✨⛵",
            ),
            (
                "💰 Джекпот на богатом",
                f"{new.count} заказов на {money} USD. Средний чек ~{avg_m}. "
                "Это уже не аперитив у причала — это полный фуршет на борту 🥂",
            ),
            (
                "🌊 Ключ от бухты",
                f"Неделя класса regatta: {money} USD"
                + (f", микс {mix}" if mix else "")
                + f". «{top_desc}» сверкает, как лак на белом корпусе яхты ⛵",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|yacht|sales|big", options)
    elif max_amount < 5000 and new.total < 25000:
        options = [
            (
                "🍹 Лёгкий коктейль",
                f"Аккуратные {new.count} поз. на {money} USD. "
                "В тропическом раю уважают и скромный фуршет — если он без суеты ✨",
            ),
            (
                "🌴 Яхта наполовину",
                f"Чек скромный (~{avg_m} USD), зато живой. "
                f"За {period} — {money} USD. Не regatta, а спокойный круиз вдоль рифа ⛵",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|yacht|sales|small", options)
    else:
        options = [
            (
                "🌊 Круиз вдоль рифа",
                f"{new.count} новых позиций на {money} USD за {period}. "
                f"Топ — «{top_desc}» ({_fmt_money(top_amt)} USD). "
                "Ровный ход, золотой закат, без лишней волны ⛵",
            ),
            (
                "🥂 Бокал на богатом",
                f"Продажи: {money} USD · {new.count} поз."
                + (f" · {mix}." if mix else ".")
                + f" Не шторм и не штиль — уверенный круиз {client} ✨",
            ),
            (
                "💙 Бирюза недели",
                f"{new.count} заказов, {money} USD, средний чек ~{avg_m}. "
                "Именно такой тон любит яхта в тропической бухте: богатство без крика ⛵",
            ),
            (
                "✨ Парус на горизонте",
                f"За неделю {money} USD. Самый яркий блик — «{top_desc}». "
                f"{client} держит курс вдоль лазурной воды 💙",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|yacht|sales|mid", options)
    cards["Новые заказы"] = {"title": title, "body": body}

    shipped = summary.shipped_orders
    late = sum(1 for r in shipped.rows if str(r.get("Поставка", "")).startswith("просрочка"))
    early = sum(1 for r in shipped.rows if str(r.get("Поставка", "")).startswith("досрочно"))
    on_time = max(shipped.count - late - early, 0)
    late_n, late_avg, late_max = _avg_late_days(shipped.rows)
    ship_money = _fmt_money(shipped.total)

    if shipped.count == 0:
        options = [
            (
                "🛥️ Яхта ещё у причала",
                "FINISHED-отгрузок за период нет. SHIPPED гуляет по лагуне как «в работе» — "
                "по правилам круиза. Парус ждёт попутного ветра ✨",
            ),
            (
                "🌙 Лагуна спокойна",
                "Пока без finished-отгрузок. Тропический остров умеет ждать красиво — "
                "фуршет накрыт, вода тёплая ⛵",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|yacht|ship|empty", options)
    elif shipped.count and late / shipped.count >= 0.5:
        options = [
            (
                "🌧️ Шторм подольше обычного",
                f"Из {shipped.count} finished-отгрузок ({ship_money} USD) "
                f"с опозданием {late}"
                + (f", в среднем ~{late_avg} дн." if late_avg else "")
                + (f", рекорд {late_max} дн." if late_max else "")
                + ". Даже яхта иногда стоит в штормовой бухте. "
                "Главное — доплыть с достоинством 🫶",
            ),
            (
                "🛟 Спокойно, капитан рядом",
                f"{late} из {shipped.count} позже плана, на {ship_money} USD всё равно у клиента. "
                "Не паникуем — поднимаем бокал за терпение 🥂",
            ),
            (
                "🌊 Волна гуляет",
                f"График плавает, но риф на месте: {shipped.count} поз. на {ship_money} USD. "
                "Лучше опоздать роскошно, чем спешить дёшево ✨",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|yacht|ship|late", options)
    elif early > late:
        options = [
            (
                "🏁 Быстрее ветра!",
                f"Досрочно {early} из {shipped.count}, в срок {on_time}. "
                "Неделя как утренний бриз в бухте: чистый кайф ☀️⛵",
            ),
            (
                "⛵ Ранний заход в порт",
                f"{early} позиций раньше срока, отгружено на {ship_money} USD. "
                "Капитан кивает: красивая швартовка ✨",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|yacht|ship|early", options)
    else:
        options = [
            (
                "🚶 Ровный круг по лагуне",
                f"FINISHED: {shipped.count} поз. на {ship_money} USD "
                f"(рано {early} / вовремя {on_time} / позже {late}). "
                "Богатый, спокойный ритм — фирменный стиль тропического отдыха ⛵",
            ),
            (
                "🥂 Тост за ровную логистику",
                f"{shipped.count} отгрузок за {period} на {ship_money} USD. "
                "Без драмы, с видом на бирюзу. Иногда лучший комментарий — «на богатом» ✨",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|yacht|ship|ok", options)
    cards["Отгруженные заказы"] = {"title": title, "body": body}

    paid = summary.paid_orders
    paid_money = _fmt_money(paid.total)
    pay_amounts = [float(r.get("Оплачено за неделю, USD") or 0) for r in paid.rows]
    max_pay = max(pay_amounts) if pay_amounts else 0.0
    pay_share = (paid.total / new.total * 100) if new.total > 0 and paid.total > 0 else 0.0

    if paid.count == 0 or paid.total <= 0:
        options = [
            (
                "🫧 Касса на якоре",
                "Платежей за период нет. Не грустим: даже на яхте "
                "сначала накрывают фуршет, потом считают кассу 🥂",
            ),
            (
                "🌙 Тишина в бухте",
                "Пока без оплат. Смотрим на закат и верим, "
                "что следующий перевод приедет на белой яхте ✨",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|yacht|pay|empty", options)
    elif paid.total >= 100000:
        options = [
            (
                "💎 Касса как полный фуршет!",
                f"За {period} пришло {paid_money} USD по {paid.count} позициям "
                f"(крупнейший ~{_fmt_money(max_pay)}). "
                "Это тот самый звон бокалов, когда вечер уже наш 🥂⛵",
            ),
            (
                "⛵ Касса на попутном ветре",
                f"{paid_money} USD — закат полностью наш. "
                f"{paid.count} платежей, и каждый блестит как золото на палубе. Спасибо, {client} ✨",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|yacht|pay|big", options)
    else:
        options = [
            (
                "🥂 Ещё по бокалу",
                f"Оплачено {paid_money} USD ({paid.count} поз., максимум {_fmt_money(max_pay)}). "
                "Не фейерверк — зато честный, солнечный приход. Касса довольна 💙",
            ),
            (
                "💳 Мягкий приход к палубе",
                f"{paid.count} платежей на {paid_money} USD за {period}."
                + (f" Это ~{pay_share:.0f}% от новых продаж недели." if pay_share else "")
                + " Капитан кивает: движение есть ⛵",
            ),
            (
                "🔑 Золотой ключ кассы",
                f"Пришло {paid_money} USD. Самый заметный платёж: {_fmt_money(max_pay)} USD. "
                f"Спасибо {client} — круиз сыграл красиво ✨",
            ),
            (
                "💰 Позывной: оплата принята",
                f"На частоте кассы — {paid_money} USD ({paid.count} поз.). "
                "Связь устойчивая, яхта у рифа, шампанское холодное ⛵🥂",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|yacht|pay|mid", options)
    cards["Оплаченные клиентом"] = {"title": title, "body": body}
    return cards


def build_bmw_mood_cards(summary: WeeklySummary, client: str) -> dict[str, dict[str, str]]:
    """BMW 8 Series on Kazan embankment + hookah lounge — original weekly copy."""
    cards: dict[str, dict[str, str]] = {}
    week_tag = summary.week_end.strftime("%Y-%m-%d")
    period = f"{summary.week_start.strftime('%d.%m')}–{summary.week_end.strftime('%d.%m')}"

    new = summary.new_orders
    amounts = [float(r.get("Сумма, USD") or 0) for r in new.rows]
    max_amount = max(amounts) if amounts else 0.0
    avg_check = (new.total / new.count) if new.count else 0.0
    top_desc, top_amt = _top_sale_line(new.rows)
    mix = _category_mix(new.rows)
    money = _fmt_money(new.total)
    max_m = _fmt_money(max_amount)
    avg_m = _fmt_money(avg_check)

    if new.count == 0:
        options = [
            (
                "🌙 Тихий круг у Кремля",
                f"За {period} новых заказов нет. Восьмёрка остывает на набережной, "
                "кальян тлеет сам по себе — иногда люкс в том, чтобы просто стоять и светиться ✨",
            ),
            (
                "🫧 Антракт в кальянной",
                "Продажи молчат, как зал между сетами. Не тушим угли — "
                "ждём следующую волну с Казанки 🚗",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|bmw|sales|empty", options)
    elif max_amount >= 30000 or new.total >= 150000:
        options = [
            (
                "🚗 Ночной заезд на Казанку",
                f"Продажи за {period}: {money} USD · {new.count} поз. "
                f"Звезда вечера — «{top_desc}» на {max_m} USD. "
                "Куул-Шариф кивает минаретами, фары BMW отвечают ✨",
            ),
            (
                "💨 Дым пошёл густой",
                f"{new.count} заказов на {money} USD. Средний чек ~{avg_m}. "
                "Это уже не разогрев углей — это полный зал кальянной на набережной 🔥",
            ),
            (
                "🔑 Ключ от восьмёрки",
                f"Неделя класса night drive: {money} USD"
                + (f", микс {mix}" if mix else "")
                + f". «{top_desc}» блестит, как лак на капоте у Кремля 💙",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|bmw|sales|big", options)
    elif max_amount < 5000 and new.total < 25000:
        options = [
            (
                "🍃 Лёгкая затяжка",
                f"Аккуратные {new.count} поз. на {money} USD. "
                "В кальянной уважают и короткий сет — если он ровный и без гари ✨",
            ),
            (
                "💙 Купе наполовину",
                f"Чек скромный (~{avg_m} USD), зато живой. "
                f"За {period} — {money} USD. Не гонка, а прогрев перед кругом по набережной 🚗",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|bmw|sales|small", options)
    else:
        options = [
            (
                "🌃 Круиз вдоль Кремля",
                f"{new.count} новых позиций на {money} USD за {period}. "
                f"Топ — «{top_desc}» ({_fmt_money(top_amt)} USD). "
                "Ровный ход, янтарь кальяна, без лишнего шума 🚗",
            ),
            (
                "🔥 Угли в норме",
                f"Продажи: {money} USD · {new.count} поз."
                + (f" · {mix}." if mix else ".")
                + f" Не дрифт и не простой — уверенный night drive {client} ✨",
            ),
            (
                "💙 Сапфир недели",
                f"{new.count} заказов, {money} USD, средний чек ~{avg_m}. "
                "Именно такой тон любит восьмёрка у Казанки: богатство без крика 🚗",
            ),
            (
                "✨ Фары на Куул-Шариф",
                f"За неделю {money} USD. Самый яркий блик — «{top_desc}». "
                f"{client} держит курс вдоль кремлёвской набережной 💙",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|bmw|sales|mid", options)
    cards["Новые заказы"] = {"title": title, "body": body}

    shipped = summary.shipped_orders
    late = sum(1 for r in shipped.rows if str(r.get("Поставка", "")).startswith("просрочка"))
    early = sum(1 for r in shipped.rows if str(r.get("Поставка", "")).startswith("досрочно"))
    on_time = max(shipped.count - late - early, 0)
    late_n, late_avg, late_max = _avg_late_days(shipped.rows)
    ship_money = _fmt_money(shipped.total)

    if shipped.count == 0:
        options = [
            (
                "🛋️ Кальянная ещё пустая",
                "FINISHED-отгрузок за период нет. SHIPPED гуляет по набережной как «в работе» — "
                "по правилам салона. Восьмёрка ждёт у Кремля ✨",
            ),
            (
                "🌙 Пауза у реки",
                "Пока без finished-отгрузок. Казань умеет ждать красиво — "
                "угли тлеют, Волга не торопится 🚗",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|bmw|ship|empty", options)
    elif shipped.count and late / shipped.count >= 0.5:
        options = [
            (
                "🌫️ Дым подольше обычного",
                f"Из {shipped.count} finished-отгрузок ({ship_money} USD) "
                f"с опозданием {late}"
                + (f", в среднем ~{late_avg} дн." if late_avg else "")
                + (f", рекорд {late_max} дн." if late_max else "")
                + ". Даже BMW иногда стоит в пробке на набережной. "
                "Главное — доехать с достоинством 🫶",
            ),
            (
                "🛟 Спокойно, кальян не гаснет",
                f"{late} из {shipped.count} позже плана, на {ship_money} USD всё равно у клиента. "
                "Не дёргаем ручник — поднимаем шланг за терпение 💨",
            ),
            (
                "🌧️ Мокрый камень Казанки",
                f"График гуляет, плитка блестит: {shipped.count} поз. на {ship_money} USD. "
                "Лучше опоздать роскошно, чем спешить дёшево ✨",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|bmw|ship|late", options)
    elif early > late:
        options = [
            (
                "🏁 Быстрее дыма!",
                f"Досрочно {early} из {shipped.count}, в срок {on_time}. "
                "Неделя как пустая набережная ночью — чистый кайф, фары в пол 🚗✨",
            ),
            (
                "💨 Ранний сет",
                f"{early} позиций раньше срока, отгружено на {ship_money} USD. "
                "Хозяин кальянной кивает: красивая комбинация 🔥",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|bmw|ship|early", options)
    else:
        options = [
            (
                "🚶 Ровный круг у Кремля",
                f"FINISHED: {shipped.count} поз. на {ship_money} USD "
                f"(рано {early} / вовремя {on_time} / позже {late}). "
                "Богатый, спокойный ритм — фирменный стиль ночной Казани 🚗",
            ),
            (
                "🔥 Тост за логистику night drive",
                f"{shipped.count} отгрузок за {period} на {ship_money} USD. "
                "Без драмы, с отражениями в Казанке. Иногда лучший комментарий — «поехали» ✨",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|bmw|ship|ok", options)
    cards["Отгруженные заказы"] = {"title": title, "body": body}

    paid = summary.paid_orders
    paid_money = _fmt_money(paid.total)
    pay_amounts = [float(r.get("Оплачено за неделю, USD") or 0) for r in paid.rows]
    max_pay = max(pay_amounts) if pay_amounts else 0.0
    pay_share = (paid.total / new.total * 100) if new.total > 0 and paid.total > 0 else 0.0

    if paid.count == 0 or paid.total <= 0:
        options = [
            (
                "🫧 Касса на паузе",
                "Платежей за период нет. Не грустим: даже в кальянной "
                "сначала разжигают угли, потом считают вечер 🔥",
            ),
            (
                "🌙 Тишина на набережной",
                "Пока без оплат. Смотрим на Куул-Шариф и верим, "
                "что следующий перевод приедет на восьмёрке ✨",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|bmw|pay|empty", options)
    elif paid.total >= 100000:
        options = [
            (
                "💎 Касса как полный зал",
                f"За {period} пришло {paid_money} USD по {paid.count} позициям "
                f"(крупнейший ~{_fmt_money(max_pay)}). "
                "Это тот самый густой дым, когда вечер уже наш 🔥🚗",
            ),
            (
                "🚗 Банк на высокой передаче",
                f"{paid_money} USD — ночь полностью наша. "
                f"{paid.count} платежей, и каждый блестит как лак BMW. Спасибо, {client} ✨",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|bmw|pay|big", options)
    else:
        options = [
            (
                "💨 Ещё по затяжке",
                f"Оплачено {paid_money} USD ({paid.count} поз., максимум {_fmt_money(max_pay)}). "
                "Не фейерверк — зато честный, янтарный приход. Касса довольна 💙",
            ),
            (
                "💳 Мягкий приход к столу",
                f"{paid.count} платежей на {paid_money} USD за {period}."
                + (f" Это ~{pay_share:.0f}% от новых продаж недели." if pay_share else "")
                + " Хозяин кальянной кивает: движение есть 🔥",
            ),
            (
                "🔑 Ключ от кассы",
                f"Пришло {paid_money} USD. Самый заметный платёж: {_fmt_money(max_pay)} USD. "
                f"Спасибо {client} — ночной круг сыграл красиво ✨",
            ),
            (
                "💙 Позывной: оплата принята",
                f"На частоте кассы — {paid_money} USD ({paid.count} поз.). "
                "Связь устойчивая, BMW у набережной, угли горячие 🚗🔥",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|bmw|pay|mid", options)
    cards["Оплаченные клиентом"] = {"title": title, "body": body}
    return cards


def build_space_mood_cards(summary: WeeklySummary, client: str = "Аэрофлот") -> dict[str, dict[str, str]]:
    """Deep space: black holes, galaxies, stars — conquering distant worlds."""
    cards: dict[str, dict[str, str]] = {}
    week_tag = summary.week_end.strftime("%Y-%m-%d")
    period = f"{summary.week_start.strftime('%d.%m')}–{summary.week_end.strftime('%d.%m')}"

    new = summary.new_orders
    amounts = [float(r.get("Сумма, USD") or 0) for r in new.rows]
    max_amount = max(amounts) if amounts else 0.0
    avg_check = (new.total / new.count) if new.count else 0.0
    top_desc, top_amt = _top_sale_line(new.rows)
    mix = _category_mix(new.rows)
    money = _fmt_money(new.total)
    max_m = _fmt_money(max_amount)
    avg_m = _fmt_money(avg_check)

    if new.count == 0:
        options = [
            (
                "🌑 Тихая орбита",
                f"За {period} новых заказов нет. Чёрная дыра молчит, "
                "звёзды мерцают сами по себе — иногда покорение миров начинается с паузы ✦",
            ),
            (
                "✨ Антракт между галактиками",
                "Продажи в гиперпространственном штиле. Не тревожим экипаж — "
                "ждём следующий сигнал с дальнего края ✦",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|space|sales|empty", options)
    elif max_amount >= 30000 or new.total >= 150000:
        options = [
            (
                "🚀 Прыжок к дальним мирам!",
                f"Продажи за {period}: {money} USD · {new.count} поз. "
                f"Звезда вечера — «{top_desc}» на {max_m} USD. "
                "Галактика аплодирует, шлюз уже открыт ✦",
            ),
            (
                "💫 Джекпот у горизонта событий",
                f"{new.count} заказов на {money} USD. Средний чек ~{avg_m}. "
                "Это уже не разведка — это полноценное покорение дальнего космоса 🌌",
            ),
            (
                "🌌 Ключ от гиперпространства",
                f"Неделя класса deep space: {money} USD"
                + (f", микс {mix}" if mix else "")
                + f". «{top_desc}» сверкает, как ядро спиральной галактики ✦",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|space|sales|big", options)
    elif max_amount < 5000 and new.total < 25000:
        options = [
            (
                "⭐ Лёгкая звезда",
                f"Аккуратные {new.count} поз. на {money} USD. "
                "В дальнем космосе уважают и скромный маяк — если он без суеты ✦",
            ),
            (
                "🛰️ Орбита наполовину",
                f"Чек скромный (~{avg_m} USD), зато живой. "
                f"За {period} — {money} USD. Не сверхновая, а спокойный дрейф вдоль рукава ✦",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|space|sales|small", options)
    else:
        options = [
            (
                "🌌 Круиз вдоль спирали",
                f"{new.count} новых позиций на {money} USD за {period}. "
                f"Топ — «{top_desc}» ({_fmt_money(top_amt)} USD). "
                "Ровный курс, золотые звёзды, без лишней турбулентности ✦",
            ),
            (
                "✦ Курс на дальний мир",
                f"Продажи: {money} USD · {new.count} поз."
                + (f" · {mix}." if mix else ".")
                + f" Не шторм и не вакуум — уверенный полёт {client} ✨",
            ),
            (
                "💙 Индиго недели",
                f"{new.count} заказов, {money} USD, средний чек ~{avg_m}. "
                "Именно такой тон любит космос: богатство без крика ✦",
            ),
            (
                "✨ Звезда на горизонте",
                f"За неделю {money} USD. Самый яркий блик — «{top_desc}». "
                f"{client} держит курс к дальним мирам 🌌",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|space|sales|mid", options)
    cards["Новые заказы"] = {"title": title, "body": body}

    shipped = summary.shipped_orders
    late = sum(1 for r in shipped.rows if str(r.get("Поставка", "")).startswith("просрочка"))
    early = sum(1 for r in shipped.rows if str(r.get("Поставка", "")).startswith("досрочно"))
    on_time = max(shipped.count - late - early, 0)
    late_n, late_avg, late_max = _avg_late_days(shipped.rows)
    ship_money = _fmt_money(shipped.total)

    if shipped.count == 0:
        options = [
            (
                "🛸 Корабль ещё у дока",
                "FINISHED-отгрузок за период нет. SHIPPED дрейфует как «в работе» — "
                "по правилам миссии. Шлюз ждёт зелёного света ✦",
            ),
            (
                "🌑 Тишина у чёрной дыры",
                "Пока без finished-отгрузок. Дальний космос умеет ждать красиво — "
                "звёзды на месте, экипаж на связи ✦",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|space|ship|empty", options)
    elif shipped.count and late / shipped.count >= 0.5:
        options = [
            (
                "☄️ Гравитация задержала",
                f"Из {shipped.count} finished-отгрузок ({ship_money} USD) "
                f"с опозданием {late}"
                + (f", в среднем ~{late_avg} дн." if late_avg else "")
                + (f", рекорд {late_max} дн." if late_max else "")
                + ". Даже у горизонта событий бывает пробка. "
                "Главное — долететь с достоинством 🫶",
            ),
            (
                "🛟 Спокойно, капитан на мостике",
                f"{late} из {shipped.count} позже плана, на {ship_money} USD всё равно у клиента. "
                "Не паникуем — поднимаем тост за терпение ✦",
            ),
            (
                "🌊 Солнечный ветер гуляет",
                f"График плавает, но звезда на месте: {shipped.count} поз. на {ship_money} USD. "
                "Лучше опоздать роскошно, чем спешить дёшево ✨",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|space|ship|late", options)
    elif early > late:
        options = [
            (
                "🏁 Быстрее света!",
                f"Досрочно {early} из {shipped.count}, в срок {on_time}. "
                "Неделя как гиперпрыжок без турбулентности: чистый кайф ✨🚀",
            ),
            (
                "🚀 Ранняя посадка на новый мир",
                f"{early} позиций раньше срока, отгружено на {ship_money} USD. "
                "Экипаж кивает: красивая стыковка ✦",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|space|ship|early", options)
    else:
        options = [
            (
                "🛰 Ровный круг по орбите",
                f"FINISHED: {shipped.count} поз. на {ship_money} USD "
                f"(рано {early} / вовремя {on_time} / позже {late}). "
                "Богатый, спокойный ритм — фирменный стиль дальнего полёта ✦",
            ),
            (
                "🥂 Тост за ровную логистику",
                f"{shipped.count} отгрузок за {period} на {ship_money} USD. "
                "Без драмы, с видом на галактику. Иногда лучший комментарий — «мир наш» ✨",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|space|ship|ok", options)
    cards["Отгруженные заказы"] = {"title": title, "body": body}

    paid = summary.paid_orders
    paid_money = _fmt_money(paid.total)
    pay_amounts = [float(r.get("Оплачено за неделю, USD") or 0) for r in paid.rows]
    max_pay = max(pay_amounts) if pay_amounts else 0.0
    pay_share = (paid.total / new.total * 100) if new.total > 0 and paid.total > 0 else 0.0

    if paid.count == 0 or paid.total <= 0:
        options = [
            (
                "🫧 Касса в криосне",
                "Платежей за период нет. Не грустим: даже у чёрной дыры "
                "сначала настраивают маяк, потом считают кассу ✦",
            ),
            (
                "🌑 Тишина в туманности",
                "Пока без оплат. Смотрим на звёзды и верим, "
                "что следующий перевод прилетит из дальнего мира ✨",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|space|pay|empty", options)
    elif paid.total >= 100000:
        options = [
            (
                "💎 Касса как сверхновая!",
                f"За {period} пришло {paid_money} USD по {paid.count} позициям "
                f"(крупнейший ~{_fmt_money(max_pay)}). "
                "Это тот самый свет, когда галактика уже наша ✦🌌",
            ),
            (
                "🚀 Касса на попутном ветре",
                f"{paid_money} USD — орбита полностью наша. "
                f"{paid.count} платежей, и каждый блестит как золото звёзд. Спасибо, {client} ✨",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|space|pay|big", options)
    else:
        options = [
            (
                "✦ Ещё по звезде",
                f"Оплачено {paid_money} USD ({paid.count} поз., максимум {_fmt_money(max_pay)}). "
                "Не сверхновая — зато честный, спокойный приход. Касса довольна 💙",
            ),
            (
                "💳 Мягкий приход к шлюзу",
                f"{paid.count} платежей на {paid_money} USD за {period}."
                + (f" Это ~{pay_share:.0f}% от новых продаж недели." if pay_share else "")
                + " Капитан кивает: движение есть ✦",
            ),
            (
                "🔑 Золотой ключ кассы",
                f"Пришло {paid_money} USD. Самый заметный платёж: {_fmt_money(max_pay)} USD. "
                f"Спасибо {client} — миссия сыграла красиво ✨",
            ),
            (
                "💫 Позывной: оплата принята",
                f"На частоте кассы — {paid_money} USD ({paid.count} поз.). "
                "Связь устойчивая, корабль у туманности, звёзды холодные ✦",
            ),
        ]
        title, body = _pick_variant(f"{week_tag}|space|pay|mid", options)
    cards["Оплаченные клиентом"] = {"title": title, "body": body}
    return cards


def lavender_cover_epigraph(report_date: date, in_work: int, shipped: int) -> str:
    options = [
        (
            "мягкий лавандовый раф: эспрессо, сливки и цветочный аромат — "
            "тот же уют, только в отчёте по заказам ♡"
        ),
        (
            f"сегодня в чашке — статус на {report_date.strftime('%d.%m')}: "
            f"{in_work} в работе, {shipped} отгружено. помешиваем осторожно 💜"
        ),
        (
            "не гонимся за идеальным графиком — наливаем заботу, "
            "добавляем терпение и щепотку лаванды ✨"
        ),
        (
            "этот отчёт пахнет перроном на закате и свежей пенкой. "
            "Аэрофлот, мы рядом — даже когда сроки гуляют ✈️♡"
        ),
        (
            "инструкция барриста нежного отчёта: "
            "смотреть цифры без паники, хвалить команду, допивать раф до конца ☕"
        ),
    ]
    digest = hashlib.md5(report_date.isoformat().encode()).hexdigest()
    return options[int(digest[:8], 16) % len(options)]


def _write_mood_card(
    ws,
    start_row: int,
    card: dict[str, str],
    image_path: Path,
    col_start: int = 15,
    theme_name: str = "lavender_raf",
) -> None:
    """Place a cute comment card to the right of a summary table."""
    if theme_name == "montecarlo_ferrari":
        title_fill = PatternFill("solid", fgColor="3B0A14")
        cream = PatternFill("solid", fgColor="F8EED9")
        body_font = Font(name="Calibri", size=10, color="3B0A14")
        title_font = Font(name="Georgia", size=12, bold=True, color="F4E7D8")
        img_size = 78
        img_col = "S"
    elif theme_name == "como_prosecco":
        title_fill = PatternFill("solid", fgColor="E8F2F4")
        cream = PatternFill("solid", fgColor="FBF4EA")
        body_font = Font(name="Calibri", size=10, color="4A5D6A")
        title_font = Font(name="Georgia", size=12, bold=True, color=COLOR_NAVY)
        img_size = 78
        img_col = "S"
    elif theme_name == "tropical_yacht":
        title_fill = PatternFill("solid", fgColor="0B3D4C")
        cream = PatternFill("solid", fgColor="FBF3E0")
        body_font = Font(name="Calibri", size=10, color="0B3D4C")
        title_font = Font(name="Georgia", size=12, bold=True, color="E0F4F7")
        img_size = 150
        img_col = "T"
    elif theme_name == "bmw_night":
        title_fill = PatternFill("solid", fgColor="0D1B2A")
        cream = PatternFill("solid", fgColor="E8EDFA")
        body_font = Font(name="Calibri", size=10, color="0D1B2A")
        title_font = Font(name="Georgia", size=12, bold=True, color="E8EEF4")
        img_size = 118
        img_col = "T"
    elif theme_name == "deep_space":
        title_fill = PatternFill("solid", fgColor="070B1A")
        cream = PatternFill("solid", fgColor="F8F1DC")
        body_font = Font(name="Calibri", size=10, color="070B1A")
        title_font = Font(name="Georgia", size=12, bold=True, color="E8ECF8")
        img_size = 150
        img_col = "T"
    else:
        title_fill = PatternFill("solid", fgColor="F3EAF8")
        cream = PatternFill("solid", fgColor="FFF8F2")
        body_font = Font(name="Calibri", size=10, color="6B5B7A")
        title_font = Font(name="Georgia", size=12, bold=True, color=COLOR_NAVY)
        img_size = 78
        img_col = "S"

    title_cell = ws.cell(start_row, col_start)
    title_cell.value = card["title"]
    ws.merge_cells(start_row=start_row, start_column=col_start, end_row=start_row, end_column=col_start + 3)
    title_cell = ws.cell(start_row, col_start)
    _style_cell(
        title_cell,
        font=title_font,
        fill=title_fill,
        alignment=Alignment(horizontal="left", vertical="center"),
        border=BORDER_THIN,
    )
    for c in range(col_start + 1, col_start + 4):
        _style_cell(ws.cell(start_row, c), fill=title_fill, border=BORDER_THIN)

    body_row = start_row + 1
    body_span = 6 if theme_name in {"bmw_night", "tropical_yacht", "deep_space"} else 4
    ws.merge_cells(
        start_row=body_row,
        start_column=col_start,
        end_row=body_row + body_span,
        end_column=col_start + 3,
    )
    body_cell = ws.cell(body_row, col_start)
    body_cell.value = card["body"]
    _style_cell(
        body_cell,
        font=body_font,
        fill=cream,
        alignment=Alignment(horizontal="left", vertical="top", wrap_text=True),
        border=BORDER_THIN,
    )
    for r in range(body_row, body_row + body_span + 1):
        for c in range(col_start, col_start + 4):
            if r == body_row and c == col_start:
                continue
            _style_cell(ws.cell(r, c), fill=cream, border=BORDER_THIN)
        ws.row_dimensions[r].height = max(ws.row_dimensions[r].height or 15, 20)

    col_widths = (("O", 20), ("P", 16), ("Q", 16), ("R", 16), ("S", 14), ("T", 18), ("U", 4))
    if theme_name in {"tropical_yacht", "deep_space"}:
        col_widths = (("O", 20), ("P", 16), ("Q", 16), ("R", 16), ("S", 14), ("T", 24), ("U", 4))
    for letter, width in col_widths:
        ws.column_dimensions[letter].width = max(ws.column_dimensions[letter].width or 0, width)

    _add_image(ws, image_path, f"{img_col}{start_row}", width=img_size, height=img_size)


def _add_image(ws, path: Path, anchor: str, width: int | None = None, height: int | None = None) -> None:
    if not path.exists():
        return
    img = XLImage(str(path))
    native_w, native_h = img.width, img.height
    if width is not None and height is not None and native_w and native_h:
        # Keep native aspect ratio: fit inside the requested box instead of stretching.
        scale = min(width / native_w, height / native_h)
        img.width = max(1, int(native_w * scale))
        img.height = max(1, int(native_h * scale))
    elif width is not None and native_w:
        img.width = width
        img.height = max(1, int(native_h * width / native_w)) if native_h else width
    elif height is not None and native_h:
        img.height = height
        img.width = max(1, int(native_w * height / native_h)) if native_w else height
    ws.add_image(img, anchor)


def write_lavender_cover_sheet(
    ws,
    client: str,
    report_date: date,
    summary_title: str,
    in_work_count: int,
    shipped_count: int,
) -> None:
    ws.sheet_view.showGridLines = False
    cream = PatternFill("solid", fgColor="FFF8F2")
    lilac = PatternFill("solid", fgColor="F3EAF8")
    for row in range(1, 36):
        ws.row_dimensions[row].height = 18
        for col in range(1, 14):
            cell = ws.cell(row, col)
            cell.fill = cream if row < 18 else lilac
    for col in range(1, 14):
        ws.column_dimensions[get_column_letter(col)].width = 11

    ws.merge_cells("A2:L2")
    title = ws["A2"]
    title.value = f"♡  {client}  ♡"
    _style_cell(
        title,
        font=Font(name="Georgia", size=28, bold=True, color="6B5B7A"),
        alignment=Alignment(horizontal="center", vertical="center"),
    )
    ws.row_dimensions[2].height = 40

    ws.merge_cells("A3:L3")
    subtitle = ws["A3"]
    subtitle.value = "лавандовый раф · нежный статус заказов"
    _style_cell(
        subtitle,
        font=Font(name="Georgia", size=14, italic=True, color="9B7EAD"),
        alignment=Alignment(horizontal="center", vertical="center"),
    )

    ws.merge_cells("A4:L4")
    meta = ws["A4"]
    meta.value = (
        f"{summary_title}  ·  {report_date.strftime('%d.%m.%Y')}  ·  "
        f"в работе {in_work_count}  ·  отгружено {shipped_count}"
    )
    _style_cell(
        meta,
        font=Font(name="Calibri", size=11, color="8A7398"),
        alignment=Alignment(horizontal="center", vertical="center"),
    )

    _add_image(ws, LAVENDER_BANNER, "B6", width=780, height=420)
    _add_image(ws, LAVENDER_CUP, "K28", width=160, height=160)

    ws.merge_cells("A30:J31")
    note = ws["A30"]
    note.value = lavender_cover_epigraph(report_date, in_work_count, shipped_count)
    _style_cell(
        note,
        font=Font(name="Georgia", size=11, italic=True, color="6B5B7A"),
        fill=PatternFill("solid", fgColor="F7F1FA"),
        alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
    )


def decorate_lavender_sheets(wb: Workbook, summary_sheet_name: str | None) -> None:
    if "Total" in wb.sheetnames:
        _add_image(wb["Total"], LAVENDER_CUP, "I1", width=110, height=110)
    if summary_sheet_name and summary_sheet_name in wb.sheetnames:
        _add_image(wb[summary_sheet_name], LAVENDER_CUP, "L1", width=100, height=100)
    if "В работе" in wb.sheetnames:
        _add_image(wb["В работе"], LAVENDER_CUP, "AT1", width=80, height=80)
    if "Отгружено" in wb.sheetnames:
        _add_image(wb["Отгружено"], LAVENDER_CUP, "AT1", width=80, height=80)


def como_cover_epigraph(report_date: date, in_work: int, shipped: int) -> str:
    options = [
        (
            "бокал просекко на набережной Комо: пастель, лимон и лёгкий золотой отсвет — "
            "тот же уют, только в отчёте по заказам 🥂"
        ),
        (
            f"сегодня на перилах — статус {report_date.strftime('%d.%m')}: "
            f"{in_work} в работе, {shipped} finished-отгрузок. cin cin ✨"
        ),
        (
            "виллы не торопятся, озеро дышит, мы считаем заказы маленькими глотками 🍋"
        ),
        (
            "SHIPPED гуляет по набережной как «в работе»; FINISHED — уже тост на террасе 🌿"
        ),
        (
            "инструкция сомелье нежного отчёта: смотреть цифры без паники, "
            "хвалить команду, не расплёскивать пузырьки 🥂"
        ),
    ]
    digest = hashlib.md5(report_date.isoformat().encode()).hexdigest()
    return options[int(digest[:8], 16) % len(options)]


def write_como_cover_sheet(
    ws,
    client: str,
    report_date: date,
    summary_title: str,
    in_work_count: int,
    shipped_count: int,
) -> None:
    ws.sheet_view.showGridLines = False
    cream = PatternFill("solid", fgColor="FBF4EA")
    mist = PatternFill("solid", fgColor="E8F2F4")
    sage = PatternFill("solid", fgColor="F3F7F2")
    for row in range(1, 42):
        ws.row_dimensions[row].height = 18
        for col in range(1, 15):
            cell = ws.cell(row, col)
            if row < 6:
                cell.fill = cream
            elif row < 22:
                cell.fill = mist
            else:
                cell.fill = sage
    for col in range(1, 15):
        ws.column_dimensions[get_column_letter(col)].width = 11

    ws.merge_cells("A2:N2")
    title = ws["A2"]
    title.value = f"🥂  {client}  🥂"
    _style_cell(
        title,
        font=Font(name="Georgia", size=28, bold=True, color="4A5D6A"),
        alignment=Alignment(horizontal="center", vertical="center"),
    )
    ws.row_dimensions[2].height = 40

    ws.merge_cells("A3:N3")
    subtitle = ws["A3"]
    subtitle.value = "просекко на набережной озера Комо · пастельный статус заказов"
    _style_cell(
        subtitle,
        font=Font(name="Georgia", size=13, italic=True, color="7A8F8A"),
        alignment=Alignment(horizontal="center", vertical="center"),
    )

    ws.merge_cells("A4:N4")
    meta = ws["A4"]
    meta.value = (
        f"{summary_title}  ·  {report_date.strftime('%d.%m.%Y')}  ·  "
        f"в работе {in_work_count}  ·  отгружено (FINISHED) {shipped_count}"
    )
    _style_cell(
        meta,
        font=Font(name="Calibri", size=11, color="6A7B78"),
        alignment=Alignment(horizontal="center", vertical="center"),
    )

    _add_image(ws, COMO_BANNER, "B6", width=720, height=400)
    _add_image(ws, COMO_GLASS, "L6", width=130, height=130)
    _add_image(ws, COMO_LEMON, "L14", width=110, height=110)
    _add_image(ws, COMO_PROMENADE, "B24", width=520, height=280)
    _add_image(ws, COMO_VILLA, "J24", width=120, height=120)
    _add_image(ws, COMO_LAKE, "L24", width=120, height=120)
    _add_image(ws, COMO_TERRACE, "J32", width=120, height=120)

    ws.merge_cells("A38:I40")
    note = ws["A38"]
    note.value = como_cover_epigraph(report_date, in_work_count, shipped_count)
    _style_cell(
        note,
        font=Font(name="Georgia", size=11, italic=True, color="4A5D6A"),
        fill=PatternFill("solid", fgColor="FBF4EA"),
        alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
    )
    ws.row_dimensions[38].height = 22
    ws.row_dimensions[39].height = 22


def decorate_como_sheets(wb: Workbook, summary_sheet_name: str | None) -> None:
    if "Total" in wb.sheetnames:
        _add_image(wb["Total"], COMO_GLASS, "I1", width=108, height=108)
        _add_image(wb["Total"], COMO_LEMON, "K1", width=72, height=72)
    if summary_sheet_name and summary_sheet_name in wb.sheetnames:
        ws = wb[summary_sheet_name]
        _add_image(ws, COMO_GLASS, "L1", width=92, height=92)
        _add_image(ws, COMO_VILLA, "M1", width=72, height=72)
        _add_image(ws, COMO_LAKE, "T4", width=88, height=88)
        _add_image(ws, COMO_TERRACE, "T20", width=88, height=88)
        _add_image(ws, COMO_LEMON, "T36", width=80, height=80)
        ws.column_dimensions["T"].width = 14
        ws.column_dimensions["U"].width = 12
    if "В работе" in wb.sheetnames:
        _add_image(wb["В работе"], COMO_LAKE, "AT1", width=78, height=78)
        _add_image(wb["В работе"], COMO_VILLA, "AV1", width=64, height=64)
    if "Отгружено" in wb.sheetnames:
        _add_image(wb["Отгружено"], COMO_GLASS, "AT1", width=78, height=78)
        _add_image(wb["Отгружено"], COMO_TERRACE, "AV1", width=64, height=64)


def montecarlo_cover_epigraph(report_date: date, in_work: int, shipped: int) -> str:
    options = [
        (
            "Ferrari 250 GTO у казино Монте-Карло: кармин, золото и влажный асфальт — "
            "та же роскошь, только в отчёте по заказам ♦"
        ),
        (
            f"сегодня на Place du Casino — статус {report_date.strftime('%d.%m')}: "
            f"{in_work} в работе, {shipped} finished-отгрузок. банк открыт ✨"
        ),
        (
            "не блефуем цифрами — ставим на ясность, терпение и высокий стиль ♠️"
        ),
        (
            "SHIPPED гуляет по залу как «в работе»; FINISHED — уже валет на столе 🏎️"
        ),
        (
            "инструкция сомелье гран-при: смотреть кассу без суеты, "
            "хвалить команду, не расплёскивать шампанское 🥂"
        ),
    ]
    digest = hashlib.md5(report_date.isoformat().encode()).hexdigest()
    return options[int(digest[:8], 16) % len(options)]


def write_montecarlo_cover_sheet(
    ws,
    client: str,
    report_date: date,
    summary_title: str,
    in_work_count: int,
    shipped_count: int,
) -> None:
    ws.sheet_view.showGridLines = False
    cream = PatternFill("solid", fgColor="F8EED9")
    crimson = PatternFill("solid", fgColor="3B0A14")
    gold = PatternFill("solid", fgColor="F0E0C8")
    for row in range(1, 44):
        ws.row_dimensions[row].height = 18
        for col in range(1, 15):
            cell = ws.cell(row, col)
            if row < 6:
                cell.fill = cream
            elif row < 23:
                cell.fill = crimson
            else:
                cell.fill = gold
    for col in range(1, 15):
        ws.column_dimensions[get_column_letter(col)].width = 11

    ws.merge_cells("A2:N2")
    title = ws["A2"]
    title.value = f"♦  {client}  ♦"
    _style_cell(
        title,
        font=Font(name="Georgia", size=28, bold=True, color="3B0A14"),
        alignment=Alignment(horizontal="center", vertical="center"),
    )
    ws.row_dimensions[2].height = 42

    ws.merge_cells("A3:N3")
    subtitle = ws["A3"]
    subtitle.value = "Ferrari 250 GTO · Casino de Monte-Carlo · weekly grand prix status"
    _style_cell(
        subtitle,
        font=Font(name="Georgia", size=13, italic=True, color="8B1E2D"),
        alignment=Alignment(horizontal="center", vertical="center"),
    )

    ws.merge_cells("A4:N4")
    meta = ws["A4"]
    meta.value = (
        f"{summary_title}  ·  {report_date.strftime('%d.%m.%Y')}  ·  "
        f"в работе {in_work_count}  ·  отгружено (FINISHED) {shipped_count}"
    )
    _style_cell(
        meta,
        font=Font(name="Calibri", size=11, color="5C3B2E"),
        alignment=Alignment(horizontal="center", vertical="center"),
    )

    _add_image(ws, MONTE_BANNER, "B6", width=720, height=400)
    _add_image(ws, MONTE_FERRARI, "L6", width=130, height=130)
    _add_image(ws, MONTE_CHAMPAGNE, "L14", width=110, height=110)
    _add_image(ws, MONTE_PROMENADE, "B24", width=520, height=280)
    _add_image(ws, MONTE_CASINO, "J24", width=120, height=120)
    _add_image(ws, MONTE_ROULETTE, "L24", width=120, height=120)
    _add_image(ws, MONTE_GOLD, "J32", width=120, height=120)

    ws.merge_cells("A38:I41")
    note = ws["A38"]
    note.value = montecarlo_cover_epigraph(report_date, in_work_count, shipped_count)
    _style_cell(
        note,
        font=Font(name="Georgia", size=11, italic=True, color="F8EED9"),
        fill=PatternFill("solid", fgColor="3B0A14"),
        alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
    )
    for r in (38, 39, 40, 41):
        ws.row_dimensions[r].height = 20


def decorate_montecarlo_sheets(wb: Workbook, summary_sheet_name: str | None) -> None:
    if "Total" in wb.sheetnames:
        _add_image(wb["Total"], MONTE_FERRARI, "I1", width=110, height=110)
        _add_image(wb["Total"], MONTE_CHAMPAGNE, "K1", width=72, height=72)
    if summary_sheet_name and summary_sheet_name in wb.sheetnames:
        ws = wb[summary_sheet_name]
        _add_image(ws, MONTE_FERRARI, "L1", width=92, height=92)
        _add_image(ws, MONTE_CASINO, "M1", width=72, height=72)
        _add_image(ws, MONTE_ROULETTE, "T4", width=88, height=88)
        _add_image(ws, MONTE_GOLD, "T20", width=88, height=88)
        _add_image(ws, MONTE_CHAMPAGNE, "T36", width=80, height=80)
        ws.column_dimensions["T"].width = 14
        ws.column_dimensions["U"].width = 12
    if "В работе" in wb.sheetnames:
        _add_image(wb["В работе"], MONTE_CASINO, "AT1", width=78, height=78)
        _add_image(wb["В работе"], MONTE_ROULETTE, "AV1", width=64, height=64)
    if "Отгружено" in wb.sheetnames:
        _add_image(wb["Отгружено"], MONTE_FERRARI, "AT1", width=78, height=78)
        _add_image(wb["Отгружено"], MONTE_GOLD, "AV1", width=64, height=64)


def yacht_cover_epigraph(report_date: date, in_work: int, shipped: int) -> str:
    options = [
        (
            "парусная яхта в тропической бухте: бирюза, золото и лёгкий бриз — "
            "та же роскошь, только в отчёте по заказам ⛵"
        ),
        (
            f"сегодня на палубе — статус {report_date.strftime('%d.%m')}: "
            f"{in_work} в работе, {shipped} finished-отгрузок. касса открыта 💰"
        ),
        (
            "не штормим цифрами — держим курс на ясность, терпение и высокий стиль ⛵"
        ),
        (
            "SHIPPED гуляет по заливу как «в работе»; FINISHED — уже фуршет на борт 🥂"
        ),
        (
            "инструкция капитана: смотреть кассу без суеты, "
            "хвалить команду, не расплёскивать шампанское на палубе 💰"
        ),
    ]
    digest = hashlib.md5(report_date.isoformat().encode()).hexdigest()
    return options[int(digest[:8], 16) % len(options)]


def write_yacht_cover_sheet(
    ws,
    client: str,
    report_date: date,
    summary_title: str,
    in_work_count: int,
    shipped_count: int,
) -> None:
    ws.sheet_view.showGridLines = False
    sand = PatternFill("solid", fgColor="FBF3E0")
    teal = PatternFill("solid", fgColor="0B3D4C")
    aqua = PatternFill("solid", fgColor="E0F4F7")
    for row in range(1, 44):
        ws.row_dimensions[row].height = 18
        for col in range(1, 15):
            cell = ws.cell(row, col)
            if row < 6:
                cell.fill = sand
            elif row < 23:
                cell.fill = teal
            else:
                cell.fill = aqua
    for col in range(1, 15):
        ws.column_dimensions[get_column_letter(col)].width = 11

    ws.merge_cells("A2:N2")
    title = ws["A2"]
    title.value = f"⛵  {client}  ⛵"
    _style_cell(
        title,
        font=Font(name="Georgia", size=28, bold=True, color="0B3D4C"),
        alignment=Alignment(horizontal="center", vertical="center"),
    )
    ws.row_dimensions[2].height = 42

    ws.merge_cells("A3:N3")
    subtitle = ws["A3"]
    subtitle.value = "парусная яхта · тропическая бухта · фуршет на богатом · weekly regatta status"
    _style_cell(
        subtitle,
        font=Font(name="Georgia", size=13, italic=True, color="1A7A8C"),
        alignment=Alignment(horizontal="center", vertical="center"),
    )

    ws.merge_cells("A4:N4")
    meta = ws["A4"]
    meta.value = (
        f"{summary_title}  ·  {report_date.strftime('%d.%m.%Y')}  ·  "
        f"в работе {in_work_count}  ·  отгружено (FINISHED) {shipped_count}"
    )
    _style_cell(
        meta,
        font=Font(name="Calibri", size=11, color="2E6B7A"),
        alignment=Alignment(horizontal="center", vertical="center"),
    )

    _add_image(ws, YACHT_BANNER, "B6", width=720, height=405)
    _add_image(ws, YACHT_SAIL, "L6", width=140, height=140)
    _add_image(ws, YACHT_BUFFET, "L14", width=120, height=120)
    _add_image(ws, YACHT_BAY, "B24", width=520, height=300)
    _add_image(ws, YACHT_MONEY, "J24", width=130, height=130)
    _add_image(ws, YACHT_CHAMPAGNE, "L24", width=130, height=130)

    ws.merge_cells("A38:I41")
    note = ws["A38"]
    note.value = yacht_cover_epigraph(report_date, in_work_count, shipped_count)
    _style_cell(
        note,
        font=Font(name="Georgia", size=11, italic=True, color="E0F4F7"),
        fill=PatternFill("solid", fgColor="0B3D4C"),
        alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
    )
    for r in (38, 39, 40, 41):
        ws.row_dimensions[r].height = 20


def decorate_yacht_sheets(wb: Workbook, summary_sheet_name: str | None) -> None:
    if "Total" in wb.sheetnames:
        _add_image(wb["Total"], YACHT_SAIL, "I1", width=120, height=120)
        _add_image(wb["Total"], YACHT_MONEY, "K1", width=88, height=88)
    # Summary: do not pin extra images to T-column — mood cards already live there.
    if "В работе" in wb.sheetnames:
        _add_image(wb["В работе"], YACHT_BUFFET, "AT1", width=96, height=96)
        _add_image(wb["В работе"], YACHT_MONEY, "AV1", width=80, height=80)
    if "Отгружено" in wb.sheetnames:
        _add_image(wb["Отгружено"], YACHT_SAIL, "AT1", width=96, height=96)
        _add_image(wb["Отгружено"], YACHT_CHAMPAGNE, "AV1", width=80, height=80)


def bmw_cover_epigraph(report_date: date, in_work: int, shipped: int) -> str:
    options = [
        (
            "BMW 8 Series на кремлёвской набережной Казани: сапфир, Куул-Шариф и лёгкий дым кальяна — "
            "тот же люкс, только в отчёте по заказам 🚗"
        ),
        (
            f"сегодня у Казанки — статус {report_date.strftime('%d.%m')}: "
            f"{in_work} в работе, {shipped} finished-отгрузок. угли горят, круиз-контроль включён ✨"
        ),
        (
            "не тормозим на цифрах — держим курс на ясность, терпение и ночной казанский стиль 🌃"
        ),
        (
            "SHIPPED едет по набережной как «в работе»; FINISHED — уже в кальянной у реки 🚗"
        ),
        (
            "инструкция night drive: смотреть кассу без суеты, "
            "хвалить команду, не тушить кальян раньше времени 🔥"
        ),
    ]
    digest = hashlib.md5(report_date.isoformat().encode()).hexdigest()
    return options[int(digest[:8], 16) % len(options)]


def write_bmw_cover_sheet(
    ws,
    client: str,
    report_date: date,
    summary_title: str,
    in_work_count: int,
    shipped_count: int,
) -> None:
    ws.sheet_view.showGridLines = False
    silver = PatternFill("solid", fgColor="E8EDFA")
    navy = PatternFill("solid", fgColor="0D1B2A")
    mist = PatternFill("solid", fgColor="E8EEF4")
    for row in range(1, 48):
        ws.row_dimensions[row].height = 18
        for col in range(1, 15):
            cell = ws.cell(row, col)
            if row < 6:
                cell.fill = silver
            elif row < 23:
                cell.fill = navy
            else:
                cell.fill = mist
    for col in range(1, 15):
        ws.column_dimensions[get_column_letter(col)].width = 11

    ws.merge_cells("A2:N2")
    title = ws["A2"]
    title.value = f"🚗  {client}  🚗"
    _style_cell(
        title,
        font=Font(name="Georgia", size=28, bold=True, color="E8EEF4"),
        fill=navy,
        alignment=Alignment(horizontal="center", vertical="center"),
    )
    ws.row_dimensions[2].height = 42

    ws.merge_cells("A3:N3")
    subtitle = ws["A3"]
    subtitle.value = "BMW 8 Series · набережная Казани · кальянная · weekly night drive"
    _style_cell(
        subtitle,
        font=Font(name="Georgia", size=13, italic=True, color="5C7AEA"),
        fill=navy,
        alignment=Alignment(horizontal="center", vertical="center"),
    )

    ws.merge_cells("A4:N4")
    meta = ws["A4"]
    meta.value = (
        f"{summary_title}  ·  {report_date.strftime('%d.%m.%Y')}  ·  "
        f"в работе {in_work_count}  ·  отгружено (FINISHED) {shipped_count}"
    )
    _style_cell(
        meta,
        font=Font(name="Calibri", size=11, color="A8B8CC"),
        fill=navy,
        alignment=Alignment(horizontal="center", vertical="center"),
    )

    # Banner keeps 16:9; thumbs stay square — no stretch.
    _add_image(ws, BMW_BANNER, "B6", width=720, height=405)
    _add_image(ws, BMW8, "L6", width=140, height=140)
    _add_image(ws, BMW_HOOKAH, "L14", width=120, height=120)
    _add_image(ws, BMW_PROMENADE, "B25", width=520, height=300)
    _add_image(ws, BMW_RIVER, "J25", width=130, height=130)
    _add_image(ws, BMW_LIGHTS, "L25", width=130, height=130)

    ws.merge_cells("A42:I45")
    note = ws["A42"]
    note.value = bmw_cover_epigraph(report_date, in_work_count, shipped_count)
    _style_cell(
        note,
        font=Font(name="Georgia", size=11, italic=True, color="E8EEF4"),
        fill=PatternFill("solid", fgColor="1B3A5C"),
        alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
    )
    for r in (42, 43, 44, 45):
        ws.row_dimensions[r].height = 20


def decorate_bmw_sheets(wb: Workbook, summary_sheet_name: str | None) -> None:
    if "Total" in wb.sheetnames:
        _add_image(wb["Total"], BMW8, "I1", width=110, height=110)
        _add_image(wb["Total"], BMW_HOOKAH, "K1", width=88, height=88)
    # Summary: do not pin extra images to T-column — mood cards already live there.
    if "В работе" in wb.sheetnames:
        _add_image(wb["В работе"], BMW_HOOKAH, "AT1", width=96, height=96)
        _add_image(wb["В работе"], BMW_RIVER, "AV1", width=80, height=80)
    if "Отгружено" in wb.sheetnames:
        _add_image(wb["Отгружено"], BMW8, "AT1", width=96, height=96)
        _add_image(wb["Отгружено"], BMW_LIGHTS, "AV1", width=80, height=80)


def space_cover_epigraph(report_date: date, in_work: int, shipped: int) -> str:
    options = [
        (
            "глубокий космос: чёрные дыры, спирали галактик и золотые звёзды — "
            "тот же люкс, только в отчёте по заказам ✦"
        ),
        (
            f"сегодня у горизонта событий — статус {report_date.strftime('%d.%m')}: "
            f"{in_work} в работе, {shipped} finished-отгрузок. курс на дальний мир включён ✨"
        ),
        (
            "не тормозим на цифрах — держим орбиту ясности, терпения и покорения дальних миров 🌌"
        ),
        (
            "SHIPPED дрейфует как «в работе»; FINISHED — уже на новой планете ✦"
        ),
        (
            "инструкция deep space: смотреть кассу без суеты, "
            "хвалить команду, не гасить маяк раньше времени ✦"
        ),
    ]
    digest = hashlib.md5(report_date.isoformat().encode()).hexdigest()
    return options[int(digest[:8], 16) % len(options)]


def write_space_cover_sheet(
    ws,
    client: str,
    report_date: date,
    summary_title: str,
    in_work_count: int,
    shipped_count: int,
) -> None:
    ws.sheet_view.showGridLines = False
    cream = PatternFill("solid", fgColor="F8F1DC")
    navy = PatternFill("solid", fgColor="070B1A")
    mist = PatternFill("solid", fgColor="E8ECF8")
    for row in range(1, 48):
        ws.row_dimensions[row].height = 18
        for col in range(1, 15):
            cell = ws.cell(row, col)
            if row < 6:
                cell.fill = cream
            elif row < 23:
                cell.fill = navy
            else:
                cell.fill = mist
    for col in range(1, 15):
        ws.column_dimensions[get_column_letter(col)].width = 11

    ws.merge_cells("A2:N2")
    title = ws["A2"]
    title.value = f"✦  {client}  ✦"
    _style_cell(
        title,
        font=Font(name="Georgia", size=28, bold=True, color="E8ECF8"),
        fill=navy,
        alignment=Alignment(horizontal="center", vertical="center"),
    )
    ws.row_dimensions[2].height = 42

    ws.merge_cells("A3:N3")
    subtitle = ws["A3"]
    subtitle.value = "космос · чёрные дыры · галактики · звёзды · покорение дальних миров"
    _style_cell(
        subtitle,
        font=Font(name="Georgia", size=13, italic=True, color="C9A227"),
        fill=navy,
        alignment=Alignment(horizontal="center", vertical="center"),
    )

    ws.merge_cells("A4:N4")
    meta = ws["A4"]
    meta.value = (
        f"{summary_title}  ·  {report_date.strftime('%d.%m.%Y')}  ·  "
        f"в работе {in_work_count}  ·  отгружено (FINISHED) {shipped_count}"
    )
    _style_cell(
        meta,
        font=Font(name="Calibri", size=11, color="A8B0CC"),
        fill=navy,
        alignment=Alignment(horizontal="center", vertical="center"),
    )

    _add_image(ws, SPACE_BANNER, "B6", width=720, height=405)
    _add_image(ws, SPACE_BLACKHOLE, "L6", width=140, height=140)
    _add_image(ws, SPACE_GALAXY, "L14", width=120, height=120)
    _add_image(ws, SPACE_NEBULA, "B25", width=520, height=300)
    _add_image(ws, SPACE_STARS, "J25", width=130, height=130)
    _add_image(ws, SPACE_VOYAGE, "L25", width=130, height=130)

    ws.merge_cells("A42:I45")
    note = ws["A42"]
    note.value = space_cover_epigraph(report_date, in_work_count, shipped_count)
    _style_cell(
        note,
        font=Font(name="Georgia", size=11, italic=True, color="E8ECF8"),
        fill=PatternFill("solid", fgColor="1A2A6B"),
        alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
    )
    for r in (42, 43, 44, 45):
        ws.row_dimensions[r].height = 20


def decorate_space_sheets(wb: Workbook, summary_sheet_name: str | None) -> None:
    if "Total" in wb.sheetnames:
        _add_image(wb["Total"], SPACE_GALAXY, "I1", width=120, height=120)
        _add_image(wb["Total"], SPACE_STARS, "K1", width=88, height=88)
    if "В работе" in wb.sheetnames:
        _add_image(wb["В работе"], SPACE_BLACKHOLE, "AT1", width=96, height=96)
        _add_image(wb["В работе"], SPACE_VOYAGE, "AV1", width=80, height=80)
    if "Отгружено" in wb.sheetnames:
        _add_image(wb["Отгружено"], SPACE_GALAXY, "AT1", width=96, height=96)
        _add_image(wb["Отгружено"], SPACE_STARS, "AV1", width=80, height=80)


def generate_report(
    input_path: Path,
    output_path: Path,
    client: str = "Utair",
    clients: list[str] | None = None,
    report_date: date | None = None,
    previous_input: Path | None = None,
    previous_report: Path | None = None,
    previous_date: date | None = None,
    week_days: int = 7,
    excluded_invoicers: set[str] | None = None,
    theme_name: str = "default",
    summary_title: str = "Сводка за неделю",
    summary_sheet_name: str = "Сводка за неделю",
) -> Path:
    report_date = report_date or date.today()
    theme = THEMES.get(theme_name, THEMES["default"])
    activate_theme(theme)

    report_clients = clients if clients else [client]
    multi_client = len(report_clients) > 1
    display_client = " + ".join(report_clients) if multi_client else client

    df = load_taz(input_path, excluded_invoicers=excluded_invoicers)
    if multi_client:
        client_df = filter_clients(df, report_clients)
        in_work_raw = filter_in_work_rows(client_df)
        shipped_raw = filter_shipped_rows(client_df)
        client_stats = [
            (
                c,
                len(filter_in_work(filter_client(client_df, c), c)),
                len(filter_shipped(filter_client(client_df, c), c)),
            )
            for c in report_clients
        ]
    else:
        client_df = filter_client(df, client)
        in_work_raw = filter_in_work(client_df, client)
        shipped_raw = filter_shipped(client_df, client)
        client_stats = None

    in_work_df = prepare_output_frame(in_work_raw)
    shipped_df = prepare_output_frame(shipped_raw)

    in_work_totals = aggregate(in_work_raw)
    shipped_totals = aggregate(shipped_raw)
    shipped_over_30 = shipped_balance_over_30_days(shipped_raw, report_date)

    wb = Workbook()
    if theme.name == "montecarlo_ferrari":
        cover = wb.active
        cover.title = "Обложка ♦"
        cover.sheet_properties.tabColor = "8B1E2D"
        write_montecarlo_cover_sheet(
            cover,
            display_client,
            report_date,
            summary_title,
            len(in_work_df),
            len(shipped_df),
        )
        total_ws = wb.create_sheet("Total", 1)
    elif theme.name == "como_prosecco":
        cover = wb.active
        cover.title = "Обложка 🥂"
        cover.sheet_properties.tabColor = "E5D5B8"
        write_como_cover_sheet(
            cover,
            display_client,
            report_date,
            summary_title,
            len(in_work_df),
            len(shipped_df),
        )
        total_ws = wb.create_sheet("Total", 1)
    elif theme.name == "tropical_yacht":
        cover = wb.active
        cover.title = "Обложка ⛵"
        cover.sheet_properties.tabColor = "1A7A8C"
        write_yacht_cover_sheet(
            cover,
            display_client,
            report_date,
            summary_title,
            len(in_work_df),
            len(shipped_df),
        )
        total_ws = wb.create_sheet("Total", 1)
    elif theme.name == "bmw_night":
        cover = wb.active
        cover.title = "Обложка 🚗"
        cover.sheet_properties.tabColor = "1B3A5C"
        write_bmw_cover_sheet(
            cover,
            display_client,
            report_date,
            summary_title,
            len(in_work_df),
            len(shipped_df),
        )
        total_ws = wb.create_sheet("Total", 1)
    elif theme.name == "deep_space":
        cover = wb.active
        cover.title = "Обложка ✦"
        cover.sheet_properties.tabColor = "1A2A6B"
        write_space_cover_sheet(
            cover,
            display_client,
            report_date,
            summary_title,
            len(in_work_df),
            len(shipped_df),
        )
        total_ws = wb.create_sheet("Total", 1)
    elif theme.name == "lavender_raf":
        cover = wb.active
        cover.title = "Обложка ♡"
        cover.sheet_properties.tabColor = "E8D5F0"
        write_lavender_cover_sheet(
            cover,
            display_client,
            report_date,
            summary_title,
            len(in_work_df),
            len(shipped_df),
        )
        total_ws = wb.create_sheet("Total", 1)
    else:
        total_ws = wb.active
        total_ws.title = "Total"
    total_ws.sheet_properties.tabColor = theme.tab_total
    write_total_sheet(
        total_ws,
        display_client,
        report_date,
        in_work_totals,
        shipped_totals,
        shipped_over_30,
        len(in_work_df),
        len(shipped_df),
        client_stats=client_stats,
    )

    include_weekly = previous_input or previous_report or previous_date is not None
    if include_weekly:
        if previous_input:
            prev_full = load_taz(previous_input, excluded_invoicers=excluded_invoicers)
        elif previous_report:
            prev_full = load_report_snapshot_as_previous(previous_report)
        else:
            prev_full = client_df.iloc[0:0].copy()
        week_end = report_date
        week_start = previous_date or (report_date - timedelta(days=week_days))
        if multi_client:
            summary = merge_weekly_summaries([
                (
                    c,
                    build_weekly_summary(
                        filter_client(client_df, c),
                        filter_client(prev_full, c),
                        week_start,
                        week_end,
                        client=c,
                    ),
                )
                for c in report_clients
            ])
        else:
            if previous_input or previous_report:
                previous_df = filter_client(prev_full, client)
            else:
                previous_df = client_df.iloc[0:0].copy()
            summary = build_weekly_summary(client_df, previous_df, week_start, week_end, client=client)
        weekly_ws = wb.create_sheet(summary_sheet_name)
        weekly_ws.sheet_properties.tabColor = theme.tab_summary
        write_weekly_summary_sheet(
            weekly_ws,
            display_client,
            summary,
            sheet_title=summary_title,
            cute_comments=(theme.name in CUTE_THEMES),
            theme_name=theme.name,
        )
    else:
        summary_sheet_name = None

    shipped_ws = wb.create_sheet("Отгружено")
    shipped_ws.sheet_properties.tabColor = theme.tab_shipped
    write_detail_sheet(shipped_ws, shipped_df, shipped_totals)

    in_work_ws = wb.create_sheet("В работе")
    in_work_ws.sheet_properties.tabColor = theme.tab_in_work
    write_detail_sheet(in_work_ws, in_work_df, in_work_totals)

    if theme.name == "montecarlo_ferrari":
        decorate_montecarlo_sheets(wb, summary_sheet_name)
    elif theme.name == "como_prosecco":
        decorate_como_sheets(wb, summary_sheet_name)
    elif theme.name == "lavender_raf":
        decorate_lavender_sheets(wb, summary_sheet_name)
    elif theme.name == "tropical_yacht":
        decorate_yacht_sheets(wb, summary_sheet_name)
    elif theme.name == "bmw_night":
        decorate_bmw_sheets(wb, summary_sheet_name)
    elif theme.name == "deep_space":
        decorate_space_sheets(wb, summary_sheet_name)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def parse_report_date(value: str) -> date:
    for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    raise argparse.ArgumentTypeError(f"Invalid date: {value}. Use DD.MM.YYYY or YYYY-MM-DD.")


def default_output_name(client: str, report_date: date, clients: list[str] | None = None) -> str:
    label = " + ".join(clients) if clients and len(clients) > 1 else client
    safe_client = re.sub(r'[\\/:*?"<>|]', "-", label.strip())
    return f"{safe_client} статус {report_date.strftime('%d.%m.%Y')}.xlsx"


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate client status Excel report from TAZ export.")
    parser.add_argument("--input", "-i", type=Path, required=True, help="Path to TAZ .xlsx file")
    parser.add_argument("--output", "-o", type=Path, help="Output .xlsx path")
    parser.add_argument("--client", "-c", default="Utair", help="Customer name (default: Utair)")
    parser.add_argument(
        "--clients",
        type=str,
        help="Several customers in one report, comma-separated (e.g. 'S7 Engineering,АК Сибирь')",
    )
    parser.add_argument(
        "--date",
        "-d",
        type=parse_report_date,
        default=date.today(),
        help="Report date for filename and >30 days calc (default: today)",
    )
    parser.add_argument(
        "--previous-input",
        type=Path,
        help="Previous TAZ export for weekly summary comparison",
    )
    parser.add_argument(
        "--previous-report",
        type=Path,
        help="Previous generated report .xlsx (fallback if previous TAZ is unavailable)",
    )
    parser.add_argument(
        "--previous-date",
        type=parse_report_date,
        help="Start date of weekly summary period (default: report date minus 7 days)",
    )
    parser.add_argument(
        "--week-days",
        type=int,
        default=7,
        help="Length of summary period in days if --previous-date is not set (default: 7)",
    )
    parser.add_argument(
        "--include-fe",
        action="store_true",
        help="Include Invoicer=ФЭ rows (needed for clients like Белавиа)",
    )
    parser.add_argument(
        "--theme",
        choices=sorted(THEMES.keys()),
        default="default",
        help="Visual theme (tropical_yacht / bmw_night / montecarlo_ferrari / como_prosecco for Aeroflot)",
    )
    parser.add_argument(
        "--summary-title",
        default="Сводка за неделю",
        help="Title for the period summary sheet",
    )
    parser.add_argument(
        "--summary-sheet",
        default="Сводка за неделю",
        help="Worksheet name for the period summary",
    )
    args = parser.parse_args()

    excluded = set() if args.include_fe else set(EXCLUDED_INVOICERS)
    clients = [c.strip() for c in args.clients.split(",") if c.strip()] if args.clients else None
    output = args.output or Path("reports") / default_output_name(args.client, args.date, clients)
    result = generate_report(
        args.input,
        output,
        client=args.client,
        clients=clients,
        report_date=args.date,
        previous_input=args.previous_input,
        previous_report=args.previous_report,
        previous_date=args.previous_date,
        week_days=args.week_days,
        excluded_invoicers=excluded,
        theme_name=args.theme,
        summary_title=args.summary_title,
        summary_sheet_name=args.summary_sheet,
    )
    print(f"Report saved: {result}")
    display = " + ".join(clients) if clients else args.client
    print(f"Client: {display}")
    loaded = filter_clients(load_taz(args.input, excluded_invoicers=excluded), clients) if clients else filter_client(load_taz(args.input, excluded_invoicers=excluded), args.client)
    if clients:
        print(f"В работе rows: {len(filter_in_work_rows(loaded))}")
        print(f"Отгружено rows: {len(filter_shipped_rows(loaded))}")
        for c in clients:
            sub = filter_client(loaded, c)
            print(f"  {c}: в работе {len(filter_in_work(sub, c))}, отгружено {len(filter_shipped(sub, c))}")
    else:
        print(f"В работе rows: {len(filter_in_work(loaded, args.client))}")
        print(f"Отгружено rows: {len(filter_shipped(loaded, args.client))}")
    if args.previous_input or args.previous_report or args.previous_date:
        print(f"Summary sheet: {args.summary_sheet}")
    if args.include_fe:
        print("Invoicer ФЭ: included")
    if args.theme != "default":
        print(f"Theme: {args.theme}")


if __name__ == "__main__":
    main()
