#!/usr/bin/env python3
"""
Гибридный анализ склада Азур (US list):
- агрегация: 1 строка на P/N, qty = число уникальных S/N (S/N уникален)
- пометка нестандартных состояний (SCRAP / BER / EXPIRED)
- тип ВС из ТУЗ/ТАЗ
- возможность ремонта по CCL (точное P/N + по смыслу), как HBS↔CCL
- ликвидность A–D как у склада Ютэйр (ТАЗ/ТУЗ)

Один лист в выходном Excel.
"""
from __future__ import annotations

import re
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import analyze_liquidity as al
import hbs_ccl_match as cclm

DATA = Path("/workspace/data")
OUT_DIR = Path("/workspace/output")
ART_DIR = Path("/opt/cursor/artifacts")
UPLOADS = Path("/home/ubuntu/.cursor/projects/workspace/uploads")

AZUR_FILE = DATA / "Azur_US_list_10.04.26.xlsx"
OUT_NAME = "Azur_US_hybrid_analysis.xlsx"

SPECIAL_CONDITIONS = {"SCRAP", "BER", "EXPIRED"}
EXCLUDE_SEMANTIC_CONCEPTS = {
    "fuel control unit",
    "flap track",
    "portable fire extinguisher",
    "engine fire extinguisher",
}

THIN = Border(
    left=Side(style="thin", color="E5E5E5"),
    right=Side(style="thin", color="E5E5E5"),
    top=Side(style="thin", color="E5E5E5"),
    bottom=Side(style="thin", color="E5E5E5"),
)
GRADE_FILL = al.GRADE_FILL
GRADE_FONT = al.GRADE_FONT
ZEBRA = PatternFill("solid", fgColor="F7F9FC")
MARK_FILL = PatternFill("solid", fgColor="FCE4D6")
EXACT_FILL = PatternFill("solid", fgColor="C6EFCE")
NAME_FILL = PatternFill("solid", fgColor="FFF2CC")
NO_REPAIR_FILL = PatternFill("solid", fgColor="F2F2F2")


HEADERS = [
    ("P/N", 18),
    ("Описание", 36),
    ("Кол-во", 8),
    ("Состояние", 14),
    ("Пометка", 28),
    ("Причина (Repair Shelf)", 36),
    ("Remarks", 24),
    ("Store", 12),
    ("Owner", 14),
    ("AC_Reg (склад)", 18),
    ("Тип ВС (рынок)", 16),
    ("Ремонт у нас", 14),
    ("Тип совпадения CCL", 28),
    ("Уверенность ремонта", 14),
    ("Примеры CCL P/N", 28),
    ("CCL наименование", 34),
    ("Комментарий (ремонт)", 42),
    ("Ликвидность", 12),
    ("Ориентир. цена USD", 14),
    ("Потенц. выручка USD", 14),
    ("Уверенность в цене", 26),
    ("Спрос (заказы/запросы)", 44),
    ("Обоснование ликвидности", 70),
]


def find_ccl_path() -> Optional[Path]:
    patterns = [
        "*Перечен*.xlsx",
        "*перечен*.xlsx",
        "*CCL*.xlsx",
        "*ccl*.xlsx",
        "*Фастэйр*.xlsx",
        "*Fastair*.xlsx",
        "*обслуживаем*.xlsx",
        "*________________*4_09*.xlsx",
        "*4_09*.xlsx",
    ]
    dirs = [DATA, UPLOADS, Path("/workspace"), Path("/opt/cursor/artifacts")]
    candidates: list[Path] = []
    for d in dirs:
        if not d.exists():
            continue
        for pat in patterns:
            candidates.extend(d.glob(pat))
    # prefer names that look like the CCL catalog
    scored = []
    for p in candidates:
        name = p.name.lower()
        if "hbs" in name and "ccl" not in name and "перечен" not in name:
            continue
        if p.name.startswith("Azur") or "US_list" in p.name or "TAZ" in p.name or "TUZ" in p.name:
            continue
        score = 0
        if "перечен" in name or "обслужиж" in name or "фаст" in name:
            score += 10
        if "4_09" in name or "рев" in name:
            score += 5
        if "ccl" in name:
            score += 3
        scored.append((score, p.stat().st_mtime, p))
    scored.sort(reverse=True)
    return scored[0][2] if scored and scored[0][0] > 0 else (scored[0][2] if scored else None)


def load_azur_rows(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
    out = []
    for r in rows_iter:
        d = dict(zip(headers, r))
        pn_raw = d.get("PN")
        pn = al.norm_pn(pn_raw)
        if not pn:
            continue
        sn = str(d.get("Serial_Number") or "").strip()
        if sn.lower() in {"none", "nan", "null", "-"}:
            sn = ""
        qty_raw = d.get("Qty")
        try:
            qty = float(qty_raw) if qty_raw not in (None, "") else None
        except (TypeError, ValueError):
            qty = None
        cond = str(d.get("Condition") or "").strip().upper().replace("U/S", "US")
        if cond == "U/S":
            cond = "US"
        out.append(
            {
                "pn": pn,
                "partno": str(pn_raw).strip() if pn_raw is not None else pn,
                "sn": sn,
                "qty": qty,
                "description": str(d.get("Description") or "").strip(),
                "condition": cond,
                "reason": str(d.get("Repair_Shelf_Reason") or "").strip(),
                "remarks": str(d.get("Remarks") or "").strip(),
                "store": str(d.get("Store") or "").strip(),
                "owner": str(d.get("Owner") or "").strip(),
                "ac_reg": str(d.get("AC_Reg") or "").strip(),
            }
        )
    wb.close()
    return out


def component_count(row: dict) -> float:
    """S/N уникален → 1 шт. Без S/N берём Qty (если есть), иначе 1."""
    if row["sn"]:
        return 1.0
    if row["qty"] is not None and row["qty"] > 0:
        return float(row["qty"])
    return 1.0


def aggregate_azur(rows: list[dict]) -> list[dict]:
    """Одна строка на P/N; qty = число уникальных компонентов (уник. S/N)."""
    # unique components first
    seen_sn: set[tuple[str, str]] = set()
    components: list[dict] = []
    no_sn_rows: list[dict] = []
    dup_sn = 0
    for r in rows:
        if r["sn"]:
            key = (r["pn"], r["sn"])
            if key in seen_sn:
                dup_sn += 1
                continue
            seen_sn.add(key)
            components.append(r)
        else:
            no_sn_rows.append(r)

    groups: dict[str, dict] = {}
    order: list[str] = []

    def ensure(pn: str, sample: dict) -> dict:
        if pn not in groups:
            groups[pn] = {
                "pn": pn,
                "partno": sample["partno"],
                "qty": 0.0,
                "descriptions": Counter(),
                "conditions": Counter(),
                "reasons": Counter(),
                "remarks": Counter(),
                "stores": Counter(),
                "owners": Counter(),
                "ac_regs": Counter(),
                "special": set(),
                "lines": 0,
            }
            order.append(pn)
        return groups[pn]

    for r in components:
        g = ensure(r["pn"], r)
        g["qty"] += 1.0
        g["lines"] += 1
        if r["description"]:
            g["descriptions"][r["description"]] += 1
        if r["condition"]:
            g["conditions"][r["condition"]] += 1
            if r["condition"] in SPECIAL_CONDITIONS:
                g["special"].add(r["condition"])
        if r["reason"]:
            g["reasons"][r["reason"]] += 1
        if r["remarks"]:
            g["remarks"][r["remarks"]] += 1
        if r["store"]:
            g["stores"][r["store"]] += 1
        if r["owner"]:
            g["owners"][r["owner"]] += 1
        if r["ac_reg"] and r["ac_reg"].upper() not in {"NA", "N/A", "-", "NONE"}:
            g["ac_regs"][r["ac_reg"]] += 1

    for r in no_sn_rows:
        g = ensure(r["pn"], r)
        g["qty"] += component_count(r)
        g["lines"] += 1
        if r["description"]:
            g["descriptions"][r["description"]] += 1
        if r["condition"]:
            g["conditions"][r["condition"]] += 1
            if r["condition"] in SPECIAL_CONDITIONS:
                g["special"].add(r["condition"])
        if r["reason"]:
            g["reasons"][r["reason"]] += 1
        if r["remarks"]:
            g["remarks"][r["remarks"]] += 1
        if r["store"]:
            g["stores"][r["store"]] += 1
        if r["owner"]:
            g["owners"][r["owner"]] += 1
        if r["ac_reg"] and r["ac_reg"].upper() not in {"NA", "N/A", "-", "NONE"}:
            g["ac_regs"][r["ac_reg"]] += 1

    out = []
    for pn in order:
        g = groups[pn]
        conds = [c for c, _ in g["conditions"].most_common()]
        special = sorted(g["special"])
        mark = ""
        if special:
            mark = "Нестандартное состояние: " + ", ".join(special)
        out.append(
            {
                "pn": pn,
                "partno": g["partno"],
                "qty": g["qty"],
                "description": g["descriptions"].most_common(1)[0][0] if g["descriptions"] else "",
                "condition": ", ".join(conds),
                "mark": mark,
                "special": special,
                "reason": "; ".join(x for x, _ in g["reasons"].most_common(3)),
                "remarks": "; ".join(x for x, _ in g["remarks"].most_common(3)),
                "store": ", ".join(x for x, _ in g["stores"].most_common()),
                "owner": ", ".join(x for x, _ in g["owners"].most_common()),
                "ac_reg": ", ".join(x for x, _ in g["ac_regs"].most_common(8)),
                "lines": g["lines"],
            }
        )
    print(
        f"Azur: raw={len(rows)}, unique SN components={len(components)}, "
        f"no-SN rows={len(no_sn_rows)}, dup SN skipped={dup_sn}, unique P/N={len(out)}"
    )
    return out


def _clean_ac(value: Any, pn: str = "") -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if not s or s.lower() in {"nan", "none", "n/a", "-", "тип вс", "a/c", "ac"}:
        return ""
    if len(s) > 40:
        return ""
    if pn and al.soft_pn_key(s.upper()) == al.soft_pn_key(pn):
        return ""
    return s


def build_ac_type_index(taz: list[al.Event], tuz_ac: dict[str, set[str]]) -> dict[str, set[str]]:
    """pn -> set of AC types from market (merged with soft/alt via resolve later)."""
    by_pn: dict[str, set[str]] = defaultdict(set)
    for e in taz:
        # ac from side channel if present
        ac = getattr(e, "ac_type", "") or ""
        if ac:
            by_pn[e.pn].add(ac)
        if e.alt:
            if ac:
                by_pn[e.alt].add(ac)
    for pn, types in tuz_ac.items():
        by_pn[pn] |= types
    return by_pn


def load_taz_with_ac(path: Path) -> tuple[list[al.Event], dict[str, set[str]]]:
    events = al.load_taz(path)
    ac_by_pn: dict[str, set[str]] = defaultdict(set)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "ORDERS" not in wb.sheetnames:
        wb.close()
        return events, ac_by_pn
    ws = wb["ORDERS"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return events, ac_by_pn
    aliases = {
        **al.TAZ_ALIASES,
        "ac": ["тип вс (продажи)", "тип вс", "a/c", "ac type", "aircraft type"],
    }
    mapping = al._header_map_from_row(rows[0], aliases)
    if "pn" not in mapping:
        for idx in range(min(3, len(rows))):
            mapping = al._header_map_from_row(rows[idx], aliases)
            if "pn" in mapping:
                break
    start = 1
    for i, row in enumerate(rows[:5]):
        if row and any(al.clean_header_cell(c) in {"p/n", "part number", "pn"} for c in row[:15] if c):
            start = i + 1
            break

    def cell(r, key, default=None):
        idx = mapping.get(key)
        if idx is None or idx >= len(r):
            return default
        return r[idx]

    for r in rows[start:]:
        if not r:
            continue
        pn = al.norm_pn(cell(r, "pn"))
        if al.is_sample_pn(pn):
            continue
        ac = _clean_ac(cell(r, "ac"), pn)
        if ac:
            ac_by_pn[pn].add(ac)
            alt = al.norm_pn(cell(r, "alt"))
            if alt and alt != pn:
                ac_by_pn[alt].add(ac)
    return events, ac_by_pn


def load_tuz_with_ac(path: Path) -> tuple[list[al.Event], dict[str, set[str]]]:
    events = al.load_tuz(path)
    ac_by_pn: dict[str, set[str]] = defaultdict(set)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sheet in wb.sheetnames:
        if sheet.strip().lower() in al.TUZ_SKIP_SHEETS:
            continue
        ws = wb[sheet]
        preview = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            preview.append(row)
            if i >= 4:
                break
        if not preview:
            continue
        aliases = {
            **al.TUZ_ALIASES,
            "ac": ["a/c", "ac", "aircraft", "тип вс", "ac type"],
        }
        hdr_idx, mapping = al.find_header_map(preview, aliases, scan=5)
        if "pn" not in mapping:
            continue
        # stream rest
        # re-open sheet iteration
        rows_all = list(ws.iter_rows(values_only=True))
        layouts = al._tuz_layouts(mapping)

        def cell_from(parsed, key):
            return parsed.get(key)

        for r in rows_all[hdr_idx + 1 :]:
            parsed = al._extract_tuz_row(r, layouts)
            if not parsed:
                continue
            pn = parsed.get("pn") or ""
            if not pn:
                continue
            # try AC from mapping directly if layouts miss it
            ac = ""
            if "ac" in mapping and mapping["ac"] < len(r):
                ac = _clean_ac(r[mapping["ac"]], pn)
            if not ac:
                # classic A/C col often 5
                for idx in (5, 7, mapping.get("ac")):
                    if idx is None or idx >= len(r):
                        continue
                    ac = _clean_ac(r[idx], pn)
                    if ac:
                        break
            if ac:
                ac_by_pn[pn].add(ac)
                alt = parsed.get("alt") or ""
                if alt and alt != pn:
                    ac_by_pn[alt].add(ac)
    wb.close()
    return events, ac_by_pn


def resolve_ac_types(pn: str, ac_by_pn: dict[str, set[str]], soft_to_pns, alt_to_pns) -> str:
    types: set[str] = set()
    candidates = {pn}
    if pn in alt_to_pns:
        candidates |= alt_to_pns[pn]
    soft = al.soft_pn_key(pn)
    candidates |= soft_to_pns.get(soft, set())
    if pn.isdigit():
        stripped = pn.lstrip("0") or "0"
        candidates.add(stripped)
    for c in candidates:
        types |= ac_by_pn.get(c, set())
    return ", ".join(sorted(types))


def match_repair(stock: list[dict], ccl_path: Optional[Path]) -> dict[str, dict]:
    """Return pn -> repair info dict."""
    empty = {
        "repair": "н/д",
        "match_type": "",
        "confidence": "",
        "ccl_pns": "",
        "ccl_name": "",
        "comment": "Каталог CCL не загружен — загрузите Перечень обслуживаемых компонентов Фастэйр",
    }
    if not ccl_path or not ccl_path.exists():
        return {r["pn"]: dict(empty) for r in stock}

    print(f"Loading CCL: {ccl_path.name}")
    ccl = cclm.load_ccl(ccl_path)
    print(f"  CCL rows: {len(ccl)}")

    # build mini HBS-like frame
    import pandas as pd

    hbs = pd.DataFrame(
        [
            {
                "pn": r["partno"],
                "name": r["description"] or "",
                "ata": "",
                "ac_type": "",
                "go_cat": "",
            }
            for r in stock
        ]
    )
    # keep original pn key
    hbs["_stock_pn"] = [r["pn"] for r in stock]

    result = cclm.match_hbs_ccl(hbs, ccl)
    out: dict[str, dict] = {}

    exact_by_idx: dict[Any, list] = defaultdict(list)
    for i, m in result["exact"]:
        exact_by_idx[i].append(m)

    semantic_by_idx = {i: matches for i, matches in result["semantic"]}

    for i, r in hbs.iterrows():
        pn = r["_stock_pn"]
        if i in exact_by_idx:
            ms = exact_by_idx[i]
            examples = [str(m["pn"]) for m in ms[:5]]
            name = str(ms[0]["name"])
            out[pn] = {
                "repair": "Да",
                "match_type": "Точное совпадение P/N",
                "confidence": "Высокая",
                "ccl_pns": ", ".join(examples),
                "ccl_name": name,
                "comment": "P/N полностью совпадает с перечнем CCL",
            }
            continue

        matches = semantic_by_idx.get(i, [])
        # filter excluded concepts
        filtered = [(c, m) for c, m in matches if c not in EXCLUDE_SEMANTIC_CONCEPTS]
        if filtered:
            concepts = sorted({c for c, _ in filtered})
            examples = []
            seen = set()
            name = ""
            for c, m in filtered:
                if m["pn"] in seen:
                    continue
                seen.add(m["pn"])
                examples.append(str(m["pn"]))
                if not name:
                    name = str(m["name"])
                if len(examples) >= 5:
                    break
            conf = "Высокая"
            if concepts == ["battery"] or any(c == "battery" for c in concepts):
                conf = "Средняя"
            out[pn] = {
                "repair": "Возможно",
                "match_type": "Совпадение по наименованию (P/N другой)",
                "confidence": conf,
                "ccl_pns": ", ".join(examples),
                "ccl_name": name,
                "comment": "Совпадение по типу: " + ", ".join(concepts),
            }
            continue

        out[pn] = {
            "repair": "Нет",
            "match_type": "",
            "confidence": "",
            "ccl_pns": "",
            "ccl_name": "",
            "comment": "В перечне CCL не найдено (ни точного P/N, ни по смыслу)",
        }
    return out


def event_global_key(e: al.Event):
    if e.kind == "order":
        inv = al.normalize_invoice(e.request_no) or (
            f"NO_INV|{al.client_key(e.client)}|{al.day_key(e.date)}|{round(e.qty or 0, 4)}"
        )
        return ("O", e.pn, inv)
    return (
        "R",
        e.pn,
        al.client_key(e.client),
        al.day_key(e.date),
        round(e.qty or 0, 4),
        round(e.price or 0, 2),
        (e.description or "")[:40],
    )


def load_market():
    taz_files = sorted(DATA.glob("TAZ*.xlsx")) + sorted(DATA.glob("TA3*.xlsx")) + sorted(
        DATA.glob("ТАЗ*.xlsx")
    )
    tuz_files = sorted(DATA.glob("TUZ*.xlsx")) + sorted(DATA.glob("ТУЗ*.xlsx"))
    # de-dupe paths
    taz_files = list(dict.fromkeys(taz_files))
    tuz_files = list(dict.fromkeys(tuz_files))

    all_events: list[al.Event] = []
    ac_by_pn: dict[str, set[str]] = defaultdict(set)

    for path in taz_files:
        part, ac = load_taz_with_ac(path)
        print(f"  TAZ {path.name}: {len(part)} events, AC P/Ns={len(ac)}")
        all_events.extend(part)
        for pn, types in ac.items():
            ac_by_pn[pn] |= types

    for path in tuz_files:
        part, ac = load_tuz_with_ac(path)
        print(f"  TUZ {path.name}: {len(part)} events, AC P/Ns={len(ac)}")
        all_events.extend(part)
        for pn, types in ac.items():
            ac_by_pn[pn] |= types

    # optional EXP if present (Utair-style)
    exp_n = 0
    for path in sorted(DATA.glob("*.csv")) + sorted(DATA.glob("EXPENDABLES*.xlsx")):
        if path.suffix.lower() == ".csv":
            part = al.load_exp_csv(path)
        else:
            part = al.load_exp(path)
        print(f"  EXP {path.name}: {len(part)}")
        all_events.extend(part)
        exp_n += len(part)

    seen = set()
    deduped = []
    for e in all_events:
        k = event_global_key(e)
        if k in seen:
            continue
        seen.add(k)
        deduped.append(e)
    print(f"  Market events deduped: {len(deduped)} (raw {len(all_events)}), EXP contrib={exp_n}")
    by_pn, soft_to_pns, alt_to_pns = al.build_market_index(deduped)
    return by_pn, soft_to_pns, alt_to_pns, ac_by_pn, len(deduped)


def score_stock(stock: list[dict], by_pn, soft_to_pns, alt_to_pns, ac_by_pn, repair_map):
    scored = []
    for item in stock:
        market = al.resolve_market(item["pn"], by_pn, soft_to_pns, alt_to_pns)
        # fake stock dict for build_row_from_stock
        stock_like = {
            "pn": item["pn"],
            "partno": item["partno"],
            "qty": item["qty"],
            "lines": item["lines"],
            "conditions": set(c.strip() for c in item["condition"].split(",") if c.strip()),
            "description": item["description"],
            "ac_typs": set(),
            "section": al.condition_section(
                next(iter(item["condition"].split(",")), "").strip() if item["condition"] else ""
            ),
        }
        row = al.build_row_from_stock(stock_like, market)
        ac_market = resolve_ac_types(item["pn"], ac_by_pn, soft_to_pns, alt_to_pns)
        rep = repair_map.get(item["pn"], {})
        scored.append(
            {
                **item,
                "ac_typ_market": ac_market,
                "liquidity_grade": row["liquidity_grade"],
                "liquidity_score": row["liquidity_score"],
                "price_ref_usd": row["price_ref_usd"],
                "potential_revenue_usd": row["potential_revenue_usd"],
                "price_confidence_unified": row["price_confidence_unified"],
                "demand_summary": row["demand_summary"],
                "rationale": row["rationale"],
                "repair": rep.get("repair", ""),
                "match_type": rep.get("match_type", ""),
                "repair_confidence": rep.get("confidence", ""),
                "ccl_pns": rep.get("ccl_pns", ""),
                "ccl_name": rep.get("ccl_name", ""),
                "repair_comment": rep.get("comment", ""),
            }
        )
    return scored


def write_excel(path: Path, scored: list[dict], meta: dict):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Анализ Азур"

    # title row then headers
    ws.merge_cells("A1:W1")
    ws["A1"] = (
        f"Склад Азур (US list) — гибридный анализ | {datetime.now():%Y-%m-%d %H:%M} | "
        f"P/N: {meta['pn_n']} | компонентов: {meta['qty_sum']:g} | "
        f"рынок событий: {meta['market_n']} | CCL: {meta['ccl_label']}"
    )
    ws["A1"].font = Font(bold=True, size=13, name="Calibri")
    ws.row_dimensions[1].height = 28

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    for col, (name, width) in enumerate(HEADERS, 1):
        cell = ws.cell(2, col, name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = THIN
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[2].height = 36
    ws.freeze_panes = "A3"

    # sort: repair yes first, then liquidity, then revenue
    repair_rank = {"Да": 0, "Возможно": 1, "Нет": 2, "н/д": 3}

    def sort_key(r):
        return (
            repair_rank.get(r.get("repair"), 9),
            al.GRADE_ORDER.get(r["liquidity_grade"], 9),
            -(r.get("potential_revenue_usd") or 0),
            -r.get("liquidity_score", 0),
            -r["qty"],
            r["partno"],
        )

    rows = sorted(scored, key=sort_key)
    conf_fill = {
        "высокая": PatternFill("solid", fgColor="1F7A4D"),
        "средняя": PatternFill("solid", fgColor="2F6FED"),
        "низкая": PatternFill("solid", fgColor="C47F00"),
        "н/п": PatternFill("solid", fgColor="8A8A8A"),
    }

    for i, r in enumerate(rows):
        values = [
            r["partno"],
            r["description"],
            r["qty"],
            r["condition"],
            r["mark"],
            r["reason"],
            r["remarks"],
            r["store"],
            r["owner"],
            r["ac_reg"],
            r["ac_typ_market"],
            r["repair"],
            r["match_type"],
            r["repair_confidence"],
            r["ccl_pns"],
            r["ccl_name"],
            r["repair_comment"],
            r["liquidity_grade"],
            r["price_ref_usd"],
            r.get("potential_revenue_usd"),
            r["price_confidence_unified"],
            r["demand_summary"],
            r["rationale"],
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(i + 3, col, val)
            cell.border = THIN
            cell.alignment = Alignment(vertical="center", wrap_text=col in {2, 5, 6, 17, 22, 23})
            if i % 2 == 1:
                cell.fill = ZEBRA
        # highlights
        if r["mark"]:
            ws.cell(i + 3, 5).fill = MARK_FILL
        grade = r["liquidity_grade"]
        gcell = ws.cell(i + 3, 18)
        gcell.fill = GRADE_FILL.get(grade, GRADE_FILL["D"])
        gcell.font = GRADE_FONT
        gcell.alignment = Alignment(horizontal="center", vertical="center")
        repair = r["repair"]
        rcell = ws.cell(i + 3, 12)
        if repair == "Да":
            rcell.fill = EXACT_FILL
        elif repair == "Возможно":
            rcell.fill = NAME_FILL
        else:
            rcell.fill = NO_REPAIR_FILL
        unified = (r.get("price_confidence_unified") or "").lower()
        for key, fill in conf_fill.items():
            if unified.startswith(key):
                ws.cell(i + 3, 21).fill = fill
                ws.cell(i + 3, 21).font = GRADE_FONT
                break
        for col in (3, 19, 20):
            cell = ws.cell(i + 3, col)
            if isinstance(cell.value, (int, float)):
                if col == 3:
                    cell.number_format = "0"
                else:
                    cell.number_format = '"$"#,##0.00'

    n = len(rows)
    if n:
        ws.auto_filter.ref = f"A2:W{n + 2}"

    # legend sheet? user asked 1 tab — keep methodology as rows at bottom? Better tiny note in row1.
    # Add second sheet would violate "1 вкладка". Stick to one.

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"Saved: {path}")
    return rows


def validate(scored: list[dict], raw_n: int, agg: list[dict]) -> list[str]:
    issues = []
    qty_sum = sum(r["qty"] for r in scored)
    if len(scored) != len(agg):
        issues.append(f"scored ({len(scored)}) != agg ({len(agg)})")
    if any(r["qty"] <= 0 for r in scored):
        issues.append("found qty<=0")
    # SN uniqueness already enforced; check no crazy qty vs raw
    if qty_sum > raw_n * 1.5:
        issues.append(f"qty_sum {qty_sum} suspiciously > raw {raw_n}")
    # special marks
    special_n = sum(1 for r in scored if r["mark"])
    print(f"Validation: P/N={len(scored)}, qty={qty_sum:g}, special marks={special_n}")
    grades = Counter(r["liquidity_grade"] for r in scored)
    print(f"  Liquidity: {dict(grades)}")
    repairs = Counter(r["repair"] for r in scored)
    print(f"  Repair: {dict(repairs)}")
    ac_filled = sum(1 for r in scored if r["ac_typ_market"])
    print(f"  AC type filled: {ac_filled}/{len(scored)}")
    price_n = sum(1 for r in scored if r["price_ref_usd"] is not None)
    print(f"  With price: {price_n}/{len(scored)}")
    return issues


def main():
    if not AZUR_FILE.exists():
        # try uploads
        up = UPLOADS / "US_list_as_of_10.04.26-export._ae8c.xlsx"
        if up.exists():
            AZUR_FILE.parent.mkdir(parents=True, exist_ok=True)
            import shutil

            shutil.copy(up, AZUR_FILE)
        else:
            print("Azur file not found", AZUR_FILE)
            sys.exit(1)

    print("Loading Azur warehouse...")
    raw = load_azur_rows(AZUR_FILE)
    stock = aggregate_azur(raw)

    print("Loading market (TAZ/TUZ)...")
    by_pn, soft_to_pns, alt_to_pns, ac_by_pn, market_n = load_market()

    ccl_path = find_ccl_path()
    if ccl_path:
        print(f"CCL found: {ccl_path}")
        ccl_label = ccl_path.name
    else:
        print("WARNING: CCL catalog not found — repair columns will be н/д")
        ccl_label = "НЕ ЗАГРУЖЕН"
    repair_map = match_repair(stock, ccl_path)

    print("Scoring liquidity...")
    scored = score_stock(stock, by_pn, soft_to_pns, alt_to_pns, ac_by_pn, repair_map)
    issues = validate(scored, len(raw), stock)
    if issues:
        print("VALIDATION ISSUES:", issues)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    ART_DIR.mkdir(parents=True, exist_ok=True)
    meta = {
        "pn_n": len(scored),
        "qty_sum": sum(r["qty"] for r in scored),
        "market_n": market_n,
        "ccl_label": ccl_label,
    }
    out = OUT_DIR / OUT_NAME
    write_excel(out, scored, meta)
    openpyxl.load_workbook(out).save(ART_DIR / OUT_NAME)

    if not ccl_path:
        print("\n*** Нужен файл CCL (Перечень обслуживаемых компонентов Фастэйр) для финализации ремонта ***")
        return 2
    return 0 if not issues else 1


if __name__ == "__main__":
    raise SystemExit(main())
