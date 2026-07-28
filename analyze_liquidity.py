#!/usr/bin/env python3
"""
Оценка ликвидности склада АТИ по рынку (ТАЗ / ТУЗ / EXPENDABLES).
"""
from __future__ import annotations

import re
import warnings
from datetime import datetime, date
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from statistics import median
from typing import Any, Iterable, Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

DATA = Path("/workspace/data")
OUT = Path("/opt/cursor/artifacts/ATI_liquidity_assessment.xlsx")
OUT_COPY = Path("/workspace/output/ATI_liquidity_assessment.xlsx")

SAMPLE_PN = {
    "SAMPLE",
    "ОБРАЗЕЦ",
    "DEFAULT",
    "P/N",
    "PN",
    "PART NUMBER",
    "PARTNUMBER",
    "-",
    "—",
    "N/A",
    "NA",
    "NONE",
    "NULL",
    "TEST",
}
SAMPLE_CLIENT = {
    "DEFAULT",
    "SAMPLE",
    "ОБРАЗЕЦ",
    "TBA",
    "TEST",
    "CLIENT",
}

# Внутренние/технические контрагенты — не считаем рыночным клиентом
INTERNAL_CLIENT_MARKERS = (
    "закупка на склад",
    "на склад",
    "warehouse",
    "stock purchase",
    "internal",
)


def is_market_client(client: str) -> bool:
    if not client:
        return False
    c = client.strip().lower()
    if client.upper() in SAMPLE_CLIENT or c in {x.lower() for x in SAMPLE_CLIENT}:
        return False
    return not any(m in c for m in INTERNAL_CLIENT_MARKERS)


def norm_pn(value: Any) -> str:
    if value is None:
        return ""
    # Excel часто отдаёт чисто числовой P/N как float 12345.0
    if isinstance(value, float):
        if value != value:  # NaN
            return ""
        if value == int(value) and abs(value) < 1e15:
            s = str(int(value))
        else:
            s = str(value).strip().upper()
    elif isinstance(value, int):
        s = str(value)
    else:
        s = str(value).strip().upper()
    if not s:
        return ""
    # unicode dashes / spaces / zero-width
    s = (
        s.replace("\u2212", "-")
        .replace("\u2013", "-")
        .replace("\u2014", "-")
        .replace("\u00a0", "")
        .replace("\u200b", "")
        .replace("\t", "")
        .replace("\n", "")
        .replace("\r", "")
    )
    s = re.sub(r"\s+", "", s)
    # 12345.0 / 12345.000 из строк
    if re.fullmatch(r"\d+\.0+", s):
        s = s.split(".", 1)[0]
    # drop trailing punctuation noise
    s = s.strip(".,;:|/\\")
    return s


def soft_pn_key(pn: str) -> str:
    """Ключ без дефисов/точек — ловит опечатки в разделителях."""
    return re.sub(r"[^A-Z0-9]", "", pn)


def pn_variant_bases(pn: str) -> list[str]:
    """Базовые формы P/N без хвостовых суффиксов (-1, -1S, S и т.п.)."""
    out = []
    cur = pn
    for _ in range(4):
        if not cur or len(soft_pn_key(cur)) < 6:
            break
        out.append(cur)
        m = re.search(r"([-_/][A-Z0-9]{1,4}|[A-Z])$", cur, re.I)
        if not m:
            break
        nxt = cur[: m.start()]
        if not nxt or nxt == cur:
            break
        cur = nxt
    # уникальные с сохранением порядка
    seen = set()
    uniq = []
    for x in out:
        if x not in seen:
            seen.add(x)
            uniq.append(x)
    return uniq


def find_near_pn_matches(pn: str, by_pn: dict[str, MarketAgg], soft_to_pns: dict[str, set[str]]) -> list[tuple[str, str]]:
    """Близкие варианты P/N только по явным суффиксам (-1, -1S, WE, S).

    Не используем «отрезание последней цифры» — иначе MS28778-20 ↔ MS28778-2.
    """
    hits: list[tuple[str, str]] = []
    seen = set()
    if is_sample_pn(pn) or pn in {"-", "—", "."}:
        return hits
    soft = soft_pn_key(pn)
    if len(soft) < 6:
        return hits

    def add(mp: str, reason: str):
        if mp == pn or mp in seen or mp not in by_pn:
            return
        if is_sample_pn(mp):
            return
        if len(soft_pn_key(mp)) < 6:
            return
        seen.add(mp)
        hits.append((mp, reason))

    bases: list[str] = []
    # буквенные pack/variant суффиксы (часто WE)
    for lit in ("WE", "S"):
        if pn.upper().endswith(lit) and len(pn) > len(lit) + 5:
            bases.append(pn[: -len(lit)])
    # один хвостовой -N / -NN / -NLetter (601R57502-1, 601R13081-1S)
    m = re.search(r"[-_/]\d{1,2}[A-Z]?$", pn, re.I)
    if m:
        bases.append(pn[: m.start()])
    # двойной шаг: -1S → сначала S уже сняли, ещё -1
    for b in list(bases):
        m2 = re.search(r"[-_/]\d{1,2}[A-Z]?$", b, re.I)
        if m2 and len(soft_pn_key(b[: m2.start()])) >= 6:
            bases.append(b[: m2.start()])

    for base in bases:
        bsoft = soft_pn_key(base)
        if len(bsoft) < 6:
            continue
        if base in by_pn:
            add(base, f"суффикс: склад {pn} ↔ рынок {base}")
        for mp in soft_to_pns.get(bsoft, set()):
            if soft_pn_key(mp) == bsoft:
                add(mp, f"суффикс: склад {pn} ↔ рынок {mp}")

    # обратное: на рынке наш P/N + WE/S (склад без суффикса)
    if not hits:
        for lit in ("WE", "S"):
            cand = pn + lit
            if cand in by_pn:
                add(cand, f"суффикс: склад {pn} ↔ рынок {cand}")
            for mp in soft_to_pns.get(soft + lit, set()):
                if soft_pn_key(mp) == soft + lit:
                    add(mp, f"суффикс: склад {pn} ↔ рынок {mp}")

    return hits


def norm_client(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    s = re.sub(r"\s+", " ", s)
    return s


def is_sample_pn(pn: str) -> bool:
    if not pn or pn in SAMPLE_PN:
        return True
    if not re.search(r"[0-9]", pn):
        return True
    if len(pn) < 2:
        return True
    return False


def parse_money(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        if value != value:  # NaN
            return None
        return float(value)
    s = str(value).strip()
    if not s or s in {"-", "—", "n/a", "N/A"}:
        return None
    # В ТУЗ Offered часто «min\nmax» в одной ячейке — берём первую строку, не склеиваем
    if "\n" in s or "\r" in s:
        s = s.splitlines()[0].strip()
    s = s.replace("\u00a0", " ").replace("\u202f", " ")
    s = re.sub(r"[^\d,.\-]", "", s)
    if not s or s in {".", ",", "-", "-.", ".-"}:
        return None
    # European: 1.234,56 or 1 234,56 already stripped spaces
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        # 100,00 or 2939,77
        parts = s.split(",")
        if len(parts[-1]) <= 2:
            s = s.replace(",", ".")
        else:
            s = s.replace(",", "")
    try:
        v = float(s)
    except ValueError:
        return None
    # защита от мусора/склеек: единичная авиазапчасть > $5M почти наверняка ошибка парсинга
    if v > 5_000_000:
        return None
    return v


def parse_qty(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value) if value == value else 0.0
    m = re.search(r"-?\d+(?:[.,]\d+)?", str(value).replace("\u00a0", ""))
    if not m:
        return 0.0
    return parse_money(m.group(0)) or 0.0


def day_key(value: Any) -> str:
    """Календарный день для уникальности запроса."""
    if value is None or value == "":
        return "NO_DATE"
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    s = str(value).strip()
    if not s:
        return "NO_DATE"
    try:
        return datetime.fromisoformat(s.replace("Z", "").split(".")[0]).date().isoformat()
    except Exception:
        pass
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s[:10], fmt).date().isoformat()
        except Exception:
            continue
    return s[:10] if len(s) >= 8 else "NO_DATE"


def client_key(client: str) -> str:
    return (client or "").strip().lower()


def normalize_invoice(value: Any) -> str:
    """Нормализация №счёта: 14082512254.0 → 14082512254."""
    if value is None or value == "":
        return ""
    if isinstance(value, float):
        if value != value:
            return ""
        if value == int(value):
            return str(int(value))
        return str(value).strip()
    if isinstance(value, int):
        return str(value)
    s = str(value).strip()
    if re.fullmatch(r"\d+\.0+", s):
        return s.split(".", 1)[0]
    return s


def clean_header_cell(h: Any) -> str:
    if h is None:
        return ""
    return re.sub(r"\s+", " ", str(h).replace("\n", " ")).strip().lower()


def parse_date(value: Any) -> Optional[date]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if not s:
        return None
    try:
        return datetime.fromisoformat(s.replace("Z", "").split(".")[0]).date()
    except Exception:
        pass
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except Exception:
            continue
    return None


@dataclass
class Event:
    source: str  # TAZ / TUZ / EXP
    kind: str  # order / request
    pn: str
    alt: str = ""
    client: str = ""
    qty: float = 0.0
    price: Optional[float] = None
    condition: str = ""
    date: Any = None
    request_no: str = ""
    sheet: str = ""
    description: str = ""
    file: str = ""


@dataclass
class MarketAgg:
    order_keys: set = field(default_factory=set)
    order_qty: float = 0.0
    order_clients: set = field(default_factory=set)
    request_keys: set = field(default_factory=set)
    request_qty: float = 0.0
    request_clients: set = field(default_factory=set)
    order_price_points: list = field(default_factory=list)
    request_price_points: list = field(default_factory=list)
    descriptions: set = field(default_factory=set)
    conditions_seen: set = field(default_factory=set)
    matched_via: set = field(default_factory=set)
    sample_clients_order: list = field(default_factory=list)
    sample_clients_request: list = field(default_factory=list)
    request_keys_tuz: set = field(default_factory=set)
    request_keys_exp: set = field(default_factory=set)
    order_dates: list = field(default_factory=list)
    request_dates: list = field(default_factory=list)

    @property
    def order_events(self) -> int:
        return len(self.order_keys)

    @property
    def request_events(self) -> int:
        return len(self.request_keys)

    @property
    def order_prices(self) -> list:
        return [p for _, p in self.order_price_points]

    @property
    def request_prices(self) -> list:
        return [p for _, p in self.request_price_points]

    def merge_event(self, e: Event):
        if e.description:
            self.descriptions.add(str(e.description).strip()[:120])
        if e.condition:
            self.conditions_seen.add(str(e.condition).strip().upper())

        d = parse_date(e.date)

        if e.kind == "order":
            ono = normalize_invoice(e.request_no)
            if not ono:
                ono = f"NO_INV|{client_key(e.client)}|{day_key(e.date)}|{round(e.qty or 0, 4)}"
            # кросс-файловый дубль того же заказа — не считаем повторно
            if ono in self.order_keys:
                return
            self.order_keys.add(ono)
            self.order_qty += e.qty or 0
            if d is not None:
                self.order_dates.append(d)
            if is_market_client(e.client):
                self.order_clients.add(e.client)
                if len(self.sample_clients_order) < 8 and e.client not in self.sample_clients_order:
                    self.sample_clients_order.append(e.client)
            if e.price is not None and e.price > 0:
                self.order_price_points.append((d, float(e.price)))
            return

        ck = client_key(e.client)
        dk = day_key(e.date)
        rkey = (ck, dk)
        is_new = rkey not in self.request_keys
        if is_new:
            self.request_keys.add(rkey)
            self.request_qty += e.qty or 0
        else:
            # тот же клиент+день из другого файла/листа — только новая цена, без +1 к запросам
            if e.price is not None and e.price > 0:
                pt = (d, round(float(e.price), 4))
                existing = {(xd, round(float(xp), 4)) for xd, xp in self.request_price_points}
                if pt not in existing:
                    self.request_price_points.append((d, float(e.price)))
            return

        if d is not None:
            self.request_dates.append(d)
        if e.source == "EXP":
            self.request_keys_exp.add(rkey)
        else:
            self.request_keys_tuz.add(rkey)
        if is_market_client(e.client):
            self.request_clients.add(e.client)
            if len(self.sample_clients_request) < 8 and e.client not in self.sample_clients_request:
                self.sample_clients_request.append(e.client)
        if e.price is not None and e.price > 0:
            self.request_price_points.append((d, float(e.price)))

    def merge_agg(self, src: "MarketAgg"):
        new_orders = src.order_keys - self.order_keys
        if new_orders and src.order_keys:
            avg_q = src.order_qty / max(len(src.order_keys), 1)
            self.order_qty += avg_q * len(new_orders)
        self.order_keys |= src.order_keys
        self.order_clients |= src.order_clients

        new_reqs = src.request_keys - self.request_keys
        if new_reqs and src.request_keys:
            avg_q = src.request_qty / max(len(src.request_keys), 1)
            self.request_qty += avg_q * len(new_reqs)
        self.request_keys |= src.request_keys
        self.request_clients |= src.request_clients
        self.request_keys_tuz |= src.request_keys_tuz
        self.request_keys_exp |= src.request_keys_exp

        self.order_price_points.extend(src.order_price_points)
        self.request_price_points.extend(src.request_price_points)
        self.order_dates.extend(src.order_dates)
        self.request_dates.extend(src.request_dates)
        self.descriptions |= src.descriptions
        self.conditions_seen |= src.conditions_seen
        for c in src.sample_clients_order:
            if c not in self.sample_clients_order and len(self.sample_clients_order) < 8:
                self.sample_clients_order.append(c)
        for c in src.sample_clients_request:
            if c not in self.sample_clients_request and len(self.sample_clients_request) < 8:
                self.sample_clients_request.append(c)


def price_summary(prices: list[float]) -> tuple[Optional[float], Optional[float], Optional[float], int]:
    if not prices:
        return None, None, None, 0
    return min(prices), median(prices), max(prices), len(prices)


# Порог "недавно" для уверенности в цене
RECENT_DAYS = 180  # 6 месяцев
AS_OF_DATE = date(2026, 7, 21)


def _days_ago(d: Optional[date], as_of: date = AS_OF_DATE) -> Optional[int]:
    if d is None:
        return None
    return (as_of - d).days


def compute_indicative_price(market: MarketAgg) -> tuple[Optional[float], str, str]:
    """Возвращает (цена, уверенность, текст деталей для обоснования)."""
    as_of = AS_OF_DATE
    order_pts = [(d, p) for d, p in market.order_price_points if p and p > 0]
    req_pts = [(d, p) for d, p in market.request_price_points if p and p > 0]

    recent_orders = [(d, p) for d, p in order_pts if d and _days_ago(d, as_of) is not None and _days_ago(d, as_of) <= RECENT_DAYS]
    older_orders = [(d, p) for d, p in order_pts if (d, p) not in recent_orders]
    recent_reqs = [(d, p) for d, p in req_pts if d and _days_ago(d, as_of) is not None and _days_ago(d, as_of) <= RECENT_DAYS]
    older_reqs = [(d, p) for d, p in req_pts if (d, p) not in recent_reqs]

    def med(pts):
        return median([p for _, p in pts]) if pts else None

    source = None
    ref = None
    detail_parts = []

    if recent_orders:
        ref = med(recent_orders)
        source = "медиана продажных цен ТАЗ за последние 6 мес."
        prices = [p for _, p in recent_orders]
        detail_parts.append(
            f"Ориентир ${ref:,.2f}: {source} (n={len(recent_orders)}, "
            f"min ${min(prices):,.2f}, max ${max(prices):,.2f})."
        )
    elif order_pts:
        ref = med(order_pts)
        source = "медиана продажных цен ТАЗ (все даты)"
        prices = [p for _, p in order_pts]
        last_d = max((d for d, _ in order_pts if d), default=None)
        detail_parts.append(
            f"Ориентир ${ref:,.2f}: {source} (n={len(order_pts)}, "
            f"min ${min(prices):,.2f}, max ${max(prices):,.2f}"
            + (f", последний заказ {last_d.isoformat()}" if last_d else "")
            + ")."
        )
    elif recent_reqs:
        ref = med(recent_reqs)
        source = "медиана Offered/Sell Price запросов ТУЗ+EXP за последние 6 мес."
        prices = [p for _, p in recent_reqs]
        detail_parts.append(
            f"Ориентир ${ref:,.2f}: {source} (n={len(recent_reqs)}, "
            f"min ${min(prices):,.2f}, max ${max(prices):,.2f})."
        )
    elif req_pts:
        ref = med(req_pts)
        source = "медиана Offered/Sell Price запросов ТУЗ+EXP (все даты)"
        prices = [p for _, p in req_pts]
        last_d = max((d for d, _ in req_pts if d), default=None)
        detail_parts.append(
            f"Ориентир ${ref:,.2f}: {source} (n={len(req_pts)}, "
            f"min ${min(prices):,.2f}, max ${max(prices):,.2f}"
            + (f", последнее предложение {last_d.isoformat()}" if last_d else "")
            + ")."
        )
    else:
        detail_parts.append(
            "Ориентировочная цена не рассчитана: нет «Продажная, ед.» в ТАЗ и нет Offered/Sell Price в ТУЗ/EXP."
        )
        return None, "н/п", " ".join(detail_parts)

    # Уверенность
    last_order = max((d for d in market.order_dates if d), default=None)
    last_req = max((d for d in market.request_dates if d), default=None)
    last_order_price = max((d for d, _ in order_pts if d), default=None)
    last_req_price = max((d for d, _ in req_pts if d), default=None)

    if recent_orders:
        conf = "высокая"
        conf_why = (
            f"уверенность высокая: есть заказ(и) ТАЗ с продажной ценой за последние {RECENT_DAYS} дн."
            + (f" (последний {last_order_price.isoformat()})" if last_order_price else "")
        )
    elif order_pts:
        # есть заказы, но не свежие
        days = _days_ago(last_order_price or last_order, as_of)
        conf = "средняя"
        conf_why = (
            "уверенность средняя: продажная цена из ТАЗ есть, но заказ не свежий"
            + (f" ({days} дн. назад)" if days is not None else "")
        )
    elif recent_reqs:
        conf = "средняя"
        conf_why = (
            f"уверенность средняя: заказов ТАЗ нет, но есть свежие Offered/Sell в ТУЗ/EXP "
            f"(≤{RECENT_DAYS} дн.)"
            + (f", последнее {last_req_price.isoformat()}" if last_req_price else "")
        )
    else:
        conf = "низкая"
        days = _days_ago(last_req_price or last_req, as_of)
        conf_why = (
            "уверенность низкая: заказов ТАЗ нет, предложения ТУЗ/EXP давние"
            + (f" (последнее {days} дн. назад)" if days is not None else " или без даты")
        )

    detail_parts.append(conf_why + ".")
    return ref, conf, " ".join(detail_parts)



def liquidity_score(m: MarketAgg) -> tuple[float, str, str]:
    """Возвращает (score, grade A-D, rationale)."""
    n_ord_cli = len(m.order_clients)
    n_req_cli = len(m.request_clients)
    n_all_cli = len(m.order_clients | m.request_clients)
    n_req = m.request_events
    n_ord = m.order_events

    score = (
        n_ord * 12.0
        + n_ord_cli * 10.0
        + min(m.order_qty, 40) * 0.4
        + n_req * 3.0
        + n_req_cli * 5.0
        + min(m.request_qty, 80) * 0.15
    )

    if n_all_cli >= 3:
        score += 8
    if n_all_cli >= 5:
        score += 10

    if n_ord >= 3 and n_ord_cli >= 2:
        grade = "A"
    elif n_ord >= 1 and (n_req >= 1 or n_ord_cli >= 1):
        grade = "A" if (n_ord >= 2 or n_all_cli >= 3) else "B"
    elif n_req >= 5 and n_req_cli >= 2:
        grade = "B"
    elif n_req >= 2 or n_req_cli >= 2:
        grade = "C"
    elif n_req >= 1 or n_ord >= 1:
        grade = "C"
    else:
        grade = "D"
        score = 0.0

    if grade != "D":
        if score >= 55 and grade in {"B", "C"}:
            grade = "A"
        elif score >= 25 and grade == "C" and (n_ord >= 1 or n_req_cli >= 2):
            grade = "B"

    parts = []
    if n_ord:
        parts.append(
            f"заказов ТАЗ: {n_ord} (уник. №заказа+P/N; клиентов: {n_ord_cli}"
            + (f" — {', '.join(m.sample_clients_order[:5])}" if m.sample_clients_order else "")
            + f", qty≈{m.order_qty:g})"
        )
    if n_req:
        parts.append(
            f"запросов ТУЗ+EXP: {n_req} (уник. P/N+клиент+день; клиентов: {n_req_cli}"
            + (f" — {', '.join(m.sample_clients_request[:5])}" if m.sample_clients_request else "")
            + f", qty≈{m.request_qty:g})"
        )
    if not parts:
        parts.append("совпадений по рынку не найдено")

    via = ", ".join(sorted(m.matched_via)) if m.matched_via else "—"
    grade_ru = {"A": "высокая", "B": "средняя", "C": "низкая", "D": "нет спроса"}.get(grade, grade)
    rationale = f"Ликвидность {grade} ({grade_ru}). " + "; ".join(parts) + f". Сопоставление: {via}."
    return round(score, 1), grade, rationale


def find_header_map(rows: list[tuple], aliases: dict[str, list[str]], scan: int = 5) -> tuple[int, dict[str, int]]:
    """Ищет строку заголовков и маппинг колонок.

    Если в превью несколько строк-заголовков (классика и сдвинутая с Sales/Purch ID),
    выбираем наиболее полную / сдвинутую — она обычно соответствует реальным данным.
    """
    def clean(h):
        if h is None:
            return ""
        return re.sub(r"\s+", " ", str(h).replace("\n", " ")).strip().lower()

    best_idx, best_map, best_score = 0, {}, -1
    for idx in range(min(scan, len(rows))):
        row = rows[idx]
        if not row:
            continue
        cleaned = [clean(c) for c in row]
        mapping = {}
        for key, names in aliases.items():
            for i, cell in enumerate(cleaned):
                if not cell:
                    continue
                for name in names:
                    if cell == name or name in cell:
                        mapping[key] = i
                        break
                if key in mapping:
                    break
        if "pn" not in mapping:
            continue
        score = len(mapping)
        if any(c == "sales id" or c.startswith("sales id") for c in cleaned):
            score += 10  # сдвинутый layout
        if "price" in mapping:
            score += 5
        if "client" in mapping:
            score += 2
        if score > best_score:
            best_idx, best_map, best_score = idx, mapping, score
    return best_idx, best_map


def iter_sheet_rows(path: Path, sheet: str) -> Iterable[tuple]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    for row in ws.iter_rows(values_only=True):
        yield row
    wb.close()


def _header_map_from_row(row: tuple, aliases: dict[str, list[str]]) -> dict[str, int]:
    cleaned = [clean_header_cell(c) for c in row]
    mapping: dict[str, int] = {}
    for key, names in aliases.items():
        # 1) точное совпадение
        for name in names:
            for i, cell in enumerate(cleaned):
                if cell == name:
                    mapping[key] = i
                    break
            if key in mapping:
                break
        if key in mapping:
            continue
        # 2) вхождение; для pn не берём alt-*
        for name in names:
            for i, cell in enumerate(cleaned):
                if not cell or name not in cell:
                    continue
                if key == "pn" and "alt" in cell:
                    continue
                if key == "sell" and "total" in cell:
                    continue
                if key == "date" and cell in {"rfq", "email thread"}:
                    continue
                mapping[key] = i
                break
            if key in mapping:
                break
    return mapping


TAZ_ALIASES = {
    "invoice": ["номер счета", "номер счёта"],
    "invoice_fe": ["номер счета фэ", "номер счёта фэ"],
    "invoice_lr": ["номер счета лр", "номер счёта лр"],
    "customer": ["customer"],
    "pn": ["p/n", "part number", "partnumber"],
    "alt": ["alt p/n", "alt. p/n"],
    "desc": ["description"],
    "qty": ["qty in po", "qtyinpo"],
    "date": ["заказ взят в работу", "заказ клиента взят"],
    "cond": ["condition"],
    "sell": ["продажная, ед"],
}


def load_taz(path: Path) -> list[Event]:
    """Загрузка ORDERS из ТАЗ с автоопределением колонок (разные годы/шаблоны)."""
    events: list[Event] = []
    seen = set()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "ORDERS" not in wb.sheetnames:
        wb.close()
        return events
    ws = wb["ORDERS"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return events

    mapping = _header_map_from_row(rows[0], TAZ_ALIASES)
    # архивный шаблон: p/n во 2-й колонке, счета ФЭ/ЛР
    if "pn" not in mapping:
        for idx in range(min(3, len(rows))):
            mapping = _header_map_from_row(rows[idx], TAZ_ALIASES)
            if "pn" in mapping:
                break
    if "pn" not in mapping:
        print(f"  WARNING: no P/N column in {path.name} ORDERS — skipped")
        return events

    start = 1
    for i, row in enumerate(rows[:5]):
        if row and any(clean_header_cell(c) in {"p/n", "part number", "pn"} for c in row[:15] if c):
            start = i + 1

    def cell(r, key, default=None):
        idx = mapping.get(key)
        if idx is None or idx >= len(r):
            return default
        return r[idx]

    for r in rows[start:]:
        if not r:
            continue
        pn = norm_pn(cell(r, "pn"))
        if is_sample_pn(pn):
            continue
        client = norm_client(cell(r, "customer"))
        if client.upper() in SAMPLE_CLIENT:
            client = ""
        invoice = normalize_invoice(cell(r, "invoice"))
        if not invoice:
            invoice = normalize_invoice(cell(r, "invoice_fe")) or normalize_invoice(cell(r, "invoice_lr"))
        qty = parse_qty(cell(r, "qty"))
        price = parse_money(cell(r, "sell"))
        cond = str(cell(r, "cond") or "").strip()
        alt = norm_pn(cell(r, "alt"))
        desc = str(cell(r, "desc") or "").strip()
        date = cell(r, "date")
        key = ("TAZ", pn, invoice or f"NO_INV|{client_key(client)}|{day_key(date)}|{round(qty, 4)}")
        if key in seen:
            continue
        seen.add(key)
        if not client and not invoice and qty == 0:
            continue
        events.append(
            Event(
                source="TAZ",
                kind="order",
                pn=pn,
                alt=alt if alt and alt != pn else "",
                client=client,
                qty=qty,
                price=price,
                condition=cond,
                date=date,
                request_no=invoice,
                sheet="ORDERS",
                description=desc,
                file=path.name,
            )
        )
    return events


TUZ_SKIP_SHEETS = {
    "suppliers list",
    "dropdown",
    "customer list",
    "transport rates and terms",
    "questions old",
    "yak 2025",
}


TUZ_ALIASES = {
    "pn": ["p/n", "part number", "partnumber"],
    "alt": ["alt. p/n", "alt p/n", "alt pn", "alt.pn"],
    "client": ["client", "customer"],
    "desc": ["description"],
    "qty": ["qty"],
    "cond": ["cond", "condition"],
    "date": ["request date"],
    "req": ["request №", "request no", "request #"],
    "price": ["offered\nper unit", "offered per unit", "offered"],
    "invoice": ["invoice to customer"],
}


def _tuz_layouts(primary: dict[str, int]) -> list[dict[str, int]]:
    """Набор возможных раскладок колонок ТУЗ (данные бывают и классика, и со сдвигом Sales/Purch ID)."""
    classic_24 = {
        "date": 1, "req": 2, "client": 4, "pn": 6, "alt": 7, "desc": 8, "qty": 9,
        "cond": 18, "price": 24, "invoice": 27,
    }
    classic_27 = {
        "date": 1, "req": 2, "client": 4, "pn": 6, "alt": 7, "desc": 8, "qty": 9,
        "cond": 21, "price": 27, "invoice": 30,
    }
    shifted = {
        "date": 3, "req": 4, "client": 6, "pn": 8, "alt": 9, "desc": 10, "qty": 11,
        "cond": 21, "price": 27, "invoice": 30,
    }
    # Questions v2 style (P/N starts at col 8)
    qv2 = {
        "date": 3, "req": 4, "client": 6, "pn": 8, "alt": 9, "desc": 10, "qty": 11,
        "cond": 20, "price": 26, "invoice": 29,
    }
    layouts = []
    if primary:
        layouts.append(dict(primary))
    for lay in (classic_27, classic_24, shifted, qv2):
        if lay not in layouts:
            layouts.append(lay)
    return layouts


def _extract_tuz_row(r: tuple, layouts: list[dict[str, int]]) -> Optional[dict]:
    """Пробует раскладки; возвращает поля строки или None."""
    for mapping in layouts:
        def get(key, default=None, _m=mapping):
            idx = _m.get(key)
            if idx is None or idx >= len(r):
                return default
            return r[idx]

        pn_raw = get("pn")
        pn = norm_pn(pn_raw)
        if is_sample_pn(pn):
            continue
        if isinstance(pn_raw, str) and " " in pn_raw.strip() and len(pn_raw) > 40:
            continue

        client = norm_client(get("client"))
        if client.upper() in SAMPLE_CLIENT:
            continue
        # client не должен выглядеть как P/N при валидном PN
        if client and not re.search(r"[A-Za-zА-Яа-я]{2,}", client) and re.search(r"\d", client):
            # похоже на номер, а не клиента — пробуем другую раскладку
            continue

        alt = norm_pn(get("alt"))
        desc = str(get("desc") or "").strip()
        qty = parse_qty(get("qty"))
        # Только явная колонка Offered — без fallback на Cert/Remarks
        price = parse_money(get("price"))
        cond = str(get("cond") or "").strip()
        if cond.lower() in {"cond", "condition"}:
            cond = ""
        date = get("date")
        # date should be date-like; if not, try nearby
        if parse_date(date) is None:
            for idx in (1, 3, mapping.get("date")):
                if idx is not None and idx < len(r) and parse_date(r[idx]) is not None:
                    date = r[idx]
                    break
        req = str(get("req") or "").strip()
        invoice = str(get("invoice") or "").strip()
        return {
            "pn": pn,
            "alt": alt if alt and alt != pn else "",
            "client": client,
            "desc": desc,
            "qty": qty,
            "price": price,
            "cond": cond,
            "date": date,
            "req": req or invoice,
        }
    return None


def load_tuz(path: Path) -> list[Event]:
    events: list[Event] = []
    seen = set()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    # берём все листы с колонкой P/N, кроме служебных
    candidate_sheets = []
    for sheet in wb.sheetnames:
        if sheet.strip().lower() in TUZ_SKIP_SHEETS:
            continue
        candidate_sheets.append(sheet)

    for sheet in candidate_sheets:
        ws = wb[sheet]
        preview = []
        all_rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            all_rows.append(row)
            if i < 5:
                preview.append(row)
        hdr_idx, mapping = find_header_map(preview, TUZ_ALIASES, scan=5)
        if "pn" not in mapping:
            continue
        mapping.setdefault("client", 4)
        mapping.setdefault("alt", mapping["pn"] + 1)
        mapping.setdefault("desc", mapping["pn"] + 2)
        mapping.setdefault("qty", mapping["pn"] + 3)
        mapping.setdefault("price", 27)

        layouts = _tuz_layouts(mapping)
        start = 0
        for i, row in enumerate(all_rows[:6]):
            if row and any(str(c).strip().lower() in {"p/n", "part number"} for c in row[:15] if c):
                start = i + 1
        start = max(start, hdr_idx + 1)

        for r in all_rows[start:]:
            if not r:
                continue
            parsed = _extract_tuz_row(r, layouts)
            if not parsed:
                continue

            dedupe_key = (
                parsed["pn"],
                client_key(parsed["client"]),
                day_key(parsed["date"]),
                parsed["req"] or "",
                round(parsed["qty"], 4),
                round(parsed["price"] or 0, 2),
                parsed["desc"][:40],
                parsed["cond"],
            )
            if dedupe_key in seen:
                continue
            seen.add(dedupe_key)

            events.append(
                Event(
                    source="TUZ",
                    kind="request",
                    pn=parsed["pn"],
                    alt=parsed["alt"],
                    client=parsed["client"],
                    qty=parsed["qty"],
                    price=parsed["price"],
                    condition=parsed["cond"],
                    date=parsed["date"],
                    request_no=parsed["req"],
                    sheet=sheet,
                    description=parsed["desc"],
                    file=path.name,
                )
            )
    wb.close()
    return events


EXP_SHEETS = ["EXP NEW", "EXP UTAIR", "EXP GR#1", "EXP GR#2", "EXP GR#3", "EXP GR#4"]

EXP_ALIASES = {
    "date": ["date of rfq", "date of"],
    "client": ["client"],
    "alt": ["alt p/n", "alt pn"],
    "pn": ["p/n", "part number"],
    "desc": ["description"],
    "qty": ["qty"],
    "cond": ["cond"],
    "sell": ["sell price ea"],
}


def _merge_csv_headers(rows: list[list[str]], max_hdr: int = 3) -> list[str]:
    """Объединяет многострочный заголовок CSV (Sell Price во 2-й строке)."""
    if not rows:
        return []
    width = max(len(r) for r in rows[:max_hdr])
    merged = [""] * width
    for r in rows[:max_hdr]:
        for i in range(width):
            cell = (r[i] if i < len(r) else "") or ""
            cell = str(cell).replace("\n", " ").strip()
            if not cell:
                continue
            if not merged[i]:
                merged[i] = cell
            elif cell.lower() not in merged[i].lower() and "sell" in cell.lower():
                merged[i] = cell
    return merged


def load_exp(path: Path) -> list[Event]:
    events: list[Event] = []
    seen = set()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sheet in EXP_SHEETS:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        for i, r in enumerate(ws.iter_rows(values_only=True)):
            if i < 2:
                continue
            if not r or len(r) < 10:
                continue
            pn = norm_pn(r[6])
            if is_sample_pn(pn):
                continue
            client = norm_client(r[3])
            if client.upper() in SAMPLE_CLIENT:
                continue
            alt = norm_pn(r[5]) if len(r) > 5 else ""
            desc = str(r[7]).strip() if r[7] is not None else ""
            qty = parse_qty(r[8])
            price = parse_money(r[24]) if len(r) > 24 else None
            cond = str(r[14]).strip() if len(r) > 14 and r[14] is not None else ""
            date = r[1] if len(r) > 1 else None
            key = (pn, client_key(client), day_key(date), round(qty, 4), round(price or 0, 2), desc[:40])
            if key in seen:
                continue
            seen.add(key)
            events.append(
                Event(
                    source="EXP",
                    kind="request",
                    pn=pn,
                    alt=alt if alt and alt != pn else "",
                    client=client,
                    qty=qty,
                    price=price,
                    condition=cond,
                    date=date,
                    sheet=sheet,
                    description=desc,
                    file=path.name,
                )
            )
    wb.close()
    return events


def load_exp_csv(path: Path) -> list[Event]:
    """Архивы EXPENDABLES в CSV с разной раскладкой колонок."""
    import csv

    events: list[Event] = []
    seen = set()
    raw_rows: list[list[str]] = []
    for enc in ("utf-8-sig", "utf-8", "cp1251", "latin1"):
        try:
            with open(path, "r", encoding=enc, newline="") as f:
                raw_rows = list(csv.reader(f))
            break
        except Exception:
            continue
    if not raw_rows:
        return events

    # ищем строку(и) заголовка
    hdr_idx = 0
    mapping: dict[str, int] = {}
    for i in range(min(4, len(raw_rows))):
        merged = _merge_csv_headers(raw_rows[i : i + 2], max_hdr=2)
        mapping = _header_map_from_row(tuple(merged), EXP_ALIASES)
        if "pn" in mapping and "client" in mapping:
            hdr_idx = i
            # если sell во второй строке — учитываем
            if "sell" not in mapping and i + 1 < len(raw_rows):
                mapping = _header_map_from_row(tuple(_merge_csv_headers(raw_rows[i : i + 2])), EXP_ALIASES)
            break
    if "pn" not in mapping:
        print(f"  WARNING: no P/N in CSV {path.name}")
        return events

    # данные начинаются после 1–2 строк заголовка
    start = hdr_idx + 1
    if start < len(raw_rows) and any("sell" in clean_header_cell(c) for c in raw_rows[start]):
        start += 1

    def cell(r, key, default=None):
        idx = mapping.get(key)
        if idx is None or idx >= len(r):
            return default
        return r[idx]

    for r in raw_rows[start:]:
        if not r:
            continue
        pn = norm_pn(cell(r, "pn"))
        if is_sample_pn(pn):
            continue
        client = norm_client(cell(r, "client"))
        if client.upper() in SAMPLE_CLIENT:
            continue
        alt = norm_pn(cell(r, "alt"))
        desc = str(cell(r, "desc") or "").strip()
        qty = parse_qty(cell(r, "qty"))
        price = parse_money(cell(r, "sell"))
        cond = str(cell(r, "cond") or "").strip()
        date = cell(r, "date")
        key = (pn, client_key(client), day_key(date), round(qty, 4), round(price or 0, 2), desc[:40])
        if key in seen:
            continue
        seen.add(key)
        events.append(
            Event(
                source="EXP",
                kind="request",
                pn=pn,
                alt=alt if alt and alt != pn else "",
                client=client,
                qty=qty,
                price=price,
                condition=cond,
                date=date,
                sheet="csv",
                description=desc,
                file=path.name,
            )
        )
    return events


def load_ati(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if not r or r[0] is None:
            continue
        pn = norm_pn(r[0])
        if is_sample_pn(pn):
            continue
        rows.append(
            {
                "partno": str(r[0]).strip(),
                "pn": pn,
                "serialno": str(r[1]).strip() if r[1] is not None else "",
                "description": str(r[2]).strip() if r[2] is not None else "",
                "ata": str(r[3]).strip() if len(r) > 3 and r[3] is not None else "",
                "ac_typ": str(r[4]).strip() if len(r) > 4 and r[4] is not None else "",
                "condition": str(r[5]).strip() if len(r) > 5 and r[5] is not None else "",
                "qty": parse_qty(r[6]) if len(r) > 6 else 1,
            }
        )
    wb.close()
    return rows


def build_market_index(events: list[Event]) -> tuple[dict[str, MarketAgg], dict[str, set[str]], dict[str, set[str]]]:
    by_pn: dict[str, MarketAgg] = defaultdict(MarketAgg)
    soft_to_pns: dict[str, set[str]] = defaultdict(set)
    alt_to_pns: dict[str, set[str]] = defaultdict(set)

    for e in events:
        by_pn[e.pn].merge_event(e)
        soft_to_pns[soft_pn_key(e.pn)].add(e.pn)
        if e.alt and not is_sample_pn(e.alt):
            alt_to_pns[e.alt].add(e.pn)
            soft_to_pns[soft_pn_key(e.alt)].add(e.pn)
            # also record alt as its own node lightly? No — map only
    return by_pn, soft_to_pns, alt_to_pns


def resolve_market(
    pn: str,
    by_pn: dict[str, MarketAgg],
    soft_to_pns: dict[str, set[str]],
    alt_to_pns: dict[str, set[str]],
) -> MarketAgg:
    """Собирает агрегат рынка для P/N склада с учётом soft/alt/вариантов."""
    result = MarketAgg()
    matched_pns = set()
    near_notes: list[str] = []

    if pn in by_pn:
        matched_pns.add(pn)
        result.matched_via.add("точное P/N")

    # ведущие нули у чисто числовых P/N (00033038 ↔ 33038)
    if pn.isdigit():
        stripped = pn.lstrip("0") or "0"
        if stripped != pn and stripped in by_pn:
            matched_pns.add(stripped)
            result.matched_via.add("ведущие нули")
        soft_stripped = stripped
        for mp in soft_to_pns.get(soft_stripped, set()):
            if mp.isdigit():
                matched_pns.add(mp)
                result.matched_via.add("ведущие нули")

    # alt references: market events where our PN was listed as ALT of another, or vice versa
    if pn in alt_to_pns:
        matched_pns |= alt_to_pns[pn]
        result.matched_via.add("ALT P/N")

    # soft match only if exact missing or to enlarge
    soft = soft_pn_key(pn)
    candidates = soft_to_pns.get(soft, set())
    if candidates:
        if pn not in by_pn:
            close = {c for c in candidates if soft_pn_key(c) == soft}
            if len(close) == 1 or all(soft_pn_key(c) == soft for c in close):
                for c in close:
                    if c != pn:
                        result.matched_via.add("нормализация (дефисы/пробелы)")
                    matched_pns.add(c)
        else:
            for c in candidates:
                if soft_pn_key(c) == soft and c != pn:
                    if abs(len(c) - len(pn)) <= 2 or soft_pn_key(c) == soft:
                        matched_pns.add(c)
                        result.matched_via.add("нормализация (дефисы/пробелы)")

    # варианты с суффиксом/близостью — если точного спроса ещё нет
    if not matched_pns:
        for mp, reason in find_near_pn_matches(pn, by_pn, soft_to_pns):
            matched_pns.add(mp)
            near_notes.append(reason)
        if near_notes:
            result.matched_via.add("вариант P/N (суффикс/близость)")

    for mp in matched_pns:
        src = by_pn.get(mp)
        if not src:
            continue
        result.merge_agg(src)

    if not result.matched_via and matched_pns:
        result.matched_via.add("точное P/N")

    # сохраняем пояснение для отчёта
    result.near_match_notes = near_notes  # type: ignore[attr-defined]
    return result


def condition_section(cond: str) -> int:
    c = (cond or "").strip().upper()
    if c == "":
        return 1
    if c in {"US", "NA"}:
        return 2
    return 3


def condition_note(cond: str) -> str:
    c = (cond or "").strip().upper()
    mapping = {
        "": "состояние не указано — риск для продажи, нужна верификация",
        "US": "UNSERVICEABLE — спрос на P/N есть/нет отдельно; для продажи обычно нужен ремонт/обмен",
        "NA": "N/A / не определено — требует уточнения пригодности",
        "N": "NEW",
        "NE": "NEW / NE",
        "SV": "SERVICEABLE",
        "OH": "OVERHAULED",
        "R": "REPAIRED / BER?",
        "S": "SERVICEABLE?",
        "IT": "INSPECTED / TESTED?",
        "I": "INSPECTED?",
        "T": "TESTED?",
        "M": "MODIFIED / MIXED?",
    }
    return mapping.get(c, c)


GRADE_ORDER = {"A": 0, "B": 1, "C": 2, "D": 3}


def fmt_price(v: Optional[float]) -> str:
    if v is None:
        return "—"
    return f"${v:,.2f}"


def demand_summary(market: MarketAgg) -> str:
    """Краткое резюме заказов/запросов в одной ячейке."""
    parts = []
    if market.order_events:
        cli = ", ".join(market.sample_clients_order[:4])
        parts.append(
            f"Заказы ТАЗ: {market.order_events}"
            + (f" · клиенты: {cli}" if cli else "")
            + (f" ({len(market.order_clients)} уник.)" if len(market.order_clients) > 1 else "")
        )
    else:
        parts.append("Заказов ТАЗ: нет")

    if market.request_events:
        cli = ", ".join(market.sample_clients_request[:4])
        parts.append(
            f"Запросы ТУЗ+EXP: {market.request_events}"
            + (f" · клиенты: {cli}" if cli else "")
            + (f" ({len(market.request_clients)} уник.)" if len(market.request_clients) > 1 else "")
        )
    else:
        parts.append("Запросов: нет")

    return " | ".join(parts)


def price_data_flag(market: MarketAgg) -> str:
    """Флаг источника ценовых/рыночных данных."""
    if market.order_events > 0:
        return "есть заказы"
    if market.request_prices:
        return "есть предложения"
    if market.request_events > 0:
        return "есть запросы, но нет предложений"
    return "нет данных"


def build_row_from_stock(stock: dict, market: MarketAgg) -> dict:
    score, grade, rationale = liquidity_score(market)
    ref_price, price_conf, price_detail = compute_indicative_price(market)
    has_demand = (market.order_events + market.request_events) > 0
    has_sell_offered = ref_price is not None
    flag = price_data_flag(market)
    summary = demand_summary(market)

    conds = sorted(stock["conditions"])
    cond_display = ", ".join(c if c else "(пусто)" for c in conds) if conds else "(пусто)"
    notes_cond = [condition_note(c) for c in conds]
    extra = "; ".join(dict.fromkeys(notes_cond))

    rationale = rationale + f" На складе: {stock['qty']:g} шт. ({stock['lines']} строк АТИ)."
    rationale = rationale + f" Состояние склада: {extra}."
    rationale = rationale + " " + price_detail
    near_notes = getattr(market, "near_match_notes", None) or []
    if near_notes:
        rationale += " ПОМЕТКА: сопоставление по варианту P/N (" + "; ".join(near_notes[:3]) + ")."
    if has_demand and not has_sell_offered:
        rationale += (
            " ВНИМАНИЕ: спрос/заказы есть, но sell/offered цены нет — денежная оценка невозможна."
        )

    desc = stock["description"] or (
        next(iter(market.descriptions), "") if market.descriptions else ""
    )
    ac = ", ".join(sorted(x for x in stock["ac_typs"] if x)) or ""
    potential_rev = None
    if ref_price is not None and stock["qty"]:
        potential_rev = round(float(ref_price) * float(stock["qty"]), 2)

    # match_via уже включает тип сопоставления
    return {
        "liquidity_grade": grade,
        "liquidity_score": score,
        "partno": stock["partno"],
        "pn": stock["pn"],
        "description": desc,
        "ac_typ": ac,
        "condition": cond_display,
        "qty": stock["qty"],
        "lines": stock["lines"],
        "taz_orders": market.order_events,
        "requests": market.request_events,
        "demand_summary": summary,
        "price_ref_usd": ref_price,
        "potential_revenue_usd": potential_rev,
        "price_confidence": price_conf,
        "price_flag": flag,
        "has_sell_offered": has_sell_offered,
        "has_demand": has_demand,
        "match_via": ", ".join(sorted(market.matched_via)) if market.matched_via else "нет",
        "rationale": rationale,
        "section": stock["section"],
        "near_match": bool(near_notes),
    }


def aggregate_ati_stock(ati_rows: list[dict]) -> list[dict]:
    """Одна строка на P/N внутри раздела Condition; qty = сумма штук."""
    groups: dict[tuple, dict] = {}
    for r in ati_rows:
        sec = condition_section(r["condition"])
        key = (r["pn"], sec)
        if key not in groups:
            groups[key] = {
                "pn": r["pn"],
                "partno": r["partno"],
                "qty": 0.0,
                "lines": 0,
                "conditions": set(),
                "description": r.get("description") or "",
                "ac_typs": set(),
                "section": sec,
            }
        g = groups[key]
        q = r["qty"] if r.get("qty") and r["qty"] > 0 else 1.0
        g["qty"] += q
        g["lines"] += 1
        g["conditions"].add((r.get("condition") or "").strip().upper())
        if r.get("description") and not g["description"]:
            g["description"] = r["description"]
        if r.get("ac_typ"):
            g["ac_typs"].add(r["ac_typ"].strip())
    return list(groups.values())


HEADERS = [
    ("Ранг", 6),
    ("Тип ВС", 12),
    ("P/N", 18),
    ("Описание", 34),
    ("Состояние", 12),
    ("Кол-во на складе Utair", 12),
    ("Ориентировочная цена USD", 14),
    ("Потенц. выручка USD", 14),
    ("Уверенность", 12),
    ("Флаг цены", 28),
    ("Спрос (заказы/запросы)", 48),
    ("Обоснование оценки", 80),
]


def style_header(ws, fill_color: str):
    fill = PatternFill("solid", fgColor=fill_color)
    font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    thin = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    for col, (name, width) in enumerate(HEADERS, 1):
        cell = ws.cell(1, col, name)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = thin
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"


GRADE_FILL = {
    "A": PatternFill("solid", fgColor="1F7A4D"),
    "B": PatternFill("solid", fgColor="2F6FED"),
    "C": PatternFill("solid", fgColor="C47F00"),
    "D": PatternFill("solid", fgColor="8A8A8A"),
}
GRADE_FONT = Font(bold=True, color="FFFFFF", name="Calibri")
ZEBRA = PatternFill("solid", fgColor="F7F9FC")
THIN = Border(
    left=Side(style="thin", color="E5E5E5"),
    right=Side(style="thin", color="E5E5E5"),
    top=Side(style="thin", color="E5E5E5"),
    bottom=Side(style="thin", color="E5E5E5"),
)


PRICE_MISSING_FILL = PatternFill("solid", fgColor="F8D7DA")
PRICE_OK_FILL = PatternFill("solid", fgColor="D4EDDA")


def write_rows(ws, rows: list[dict], header_color: str):
    style_header(ws, header_color)
    conf_fill = {
        "высокая": PatternFill("solid", fgColor="1F7A4D"),
        "средняя": PatternFill("solid", fgColor="2F6FED"),
        "низкая": PatternFill("solid", fgColor="C47F00"),
        "н/п": PatternFill("solid", fgColor="8A8A8A"),
    }
    flag_fill = {
        "есть заказы": PatternFill("solid", fgColor="1F7A4D"),
        "есть предложения": PatternFill("solid", fgColor="2F6FED"),
        "есть запросы, но нет предложений": PatternFill("solid", fgColor="C47F00"),
        "нет данных": PatternFill("solid", fgColor="8A8A8A"),
    }
    for i, r in enumerate(rows, 1):
        values = [
            i,
            r["ac_typ"],
            r["partno"],
            r["description"],
            r["condition"],
            r["qty"],
            r["price_ref_usd"],
            r.get("potential_revenue_usd"),
            r["price_confidence"],
            r["price_flag"],
            r["demand_summary"],
            r["rationale"],
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(i + 1, col, val)
            cell.border = THIN
            cell.alignment = Alignment(vertical="center", wrap_text=(col in {4, 11, 12}))
            if i % 2 == 0:
                cell.fill = ZEBRA
            if col == 6:
                cell.font = Font(bold=True, name="Calibri", size=11)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if col in {7, 8} and isinstance(val, (int, float)):
                cell.number_format = '"$"#,##0.00'
            if col == 9:
                cell.fill = conf_fill.get(r["price_confidence"], conf_fill["н/п"])
                cell.font = Font(bold=True, color="FFFFFF", name="Calibri")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if col == 10:
                cell.fill = flag_fill.get(r["price_flag"], flag_fill["нет данных"])
                cell.font = Font(bold=True, color="FFFFFF", name="Calibri")
                cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[i + 1].height = 56
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(rows)+1}"


def write_summary(ws, ati_rows, scored, taz_n, tuz_n, exp_n, notes: list[str], source_lines: list[str] | None = None):
    ws["A1"] = "Оценка ликвидности склада АТИ (Utair)"
    ws["A1"].font = Font(bold=True, size=16, color="1A1A1A", name="Calibri")
    ws["A2"] = (
        f"Дата отчёта: {datetime.now():%Y-%m-%d %H:%M} | "
        "Источники: ТАЗ (несколько периодов), ТУЗ, EXPENDABLES/CSV, АТИ"
    )
    ws.merge_cells("A2:F2")

    with_rev = [r for r in scored if r.get("potential_revenue_usd") is not None]
    total_rev = sum(r["potential_revenue_usd"] for r in with_rev)
    matched = sum(1 for r in scored if r["liquidity_grade"] != "D")

    summary = [
        ("Строк в исходном АТИ", len(ati_rows)),
        ("Позиций в отчёте (P/N после агрегации)", len(scored)),
        ("Уникальных P/N на складе", len({r["pn"] for r in scored})),
        ("Суммарное кол-во на складе Utair, шт.", sum(r["qty"] for r in scored)),
        ("Позиций с найденным спросом (A/B/C)", matched),
        ("Позиций без спроса (D)", sum(1 for r in scored if r["liquidity_grade"] == "D")),
        ("Раздел 1 — Condition пусто", sum(1 for r in scored if r["section"] == 1)),
        ("Раздел 2 — Condition US/NA", sum(1 for r in scored if r["section"] == 2)),
        ("Раздел 3 — прочие Condition", sum(1 for r in scored if r["section"] == 3)),
        ("Ликвидность A (высокая)", sum(1 for r in scored if r["liquidity_grade"] == "A")),
        ("Ликвидность B (средняя)", sum(1 for r in scored if r["liquidity_grade"] == "B")),
        ("Ликвидность C (низкая)", sum(1 for r in scored if r["liquidity_grade"] == "C")),
        ("Ликвидность D (нет спроса)", sum(1 for r in scored if r["liquidity_grade"] == "D")),
        ("Цена: уверенность высокая", sum(1 for r in scored if r["price_confidence"] == "высокая")),
        ("Цена: уверенность средняя", sum(1 for r in scored if r["price_confidence"] == "средняя")),
        ("Цена: уверенность низкая", sum(1 for r in scored if r["price_confidence"] == "низкая")),
        ("Цена: н/п (нет sell/offered)", sum(1 for r in scored if r["price_confidence"] == "н/п")),
        ("Позиций с потенциальной выручкой", len(with_rev)),
        ("Суммарная потенц. выручка (qty×цена), USD", round(total_rev, 2)),
        ("Событий ТАЗ после дедупа (P/N+№заказа)", taz_n),
        ("Событий ТУЗ после дедупа (сырые строки)", tuz_n),
        ("Событий EXPENDABLES после дедупа", exp_n),
    ]
    ws["A4"] = "Сводка"
    ws["A4"].font = Font(bold=True, size=13)
    for i, (k, v) in enumerate(summary, 5):
        ws.cell(i, 1, k).font = Font(name="Calibri", size=11)
        cell = ws.cell(i, 2, v)
        cell.font = Font(bold=True, name="Calibri", size=11)
        if "выручка" in k.lower() and isinstance(v, (int, float)):
            cell.number_format = '"$"#,##0.00'

    row = 5 + len(summary) + 1
    ws.cell(row, 1, "Методика оценки").font = Font(bold=True, size=13)
    method = [
        "A — высокая: подтверждённые заказы и/или устойчивый спрос у нескольких клиентов.",
        "B — средняя: есть заказы или заметные повторные запросы.",
        "C — низкая: единичные/редкие запросы без сильной истории продаж.",
        "D — нет спроса: P/N не найден в ТАЗ/ТУЗ/EXPENDABLES (нормализация, soft-ключ, ALT).",
        "Заказы ТАЗ: 1 заказ = P/N + номер заказа (счёта); дубли между файлами периодов снимаются.",
        "Запросы ТУЗ+EXP: 1 запрос = P/N + клиент + календарный день; пересечения периодов дедуплицируются.",
        "«Кол-во на складе Utair» = сумма штук по P/N в разделе Condition.",
        "Потенц. выручка = ориентировочная цена × кол-во на складе Utair (только где есть sell/offered).",
        "Ориентир цены: свежая медиана ТАЗ (≤6 мес.) → любая ТАЗ → свежие Offered/Sell → любые запросы.",
        "В ориентир НЕ входят Закупка/Supplier/Root/Market Price EA.",
        "US/NA и пустое состояние выделены: спрос на P/N ≠ готовность партии к продаже.",
    ]
    for j, t in enumerate(method):
        ws.cell(row + 1 + j, 1, t)
        ws.merge_cells(start_row=row + 1 + j, start_column=1, end_row=row + 1 + j, end_column=6)

    row2 = row + 1 + len(method) + 1
    ws.cell(row2, 1, "Замечания по качеству данных (авто)").font = Font(bold=True, size=13)
    for i, t in enumerate(notes):
        ws.cell(row2 + 1 + i, 1, f"• {t}")
        ws.merge_cells(start_row=row2 + 1 + i, start_column=1, end_row=row2 + 1 + i, end_column=6)

    if source_lines:
        row3 = row2 + 1 + len(notes) + 1
        ws.cell(row3, 1, "Загруженные источники").font = Font(bold=True, size=13)
        for i, t in enumerate(source_lines):
            ws.cell(row3 + 1 + i, 1, f"• {t}")
            ws.merge_cells(start_row=row3 + 1 + i, start_column=1, end_row=row3 + 1 + i, end_column=6)

    ws.column_dimensions["A"].width = 62
    ws.column_dimensions["B"].width = 18


def build_seller_priority(scored: list[dict], limit: int = 60) -> list[dict]:
    """Позиции 1-й очереди для уточнения у продавца: ликвидные + высокая потенц. выручка.

    Включаем US/NA и пустое состояние — как раз там нужна верификация.
    """
    eligible = [
        r
        for r in scored
        if r["liquidity_grade"] in {"A", "B"}
        or (
            r["liquidity_grade"] == "C"
            and (r.get("potential_revenue_usd") or 0) >= 5000
            and (r["taz_orders"] + r["requests"]) >= 2
        )
    ]

    def prio_key(r):
        rev = r.get("potential_revenue_usd") or 0
        needs_info = 1 if r["section"] in {1, 2} else 0
        return (
            GRADE_ORDER[r["liquidity_grade"]],
            -needs_info,
            -rev,
            -r["liquidity_score"],
            -r["qty"],
            r["partno"],
        )

    return sorted(eligible, key=prio_key)[:limit]


def audit_unmatched(scored: list[dict], by_pn: dict, soft_to_pns: dict, alt_to_pns: dict) -> dict:
    """Глубокая проверка D-позиций: нет ли скрытых совпадений."""
    d_rows = [r for r in scored if r["liquidity_grade"] == "D"]
    soft_hits = 0
    alt_hits = 0
    contains_hits = []
    market_soft = set(soft_to_pns.keys())
    market_pns = set(by_pn.keys())

    for r in d_rows:
        pn = r["pn"]
        soft = soft_pn_key(pn)
        if soft in market_soft and soft_to_pns[soft]:
            soft_hits += 1
        if pn in alt_to_pns:
            alt_hits += 1
        # очень осторожный contains: только если складской PN длинный и целиком входит
        if len(soft) >= 8:
            for mp in market_pns:
                ms = soft_pn_key(mp)
                if soft != ms and (soft in ms or ms in soft) and abs(len(soft) - len(ms)) <= 3:
                    contains_hits.append((pn, mp))
                    break

    return {
        "d_count": len(d_rows),
        "soft_cluster_nonempty": soft_hits,
        "alt_hits": alt_hits,
        "near_contains": contains_hits[:20],
        "near_contains_n": len(contains_hits),
    }


def main():
    taz_files = [
        DATA / "TAZ_17.07.2026.xlsx",
        DATA / "TA3 2025.xlsx",
        DATA / "TA3-Архив 2024-01-26.xlsx",
    ]
    tuz_files = [
        DATA / "TUZ_17.07.2026.xlsx",
        DATA / "ТУЗ 2025 Jan - June.xlsx",
        DATA / "ТУЗ 6_19.xlsx",
    ]
    exp_xlsx = [DATA / "EXPENDABLES.xlsx"]
    exp_csv = sorted(DATA.glob("*.csv"))

    source_lines: list[str] = []
    taz: list[Event] = []
    print("Loading TAZ files...")
    for path in taz_files:
        if not path.exists():
            print(f"  MISSING {path.name}")
            continue
        part = load_taz(path)
        print(f"  {path.name}: {len(part)} orders")
        source_lines.append(f"ТАЗ {path.name}: {len(part)} строк заказов")
        taz.extend(part)

    tuz: list[Event] = []
    print("Loading TUZ files...")
    for path in tuz_files:
        if not path.exists():
            print(f"  MISSING {path.name}")
            continue
        part = load_tuz(path)
        print(f"  {path.name}: {len(part)} requests")
        source_lines.append(f"ТУЗ {path.name}: {len(part)} строк запросов")
        tuz.extend(part)

    exp: list[Event] = []
    print("Loading EXPENDABLES...")
    for path in exp_xlsx:
        if path.exists():
            part = load_exp(path)
            print(f"  {path.name}: {len(part)} requests")
            source_lines.append(f"EXP {path.name}: {len(part)} строк")
            exp.extend(part)
        else:
            print(f"  MISSING {path.name} (июльский xlsx — используем архивные CSV)")
            source_lines.append(f"ВНИМАНИЕ: {path.name} отсутствует — нет вкладки EXP 2026 xlsx")
    for path in exp_csv:
        part = load_exp_csv(path)
        print(f"  {path.name}: {len(part)} requests")
        source_lines.append(f"EXP CSV {path.name}: {len(part)} строк")
        exp.extend(part)

    print("Loading ATI...")
    ati = load_ati(DATA / "ATI.xlsx")
    print(f"  ATI rows: {len(ati)}")

    # глобальный дедуп событий между файлами до индекса
    def event_global_key(e: Event):
        if e.kind == "order":
            inv = normalize_invoice(e.request_no) or f"NO_INV|{client_key(e.client)}|{day_key(e.date)}|{round(e.qty or 0, 4)}"
            return ("O", e.pn, inv)
        return ("R", e.pn, client_key(e.client), day_key(e.date), round(e.qty or 0, 4), round(e.price or 0, 2), (e.description or "")[:40])

    all_raw = taz + tuz + exp
    seen_ev = set()
    all_events: list[Event] = []
    for e in all_raw:
        k = event_global_key(e)
        if k in seen_ev:
            continue
        seen_ev.add(k)
        all_events.append(e)
    print(f"  Events raw={len(all_raw)} after cross-file dedupe={len(all_events)}")

    by_pn, soft_to_pns, alt_to_pns = build_market_index(all_events)
    print(f"  Market unique PNs: {len(by_pn)}")

    ati_pn_counts = defaultdict(int)
    for r in ati:
        ati_pn_counts[r["pn"]] += 1
    dup_pn = sum(1 for v in ati_pn_counts.values() if v > 1)
    stock = aggregate_ati_stock(ati)

    scored = []
    for item in stock:
        market = resolve_market(item["pn"], by_pn, soft_to_pns, alt_to_pns)
        scored.append(build_row_from_stock(item, market))

    audit = audit_unmatched(scored, by_pn, soft_to_pns, alt_to_pns)
    print(
        "  Unmatched audit D=",
        audit["d_count"],
        "soft_nonempty=",
        audit["soft_cluster_nonempty"],
        "alt=",
        audit["alt_hits"],
        "near_contains=",
        audit["near_contains_n"],
    )
    if audit["near_contains"]:
        print("  near-contains samples:", audit["near_contains"][:8])

    near_variant_n = sum(1 for r in scored if "вариант P/N" in (r.get("match_via") or ""))
    notes = [
        f"В АТИ {dup_pn} P/N встречались более одного раза — в отчёте одна строка на P/N внутри раздела Condition.",
        f"Исходных строк АТИ: {len(ati)}; после агрегации позиций: {len(stock)}.",
        "Дедуп заказов: P/N + №счёта across всех файлов ТАЗ.",
        "Дедуп запросов: P/N + клиент + день (+ qty/price/desc для сырых строк) across ТУЗ/EXP.",
        "Сопоставление P/N: регистр, пробелы, тире; soft-ключ; ALT; ведущие нули; варианты с суффиксом.",
        f"Вариантное сопоставление (напр. 601R57502-1↔601R57502): {near_variant_n} позиций; "
        "в обосновании есть пометка «ПОМЕТКА: сопоставление по варианту P/N».",
        "Цена ориентир только из «Продажная, ед.» / «Offered per unit $» / «Sell Price EA».",
        "«Закупка на склад» не считается рыночным клиентом.",
        f"Позиций без спроса (D): {audit['d_count']} — после полного прохода по рынку (включая варианты).",
        "Offered в ТУЗ иногда содержит два числа в одной ячейке (две строки) — сейчас берём верхнее число.",
        "АТИ: исходный файл «АТИ для реализации» (serial-level).",
    ]
    exp_utair = list(DATA.glob("*1 DECEMBER 2025*Utair*.csv")) + list(DATA.glob("*DECEMBER 2025*Exp*.csv"))
    if exp_utair:
        notes.insert(0, f"Подключён EXPENDABLES Utair CSV: {exp_utair[0].name}.")
    elif not (DATA / "EXPENDABLES.xlsx").exists():
        notes.insert(0, "Нет EXPENDABLES.xlsx — учтены архивные CSV.")

    def sort_key(r):
        return (GRADE_ORDER[r["liquidity_grade"]], -r["liquidity_score"], -r["qty"], r["partno"])

    sec1 = sorted([r for r in scored if r["section"] == 1], key=sort_key)
    sec2 = sorted([r for r in scored if r["section"] == 2], key=sort_key)
    sec3 = sorted([r for r in scored if r["section"] == 3], key=sort_key)

    top_by_pn: dict[str, dict] = {}
    for r in scored:
        if r["liquidity_grade"] not in {"A", "B"}:
            continue
        pn = r["pn"]
        if pn not in top_by_pn:
            top_by_pn[pn] = dict(r)
        else:
            cur = top_by_pn[pn]
            cur["qty"] += r["qty"]
            # пересчитать выручку по суммарному qty
            if cur.get("price_ref_usd") is not None:
                cur["potential_revenue_usd"] = round(float(cur["price_ref_usd"]) * float(cur["qty"]), 2)
            cur["lines"] = cur.get("lines", 0) + r.get("lines", 0)
            conds = {x.strip() for x in cur["condition"].split(",") if x.strip()}
            conds |= {x.strip() for x in r["condition"].split(",") if x.strip()}
            cur["condition"] = ", ".join(sorted(conds))
            if GRADE_ORDER[r["liquidity_grade"]] < GRADE_ORDER[cur["liquidity_grade"]] or (
                r["liquidity_grade"] == cur["liquidity_grade"] and r["liquidity_score"] > cur["liquidity_score"]
            ):
                for k in (
                    "liquidity_grade",
                    "liquidity_score",
                    "taz_orders",
                    "requests",
                    "demand_summary",
                    "has_sell_offered",
                    "has_demand",
                    "price_ref_usd",
                    "price_confidence",
                    "price_flag",
                    "match_via",
                    "rationale",
                    "description",
                    "ac_typ",
                ):
                    cur[k] = r[k]
                if cur.get("price_ref_usd") is not None:
                    cur["potential_revenue_usd"] = round(float(cur["price_ref_usd"]) * float(cur["qty"]), 2)
                cur["rationale"] = (
                    f"{r['rationale']} (в ТОП qty суммирован по всем состояниям склада: {cur['qty']:g} шт.)"
                )
    top = sorted(top_by_pn.values(), key=sort_key)[:80]
    seller_prio = build_seller_priority(scored, limit=60)

    taz_n = sum(1 for e in all_events if e.source == "TAZ")
    tuz_n = sum(1 for e in all_events if e.source == "TUZ")
    exp_n = sum(1 for e in all_events if e.source == "EXP")

    print("Writing Excel...")
    wb = openpyxl.Workbook()
    ws0 = wb.active
    ws0.title = "0. Сводка"
    write_summary(ws0, ati, scored, taz_n, tuz_n, exp_n, notes, source_lines)

    ws_prio = wb.create_sheet("1 очередь у продавца")
    write_rows(ws_prio, seller_prio, "8B4513")
    ws_prio["A1"].value  # header already set
    # пояснение над таблицей нельзя без сдвига — добавим примечание в легенду

    ws_top = wb.create_sheet("ТОП ликвидных")
    write_rows(ws_top, top, "1F7A4D")

    ws1 = wb.create_sheet("1. Condition пусто")
    write_rows(ws1, sec1, "6B4C9A")

    ws2 = wb.create_sheet("2. Condition US_NA")
    write_rows(ws2, sec2, "B33B3B")

    ws3 = wb.create_sheet("3. Прочие Condition")
    write_rows(ws3, sec3, "2F5D9F")

    wsl = wb.create_sheet("Легенда")
    wsl["A1"] = "Цвета ликвидности"
    wsl["A1"].font = Font(bold=True, size=13)
    for i, (g, name) in enumerate([("A", "Высокая"), ("B", "Средняя"), ("C", "Низкая"), ("D", "Нет спроса")], 3):
        c = wsl.cell(i, 1, g)
        c.fill = GRADE_FILL[g]
        c.font = GRADE_FONT
        wsl.cell(i, 2, name)
    wsl["A8"] = "Уверенность в цене"
    wsl["A8"].font = Font(bold=True, size=13)
    for i, (name, color, desc) in enumerate([
        ("высокая", "1F7A4D", "есть заказ ТАЗ с продажной ценой за последние 6 мес."),
        ("средняя", "2F6FED", "есть ТАЗ (не свежий) или свежие Offered/Sell в ТУЗ/EXP"),
        ("низкая", "C47F00", "заказов ТАЗ нет, предложения ТУЗ/EXP давние"),
        ("н/п", "8A8A8A", "нет sell/offered цены — ориентир не рассчитан"),
    ], 9):
        c = wsl.cell(i, 1, name)
        c.fill = PatternFill("solid", fgColor=color)
        c.font = Font(bold=True, color="FFFFFF", name="Calibri")
        wsl.cell(i, 2, desc)
    wsl["A14"] = "Флаг цены"
    wsl["A14"].font = Font(bold=True, size=13)
    for i, (name, color, desc) in enumerate([
        ("есть заказы", "1F7A4D", "в ТАЗ есть заказы по P/N"),
        ("есть предложения", "2F6FED", "заказов нет, но в ТУЗ/EXP есть Offered/Sell Price"),
        ("есть запросы, но нет предложений", "C47F00", "запросы ТУЗ/EXP есть, но Offered/Sell ни разу не заполняли"),
        ("нет данных", "8A8A8A", "нет заказов и запросов по P/N"),
    ], 15):
        c = wsl.cell(i, 1, name)
        c.fill = PatternFill("solid", fgColor=color)
        c.font = Font(bold=True, color="FFFFFF", name="Calibri")
        wsl.cell(i, 2, desc)
    wsl["A20"] = "Ликвидность и балл скрыты: сортировка по ним сохранена в ранге. Детали — в обосновании."
    wsl["A21"] = "В ориентир НЕ входят: Закупка, Supplier/Root Price, Market Price EA."
    wsl["A22"] = "«Кол-во на складе Utair» — остаток на складе; потенц. выручка = цена × это кол-во."
    wsl["A23"] = (
        "Лист «1 очередь у продавца»: A/B и сильные C с выручкой; приоритет US/NA и пустому состоянию."
    )
    wsl.column_dimensions["A"].width = 36
    wsl.column_dimensions["B"].width = 70

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT_COPY.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    wb.save(OUT_COPY)
    print(f"Saved: {OUT}")
    print(f"Saved: {OUT_COPY}")
    total_rev = sum(r["potential_revenue_usd"] for r in scored if r.get("potential_revenue_usd") is not None)
    print(
        "Counts:",
        f"sec1={len(sec1)} sec2={len(sec2)} sec3={len(sec3)} top={len(top)} seller={len(seller_prio)}",
        f"A={sum(1 for r in scored if r['liquidity_grade']=='A')}",
        f"B={sum(1 for r in scored if r['liquidity_grade']=='B')}",
        f"C={sum(1 for r in scored if r['liquidity_grade']=='C')}",
        f"D={sum(1 for r in scored if r['liquidity_grade']=='D')}",
        f"rev_total=${total_rev:,.0f}",
        f"price_conf high/med/low/na="
        f"{sum(1 for r in scored if r['price_confidence']=='высокая')}/"
        f"{sum(1 for r in scored if r['price_confidence']=='средняя')}/"
        f"{sum(1 for r in scored if r['price_confidence']=='низкая')}/"
        f"{sum(1 for r in scored if r['price_confidence']=='н/п')}",
    )


if __name__ == "__main__":
    main()
